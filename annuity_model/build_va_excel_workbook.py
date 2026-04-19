"""VA Excel workbook builder."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import pricing_projection as sp
import va_projection as va
from excel_builder_helpers import (
    ExcelPythonSnapshot,
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from excel_workbook_validator import validate_workbook_or_raise
from recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_INDEX = "IndexScenario"
VA_PROJ_MAX_ROWS = 600

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_SP = 4
_IN_ROW_ME = 5
_IN_ROW_FREQ = 6
_IN_ROW_VAL = 7
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_HORIZON_YEARS = 10
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


@dataclass(frozen=True)
class VAExcelBuildSpec:
    contract: va.VAContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float
    index_s0: float
    index_levels_payment: np.ndarray


def va_excel_spec_from_launcher(
    *,
    contract: va.VAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
    **_kw: object,
) -> VAExcelBuildSpec:
    return VAExcelBuildSpec(
        contract=contract, yield_curve=yield_curve, mortality=mortality,
        horizon_age=int(horizon_age), spread=float(spread), valuation_year=int(valuation_year),
        expenses=expenses, yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
        index_s0=float(index_s0),
        index_levels_payment=np.asarray(index_levels_at_payment, dtype=float),
    )


def build_va_workbook_from_spec(
    spec: VAExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    res = va.price_va_single_premium(
        contract=spec.contract, yield_curve=spec.yield_curve, mortality=spec.mortality,
        horizon_age=spec.horizon_age, spread=spec.spread,
        valuation_year=(spec.valuation_year if not isinstance(spec.mortality, sp.MortalityTableQx) else None),
        expenses=spec.expenses, index_s0=float(spec.index_s0),
        index_levels_payment=np.asarray(spec.index_levels_payment, dtype=float),
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Single premium", float(spec.contract.single_premium)),
        ("M&E charge (annual)", float(spec.contract.me_charge_annual)),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Horizon years", int(spec.contract.horizon_years)),
        ("Monthly expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("GMDB basis", spec.contract.gmdb_basis),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(ws_in, InputsSheetSpec(title="VA Inputs (matches model launcher / Python)", rows=rows))
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{_in_addr('B', _IN_ROW_HORIZON_YEARS)}*{_in_addr('B', _IN_ROW_FREQ)})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame({
        "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
        "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
    })
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=VA_PROJ_MAX_ROWS, y_last_row=y_last_row)

    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx_arr = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx_arr[i]))

    ws_ix = wb.create_sheet(SHEET_INDEX)
    ws_ix["A1"] = "month"
    ws_ix["B1"] = "index_level"
    s0_py = float(spec.index_s0)
    levels_py = np.asarray(spec.index_levels_payment, dtype=float)
    L = np.zeros(levels_py.shape[0] + 1, dtype=float)
    L[0] = s0_py
    L[1:] = levels_py
    for j in range(L.shape[0]):
        ws_ix.cell(row=2 + j, column=1, value=int(j))
        ws_ix.cell(row=2 + j, column=2, value=float(L[j]))

    nm_ref = _in_addr("B", _IN_ROW_NMONTHS)
    last_cap_row = 3 + VA_PROJ_MAX_ROWS
    first = 4
    sp_ref = _in_addr("B", _IN_ROW_SP)
    issue_age_ref = _in_addr("B", _IN_ROW_ISSUE_AGE)
    freq_ref = _in_addr("B", _IN_ROW_FREQ)
    me_ref = _in_addr("B", _IN_ROW_ME)
    exp_m_ref = _in_addr("B", 11)
    exp_inf_ref = _in_addr("B", 12)
    horizon_months_ref = f"{_in_addr('B', _IN_ROW_HORIZON_YEARS)}*{_in_addr('B', _IN_ROW_FREQ)}"
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"
    idx_b = f"{SHEET_INDEX}!$B$2:$B$10000"
    idx_a = f"{SHEET_INDEX}!$A$2:$A$10000"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "VA liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={issue_age_ref}"

    hdr = (
        "Month",          # A
        "t_years",        # B
        "AttainedAge",    # C
        "SurvivalEnd",    # D
        "SurvivalStart",  # E
        "MonthDeathProb", # F
        "IndexLevel_m",   # G
        "AV_end",         # H
        "DB_end",         # I (max(AV, premium))
        "DeathCF",        # J
        "MaturityCF",     # K
        "ExpExpenseCF",   # L
        "ExpTotalCF",     # M  <-- layout column = M
        "",               # N (filler)
        "DiscountFactor", # O
        "PVBenefitCF",    # P
        "PVExpenseCF",    # Q
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    # Seed the AV recursion: row above the first data row needs AV[0] = SP.
    ws_pr.cell(row=3, column=8, value=float(spec.contract.single_premium))

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{freq_ref})')
        ws_pr.cell(row=r, column=3, value=f'=IF({a}="","",{issue_age_ref}+({a}-1)/{freq_ref})')
        ws_pr.cell(
            row=r, column=4,
            value=(
                f'=IF({a}="","",IFERROR(POWER(1-MIN(MAX(IFERROR(INDEX({SHEET_QX}!$B$2:$B$200,'
                f'MATCH(INT({issue_age_ref}+({a}-1)/12),{SHEET_QX}!$A$2:$A$200,0)),0),0),0.999),'
                f'{a}/12)*1,0))'
            ),
        )
        if r == first:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",1)')
        else:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",D{r - 1})')
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(
            row=r, column=7,
            value=f'=IF({a}="","",IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0)),""))',
        )
        # AV[t] = AV[t-1] * (S[t]/S[t-1]) * (1 - me_monthly)
        prev_idx = (
            f'IFERROR(INDEX({idx_b},MATCH({a}-1,{idx_a},0)),{idx_b}_____)'  # placeholder
        )
        # Use month-0 value (level at row 2 of IndexScenario) for first month.
        ws_pr.cell(
            row=r, column=8,
            value=(
                f'=IF({a}="",0,'
                f'IF({a}=1,{sp_ref}*(G{r}/INDEX({idx_b},1))*(1-{me_ref}/12),'
                f'H{r-1}*(G{r}/INDEX({idx_b},MATCH({a}-1,{idx_a},0)))*(1-{me_ref}/12)))'
            ),
        )
        ws_pr.cell(row=r, column=9, value=f'=IF({a}="",0,MAX(H{r},{sp_ref}))')
        ws_pr.cell(row=r, column=10, value=f'=IF({a}="",0,I{r}*F{r})')
        ws_pr.cell(row=r, column=11, value=f'=IF({a}="",0,IF({a}={horizon_months_ref},H{r}*D{r},0))')
        ws_pr.cell(row=r, column=12, value=f'=IF({a}="",0,{exp_m_ref}*POWER(1+(POWER(1+{exp_inf_ref},1/12)-1),{a}-1)*E{r})')
        ws_pr.cell(row=r, column=13, value=f'=IF({a}="",0,J{r}+K{r}+L{r})')
        ws_pr.cell(
            row=r, column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=16, value=f'=IF({a}="",0,(J{r}+K{r})*O{r})')
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,L{r}*O{r})')

    money_cols = (7, 8, 9, 10, 11, 12, 13, 16, 17)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 15):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV benefits (death+maturity)", f"=SUM(P{first}:P{last_cap_row})"),
                (5, "PV expenses", f"=SUM(Q{first}:Q{last_cap_row})"),
                (6, "Σ l_end · v (annuity-style factor)",
                 f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})"),
                (7, "PV total (ben+exp)", "=X4+X5"),
                (8, "Single premium (input)", f"={sp_ref}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    snap_py = ExcelPythonSnapshot(
        pv_benefit=pv_b, pv_monthly_expenses=pv_e, pv_monthly_total=pv_t,
        single_premium=float(res.single_premium), annuity_factor=float(res.annuity_factor),
    )
    va_rows = [
        ("PV benefits", pv_b, f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV expenses", pv_e, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV total (ben+exp)", pv_t, f"={LIABILITY_SHEET_NAME}!X7", "money"),
        ("Single premium", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Annuity factor", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb, snap_py, alm_layout=None, alm_snapshot=None, pricing_rows=va_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=(
            "VA: AV walks monthly with sub-account return * (1 - M&E_monthly). "
            "GMDB = max(AV, single_premium). Maturity cashflow at horizon month."
        ),
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
