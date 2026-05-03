"""FIA Excel workbook builder."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import fia_projection as fp
import pricing_projection as sp
from excel_builder_helpers import (
    ExcelPythonSnapshot,
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from excel_workbook_validator import validate_workbook_or_raise
from recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_INDEX = "IndexScenario"
FIA_PROJ_MAX_ROWS = 600

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_SP = 4
_IN_ROW_PART = 5
_IN_ROW_FREQ = 6
_IN_ROW_VAL = 7
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_CAP = 10
_IN_ROW_FLOOR = 11
_IN_ROW_HORIZON_YEARS = 12
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


@dataclass(frozen=True)
class FIAExcelBuildSpec:
    contract: fp.FIAContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float
    index_s0: float
    index_levels_payment: np.ndarray


def fia_excel_spec_from_launcher(
    *,
    contract: fp.FIAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
    **_kw: object,
) -> FIAExcelBuildSpec:
    return FIAExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
        index_s0=float(index_s0),
        index_levels_payment=np.asarray(index_levels_at_payment, dtype=float),
    )


def build_fia_workbook_from_spec(
    spec: FIAExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    res = fp.price_fia_single_premium(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=(
            spec.valuation_year if not isinstance(spec.mortality, sp.MortalityTableQx) else None
        ),
        expenses=spec.expenses,
        index_s0=float(spec.index_s0),
        index_levels_payment=np.asarray(spec.index_levels_payment, dtype=float),
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Single premium", float(spec.contract.single_premium)),
        ("Participation", float(spec.contract.participation)),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Cap (decimal / yr segment)", float(spec.contract.cap)),
        ("Floor (decimal / yr segment)", float(spec.contract.floor)),
        ("Horizon years", int(spec.contract.horizon_years)),
        ("Monthly expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(
        ws_in, InputsSheetSpec(title="FIA Inputs (matches model launcher / Python)", rows=rows)
    )
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{_in_addr('B', _IN_ROW_HORIZON_YEARS)}*{_in_addr('B', _IN_ROW_FREQ)})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=FIA_PROJ_MAX_ROWS, y_last_row=y_last_row)

    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx[i]))

    ws_ix = wb.create_sheet(SHEET_INDEX)
    ws_ix["A1"] = "month"
    ws_ix["B1"] = "index_level"
    s0_py = float(spec.index_s0)
    levels_py = np.asarray(spec.index_levels_payment, dtype=float)
    L = np.zeros(levels_py.shape[0] + 1, dtype=float)
    L[0] = s0_py
    L[1:] = levels_py
    for j in range(L.shape[0]):
        ws_ix.cell(row=2 + j, column=1, value=int(j))
        ws_ix.cell(row=2 + j, column=2, value=float(L[j]))

    nm_ref = _in_addr("B", _IN_ROW_NMONTHS)
    last_cap_row = 3 + FIA_PROJ_MAX_ROWS
    first = 4
    sp_ref = _in_addr("B", _IN_ROW_SP)
    issue_age_ref = _in_addr("B", _IN_ROW_ISSUE_AGE)
    freq_ref = _in_addr("B", _IN_ROW_FREQ)
    cap_ref = _in_addr("B", _IN_ROW_CAP)
    floor_ref = _in_addr("B", _IN_ROW_FLOOR)
    part_ref = _in_addr("B", _IN_ROW_PART)
    exp_m_ref = _in_addr("B", 13)
    exp_inf_ref = _in_addr("B", 14)
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"
    horizon_months_ref = f"{_in_addr('B', _IN_ROW_HORIZON_YEARS)}*{_in_addr('B', _IN_ROW_FREQ)}"
    idx_b = f"{SHEET_INDEX}!$B$2:$B$10000"
    idx_a = f"{SHEET_INDEX}!$A$2:$A$10000"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "FIA liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={issue_age_ref}"

    hdr = (
        "Month",  # A
        "t_years",  # B
        "AttainedAge",  # C
        "SurvivalEnd",  # D
        "SurvivalStart",  # E
        "MonthDeathProb",  # F
        "IndexLevel_m",  # G
        "RawSegReturn",  # H
        "CreditedSeg",  # I
        "AV_end",  # J
        "DeathCF",  # K
        "ExpExpenseCF",  # L
        "ExpTotalCF",  # M  <-- layout column = M
        "MaturityCF",  # N
        "DiscountFactor",  # O
        "PVBenefitCF",  # P
        "PVExpenseCF",  # Q
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    ws_pr.cell(row=3, column=10, value=0.0)
    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{freq_ref})')
        ws_pr.cell(row=r, column=3, value=f'=IF({a}="","",{issue_age_ref}+({a}-1)/{freq_ref})')
        qx_expr = (
            f"MIN(MAX(IFERROR(INDEX({SHEET_QX}!$B$2:$B$200,"
            f"MATCH(INT({issue_age_ref}+({a}-1)/12),{SHEET_QX}!$A$2:$A$200,0)),0),0),0.999)"
        )
        p_m_expr = f"EXP(-(-LN(1-{qx_expr}))/12)"
        if r == first:
            surv_end_formula = f'=IF({a}="","",{p_m_expr})'
        else:
            surv_end_formula = f'=IF({a}="","",D{r - 1}*{p_m_expr})'
        ws_pr.cell(row=r, column=4, value=surv_end_formula)
        if r == first:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",1)')
        else:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",D{r - 1})')
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(
            row=r, column=7, value=f'=IF({a}="","",IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0)),""))'
        )
        ws_pr.cell(
            row=r,
            column=8,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0))/INDEX({idx_b},MATCH({a}-12,{idx_a},0))-1,0),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=9,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"MAX({floor_ref},MIN({cap_ref},{part_ref}*H{r})),0))"
            ),
        )
        ws_pr.cell(
            row=r, column=10, value=f'=IF({a}="","",IF({a}=1,{sp_ref}*(1+I{r}),J{r - 1}*(1+I{r})))'
        )
        ws_pr.cell(row=r, column=11, value=f'=IF({a}="",0,J{r}*F{r})')
        ws_pr.cell(
            row=r,
            column=12,
            value=f'=IF({a}="",0,{exp_m_ref}*POWER(1+(POWER(1+{exp_inf_ref},1/12)-1),{a}-1)*E{r})',
        )
        # MaturityCF in column N (paid at the horizon month only)
        ws_pr.cell(
            row=r,
            column=14,
            value=f'=IF({a}="",0,IF({a}={horizon_months_ref},J{r}*D{r},0))',
        )
        # ExpTotalCF (column M) = DeathCF + Expense + Maturity
        ws_pr.cell(row=r, column=13, value=f'=IF({a}="",0,K{r}+L{r}+N{r})')
        ws_pr.cell(
            row=r,
            column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=16, value=f'=IF({a}="",0,(K{r}+N{r})*O{r})')
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,L{r}*O{r})')

    money_cols = (7, 8, 9, 10, 11, 12, 13, 14, 16, 17)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 15):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV benefits (claims+maturity)", f"=SUM(P{first}:P{last_cap_row})"),
                (5, "PV expenses", f"=SUM(Q{first}:Q{last_cap_row})"),
                (
                    6,
                    "Σ l_end · v (annuity-style factor)",
                    f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})",
                ),
                (7, "PV total (ben+exp)", "=X4+X5"),
                (8, "Single premium (input)", f"={sp_ref}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    snap_py = ExcelPythonSnapshot(
        pv_benefit=pv_b,
        pv_monthly_expenses=pv_e,
        pv_monthly_total=pv_t,
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    fia_rows: list[tuple[str, float, str, str]] = [
        ("PV benefits", pv_b, f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV expenses", pv_e, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV total (ben+exp)", pv_t, f"={LIABILITY_SHEET_NAME}!X7", "money"),
        ("Single premium", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Annuity factor", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=None,
        alm_snapshot=None,
        pricing_rows=fia_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=(
            "FIA: column B is Python at export; column C references Liabilities summary. "
            "IndexScenario holds L[month] for end-of-month index levels. "
            "AV_end accumulates by capped credit on segment anniversaries; floor 0 means "
            "AV cannot decrease."
        ),
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
