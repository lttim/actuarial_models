from __future__ import annotations

import math
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import pricing_projection as sp
import term_projection as tp

from build_pricing_excel_workbook import (
    ALM_ENGINE_STEP_MONTHS,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_PROJECTION_FIRST_DATA_ROW,
    ALMExcelSnapshot,
    ExcelPythonSnapshot,
    LIABILITY_SHEET_NAME,
    _write_alm_projection_sheet,
    _write_model_check_sheet,
    alm_excel_downsample_snapshot,
    alm_excel_truncate_snapshot,
    inject_alm_projection_formula_cached_values,
)
from recalc_excel_shared import RECALC_MONTHLY_CURVE_SHEET, write_monthly_curve_logdf, write_yield_curve_sheet

TERM_PROJ_MAX_ROWS = 600

SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_MTH_QX = "MortalMonthly"

# Inputs rows (align with SPIA: B6 = payments/year, B9 = spread for MonthlyCurve + ALM_Engine).
_IN_ROW_ISSUE_AGE = 3
_IN_ROW_DEATH_BEN = 5
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_PREMIUM = 10
_IN_ROW_TERM_Y = 11
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


def _n_months_cell() -> str:
    return _in_addr("B", _IN_ROW_NMONTHS)


def _fill_mortal_monthly_rpmp(
    ws_m,
    *,
    mort: sp.MortalityTableRP2014MP2016,
    issue_age: int,
    valuation_year: int,
    n_months: int,
) -> None:
    dt = 1.0 / 12.0
    for k in range(1, n_months + 1):
        r = 1 + k
        m_index = k - 1
        age_start = issue_age + m_index * dt
        age_int = int(math.floor(age_start))
        calendar_year_start = valuation_year + 1 + (m_index // 12)
        qxv = float(
            mort.qx_at_int_age_and_calendar_year(age_int=age_int, calendar_year=calendar_year_start)
        )
        ws_m.cell(row=r, column=1, value=f"=IF(ROW()-1>{_n_months_cell()},\"\",ROW()-1)")
        ws_m.cell(row=r, column=2, value=int(age_int))
        ws_m.cell(row=r, column=3, value=int(calendar_year_start))
        ws_m.cell(row=r, column=4, value=float(qxv))
    last_data = 1 + n_months
    last_cap = 1 + TERM_PROJ_MAX_ROWS
    for r in range(last_data + 1, last_cap + 1):
        ws_m.cell(row=r, column=1, value=f"=IF(ROW()-1>{_n_months_cell()},\"\",ROW()-1)")
        for c in (2, 3, 4):
            ws_m.cell(row=r, column=c, value="")


@dataclass(frozen=True)
class TermExcelBuildSpec:
    contract: tp.TermLifeContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str


def term_excel_spec_from_launcher(
    *,
    contract: tp.TermLifeContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
) -> TermExcelBuildSpec:
    return TermExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
    )


def _qx_lookup_expr(acell: str, mode: Literal["qx_table", "mortal_monthly"]) -> str:
    clamp_inner = "MIN(MAX({inner},0),0.999999)"
    if mode == "qx_table":
        inner = (
            f"INDEX({SHEET_QX}!$B$2:$B$50000,"
            f"MATCH(INT({_in_addr('B', _IN_ROW_ISSUE_AGE)}+({acell}-1)/12),"
            f"{SHEET_QX}!$A$2:$A$50000,0))"
        )
        return clamp_inner.format(inner=inner)
    inner = (
        f"INDEX({SHEET_MTH_QX}!$D$2:$D$50000,"
        f"MATCH({acell},{SHEET_MTH_QX}!$A$2:$A$50000,0))"
    )
    return clamp_inner.format(inner=inner)


def _survival_end_formula(
    r: int, acell: str, mode: Literal["qx_table", "mortal_monthly"], *, first_row: int
) -> str:
    qx_e = _qx_lookup_expr(acell, mode)
    p_m = f"EXP(-(-LN(1-{qx_e}))/12)"
    if r == first_row:
        return f"=IF({acell}=\"\",\"\",{p_m})"
    prev = f"D{r-1}"
    return f"=IF({acell}=\"\",\"\",{prev}*{p_m})"


def build_term_workbook_from_spec(
    spec: TermExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
    alm_snapshot: ALMExcelSnapshot | None = None,
    alm_assumptions: sp.ALMAssumptions | None = None,
) -> bytes:
    res = tp.price_term_life_level_monthly(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=spec.valuation_year
        if not isinstance(spec.mortality, sp.MortalityTableQx)
        else None,
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    ws_in["A1"] = "Term Life Inputs (matches model launcher / Python)"
    ws_in["A1"].font = Font(bold=True, size=12)
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Sex", spec.contract.sex),
        ("Death benefit", spec.contract.death_benefit),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Monthly premium", spec.contract.monthly_premium),
        ("Term years", spec.contract.term_years),
        ("Benefit timing", spec.contract.benefit_timing),
        ("Premium mode", spec.contract.premium_mode),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws_in[f"A{i}"] = k
        ws_in[f"B{i}"] = v
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', 6)},0)),"
        f"{_in_addr('B', _IN_ROW_TERM_Y)}*{_in_addr('B', 6)})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=TERM_PROJ_MAX_ROWS, y_last_row=y_last_row)

    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx[i]))
        mort_mode: Literal["qx_table", "mortal_monthly"] = "qx_table"
    else:
        mort_mode = "mortal_monthly"
        ws_m = wb.create_sheet(SHEET_MTH_QX)
        ws_m["A1"] = "month"
        ws_m["B1"] = "age_int"
        ws_m["C1"] = "calendar_year_start"
        ws_m["D1"] = "qx_annual"
        _fill_mortal_monthly_rpmp(
            ws_m,
            mort=spec.mortality,
            issue_age=spec.contract.issue_age,
            valuation_year=int(spec.valuation_year),
            n_months=min(int(res.months.shape[0]), TERM_PROJ_MAX_ROWS),
        )

    nm_ref = _n_months_cell()
    last_cap_row = 3 + TERM_PROJ_MAX_ROWS
    first = 4
    ben = _in_addr("B", _IN_ROW_DEATH_BEN)
    prem = _in_addr("B", _IN_ROW_PREMIUM)
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "Term life liability cashflows & pricing (formula-driven; not asset ALM)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={_in_addr('B', _IN_ROW_ISSUE_AGE)}"
    ws_pr["V2"] = "=X9"

    hdr = (
        "Month",
        "t_years",
        "AttainedAge",
        "SurvivalEnd",
        "SurvivalStart",
        "MonthDeathProb",
        "ExpClaims",
        "ExpPremiums",
        "ExpNetOutflow",
        "",
        "",
        "",
        "",
        "",
        "DiscountFactor",
        "ImpliedZeroFromDF",
        "ExpBenefitCF",
        "ExpExpenseCF",
        "ExpTotalCF",
        "PVBenefitCF",
        "PVExpenseCF",
        "PVNetOutflow",
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f"=IF(ROW()-3>{nm_ref},\"\",ROW()-3)")
        ws_pr.cell(row=r, column=2, value=f"=IF({a}=\"\",\"\",{a}/Inputs!$B$6)")
        ws_pr.cell(row=r, column=3, value=f"=IF({a}=\"\",\"\",Inputs!$B$3+({a}-1)/Inputs!$B$6)")
        ws_pr.cell(row=r, column=4, value=_survival_end_formula(r, a, mort_mode, first_row=first))
        if r == first:
            d_surv_start = f"=IF({a}=\"\",\"\",1)"
        else:
            d_surv_start = f"=IF({a}=\"\",\"\",D{r-1})"
        ws_pr.cell(row=r, column=5, value=d_surv_start)
        ws_pr.cell(row=r, column=6, value=f"=IF({a}=\"\",\"\",MAX(0,MIN(1,E{r}-D{r})))")
        ws_pr.cell(row=r, column=7, value=(
            f"=IF({a}=\"\",0,IF(MOD({a},12)=0,{ben}*SUM(OFFSET(F{r},-11,0,12,1)),0))"
        ))
        ws_pr.cell(row=r, column=8, value=f"=IF({a}=\"\",0,{prem}*E{r})")
        ws_pr.cell(row=r, column=9, value=f"=IF({a}=\"\",0,G{r}-H{r})")
        ws_pr.cell(row=r, column=15, value=f"=IF({a}=\"\",\"\",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),\"\"))")
        ws_pr.cell(row=r, column=16, value=f"=IF({a}=\"\",\"\",IF(B{r}>0,-LN(O{r})/B{r},\"\"))")
        ws_pr.cell(row=r, column=17, value=f"=IF({a}=\"\",0,G{r})")
        ws_pr.cell(row=r, column=18, value=f"=IF({a}=\"\",0,H{r})")
        ws_pr.cell(row=r, column=19, value=f"=IF({a}=\"\",0,I{r})")
        ws_pr.cell(row=r, column=20, value=f"=IF({a}=\"\",0,G{r}*O{r})")
        ws_pr.cell(row=r, column=21, value=f"=IF({a}=\"\",0,H{r}*O{r})")
        ws_pr.cell(row=r, column=22, value=f"=IF({a}=\"\",0,I{r}*O{r})")

    money_cols = (7, 8, 9, 17, 18, 19, 20, 21, 22)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 15, 16):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    ws_pr["W3"] = "Summary"
    ws_pr["W3"].font = Font(bold=True)
    ws_pr["W4"] = "PV claims"
    ws_pr["X4"] = f"=SUM(T{first}:T{last_cap_row})"
    ws_pr["W5"] = "PV premiums"
    ws_pr["X5"] = f"=SUM(U{first}:U{last_cap_row})"
    ws_pr["W6"] = "Σ l_start · v (annuity-style factor)"
    ws_pr["X6"] = f"=SUMPRODUCT(E{first}:E{last_cap_row},O{first}:O{last_cap_row})"
    ws_pr["W7"] = "PV net (claims − premiums)"
    ws_pr["X7"] = f"=X4-X5"
    ws_pr["W8"] = "Actuarial present value (pricing)"
    ws_pr["X8"] = "=X7"
    ws_pr["W9"] = "Reserve at t=0"
    ws_pr["X9"] = "=X7"

    alm_layout = None
    alm_snap_for_book = None
    if alm_snapshot is not None:
        if alm_assumptions is None:
            raise ValueError("alm_assumptions is required when alm_snapshot is provided.")
        alm_snap_for_book = alm_excel_downsample_snapshot(alm_snapshot, int(ALM_ENGINE_STEP_MONTHS))
        alm_snap_for_book = alm_excel_truncate_snapshot(alm_snap_for_book, ALM_EXCEL_PATH_MONTH_CAP)
        alm_layout = _write_alm_projection_sheet(
            wb,
            alm_snap_for_book,
            alm_assumptions,
            n_months=int(res.months.shape[0]),
            y_last_row=int(y_last_row),
            engine_step_months=int(ALM_ENGINE_STEP_MONTHS),
        )

    snap_py = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    prem_display = float(-res.pv_monthly_expenses)
    term_rows: list[tuple[str, float, str, str]] = [
        ("PV claims", float(res.pv_benefit), f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV premiums", prem_display, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV net (claims − premiums)", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X7", "money"),
        ("Actuarial present value (pricing)", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Σ survival start · discount (annuity-style factor)", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    _write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=alm_layout,
        alm_snapshot=alm_snap_for_book,
        pricing_rows=term_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME}; optional ALM_Projection)",
        subtitle=(
            "Column B is the Python snapshot at export. Column C aggregates "
            f"{LIABILITY_SHEET_NAME} summary formulas; column D should be ~0 after a full recalc. "
            "Edit Inputs, YieldCurve, MonthlyCurve, and mortality tabs as for SPIA exports. "
            f"ALM sheets mirror the SPIA workbook and link {LIABILITY_SHEET_NAME} column S (ExpTotalCF) and O (discount)."
        ),
    )

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if alm_snap_for_book is not None:
        data = inject_alm_projection_formula_cached_values(
            data,
            first_data_row=int(ALM_PROJECTION_FIRST_DATA_ROW),
            snap=alm_snap_for_book,
        )

    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
