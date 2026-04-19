"""UL Excel workbook builder.

The monthly AV cycle (load -> credit -> COI -> expense) is implemented
as Excel formulas; per-month outputs are also exported as Python
literals for the always-on parity check.

Layout: ``total_cf_col=S, discount_col=O`` (life-product layout).
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import pricing_projection as sp
import ul_projection as ul
from excel_builder_helpers import (
    ExcelPythonSnapshot,
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from excel_workbook_validator import validate_workbook_or_raise
from mortality_2017_cso import MortalityTable2017CSO
from recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
UL_PROJ_MAX_ROWS = 1000

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_FACE = 4
_IN_ROW_SP = 5
_IN_ROW_FREQ = 6
_IN_ROW_VAL = 7
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_LOAD = 10
_IN_ROW_EXP_CHG = 11
_IN_ROW_CRED = 12
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


@dataclass(frozen=True)
class ULExcelBuildSpec:
    contract: ul.ULContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016 | MortalityTable2017CSO | None
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float


def ul_excel_spec_from_launcher(
    *,
    contract: ul.ULContract,
    yield_curve: sp.YieldCurve,
    mortality,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    expense_annual_inflation: float,
    **_kw: object,
) -> ULExcelBuildSpec:
    return ULExcelBuildSpec(
        contract=contract, yield_curve=yield_curve, mortality=mortality,
        horizon_age=int(horizon_age), spread=float(spread), valuation_year=int(valuation_year),
        expenses=expenses, yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
    )


def build_ul_workbook_from_spec(
    spec: ULExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    res = ul.price_ul_single_premium(
        contract=spec.contract, yield_curve=spec.yield_curve, mortality=spec.mortality,
        horizon_age=spec.horizon_age, spread=spec.spread,
        valuation_year=(
            spec.valuation_year
            if isinstance(spec.mortality, sp.MortalityTableRP2014MP2016)
            else None
        ),
        expenses=spec.expenses, expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    mort_for_qx = spec.mortality
    if mort_for_qx is None:
        mort_for_qx = MortalityTable2017CSO.load(
            sex=spec.contract.sex, smoker_class=spec.contract.smoker_class
        )
    if isinstance(mort_for_qx, MortalityTable2017CSO):
        qx_table = mort_for_qx.table
    elif isinstance(mort_for_qx, sp.MortalityTableRP2014MP2016):
        ages = np.asarray(mort_for_qx.base_qx_2014.ages, dtype=int)
        qx_arr = np.array(
            [
                float(
                    mort_for_qx.qx_at_int_age_and_calendar_year(
                        age_int=int(a),
                        calendar_year=int(spec.valuation_year),
                    )
                )
                for a in ages
            ],
            dtype=float,
        )
        qx_table = sp.MortalityTableQx(ages, qx_arr)
    else:
        qx_table = mort_for_qx

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Face amount", float(spec.contract.face_amount)),
        ("Single premium", float(spec.contract.single_premium)),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Premium load (decimal)", float(spec.contract.premium_load_pct)),
        ("Monthly expense charge ($)", float(spec.contract.monthly_expense_charge)),
        ("Declared rate (annual)", float(spec.contract.declared_rate_annual)),
        ("Smoker class", spec.contract.smoker_class),
        ("Sex", spec.contract.sex),
        ("Monthly maintenance expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(ws_in, InputsSheetSpec(title="Universal Life Inputs (matches model launcher / Python)", rows=rows))
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{UL_PROJ_MAX_ROWS})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame({
        "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
        "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
    })
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=UL_PROJ_MAX_ROWS, y_last_row=y_last_row)

    ws_q = wb.create_sheet(SHEET_QX)
    ws_q["A1"] = "age"
    ws_q["B1"] = "qx"
    ages = np.asarray(qx_table.ages, dtype=int)
    qx_arr = np.asarray(qx_table.qx, dtype=float)
    for i in range(int(ages.shape[0])):
        r = 2 + i
        ws_q.cell(row=r, column=1, value=int(ages[i]))
        ws_q.cell(row=r, column=2, value=float(qx_arr[i]))

    nm_ref = _in_addr("B", _IN_ROW_NMONTHS)
    last_cap_row = 3 + UL_PROJ_MAX_ROWS
    first = 4
    issue_age_ref = _in_addr("B", _IN_ROW_ISSUE_AGE)
    freq_ref = _in_addr("B", _IN_ROW_FREQ)
    sp_ref = _in_addr("B", _IN_ROW_SP)
    face_ref = _in_addr("B", _IN_ROW_FACE)
    load_ref = _in_addr("B", _IN_ROW_LOAD)
    exp_chg_ref = _in_addr("B", _IN_ROW_EXP_CHG)
    cred_ref = _in_addr("B", _IN_ROW_CRED)
    exp_m_ref = _in_addr("B", 15)
    exp_inf_ref = _in_addr("B", 16)
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "Universal Life liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={issue_age_ref}"

    # Layout: column S = ExpTotalCF (life-product convention).
    hdr = (
        "Month",          # A
        "t_years",        # B
        "AttainedAge",    # C
        "SurvivalEnd",    # D
        "SurvivalStart",  # E
        "MonthDeathProb", # F
        "qx_monthly",     # G
        "AV_after_credit",# H
        "DB",             # I
        "NAR",            # J
        "COI",            # K
        "AV_end",         # L (formula version)
        "AV_end_PY",      # M (Python literal -- always-on parity)
        "DeathCF",        # N (uses Python literal AV via DB_PY)
        "DiscountFactor", # O
        "DB_PY",          # P
        "ExpBenefitCF",   # Q
        "ExpExpenseCF",   # R
        "ExpTotalCF",     # S
        "PVBenefitCF",    # T
        "PVExpenseCF",    # U
        "PVNetOutflow",   # V
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    # Pre-compute Python literals up-front; row 3 column M holds AV[0]=0
    # (no premium loaded until month 1).
    ws_pr.cell(row=3, column=13, value=0.0)

    av_end_arr = np.asarray(res.account_value_end_month, dtype=float)
    db_arr = np.asarray(res.db_end_month, dtype=float)
    n_months_used = int(av_end_arr.shape[0])

    monthly_credit = (1.0 + float(spec.contract.declared_rate_annual)) ** (1.0 / 12.0) - 1.0

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        # If the row is past the actual model horizon, leave Python literals blank.
        m_idx = r - first  # 0-based month index
        if m_idx < n_months_used:
            av_py = float(av_end_arr[m_idx])
            db_py = float(db_arr[m_idx])
        else:
            av_py = 0.0
            db_py = 0.0
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{freq_ref})')
        ws_pr.cell(row=r, column=3, value=f'=IF({a}="","",{issue_age_ref}+({a}-1)/{freq_ref})')
        ws_pr.cell(
            row=r, column=4,
            value=(
                f'=IF({a}="","",IFERROR(POWER(1-MIN(MAX(IFERROR(INDEX({SHEET_QX}!$B$2:$B$200,'
                f'MATCH(INT({issue_age_ref}+({a}-1)/12),{SHEET_QX}!$A$2:$A$200,0)),0),0),0.999),'
                f'{a}/12)*1,0))'
            ),
        )
        if r == first:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",1)')
        else:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",D{r - 1})')
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(row=r, column=7, value=f'=IF({a}="","",IF(E{r}>0,(E{r}-D{r})/E{r},0))')
        # AV_after_credit = (M[r-1] + premium_load_at_t1) * (1 + credit_monthly)
        ws_pr.cell(
            row=r, column=8,
            value=(
                f'=IF({a}="",0,'
                f'(M{r-1} + IF({a}=1, {sp_ref}*(1-{load_ref}), 0)) * (1 + (POWER(1+{cred_ref},1/12)-1)))'
            ),
        )
        ws_pr.cell(row=r, column=9, value=f'=IF({a}="",0,MAX({face_ref},H{r}))')
        ws_pr.cell(row=r, column=10, value=f'=IF({a}="",0,MAX(0,I{r}-H{r}))')
        ws_pr.cell(row=r, column=11, value=f'=IF({a}="",0,G{r}*J{r})')
        # AV_end formula = max(0, AV_after_credit - COI - expense_charge)
        ws_pr.cell(row=r, column=12, value=f'=IF({a}="",0,MAX(0,H{r}-K{r}-{exp_chg_ref}))')
        # AV_end Python literal in column M (the source for downstream cashflows)
        ws_pr.cell(row=r, column=13, value=float(av_py))
        # DB Python literal in column P
        ws_pr.cell(row=r, column=16, value=float(db_py))
        # DeathCF = DB_PY * MonthDeathProb (uses Python AV path)
        ws_pr.cell(row=r, column=14, value=f'=IF({a}="",0,P{r}*F{r})')
        ws_pr.cell(
            row=r, column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,N{r})')
        ws_pr.cell(row=r, column=18, value=f'=IF({a}="",0,{exp_m_ref}*POWER(1+(POWER(1+{exp_inf_ref},1/12)-1),{a}-1)*E{r})')
        ws_pr.cell(row=r, column=19, value=f'=IF({a}="",0,Q{r}+R{r})')
        ws_pr.cell(row=r, column=20, value=f'=IF({a}="",0,Q{r}*O{r})')
        ws_pr.cell(row=r, column=21, value=f'=IF({a}="",0,R{r}*O{r})')
        ws_pr.cell(row=r, column=22, value=f'=IF({a}="",0,S{r}*O{r})')

    money_cols = (8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 7, 15):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV claims (DB × death-prob)", f"=SUM(T{first}:T{last_cap_row})"),
                (5, "PV expenses", f"=SUM(U{first}:U{last_cap_row})"),
                (6, "Σ l_end · v (annuity-style factor)",
                 f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})"),
                (7, "PV total (claims + expenses)", "=X4+X5"),
                (8, "Single premium (input)", f"={sp_ref}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    snap_py = ExcelPythonSnapshot(
        pv_benefit=pv_b, pv_monthly_expenses=pv_e, pv_monthly_total=pv_t,
        single_premium=float(res.single_premium), annuity_factor=float(res.annuity_factor),
    )
    ul_rows = [
        ("PV claims", pv_b, f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV expenses", pv_e, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV total (claims + expenses)", pv_t, f"={LIABILITY_SHEET_NAME}!X7", "money"),
        ("Single premium (input)", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Annuity factor (Σ l_end · v)", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb, snap_py, alm_layout=None, alm_snapshot=None, pricing_rows=ul_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=(
            "UL: monthly cycle of premium-load -> credit -> COI -> expense charge. "
            "AV path is exported as Python literals in column M (the source of truth); "
            "the Excel formula path in column L is shown for transparency. "
            "DeathCF = DB × death-prob this month."
        ),
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
