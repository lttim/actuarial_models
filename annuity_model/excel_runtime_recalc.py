"""LibreOffice-headless runtime Excel recalc helper (P0 hardening, 2026-04).

The static parity stack (``tests/parity/excel_formula_sim.py`` plus the
AST-walking ``excel_workbook_validator.py``) catches Python<->Excel logic
drift and structural breaks. The remaining gap is the case where an emitted
formula string differs from what the simulation expects AND Excel itself
would compute a different answer. To close that gap we need an actual
spreadsheet engine to recalculate the workbook bytes we ship.

Pure-Python alternatives (``xlcalculator``, ``formulas``, ``pycel``) all
have caveats:

* ``xlcalculator==0.5.0`` transitively pins ``yearfrac<2`` which is
  incompatible with our ``numpy==2.4.4`` lock (parked since 2026-04, see
  ``docs/runbooks/runtime_excel_recalc_gate.md``).
* ``formulas==1.3.4`` installs cleanly but takes >3 minutes to load and
  recalculate the SPIA workbook because it builds the full dependency graph
  for thousands of cells. Unusable in CI.
* ``pycel`` has narrower function coverage and a similar perf profile.

LibreOffice headless is the pragmatic answer: it's the same engine Excel
users have been opening these workbooks in for years, install size on Linux
is ~250 MB via apt, and recalculating a SPIA workbook takes <5 seconds.

API
---
``ensure_libreoffice_available()`` returns the resolved soffice executable
or raises ``LibreOfficeNotAvailableError`` with a clear install hint. Tests use
``libreoffice_available()`` to decide whether to skip.

``recalc_workbook(blob, *, timeout=60.0)`` writes the bytes to a temp
directory, invokes soffice with ``--headless --calc --convert-to xlsx``,
and returns the recalculated workbook bytes. Cached cell values are then
readable via ``openpyxl.load_workbook(..., data_only=True)``.

This module imports only from the standard library + openpyxl (a runtime
dep) so it is safe to call from the parity gate, ``parity_trace.py``, and
ad-hoc debugging scripts.
"""

from __future__ import annotations

import io
import os
import platform
import shutil
import subprocess
import tempfile
import time
from contextlib import contextmanager
from collections.abc import Iterable
from functools import lru_cache
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook, load_workbook

__all__ = [
    "LIBREOFFICE_INSTALL_HINT",
    "LibreOfficeNotAvailable",
    "LibreOfficeNotAvailableError",
    "MACOS_LIBREOFFICE_RECALC_ENV",
    "RecalcTimeout",
    "RecalcTimeoutError",
    "ensure_libreoffice_available",
    "libreoffice_available",
    "libreoffice_recalc_disabled_reason",
    "libreoffice_runtime_recalc_available",
    "read_recalculated_cells",
    "recalc_workbook",
    "resolve_soffice",
]


LIBREOFFICE_INSTALL_HINT = (
    "LibreOffice (`soffice`) is required to recalculate workbook bytes. "
    "Install it with:\n"
    "  - Linux (Debian/Ubuntu):  sudo apt-get install -y libreoffice-calc\n"
    "  - macOS:                  brew install --cask libreoffice  "
    "(or download from https://www.libreoffice.org/download/)\n"
    "  - Windows:                winget install TheDocumentFoundation.LibreOffice\n"
    "After installing, ensure `soffice` is on PATH (or set "
    "$LIBREOFFICE_SOFFICE to its absolute path)."
)

MACOS_LIBREOFFICE_RECALC_ENV = "ANNUITY_MODEL_ENABLE_MACOS_LIBREOFFICE_RECALC"
"""Opt-in variable for real LibreOffice runtime recalc on macOS.

macOS LibreOffice may surface desktop crash/reopen dialogs when automated
headless Calc quits unexpectedly. The parity suite therefore skips runtime
LibreOffice recalc on macOS by default. Linux CI still runs it when soffice is
installed, and macOS users can opt in explicitly by setting this variable to a
truthy value.
"""


class LibreOfficeNotAvailableError(RuntimeError):
    """Raised when LibreOffice is needed but not installed / not on PATH."""


class RecalcTimeoutError(RuntimeError):
    """Raised when soffice does not finish recalculation within the deadline."""


LibreOfficeNotAvailable = LibreOfficeNotAvailableError
RecalcTimeout = RecalcTimeoutError
LOCK_PATH = Path(tempfile.gettempdir()) / "annuity_model_libreoffice_recalc.lock"
"""Host-wide lock used to serialize LibreOffice recalc across pytest agents.

LibreOffice headless is reliable when one Calc process owns a user profile.
On macOS it can abort or surface desktop crash dialogs when multiple headless
``soffice`` processes launch concurrently, even with different ``HOME`` values.
The lock is intentionally outside the repo so cloned worktrees and subagents on
the same host still serialize through the same gate.
"""


def _candidate_paths() -> Iterable[str]:
    import os

    env = os.environ.get("LIBREOFFICE_SOFFICE")
    if env:
        yield env
    yield "soffice"
    yield "libreoffice"
    # Common macOS install location (Homebrew cask + manual download).
    yield "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    # Common Linux install locations (apt, snap, flatpak).
    yield "/usr/bin/soffice"
    yield "/usr/lib/libreoffice/program/soffice"
    yield "/snap/libreoffice/current/usr/bin/soffice"


def resolve_soffice() -> str | None:
    """Return the first runnable soffice path, or None if none found."""
    for cand in _candidate_paths():
        if not cand:
            continue
        # Absolute paths: check existence directly.
        if cand.startswith("/"):
            if Path(cand).is_file():
                return cand
            continue
        # Bare names: check PATH.
        which = shutil.which(cand)
        if which is not None:
            return which
    return None


def _running_in_codex_seatbelt_sandbox() -> bool:
    """True when LibreOffice conversion is known to abort with a macOS dialog."""

    return bool(os.environ.get("CODEX_SANDBOX"))


@lru_cache(maxsize=8)
def _soffice_can_convert(soffice: str) -> bool:
    """Return True only if this process can actually run headless conversion.

    On macOS inside some sandboxed runners, ``soffice --headless --version``
    succeeds but any conversion aborts with signal 6. The runtime recalc tests
    care about the latter capability, so availability must probe conversion
    rather than binary presence.
    """
    if _running_in_codex_seatbelt_sandbox():
        return False
    try:
        with tempfile.TemporaryDirectory(prefix="annuity_lo_probe_") as tmpdir:
            tmp = Path(tmpdir)
            in_path = tmp / "probe.xlsx"
            out_dir = tmp / "out"
            out_dir.mkdir()
            wb = Workbook()
            ws = wb.active
            ws["A1"] = 1
            ws["A2"] = "=A1+1"
            wb.save(in_path)
            proc = subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--calc",
                    "--norestore",
                    "--nologo",
                    "--nofirststartwizard",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(out_dir),
                    str(in_path),
                ],
                capture_output=True,
                timeout=20.0,
                check=False,
                env={
                    **os.environ,
                    "HOME": str(tmp),
                    "TMPDIR": str(tmp),
                },
            )
            return proc.returncode == 0 and any(out_dir.glob("*.xlsx"))
    except (OSError, subprocess.SubprocessError):
        return False


def libreoffice_available() -> bool:
    if libreoffice_recalc_disabled_reason() is not None:
        return False
    soffice = resolve_soffice()
    return soffice is not None and _soffice_can_convert(soffice)


def _truthy_env(name: str) -> bool:
    return str(os.environ.get(name, "")).strip().lower() in {"1", "true", "yes", "on"}


def libreoffice_recalc_disabled_reason() -> str | None:
    if platform.system() == "Darwin" and not _truthy_env(MACOS_LIBREOFFICE_RECALC_ENV):
        return (
            "LibreOffice runtime recalc is disabled by default on macOS because "
            "headless soffice can trigger desktop crash/reopen dialogs. Set "
            f"{MACOS_LIBREOFFICE_RECALC_ENV}=1 to opt in for a single controlled run. "
            "Static workbook validation and Python-vs-Excel snapshot parity still run."
        )
    return None


def libreoffice_runtime_recalc_available() -> bool:
    return libreoffice_recalc_disabled_reason() is None and libreoffice_available()


def ensure_libreoffice_available() -> str:
    disabled = libreoffice_recalc_disabled_reason()
    if disabled is not None:
        raise LibreOfficeNotAvailable(disabled)
    soffice = resolve_soffice()
    if soffice is None:
        raise LibreOfficeNotAvailable(LIBREOFFICE_INSTALL_HINT)
    if _running_in_codex_seatbelt_sandbox():
        raise LibreOfficeNotAvailable(
            "LibreOffice (`soffice`) is installed, but this Codex seatbelt "
            "sandbox cannot run headless workbook conversion without macOS "
            "showing a crash dialog. Re-run the runtime recalc gate with "
            "sandbox escalation or from a normal Terminal."
        )
    if not _soffice_can_convert(soffice):
        raise LibreOfficeNotAvailable(
            "LibreOffice (`soffice`) is installed but cannot perform a "
            "headless workbook conversion in the current process sandbox. "
            "Run the recalc gate outside the sandbox, or fix local app "
            "permissions so `soffice --headless --convert-to xlsx` succeeds."
        )
    return soffice


@contextmanager
def _libreoffice_recalc_lock(*, timeout: float) -> Iterator[None]:
    """Cross-process advisory lock around the fragile LibreOffice subprocess."""
    deadline = time.monotonic() + max(float(timeout), 1.0)
    LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LOCK_PATH.open("a+b") as fh:
        try:
            import fcntl
        except ImportError:  # pragma: no cover - Windows fallback is exercised manually.
            yield
            return
        while True:
            try:
                fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                break
            except BlockingIOError:
                if time.monotonic() >= deadline:
                    raise RecalcTimeout(
                        f"Timed out waiting for LibreOffice recalc lock at {LOCK_PATH}. "
                        "Another pytest/subagent process is still recalculating a workbook."
                    )
                time.sleep(0.25)
        try:
            yield
        finally:
            fcntl.flock(fh.fileno(), fcntl.LOCK_UN)


def _soffice_command(
    *,
    soffice: str,
    profile_dir: Path,
    out_dir: Path,
    in_path: Path,
) -> list[str]:
    return [
        soffice,
        f"-env:UserInstallation={profile_dir.as_uri()}",
        "--headless",
        "--calc",
        "--norestore",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(out_dir),
        str(in_path),
    ]


def _soffice_env(tmp: Path) -> dict[str, str]:
    return {
        **os.environ,
        "HOME": str(tmp),
        "TMPDIR": str(tmp),
        "SAL_USE_VCLPLUGIN": os.environ.get("SAL_USE_VCLPLUGIN", "svp"),
        "JAVA_TOOL_OPTIONS": os.environ.get("JAVA_TOOL_OPTIONS", "-Djava.awt.headless=true"),
    }


def recalc_workbook(blob: bytes, *, timeout: float = 60.0) -> bytes:
    """Recalculate *blob* via LibreOffice headless and return the new bytes.

    The cached cell values in the returned bytes are what Excel would show
    if a user opened the workbook and pressed F9. Read them with
    ``openpyxl.load_workbook(..., data_only=True)``.

    Parameters
    ----------
    blob:
        The workbook bytes to recalculate. Typically the output of
        ``build_*_excel_workbook.build_workbook_from_spec(...)``.
    timeout:
        Wall-clock seconds to wait for ``soffice`` to finish. Most SPIA /
        Term / RILA workbooks complete in 2-5 seconds; the default 60 s
        leaves ample margin for slower CI runners.

    Raises
    ------
    LibreOfficeNotAvailable
        When ``soffice`` is not installed.
    RecalcTimeout
        When soffice does not return within *timeout* seconds.
    RuntimeError
        When soffice exits non-zero or does not produce an output file.
    """
    soffice = ensure_libreoffice_available()
    with tempfile.TemporaryDirectory(prefix="annuity_recalc_") as tmpdir:
        tmp = Path(tmpdir)
        in_path = tmp / "in.xlsx"
        out_dir = tmp / "out"
        profile_dir = tmp / "lo_profile"
        out_dir.mkdir()
        profile_dir.mkdir()
        in_path.write_bytes(blob)
        # NOTE: soffice rewrites the output filename to match the input stem,
        # so we get out_dir / "in.xlsx".
        cmd = _soffice_command(
            soffice=soffice,
            profile_dir=profile_dir,
            out_dir=out_dir,
            in_path=in_path,
        )
        try:
            with _libreoffice_recalc_lock(timeout=timeout):
                proc = subprocess.run(
                    cmd,
                    capture_output=True,
                    timeout=timeout,
                    check=False,
                    # Use a per-invocation user profile dir and a host-wide
                    # lock. The profile prevents default ~/.config collisions;
                    # the lock prevents macOS LibreOffice crash dialogs caused
                    # by concurrent headless Calc startup.
                    env=_soffice_env(tmp),
                )
        except subprocess.TimeoutExpired as exc:
            raise RecalcTimeout(
                f"soffice did not complete within {timeout}s. "
                f"stdout={exc.stdout!r} stderr={exc.stderr!r}"
            ) from exc
        if proc.returncode != 0:
            raise RuntimeError(
                f"soffice exited {proc.returncode}: "
                f"stdout={proc.stdout.decode(errors='replace')!r} "
                f"stderr={proc.stderr.decode(errors='replace')!r}"
            )
        out_path = out_dir / "in.xlsx"
        if not out_path.is_file():
            # Some LibreOffice versions write to the input dir or change
            # the extension; fall back to scanning out_dir for any xlsx.
            xlsx_files = sorted(out_dir.glob("*.xlsx"))
            if not xlsx_files:
                raise RuntimeError(
                    f"soffice produced no output xlsx in {out_dir}; "
                    f"stdout={proc.stdout.decode(errors='replace')!r} "
                    f"stderr={proc.stderr.decode(errors='replace')!r}"
                )
            out_path = xlsx_files[0]
        return out_path.read_bytes()


def read_recalculated_cells(
    recalculated_blob: bytes, addresses: Iterable[str]
) -> dict[str, object]:
    """Convenience: return ``{address: cached_value}`` from a recalc'd blob.

    *addresses* are sheet-qualified, e.g. ``"ModelCheck!B5"``. Missing
    sheets / cells return ``None`` so callers can compose a clear failure
    message rather than swallowing a KeyError.
    """
    wb = load_workbook(io.BytesIO(recalculated_blob), data_only=True)
    out: dict[str, object] = {}
    for addr in addresses:
        if "!" not in addr:
            raise ValueError(
                f"Cell address {addr!r} must be sheet-qualified (e.g. 'ModelCheck!B5')."
            )
        sheet_name, cell_ref = addr.split("!", 1)
        if sheet_name not in wb.sheetnames:
            out[addr] = None
            continue
        ws = wb[sheet_name]
        out[addr] = ws[cell_ref].value
    return out
