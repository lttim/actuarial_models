"""
Defensive validation for Excel workbooks built by this project.

Goal
----
Prevent two distinct classes of regression that have hit users in the field:

1. **Excel "Removed Records: Formula from /xl/worksheets/sheetN.xml part"**
   repair messages. Root causes: wrong argument counts, unbalanced parens/quotes,
   embedded error literals (``#REF!`` / ``#NAME?``).

2. **Silent reconciliation failures** where Excel opens cleanly but ALM /
   Liabilities cells produce zero or garbage because a formula points at a
   column that does not exist on the target sheet (e.g. ``Liabilities!S`` when
   the RILA liability grid puts ``ExpTotalCF`` in column ``M``). Excel happily
   coerces blanks to zero in ``SUMPRODUCT`` / ``INDEX``, so the only visible
   symptom is a non-zero ``Difference (Excel − Python)`` row in ``ModelCheck``
   after a full recalc.

Both classes are now covered by ``validate_workbook_or_raise``:

* Static formula checks (paren/quote balance, function arity, error literals,
  trailing empty arguments).
* **Cross-sheet column reference resolution** — every direct ``Sheet!Col``
  reference *and* every ``Sheet!Col`` literal embedded inside a string passed
  to ``INDIRECT(...)`` is checked against the actual columns populated on the
  target sheet at workbook-save time. A reference to a fully empty column on
  another sheet is now a hard error.

Anyone exporting a workbook from this project should call
:func:`validate_workbook_or_raise` immediately before saving. The provided
pytest tests in ``tests/test_excel_export_validation.py`` run this validator
against freshly built SPIA, RILA, and Term workbooks so regressions are caught
in CI.

Public API
----------
* :func:`validate_workbook` — returns a list of :class:`FormulaIssue`.
* :func:`validate_workbook_or_raise` — raises :class:`ExcelWorkbookValidationError`
  if any issues are found.
* :func:`validate_formula` — single-formula static check (no workbook context,
  so no cross-sheet checks).
* :class:`ExcelWorkbookValidationError` — exception with a human-readable summary.

Validation philosophy
---------------------
We deliberately apply a *whitelist* of known functions with their accepted
arity. Unknown function names are ignored (no false positives). When you add a
new function in any workbook builder, register it in :data:`FUNCTION_ARITIES`
below. The cross-sheet column resolver is also conservative: it only flags
references to sheets present in the workbook with columns that are completely
empty (no data, no header).
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass

try:  # openpyxl is the workbook substrate; only imported lazily for type checks.
    from openpyxl.utils import column_index_from_string, get_column_letter
except Exception:  # pragma: no cover - openpyxl is a hard dep of this project
    column_index_from_string = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]

# Excel function name -> (min_args, max_args). ``None`` for max means unbounded.
# Keep this list aligned with functions actually used by the workbook builders in
# this repo. Adding a function here is required when a builder starts using it.
FUNCTION_ARITIES: dict[str, tuple[int, int | None]] = {
    "IF": (3, 3),  # Strict: always require an explicit false branch (avoids the Excel repair).
    "IFERROR": (2, 2),
    "AND": (1, None),
    "OR": (1, None),
    "NOT": (1, 1),
    "MIN": (1, None),
    "MAX": (1, None),
    "SUM": (1, None),
    "SUMPRODUCT": (1, None),
    "AVERAGE": (1, None),
    "COUNT": (1, None),
    "COUNTA": (1, None),
    "INDEX": (2, 4),
    "MATCH": (2, 3),
    "VLOOKUP": (3, 4),
    "HLOOKUP": (3, 4),
    "OFFSET": (3, 5),
    "ROW": (0, 1),
    "COLUMN": (0, 1),
    "ROWS": (1, 1),
    "COLUMNS": (1, 1),
    "EXP": (1, 1),
    "LN": (1, 1),
    "LOG": (1, 2),
    "POWER": (2, 2),
    "ABS": (1, 1),
    "MOD": (2, 2),
    "ROUND": (2, 2),
    "CEILING": (2, 2),
    "FLOOR": (2, 2),
    "SQRT": (1, 1),
    "INT": (1, 1),
    "TRUNC": (1, 2),
    "ISNUMBER": (1, 1),
    "ISBLANK": (1, 1),
    "ISERROR": (1, 1),
    "ISNA": (1, 1),
    "TEXT": (2, 2),
    "VALUE": (1, 1),
    "CONCATENATE": (1, None),
    "CONCAT": (1, None),
    "LEN": (1, 1),
    "LEFT": (1, 2),
    "RIGHT": (1, 2),
    "MID": (3, 3),
    "TRIM": (1, 1),
    "UPPER": (1, 1),
    "LOWER": (1, 1),
    "PROPER": (1, 1),
    "DATE": (3, 3),
    "YEAR": (1, 1),
    "MONTH": (1, 1),
    "DAY": (1, 1),
    "EDATE": (2, 2),
    "EOMONTH": (2, 2),
    "NPV": (2, None),
    "IRR": (1, 2),
    "PV": (3, 5),
    "FV": (3, 5),
    "PMT": (3, 5),
    "RATE": (3, 6),
    "NPER": (3, 5),
    # Lookup helpers commonly used here:
    "XLOOKUP": (3, 6),
    "XMATCH": (2, 4),
    "FILTER": (2, 3),
    "UNIQUE": (1, 3),
    "SORT": (1, 4),
}

# Excel error tokens that should never be embedded as literal text inside a formula
# (they typically mean a string substitution went wrong).
ERROR_LITERALS: tuple[str, ...] = (
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#DIV/0!",
    "#NUM!",
    "#VALUE!",
    "#N/A",
)


@dataclass(frozen=True)
class FormulaIssue:
    """One validation problem detected in a workbook formula."""

    sheet: str
    cell: str
    formula: str
    message: str

    def __str__(self) -> str:  # pragma: no cover - human-readable rendering
        return f"{self.sheet}!{self.cell}: {self.message}\n    formula = {self.formula}"


class ExcelWorkbookValidationError(ValueError):
    """Raised when one or more formulas in a workbook fail static validation."""

    def __init__(self, issues: list[FormulaIssue]) -> None:
        self.issues = list(issues)
        head = (
            f"{len(self.issues)} Excel formula issue(s) detected. Excel will likely "
            "show a 'Removed Records: Formula from /xl/worksheets/sheetN.xml part' "
            "repair message for this file. Fix all issues before saving:\n"
        )
        body = "\n".join(f"  - {iss}" for iss in self.issues[:25])
        more = "" if len(self.issues) <= 25 else f"\n  ... and {len(self.issues) - 25} more."
        super().__init__(head + body + more)


# ---------------------------------------------------------------------------
# Tokenizer
# ---------------------------------------------------------------------------


def _strip_outer_eq(formula: str) -> str:
    return formula[1:] if formula.startswith("=") else formula


def _check_balanced_parens_and_quotes(formula: str) -> str | None:
    """Return an error message if parentheses or string quotes are not balanced, else None."""
    depth = 0
    in_string = False
    i = 0
    n = len(formula)
    while i < n:
        ch = formula[i]
        if in_string:
            if ch == '"':
                if i + 1 < n and formula[i + 1] == '"':
                    i += 2
                    continue
                in_string = False
        else:
            if ch == '"':
                in_string = True
            elif ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth < 0:
                    return f"unbalanced ')' at position {i}"
        i += 1
    if in_string:
        return "unterminated string literal"
    if depth != 0:
        return f"unbalanced parentheses (net depth {depth})"
    return None


def _strip_strings_and_brackets(formula: str) -> str:
    """Replace string literals and bracketed sheet refs with spaces (preserves length).

    Hot path. ``validate_workbook`` calls this function many times per formula in a tight
    loop, so the implementation is written defensively to avoid quadratic costs:

    * Use ``str.find`` to skip directly to the next interesting character rather than
      walking the formula character by character in pure Python.
    * Build the output as a single ``"".join(parts)`` of slices + replacement spaces.
    """
    if '"' not in formula and "[" not in formula:
        # Hot path: most formulas have no quoted strings or table-style brackets and can
        # be returned as-is without any allocation.
        return formula
    out: list[str] = []
    i = 0
    n = len(formula)
    while i < n:
        ch = formula[i]
        if ch == '"':
            # Find the matching closing quote, treating "" as escape.
            j = i + 1
            while j < n:
                if formula[j] == '"':
                    if j + 1 < n and formula[j + 1] == '"':
                        j += 2
                        continue
                    break
                j += 1
            # Replace ``i..j`` (inclusive) with spaces.
            out.append(" " * (min(j, n - 1) - i + 1))
            i = j + 1
        elif ch == "[":
            j = formula.find("]", i + 1)
            if j == -1:
                j = n - 1
            out.append(" " * (j - i + 1))
            i = j + 1
        else:
            # Append a contiguous run of "safe" characters (no quotes/brackets) at once.
            j = i
            while j < n and formula[j] != '"' and formula[j] != "[":
                j += 1
            out.append(formula[i:j])
            i = j
    return "".join(out)


def _function_calls(formula: str, stripped: str | None = None) -> Iterable[tuple[str, int, int]]:
    """Yield (function_name_uppercase, open_paren_index, close_paren_index) for each call.

    ``stripped`` may be supplied by callers that have already computed
    ``_strip_strings_and_brackets(formula)`` to avoid recomputing it. The yielded
    indices are valid against both ``formula`` and ``stripped`` (the strip preserves
    length).
    """
    s = stripped if stripped is not None else _strip_strings_and_brackets(formula)
    n = len(s)
    i = 0
    stack: list[tuple[str | None, int]] = []  # (function_name, open_paren_index)
    pending_name_chars: list[str] = []
    pending_name_start = -1

    def flush_name() -> str | None:
        nonlocal pending_name_chars, pending_name_start
        if not pending_name_chars:
            return None
        name = "".join(pending_name_chars)
        pending_name_chars = []
        pending_name_start = -1
        # Strip leading sheet refs like "Sheet1!" so we only consider function names that begin a call.
        if "!" in name:
            return None
        if not name or not (name[0].isalpha() or name[0] == "_"):
            return None
        return name.upper()

    while i < n:
        ch = s[i]
        if ch.isalnum() or ch in "._":
            if not pending_name_chars:
                pending_name_start = i
            pending_name_chars.append(ch)
            i += 1
            continue
        if ch == "(":
            name = flush_name()
            stack.append((name, i))
            i += 1
            continue
        if ch == ")":
            if stack:
                name, open_idx = stack.pop()
                if name is not None:
                    yield name, open_idx, i
            i += 1
            continue
        # Other separators reset the name buffer.
        flush_name()
        i += 1


def _arg_count(formula: str, open_idx: int, close_idx: int, stripped: str | None = None) -> int:
    """Count top-level args between open_idx and close_idx in the (paren-stripped) formula.

    ``stripped`` may be passed in to avoid re-stripping; ``formula`` is otherwise unused
    apart from defaulting the strip computation.
    """
    s = stripped if stripped is not None else _strip_strings_and_brackets(formula)
    inner = s[open_idx + 1 : close_idx]
    if not inner.strip():
        return 0
    depth = 0
    args = 1
    for ch in inner:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "," and depth == 0:
            args += 1
    return args


def _has_trailing_empty_arg(formula: str, open_idx: int, close_idx: int) -> bool:
    """
    Return True when the call ends with a literal ``,)`` (no intervening token).

    We must use the **original** formula text (not the string-stripped form) here:
    a legitimate ``IFERROR(x, "")`` would look like ``IFERROR(x,   )`` after
    string stripping and would trip the heuristic.

    Forbidding this pattern catches the real bug — an f-string that lost its
    substitution and produced ``IF(a, b, )`` / ``IFERROR(x, )`` — without false
    positives on intentional empty-string defaults.
    """
    inner = formula[open_idx + 1 : close_idx]
    stripped = inner.rstrip()
    return stripped.endswith(",")


def _split_top_level_args(formula_inner: str) -> list[str]:
    """Split paren-stripped inner-call text on top-level commas (preserves order)."""
    parts: list[str] = []
    buf: list[str] = []
    depth = 0
    for ch in formula_inner:
        if ch == "(":
            depth += 1
            buf.append(ch)
        elif ch == ")":
            depth -= 1
            buf.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    parts.append("".join(buf))
    return parts


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def validate_formula(
    sheet: str,
    cell: str,
    formula: object,
    *,
    _stripped: str | None = None,
) -> list[FormulaIssue]:
    """Statically validate a single formula. Returns a list of issues (possibly empty).

    ``_stripped`` is an internal optimization: the workbook walker passes the
    pre-computed string-stripped body to avoid recomputing it for every check.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    issues: list[FormulaIssue] = []
    body = _strip_outer_eq(formula)

    err = _check_balanced_parens_and_quotes(body)
    if err is not None:
        issues.append(FormulaIssue(sheet=sheet, cell=cell, formula=formula, message=err))
        return issues

    cleaned = _stripped if _stripped is not None else _strip_strings_and_brackets(body)

    for tok in ERROR_LITERALS:
        pos = cleaned.find(tok)
        if pos != -1:
            issues.append(
                FormulaIssue(
                    sheet=sheet,
                    cell=cell,
                    formula=formula,
                    message=f"embedded Excel error literal {tok!r} at position {pos}",
                )
            )

    for name, open_idx, close_idx in _function_calls(body, stripped=cleaned):
        bounds = FUNCTION_ARITIES.get(name)
        if bounds is not None:
            lo, hi = bounds
            n_args = _arg_count(body, open_idx, close_idx, stripped=cleaned)
            if n_args < lo or (hi is not None and n_args > hi):
                hi_str = "any" if hi is None else str(hi)
                issues.append(
                    FormulaIssue(
                        sheet=sheet,
                        cell=cell,
                        formula=formula,
                        message=(
                            f"{name} called with {n_args} arg(s); expected "
                            f"{lo}..{hi_str}. Excel may flag this as a corrupt formula."
                        ),
                    )
                )
        if _has_trailing_empty_arg(body, open_idx, close_idx):
            issues.append(
                FormulaIssue(
                    sheet=sheet,
                    cell=cell,
                    formula=formula,
                    message=(
                        f"{name} ends with an implicit empty argument (', )'). Write the "
                        "intended value explicitly (e.g. ',\"\")' or ',0)'). Empty trailing "
                        "args are usually a sign that an f-string lost its substitution."
                    ),
                )
            )

    return issues


# ---------------------------------------------------------------------------
# Cross-sheet column reference resolution
# ---------------------------------------------------------------------------

# Matches ``Sheet!A1``, ``Sheet!$A$1``, ``Sheet!A:A``, ``Sheet!$A:$A``,
# ``Sheet!A1:B10``, ``Sheet!$A$1:$B$10``, ``Sheet!$A:$B``. Sheet names are limited
# to the unquoted form ``[A-Za-z_][A-Za-z0-9_]*`` since this project never uses
# spaces or punctuation in sheet names — quoted forms are intentionally ignored
# rather than risk false positives.
_CROSS_SHEET_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_])"
    r"([A-Za-z_][A-Za-z0-9_]*)!"
    r"\$?([A-Z]+)(?:\$?\d+)?"
    r"(?::\$?([A-Z]+)(?:\$?\d+)?)?"
)

# Functions whose first argument, when a literal string, contains a cell address
# we should still validate (e.g. ``INDIRECT("Liabilities!S5:S543")``).
_INDIRECT_LIKE_FUNCTIONS = ("INDIRECT",)


def _column_letters_in_range(start_col: str, end_col: str | None) -> list[str]:
    """Expand a column-letter range like ``A`` or ``A..C`` into individual letters."""
    if column_index_from_string is None or get_column_letter is None:
        return [start_col.upper()] if not end_col else []
    a = column_index_from_string(start_col.upper())
    b = column_index_from_string((end_col or start_col).upper())
    if b < a:
        a, b = b, a
    return [get_column_letter(i) for i in range(a, b + 1)]


def _populated_columns_per_sheet(workbook: object) -> dict[str, set[str]]:
    """
    Return ``{sheet_title: {column_letter, ...}}`` for every sheet in the workbook.

    A column counts as populated if any cell in that column has a value (header,
    formula, or data) — that is enough to know the reference will resolve to
    something rather than to a fully empty stripe.

    Implementation uses openpyxl's sparse cell storage (``ws._cells``) so we visit
    only the cells that actually exist instead of iterating the entire bounding
    rectangle (which would make a ~600 row × 150 col ALM_Engine sheet 90,000
    iterations of empty cells).
    """
    if get_column_letter is None:
        return {}
    out: dict[str, set[str]] = {}
    for ws in getattr(workbook, "worksheets", []):
        title = str(getattr(ws, "title", "?"))
        cols: set[str] = set()
        sparse = getattr(ws, "_cells", None)
        if isinstance(sparse, dict):
            col_indices = {c for (_r, c) in sparse.keys()}
            cols = {get_column_letter(c) for c in col_indices}
        else:  # pragma: no cover - defensive fallback
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is None:
                        continue
                    cols.add(get_column_letter(cell.column))
        out[title] = cols
    return out


def _iter_cross_sheet_refs(
    formula_body: str, stripped: str | None = None
) -> Iterable[tuple[str, str, str | None]]:
    """
    Yield ``(sheet_name, start_col_letter, end_col_letter_or_None)`` for every
    cross-sheet reference in ``formula_body`` — both direct references and any
    references embedded inside string literals consumed by ``INDIRECT(...)``.

    Skips the work entirely if ``formula_body`` does not contain ``"!"`` (Excel's
    sheet-qualified separator); the vast majority of cells on a given sheet are
    intra-sheet references and need no cross-sheet check.
    """
    if "!" not in formula_body:
        return
    seen: set[tuple[str, str, str | None]] = set()
    cleaned = stripped if stripped is not None else _strip_strings_and_brackets(formula_body)

    for m in _CROSS_SHEET_REF_RE.finditer(cleaned):
        triple = (m.group(1), m.group(2).upper(), (m.group(3).upper() if m.group(3) else None))
        if triple not in seen:
            seen.add(triple)
            yield triple

    if "INDIRECT" not in formula_body.upper():
        return
    for name, open_idx, close_idx in _function_calls(formula_body, stripped=cleaned):
        if name not in _INDIRECT_LIKE_FUNCTIONS:
            continue
        inner = formula_body[open_idx + 1 : close_idx]
        for m in _CROSS_SHEET_REF_RE.finditer(inner):
            triple = (m.group(1), m.group(2).upper(), (m.group(3).upper() if m.group(3) else None))
            if triple not in seen:
                seen.add(triple)
                yield triple


def _check_cross_sheet_columns(
    *,
    sheet: str,
    cell_addr: str,
    formula: str,
    populated: dict[str, set[str]],
    stripped: str | None = None,
) -> list[FormulaIssue]:
    """Flag any cross-sheet reference whose column does not exist on the target sheet."""
    if not populated:
        return []
    body = _strip_outer_eq(formula)
    if "!" not in body:
        return []
    issues: list[FormulaIssue] = []
    for ref_sheet, col_a, col_b in _iter_cross_sheet_refs(body, stripped=stripped):
        if ref_sheet not in populated:
            # Could be a defined name / table — don't false-positive.
            continue
        cols = _column_letters_in_range(col_a, col_b)
        present = populated[ref_sheet]
        missing = [c for c in cols if c not in present]
        if missing:
            range_label = f"{col_a}" if col_b is None or col_b == col_a else f"{col_a}:{col_b}"
            issues.append(
                FormulaIssue(
                    sheet=sheet,
                    cell=cell_addr,
                    formula=formula,
                    message=(
                        f"references {ref_sheet}!{range_label} but column(s) "
                        f"{','.join(missing)} on {ref_sheet!r} are completely empty. "
                        "Excel will silently treat the reference as zero/blank, "
                        "breaking ModelCheck reconciliation."
                    ),
                )
            )
    return issues


# Pattern that collapses every digit run in a formula to a single ``#`` placeholder.
# Two formulas that differ only in row numbers (``A14`` vs ``A15``, ``S543`` vs ``S544``)
# share the same template and therefore the same validation outcome.
_DIGIT_RUN_RE = re.compile(r"\d+")


def _formula_template(formula: str) -> str:
    return _DIGIT_RUN_RE.sub("#", formula)


def validate_workbook(workbook: object) -> list[FormulaIssue]:
    """Walk every cell in an openpyxl Workbook; return all formula issues found.

    The hot path applies two layers of caching to keep validation under a second
    even on RILA workbooks with 75 000+ formulas:

    * ``stripped`` — string-stripped form of each formula, computed exactly once
      per cell and reused across the syntax checks and the cross-sheet column
      resolver (without this they each re-stripped, costing 5–10× extra work).
    * ``template_cache`` — per-workbook dedup that keys on the formula's
      digit-stripped *template*. Excel models generated by this project emit
      huge blocks of structurally identical formulas (a column of 600 rows that
      differ only in row numbers shares one template), so we only fully
      validate each distinct template once and replay the resulting issue
      messages for subsequent cells. Issue ``cell`` and ``formula`` fields stay
      cell-specific so error reports remain precise.
    """
    issues: list[FormulaIssue] = []
    populated = _populated_columns_per_sheet(workbook)
    template_cache: dict[str, list[str]] = {}
    for ws in getattr(workbook, "worksheets", []):
        sheet = str(getattr(ws, "title", "?"))
        sparse = getattr(ws, "_cells", None)
        if isinstance(sparse, dict):
            cells_iter: Iterable = sparse.values()
        else:  # pragma: no cover - defensive fallback
            cells_iter = (c for row in ws.iter_rows() for c in row)
        for cell in cells_iter:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            template = _formula_template(v)
            cached = template_cache.get(template)
            if cached is not None:
                if cached:
                    issues.extend(
                        FormulaIssue(sheet=sheet, cell=cell.coordinate, formula=v, message=msg)
                        for msg in cached
                    )
                continue
            body = _strip_outer_eq(v)
            stripped = _strip_strings_and_brackets(body)
            cell_issues = list(validate_formula(sheet, cell.coordinate, v, _stripped=stripped))
            cell_issues.extend(
                _check_cross_sheet_columns(
                    sheet=sheet,
                    cell_addr=cell.coordinate,
                    formula=v,
                    populated=populated,
                    stripped=stripped,
                )
            )
            template_cache[template] = [iss.message for iss in cell_issues]
            issues.extend(cell_issues)
    return issues


def validate_workbook_or_raise(workbook: object) -> None:
    """Validate a workbook; raise :class:`ExcelWorkbookValidationError` if any issue found."""
    issues = validate_workbook(workbook)
    if issues:
        raise ExcelWorkbookValidationError(issues)
