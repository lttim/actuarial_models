from __future__ import annotations

import io
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Literal, NamedTuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import spia_projection as sp

BASE_DIR = Path(__file__).resolve().parent
OUT_XLSX = BASE_DIR / "spia_projection_model.xlsx"


@dataclass(frozen=True)
class ExcelPythonSnapshot:
    """Pricing metrics from the Python run used to build the workbook (for cross-checking formulas)."""

    pv_benefit: float
    pv_monthly_expenses: float
    pv_monthly_total: float
    single_premium: float
    annuity_factor: float


@dataclass
class MCExcelSnapshot:
    """Pre-computed Monte Carlo statistics for embedding in the Excel workbook.

    Build with ``mc_excel_snapshot_from_result`` after running
    ``price_spia_single_premium_monte_carlo``.
    """

    n_sims: int
    annual_drift: float
    annual_vol: float
    s0: float
    summary_df: pd.DataFrame  # columns: Metric, Mean, Std Dev, P5, P25, Median, P75, P95
    prem_hist_mids: list  # bin midpoints (float)
    prem_hist_counts: list  # counts (int)


@dataclass(frozen=True)
class ALMExcelSnapshot:
    """Monthly ALM path from the latest Python ALM run; workbook embeds series as values with derived formulas."""

    initial_asset_market_value: float
    bucket_names: tuple[str, ...]
    weights: tuple[float, ...]
    duration_gap: float
    duration_assets_mac: float
    duration_liabilities_mac: float
    pv01_net: float
    month_index: np.ndarray  # 0..n-1 paid months
    asset_market_value: np.ndarray
    liability_pv: np.ndarray
    borrowing_balance: np.ndarray  # shape (n_months,)
    bucket_asset_mv: np.ndarray  # shape (n_buckets, n_months)


class ALMDashboardLayout(NamedTuple):
    """Row references on ``ALM_Projection`` for Dashboard links and charts."""

    header_row: int
    first_data_row: int
    last_data_row: int


ALM_SHEET_NAME = "ALM_Projection"
# First month of ALM path data on ALM_Projection (below header row 12). Keep in sync with _write_alm_projection_sheet.
ALM_PROJECTION_FIRST_DATA_ROW = 13


def alm_excel_snapshot_from_result(
    alm: "sp.ALMResult",
    asm: "sp.ALMAssumptions | None",
    *,
    initial_asset_market_value: float | None = None,
) -> ALMExcelSnapshot:
    """Build an embeddable ALM snapshot; prefers ``asm.allocation`` for bucket labels and weights."""
    if asm is None:
        raise ValueError("ALMAssumptions required to label ALM workbook columns.")
    spec = asm.allocation
    names = tuple(str(b.name) for b in spec.buckets)
    w = tuple(float(x) for x in np.asarray(spec.weights, dtype=float).tolist())
    bmv = np.asarray(alm.bucket_asset_mv, dtype=float)
    if bmv.shape != (len(names), alm.asset_market_value.shape[0]):
        raise ValueError("bucket_asset_mv shape mismatch.")
    if initial_asset_market_value is not None:
        aum0 = float(initial_asset_market_value)
    else:
        aum0 = float(np.sum(bmv[:, 0])) if bmv.size else float("nan")
    return ALMExcelSnapshot(
        initial_asset_market_value=aum0,
        bucket_names=names,
        weights=w,
        duration_gap=float(alm.duration_gap),
        duration_assets_mac=float(alm.duration_assets_mac),
        duration_liabilities_mac=float(alm.duration_liabilities_mac),
        pv01_net=float(alm.pv01_net),
        month_index=np.asarray(alm.month_index, dtype=int),
        asset_market_value=np.asarray(alm.asset_market_value, dtype=float),
        liability_pv=np.asarray(alm.liability_pv, dtype=float),
        borrowing_balance=np.asarray(alm.borrowing_balance, dtype=float),
        bucket_asset_mv=bmv,
    )


def mc_excel_snapshot_from_result(
    mc: "sp.SPIAMonteCarloResult",
    *,
    annual_drift: float,
    annual_vol: float,
    s0: float,
    n_hist_bins: int = 40,
) -> MCExcelSnapshot:
    """Build an MCExcelSnapshot from a SPIAMonteCarloResult for embedding in Excel."""
    metrics = [
        ("Single Premium", mc.single_premium),
        ("PV Benefit", mc.pv_benefit),
        ("PV Monthly Expenses", mc.pv_monthly_expenses),
        ("PV Monthly Total", mc.pv_monthly_total),
        ("Annuity Factor", mc.annuity_factor),
    ]
    rows = []
    for name, arr in metrics:
        rows.append(
            {
                "Metric": name,
                "Mean": float(np.mean(arr)),
                "Std Dev": float(np.std(arr)),
                "P5": float(np.percentile(arr, 5)),
                "P25": float(np.percentile(arr, 25)),
                "Median": float(np.median(arr)),
                "P75": float(np.percentile(arr, 75)),
                "P95": float(np.percentile(arr, 95)),
            }
        )
    summary_df = pd.DataFrame(rows)

    counts, edges = np.histogram(mc.single_premium, bins=n_hist_bins)
    mids = 0.5 * (edges[:-1] + edges[1:])

    return MCExcelSnapshot(
        n_sims=mc.n_sims,
        annual_drift=float(annual_drift),
        annual_vol=float(annual_vol),
        s0=float(s0),
        summary_df=summary_df,
        prem_hist_mids=mids.tolist(),
        prem_hist_counts=counts.tolist(),
    )


@dataclass(frozen=True)
class ExcelBuildSpec:
    """Inputs for a formula-driven SPIA workbook aligned with `spia_ui.py` / `spia_projection.py`."""

    issue_age: int
    sex: str
    benefit_annual: float
    payment_freq_per_year: int
    valuation_year: int
    horizon_age: int
    spread: float
    yield_curve_df: pd.DataFrame  # columns: maturity_years, zero_rate
    mortality_excel_mode: Literal["rp_mp", "static"]  # static = no MP layer (synthetic or qx CSV)
    base_qx_df: pd.DataFrame  # columns: age, qx
    mp_improvement_long_df: pd.DataFrame | None  # columns: age, year, improvement_rate (long); None if static
    policy_expense_dollars: float
    premium_expense_rate: float  # decimal, e.g. 0.02 for 2%
    monthly_expense_dollars: float
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    n_months: int
    expense_annual_inflation: float  # decimal, e.g. 0.025 for 2.5%/year on maintenance expenses
    index_scenario_df: pd.DataFrame  # columns: month (0..n_months), sp500_level

    def __post_init__(self) -> None:
        if self.payment_freq_per_year != 12:
            raise ValueError("Excel workbook currently assumes monthly payments (12 per year).")
        if self.n_months < 1:
            raise ValueError("n_months must be positive.")


def _n_months_for_contract(issue_age: int, horizon_age: int) -> int:
    dt = 1.0 / 12.0
    return max(1, int(round((horizon_age - issue_age) / dt)))


def _mp_table_to_long_df(mortality: sp.MortalityTableRP2014MP2016) -> pd.DataFrame:
    ages = np.asarray(mortality.mp2016_ages, dtype=int)
    years = np.asarray(mortality.mp2016_years, dtype=int)
    mat = np.asarray(mortality.mp2016_i_matrix, dtype=float)
    rows: list[dict[str, float | int]] = []
    for i, a in enumerate(ages):
        for j, y in enumerate(years):
            rows.append({"age": int(a), "year": int(y), "improvement_rate": float(mat[i, j])})
    return pd.DataFrame(rows)


def excel_spec_from_launcher(
    *,
    contract: sp.SPIAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
) -> ExcelBuildSpec:
    """
    Build a spec from the same objects used in `price_spia_single_premium` / the Streamlit launcher.

    `valuation_year` should match the launcher field (Inputs / calendar column), including for static
    mortality where Python pricing ignores it.
    """
    if contract.payment_freq_per_year != 12:
        raise ValueError("Excel export supports monthly payment frequency only.")

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(yield_curve.zero_rates, dtype=float),
        }
    )

    if isinstance(mortality, sp.MortalityTableRP2014MP2016):
        base_qx_df = pd.DataFrame(
            {"age": np.asarray(mortality.base_qx_2014.ages, dtype=int), "qx": np.asarray(mortality.base_qx_2014.qx, dtype=float)}
        )
        mp_long = _mp_table_to_long_df(mortality)
        excel_mort = "rp_mp"
    else:
        base_qx_df = pd.DataFrame(
            {"age": np.asarray(mortality.ages, dtype=int), "qx": np.asarray(mortality.qx, dtype=float)}
        )
        mp_long = None
        excel_mort = "static"

    vy = int(valuation_year)

    n_m = _n_months_for_contract(contract.issue_age, horizon_age)
    idx_rows = [{"month": 0, "sp500_level": float(index_s0)}]
    for k in range(1, n_m + 1):
        idx_rows.append({"month": int(k), "sp500_level": float(index_levels_at_payment[k - 1])})
    idx_df = pd.DataFrame(idx_rows)

    return ExcelBuildSpec(
        issue_age=int(contract.issue_age),
        sex=str(contract.sex),
        benefit_annual=float(contract.benefit_annual),
        payment_freq_per_year=int(contract.payment_freq_per_year),
        valuation_year=vy,
        horizon_age=int(horizon_age),
        spread=float(spread),
        yield_curve_df=ycdf,
        mortality_excel_mode=excel_mort,
        base_qx_df=base_qx_df,
        mp_improvement_long_df=mp_long,
        policy_expense_dollars=float(expenses.policy_expense_dollars),
        premium_expense_rate=float(expenses.premium_expense_rate),
        monthly_expense_dollars=float(expenses.monthly_expense_dollars),
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        n_months=n_m,
        expense_annual_inflation=float(expense_annual_inflation),
        index_scenario_df=idx_df,
    )


def _read_text(path: Path, fallback: str = "") -> str:
    try:
        return path.read_text(encoding="utf-8")
    except Exception:
        return fallback


def _write_inputs(ws, spec: ExcelBuildSpec) -> None:
    ws.title = "Inputs"
    ws["A1"] = "SPIA Inputs (matches model launcher / Python)"
    ws["A1"].font = Font(bold=True, size=12)

    mort_cell = "rp_mp" if spec.mortality_excel_mode == "rp_mp" else "static"

    rows: list[tuple[str, object]] = [
        ("Issue Age", spec.issue_age),
        ("Sex", spec.sex),
        ("Annual Benefit", spec.benefit_annual),
        ("Payment Frequency", spec.payment_freq_per_year),
        ("Valuation Year", spec.valuation_year),
        ("Horizon Age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Policy Expense Dollars", spec.policy_expense_dollars),
        ("Premium Expense Rate (decimal)", spec.premium_expense_rate),
        ("Monthly Expense Dollars", spec.monthly_expense_dollars),
        ("Mortality Excel Mode", mort_cell),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
        ("Expense annual inflation (decimal)", spec.expense_annual_inflation),
    ]

    # Data starts row 3: labels col A, values col B
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    ws["A21"] = "Derived"
    ws["A21"].font = Font(bold=True)
    ws["A22"] = "Monthly base benefit (pre-index)"
    ws["B22"] = "=B5/B6"
    ws["A23"] = "Projection Months (info; grid is fixed at build)"
    ws["B23"] = "=MAX(1,ROUND((B8-B3)*B6,0))"

    ws["D3"] = "Notes"
    ws["D4"] = "Rates are decimals unless labeled otherwise."
    ws["D5"] = "Valuation date: 12/31 of Valuation Year (RP+MP path)."
    ws["D6"] = "Mortality Excel Mode: rp_mp uses MP-2016 sums; static uses Base Qx only."
    ws["D7"] = "Discount factors use log-linear interpolation on DF nodes (matches Python YieldCurve)."
    ws["D8"] = "Benefits: return indexation from IndexScenario; expenses: monthly CPI-style from B17."
    ws["D9"] = "Changing horizon/issue age does not auto-resize sheets; regenerate from the launcher."
    ws["D10"] = "Spread B9 is added to zero rates. Negative B9 lowers discount yields and raises PV—must match launcher."
    ws["D11"] = "See ModelCheck: Python snapshot vs Projection! formulas; large |Difference| means inputs/recalc issues."


def _write_simple_table(ws, title: str, df: pd.DataFrame) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    for c, col in enumerate(df.columns, start=1):
        ws.cell(row=3, column=c, value=col).font = Font(bold=True)
    for r, row in enumerate(df.itertuples(index=False), start=4):
        for c, val in enumerate(row, start=1):
            if isinstance(val, (int, float, np.integer, np.floating)):
                ws.cell(row=r, column=c, value=float(val))
            else:
                ws.cell(row=r, column=c, value=val)


def _write_monthly_curve_logdf(ws, n_months: int, y_last_row: int) -> None:
    """
    Monthly discount factors consistent with Python `YieldCurve.discount_factors`:
    log-linear interpolation on DF between curve nodes; flat zero-rate extrapolation beyond endpoints.
    """
    ws.title = "MonthlyCurve"
    ws["A1"] = "Monthly Discount Factors (log-linear on DF)"
    ws["A1"].font = Font(bold=True, size=12)

    headers = [
        "Month",
        "t_years",
        "BracketIndex",
        "LowerMat",
        "UpperMat",
        "LowerZero",
        "UpperZero",
        "InterpWeight",
        "LogDF_lower_node",
        "LogDF_upper_node",
        "LogDF_t",
        "DiscountFactor",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)

    first = 4
    last = first + n_months - 1
    y_rng = f"YieldCurve!$A$4:$A${y_last_row}"
    z_rng = f"YieldCurve!$B$4:$B${y_last_row}"

    for r in range(first, last + 1):
        ws[f"A{r}"] = r - first + 1
        ws[f"B{r}"] = f"=A{r}/Inputs!$B$6"
        ws[f"C{r}"] = (
            f"=IF(B{r}<=INDEX({y_rng},1),1,"
            f"IF(B{r}>=INDEX({y_rng},ROWS({y_rng})),ROWS({y_rng})-1,"
            f"MATCH(B{r},{y_rng},1)))"
        )
        ws[f"D{r}"] = f"=INDEX({y_rng},C{r})"
        ws[f"E{r}"] = f"=INDEX({y_rng},C{r}+1)"
        ws[f"F{r}"] = f"=INDEX({z_rng},C{r})"
        ws[f"G{r}"] = f"=INDEX({z_rng},C{r}+1)"
        ws[f"H{r}"] = f"=IF(E{r}=D{r},0,(B{r}-D{r})/(E{r}-D{r}))"
        ws[f"I{r}"] = f"=-(F{r}+Inputs!$B$9)*D{r}"
        ws[f"J{r}"] = f"=-(G{r}+Inputs!$B$9)*E{r}"
        ws[f"K{r}"] = (
            f"=IF(B{r}<=INDEX({y_rng},1),-(INDEX({z_rng},1)+Inputs!$B$9)*B{r},"
            f"IF(B{r}>=INDEX({y_rng},ROWS({y_rng})),-(INDEX({z_rng},ROWS({z_rng}))+Inputs!$B$9)*B{r},"
            f"I{r}+H{r}*(J{r}-I{r})))"
        )
        ws[f"L{r}"] = f"=EXP(K{r})"


def _write_projection(ws, n_months: int, y_last_row: int, idx_last_row: int) -> None:
    ws.title = "Projection"
    ws["A1"] = "SPIA Monthly Projection (Formula Driven)"
    ws["A1"].font = Font(bold=True, size=12)

    headers = [
        "Month",
        "t_years",
        "AttainedAgeQx",
        "IntAge",
        "CalendarYear",
        "IndexLevel",
        "BenefitNominal",
        "ExpenseNominal",
        "BaseQx",
        "CumMP_log_sum",
        "QxAnnual",
        "MuYear",
        "SurvInterval",
        "SurvivalToPay",
        "DiscountFactor",
        "ImpliedZeroFromDF",
        "ExpBenefitCF",
        "ExpExpenseCF",
        "ExpTotalCF",
        "PVBenefitCF",
        "PVExpenseCF",
        "ReserveAfterPayment",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)

    first = 4
    last = first + n_months - 1
    ir = f"IndexScenario!$A$4:$A${idx_last_row}"
    br = f"IndexScenario!$B$4:$B${idx_last_row}"

    for r in range(first, last + 1):
        ws[f"A{r}"] = r - first + 1
        ws[f"B{r}"] = f"=A{r}/Inputs!$B$6"
        # Fractional age at start of (month A) interval — matches Python monthly_survival_to_payment q_x lookup.
        ws[f"C{r}"] = f"=Inputs!$B$3+(A{r}-1)/Inputs!$B$6"
        ws[f"D{r}"] = f"=INT(C{r})"
        ws[f"E{r}"] = f"=Inputs!$B$7+1+INT((A{r}-1)/Inputs!$B$6)"
        ws[f"F{r}"] = f'=IFERROR(INDEX({br},MATCH(A{r},{ir},0)),"")'
        if r == first:
            ws[f"G{r}"] = f"=Inputs!$B$22*F{r}/INDEX({br},MATCH(0,{ir},0))"
            ws[f"H{r}"] = "=Inputs!$B$12"
        else:
            ws[f"G{r}"] = f"=G{r-1}*F{r}/F{r-1}"
            ws[f"H{r}"] = f"=H{r-1}*(1+Inputs!$B$17)^(1/12)"

        ws[f"I{r}"] = f'=IFERROR(INDEX(BaseQx!$B:$B,MATCH(D{r},BaseQx!$A:$A,0)),"")'
        ws[f"J{r}"] = (
            f'=IF(Inputs!$B$13="rp_mp",'
            f'SUMIFS(MP2016_Long!$C:$C,MP2016_Long!$A:$A,D{r},MP2016_Long!$B:$B,">=2014",MP2016_Long!$B:$B,"<"&E{r}),0)'
        )
        ws[f"K{r}"] = f"=MIN(0.999999,MAX(0,I{r}*EXP(J{r})))"
        ws[f"L{r}"] = f"=-LN(1-K{r})"
        ws[f"M{r}"] = f"=EXP(-L{r}/Inputs!$B$6)"
        if r == first:
            ws[f"N{r}"] = f"=M{r}"
        else:
            ws[f"N{r}"] = f"=N{r-1}*M{r}"

        ws[f"O{r}"] = f'=IFERROR(INDEX(MonthlyCurve!$L:$L,MATCH(A{r},MonthlyCurve!$A:$A,0)),"")'
        ws[f"P{r}"] = f"=IF(B{r}>0,-LN(O{r})/B{r},\"\")"
        ws[f"Q{r}"] = f"=G{r}*N{r}"
        ws[f"R{r}"] = f"=H{r}*N{r}"
        ws[f"S{r}"] = f"=Q{r}+R{r}"
        ws[f"T{r}"] = f"=Q{r}*O{r}"
        ws[f"U{r}"] = f"=R{r}*O{r}"

        if r == last:
            ws[f"V{r}"] = 0.0
        else:
            ws[f"V{r}"] = f"=SUMPRODUCT(S{r+1}:S{last},O{r+1}:O{last})/(N{r}*O{r})"

    ws["W3"] = "Summary"
    ws["W3"].font = Font(bold=True)
    ws["W4"] = "PV Benefits"
    ws["X4"] = f"=SUM(T{first}:T{last})"
    ws["W5"] = "PV Monthly Expenses"
    ws["X5"] = f"=SUM(U{first}:U{last})"
    ws["W6"] = "Annuity Factor"
    ws["X6"] = f"=SUMPRODUCT(N{first}:N{last},O{first}:O{last})"
    ws["W7"] = "PV Monthly Total"
    ws["X7"] = f"=X4+X5"
    ws["W8"] = "Single Premium"
    ws["X8"] = "=(Inputs!$B$10+X7)/(1-Inputs!$B$11)"
    ws["W9"] = "Reserve at t=0"
    ws["X9"] = "=X7"

    ws["A2"] = "ReserveAtT0"
    ws["B2"] = 0
    ws["C2"] = "=Inputs!$B$3"
    ws["V2"] = "=X9"


def _alm_liability_pv_cell_formula(*, excel_row: int, proj_last_row: int) -> str:
    """
    Liability PV at end of month A{excel_row} (month numbering matches Projection column A).
    Mirrors Python ``liab_pv_path``: sum_{j>M} S_j O_j / O_M with S=Projection!S, O=Projection!O.
    """
    pl = int(proj_last_row)
    r = int(excel_row)
    return (
        f'=IF(INDEX(Projection!$O:$O,3+A{r})<=0,NA(),'
        f'IF(4+A{r}>{pl},0,'
        f'SUMPRODUCT(INDIRECT("Projection!S" & (4+A{r}) & ":S{pl}"),'
        f'INDIRECT("Projection!O" & (4+A{r}) & ":O{pl}"))'
        f'/INDEX(Projection!$O:$O,3+A{r})))'
    )


def _write_alm_projection_sheet(wb: Workbook, snap: ALMExcelSnapshot) -> ALMDashboardLayout:
    ws = wb.create_sheet(ALM_SHEET_NAME)
    ws["A1"] = "ALM monthly path (mixed: liability from Projection formulas; assets from Python ladder)"
    ws["A1"].font = Font(bold=True, size=12)
    n = int(snap.asset_market_value.shape[0])
    n_b = len(snap.bucket_names)
    proj_last = 3 + n
    ws["A2"] = (
        f"Liability PV (column D) is calculated as SUMPRODUCT(Projection!S×O for months after row A) ÷ Projection!O, "
        f"matching the Python ALM liability measure (rows 4–{proj_last} on Projection). "
        "Asset MV (C) and bucket columns still use the Python Treasury-ladder path (reinvest / rebalance / "
        "disinvest / borrowing rules are not rebuilt as native formulas). Borrowing (E) is from that Python path. "
        "Surplus F = C−D−E; funding G = C/(D+E). Re-run ALM in the app to refresh embedded asset columns."
    )
    last_hdr_col = get_column_letter(7 + n_b)
    ws.merge_cells(f"A2:{last_hdr_col}2")

    ws["A3"], ws["B3"] = "Initial AUM ($)", float(snap.initial_asset_market_value)
    ws["A4"], ws["B4"] = "Duration gap (y) (issue-time Macaulay)", float(snap.duration_gap)
    ws["A5"], ws["B5"] = "Asset Macaulay duration (y)", float(snap.duration_assets_mac)
    ws["A6"], ws["B6"] = "Liability Macaulay duration (y)", float(snap.duration_liabilities_mac)
    ws["A7"], ws["B7"] = "PV01 net ($/bp) (issue-time)", float(snap.pv01_net)
    ws["A9"] = "Target allocation (issue)"
    alloc_txt = "; ".join(f"{snap.bucket_names[i]} {snap.weights[i]:.2%}" for i in range(len(snap.bucket_names)))
    ws["B9"] = alloc_txt

    header_row = ALM_PROJECTION_FIRST_DATA_ROW - 1
    headers = [
        "Month",
        "Attained age",
        "Asset MV",
        "Liability PV",
        "Borrowing",
        "Surplus",
        "Funding ratio",
    ] + list(snap.bucket_names)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h).font = Font(bold=True)

    first_data = ALM_PROJECTION_FIRST_DATA_ROW
    debt = np.asarray(snap.borrowing_balance, dtype=float).reshape(-1)
    if debt.shape[0] != n:
        raise ValueError("borrowing_balance length must match asset path.")
    for i in range(n):
        r = first_data + i
        ws.cell(row=r, column=1, value=int(snap.month_index[i] + 1))
        ws.cell(row=r, column=2, value=f'=IFERROR(INDEX(Projection!$C:$C,MATCH(A{r},Projection!$A:$A,0)),"")')
        ws.cell(row=r, column=3, value=float(snap.asset_market_value[i]))
        ws.cell(row=r, column=4, value=_alm_liability_pv_cell_formula(excel_row=r, proj_last_row=proj_last))
        ws.cell(row=r, column=5, value=float(debt[i]))
        ws.cell(row=r, column=6, value=f"=C{r}-D{r}-E{r}")
        ws.cell(row=r, column=7, value=f'=IF((D{r}+E{r})>0,C{r}/(D{r}+E{r}),"")')
        for b in range(n_b):
            ws.cell(row=r, column=8 + b, value=float(snap.bucket_asset_mv[b, i]))

    last_data = first_data + n - 1
    ws["B3"].number_format = "#,##0.00"
    for addr in ("B4", "B5", "B6"):
        ws[addr].number_format = "0.00"
    ws["B7"].number_format = "#,##0.00"
    for r in range(first_data, last_data + 1):
        for col in range(3, 8 + n_b):
            cell = ws.cell(row=r, column=col)
            if col in (3, 4, 5, 6) or col >= 8:
                cell.number_format = "#,##0.00"
            elif col == 7:
                cell.number_format = "0.000"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for c in range(3, 8 + n_b + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    return ALMDashboardLayout(header_row=header_row, first_data_row=first_data, last_data_row=last_data)


def _write_mc_summary_sheet(wb: Workbook, mc: MCExcelSnapshot) -> None:
    """Embed Monte Carlo summary statistics and a premium distribution chart in MC_Summary sheet."""
    ws = wb.create_sheet("MC_Summary")

    ws["A1"] = (
        f"Monte Carlo Summary — {mc.n_sims:,} simulations | "
        f"GBM drift {mc.annual_drift * 100:.1f}% | vol {mc.annual_vol * 100:.1f}% | S\u2080 {mc.s0:.2f}"
    )
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "Index return paths simulated via GBM. Mortality, yield curve, and expense inflation remain "
        "deterministic. Statistics span the full distribution of pricing outcomes across all paths."
    )
    ws.merge_cells("A2:H2")

    # --- Summary statistics table (rows 4–9) ---
    df = mc.summary_df
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=4, column=c_idx, value=col_name).font = Font(bold=True)

    for r_idx, row in enumerate(df.itertuples(index=False), start=5):
        vals = list(row)
        metric_name = str(vals[0])
        ws.cell(row=r_idx, column=1, value=metric_name)
        is_factor = "Factor" in metric_name
        num_fmt = "0.000000" if is_factor else "#,##0.00"
        for c_idx, val in enumerate(vals[1:], start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=float(val))
            cell.number_format = num_fmt

    # --- Premium distribution histogram data (rows 11 onward, columns A–B) ---
    n_bins = len(mc.prem_hist_mids)
    h_title_row = 11
    h_header_row = 12
    h_data_start = 13
    h_data_end = h_data_start + n_bins - 1

    ws.cell(row=h_title_row, column=1, value="Premium Distribution").font = Font(bold=True)
    ws.cell(row=h_header_row, column=1, value="Bin Midpoint ($)").font = Font(bold=True)
    ws.cell(row=h_header_row, column=2, value="Count").font = Font(bold=True)

    for i, (mid, cnt) in enumerate(zip(mc.prem_hist_mids, mc.prem_hist_counts)):
        ws.cell(row=h_data_start + i, column=1, value=round(float(mid), 0)).number_format = "#,##0"
        ws.cell(row=h_data_start + i, column=2, value=int(cnt))

    # --- Embedded bar chart ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Single Premium Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Premium ($)"
    chart.grouping = "clustered"
    chart.overlap = 100
    chart.legend = None
    chart.height = 12
    chart.width = 20

    counts_ref = Reference(ws, min_col=2, min_row=h_data_start, max_row=h_data_end)
    cats_ref = Reference(ws, min_col=1, min_row=h_data_start, max_row=h_data_end)
    chart.add_data(counts_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "D11")

    # Column widths
    ws.column_dimensions["A"].width = 24
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 14


def _write_dashboard(wb: Workbook, n_months: int, *, alm_layout: ALMDashboardLayout | None = None) -> None:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "SPIA Policy Projection Dashboard"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Policy Inputs"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Issue Age"
    ws["B4"] = "=Inputs!B3"
    ws["A5"] = "Sex"
    ws["B5"] = "=Inputs!B4"
    ws["A6"] = "Annual Benefit"
    ws["B6"] = "=Inputs!B5"
    ws["A7"] = "Valuation Year (12/31)"
    ws["B7"] = "=Inputs!B7"
    ws["A8"] = "Horizon Age"
    ws["B8"] = "=Inputs!B8"
    ws["A9"] = "Spread"
    ws["B9"] = "=Inputs!B9"

    ws["D3"] = "Expense Assumptions"
    ws["D3"].font = Font(bold=True)
    ws["D4"] = "Policy Expense ($)"
    ws["E4"] = "=Inputs!B10"
    ws["D5"] = "Premium Expense Rate (decimal)"
    ws["E5"] = "=Inputs!B11"
    ws["D6"] = "Monthly Expense ($)"
    ws["E6"] = "=Inputs!B12"

    ws["A11"] = "Pricing & Reserve Summary"
    ws["A11"].font = Font(bold=True)
    ws["A12"] = "Single Premium"
    ws["B12"] = "=Projection!X8"
    ws["A13"] = "PV Benefits"
    ws["B13"] = "=Projection!X4"
    ws["A14"] = "PV Monthly Expenses"
    ws["B14"] = "=Projection!X5"
    ws["A15"] = "PV Monthly Total"
    ws["B15"] = "=Projection!X7"
    ws["A16"] = "Annuity Factor"
    ws["B16"] = "=Projection!X6"
    ws["A17"] = "Economic Reserve at t=0"
    ws["B17"] = "=Projection!V2"

    for cell in ("B6", "E4", "E6", "B12", "B13", "B14", "B15", "B17"):
        ws[cell].number_format = "#,##0.00"
    ws["E5"].number_format = "0.00%"
    ws["B16"].number_format = "0.000000"

    proj_start = 4
    proj_end = proj_start + n_months - 1

    chart_surv = LineChart()
    chart_surv.title = "Survival to Monthly Payment"
    chart_surv.y_axis.title = "P(alive)"
    chart_surv.x_axis.title = "Attained Age"
    data_surv = Reference(wb["Projection"], min_col=14, min_row=proj_start, max_row=proj_end)
    cats_age = Reference(wb["Projection"], min_col=3, min_row=proj_start, max_row=proj_end)
    chart_surv.add_data(data_surv, titles_from_data=False)
    chart_surv.set_categories(cats_age)
    chart_surv.height = 6
    chart_surv.width = 9
    ws.add_chart(chart_surv, "A20")

    chart_cf = LineChart()
    chart_cf.title = "Expected Monthly Cashflows"
    chart_cf.y_axis.title = "Expected Cashflow ($)"
    chart_cf.x_axis.title = "Attained Age"
    data_cf = Reference(wb["Projection"], min_col=17, max_col=18, min_row=3, max_row=proj_end)
    chart_cf.add_data(data_cf, titles_from_data=True)
    chart_cf.set_categories(cats_age)
    chart_cf.height = 6
    chart_cf.width = 9
    ws.add_chart(chart_cf, "J20")

    chart_res = LineChart()
    chart_res.title = "Economic Reserve"
    chart_res.y_axis.title = "Reserve ($)"
    chart_res.x_axis.title = "Attained Age"
    data_res = Reference(wb["Projection"], min_col=22, min_row=2, max_row=proj_end)
    cats_res_age = Reference(wb["Projection"], min_col=3, min_row=2, max_row=proj_end)
    chart_res.add_data(data_res, titles_from_data=False)
    chart_res.set_categories(cats_res_age)
    chart_res.height = 6
    chart_res.width = 18
    ws.add_chart(chart_res, "A36")

    ws["A54"] = "Checkpoint Metrics (Nearest Monthly Projection Point)"
    ws["A54"].font = Font(bold=True)
    cp_headers = ["Target Age", "Projection Row", "Survival", "Exp Benefit ($)", "Exp Expense ($)", "Reserve ($)"]
    for c, h in enumerate(cp_headers, start=1):
        ws.cell(row=55, column=c, value=h).font = Font(bold=True)

    checkpoint_ages = [65, 70, 75, 80, 85, 90, 100, 110]
    for i, age in enumerate(checkpoint_ages, start=56):
        ws[f"A{i}"] = age
        ws[f"B{i}"] = (
            f"=IFERROR(MATCH(A{i},Projection!$C$4:$C${3+n_months},0)+3,"
            f"IF(A{i}<INDEX(Projection!$C$4:$C${3+n_months},1),4,"
            f"MATCH(A{i},Projection!$C$4:$C${3+n_months},1)+3))"
        )
        ws[f"C{i}"] = f"=INDEX(Projection!$N:$N,B{i})"
        ws[f"D{i}"] = f"=INDEX(Projection!$Q:$Q,B{i})"
        ws[f"E{i}"] = f"=INDEX(Projection!$R:$R,B{i})"
        ws[f"F{i}"] = f"=INDEX(Projection!$V:$V,B{i})"

    for r in range(56, 56 + len(checkpoint_ages)):
        ws[f"C{r}"].number_format = "0.000000"
        ws[f"D{r}"].number_format = "#,##0.00"
        ws[f"E{r}"].number_format = "#,##0.00"
        ws[f"F{r}"].number_format = "#,##0.00"

    if alm_layout is not None:
        wsa = wb[ALM_SHEET_NAME]
        dr = alm_layout.first_data_row
        lr = alm_layout.last_data_row
        hr = alm_layout.header_row
        base = 66
        ws.cell(row=base, column=1, value="ALM summary (latest run — detail on ALM_Projection)").font = Font(bold=True)
        rows_meta = [
            ("Initial AUM ($)", f"={ALM_SHEET_NAME}!B3"),
            ("Funding ratio (month 1)", f"={ALM_SHEET_NAME}!G{dr}"),
            ("Min surplus ($)", f"=MIN({ALM_SHEET_NAME}!F{dr}:F{lr})"),
            ("Ending surplus ($)", f"={ALM_SHEET_NAME}!F{lr}"),
            ("Ending funding ratio", f"={ALM_SHEET_NAME}!G{lr}"),
            ("Duration gap (y)", f"={ALM_SHEET_NAME}!B4"),
            ("PV01 net ($/bp)", f"={ALM_SHEET_NAME}!B7"),
        ]
        for i, (lab, formula) in enumerate(rows_meta, start=1):
            rr = base + i
            ws.cell(row=rr, column=1, value=lab)
            ws.cell(row=rr, column=2, value=formula)
        ws[f"B{base + 1}"].number_format = "#,##0.00"
        ws[f"B{base + 2}"].number_format = "0.000"
        ws[f"B{base + 3}"].number_format = "#,##0.00"
        ws[f"B{base + 4}"].number_format = "#,##0.00"
        ws[f"B{base + 5}"].number_format = "0.000"
        ws[f"B{base + 6}"].number_format = "0.00"
        ws[f"B{base + 7}"].number_format = "#,##0.00"

        chart_am = LineChart()
        chart_am.title = "ALM — Asset MV and liability PV"
        chart_am.y_axis.title = "$"
        chart_am.x_axis.title = "Attained age"
        chart_am.add_data(Reference(wsa, min_col=3, max_col=4, min_row=hr, max_row=lr), titles_from_data=True)
        chart_am.set_categories(Reference(wsa, min_col=2, min_row=dr, max_row=lr))
        chart_am.height = 6
        chart_am.width = 10
        ws.add_chart(chart_am, f"A{base + 9}")

        chart_s = LineChart()
        chart_s.title = "ALM — Surplus"
        chart_s.y_axis.title = "Surplus ($)"
        chart_s.x_axis.title = "Attained age"
        chart_s.add_data(Reference(wsa, min_col=6, min_row=hr, max_row=lr), titles_from_data=True)
        chart_s.set_categories(Reference(wsa, min_col=2, min_row=dr, max_row=lr))
        chart_s.height = 6
        chart_s.width = 10
        ws.add_chart(chart_s, f"J{base + 9}")


def _write_model_check_sheet(
    wb: Workbook,
    snap: ExcelPythonSnapshot,
    *,
    alm_layout: ALMDashboardLayout | None = None,
    alm_snapshot: ALMExcelSnapshot | None = None,
) -> None:
    """Embed Python pricing outputs next to Projection summary formulas; optional ALM row checks vs ALM_Projection."""
    ws = wb.create_sheet("ModelCheck")
    ws["A1"] = "Python snapshot vs Excel (Projection; optional ALM_Projection)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "Column B is the Python snapshot at export. Column C points at workbook formulas or embedded ALM cells; "
        "column D should be ~0 after recalc if Inputs match the launcher. "
        "ALM: liability PV on ALM_Projection is formula-linked to Projection; surplus and funding use sheet formulas; "
        "asset column remains the Python ladder snapshot."
    )
    ws.merge_cells("A2:D2")

    headers = ("Metric", "Python snapshot", "Excel (formula)", "Difference (Excel − Python)")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)

    pricing_rows: list[tuple[str, float, str, str]] = [
        ("PV benefits", snap.pv_benefit, "=Projection!X4", "money"),
        ("PV monthly expenses", snap.pv_monthly_expenses, "=Projection!X5", "money"),
        ("PV monthly total (ben+exp)", snap.pv_monthly_total, "=Projection!X7", "money"),
        ("Single premium", snap.single_premium, "=Projection!X8", "money"),
        ("Annuity factor", snap.annuity_factor, "=Projection!X6", "factor"),
    ]
    row_idx = 5
    for label, val, xls_ref, kind in pricing_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=float(val))
        ws.cell(row=row_idx, column=3, value=xls_ref)
        ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}")
        if kind == "factor":
            ws.cell(row=row_idx, column=2).number_format = "0.000000"
            ws.cell(row=row_idx, column=4).number_format = "0.000000"
        else:
            ws.cell(row=row_idx, column=2).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
        row_idx += 1

    if (alm_layout is None) ^ (alm_snapshot is None):
        raise ValueError("alm_layout and alm_snapshot must both be set or both omitted.")

    if alm_layout is not None and alm_snapshot is not None:
        dr = alm_layout.first_data_row
        lr = alm_layout.last_data_row
        sh = ALM_SHEET_NAME
        n_path = int(alm_snapshot.asset_market_value.shape[0])
        if lr - dr + 1 != n_path:
            raise ValueError("ALM ModelCheck layout rows do not match snapshot length.")
        a_mv = np.asarray(alm_snapshot.asset_market_value, dtype=float)
        l_pv = np.asarray(alm_snapshot.liability_pv, dtype=float)
        debt_v = np.asarray(alm_snapshot.borrowing_balance, dtype=float)
        path_surp = a_mv - l_pv - debt_v
        a0, l0 = float(a_mv[0]), float(l_pv[0])
        d0 = float(debt_v[0])
        s0 = float(path_surp[0])
        s_end = float(path_surp[-1])
        s_min = float(np.min(path_surp))
        denom0 = l0 + d0
        f0 = float(a0 / denom0) if denom0 > 1e-12 else float("nan")

        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        ws.cell(row=row_idx, column=1, value="ALM checks (ALM_Projection sheet)").font = Font(bold=True)
        row_idx += 1

        min_rng = f"=MIN({sh}!F{dr}:F{lr})"
        alm_rows: list[tuple[str, float, str, str]] = [
            ("ALM initial AUM (meta)", float(alm_snapshot.initial_asset_market_value), f"={sh}!B3", "money"),
            ("ALM asset MV (month 1)", a0, f"={sh}!C{dr}", "money"),
            ("ALM liability PV (month 1)", l0, f"={sh}!D{dr}", "money"),
            ("ALM surplus (month 1)", s0, f"={sh}!F{dr}", "money"),
        ]
        if math.isfinite(f0):
            alm_rows.append(("ALM funding ratio (month 1)", f0, f"={sh}!G{dr}", "fr"))
        alm_rows.extend(
            [
                ("ALM surplus (final month)", s_end, f"={sh}!F{lr}", "money"),
                ("ALM min surplus (path)", s_min, min_rng, "money"),
                ("ALM duration gap (y)", float(alm_snapshot.duration_gap), f"={sh}!B4", "dur"),
                ("ALM PV01 net ($/bp)", float(alm_snapshot.pv01_net), f"={sh}!B7", "money"),
            ]
        )
        for label, val, xls_ref, kind in alm_rows:
            ws.cell(row=row_idx, column=1, value=label)
            b_cell = ws.cell(row=row_idx, column=2)
            if isinstance(val, float) and math.isfinite(val):
                b_cell.value = float(val)
            else:
                b_cell.value = ""
            ws.cell(row=row_idx, column=3, value=xls_ref)
            ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}")
            if kind == "factor":
                b_cell.number_format = "0.000000"
                ws.cell(row=row_idx, column=4).number_format = "0.000000"
            elif kind == "fr":
                b_cell.number_format = "0.000"
                ws.cell(row=row_idx, column=4).number_format = "0.000"
            elif kind == "dur":
                b_cell.number_format = "0.00"
                ws.cell(row=row_idx, column=4).number_format = "0.00"
            else:
                b_cell.number_format = "#,##0.00"
                ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
            row_idx += 1

    tr = row_idx + 1
    ws.cell(row=tr, column=1, value="Troubleshooting").font = Font(bold=True)
    ws.cell(
        row=tr + 1,
        column=1,
        value=(
            "A ~3–4% higher Excel PV vs column B usually means Inputs!B9 (spread) is negative or does not match "
            "the Streamlit run (e.g. spreadsheet edited after download). Confirm valuation year B7 and all Inputs. "
            "Large ALM differences mean ALM_Projection cells were edited or the workbook predates a new ALM run."
        ),
    )
    ws.merge_cells(f"A{tr + 1}:D{tr + 3}")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 26


def _write_python_tab(wb: Workbook) -> None:
    ws = wb.create_sheet("Python_Runbook")
    ws["A1"] = "Python / Launcher Runbook"
    ws["A1"].font = Font(bold=True, size=14)

    lines: list[tuple[str, str]] = [
        ("Purpose", "Regenerate inputs and this workbook from the Python SPIA scaffold."),
        ("Working folder", "annuity_model"),
        ("Streamlit UI", "streamlit run spia_ui.py — Run & results, then download Excel for the same assumptions."),
        ("", ""),
        ("Core commands", "From the annuity_model folder:"),
        ("1", "python spia_projection.py"),
        ("2", "python illustrate_spia_projection.py"),
        ("3", "python build_spia_excel_workbook.py"),
        ("", ""),
        ("Typical CSV inputs", "treasury_zero_rate_curve_latest.csv (or par curve + bootstrap in UI)"),
        ("", "rp2014_male_healthy_annuitant_qx_2014.csv / custom q_x CSV"),
        ("", "mp2016_male_improvement_rates.csv (long: age, year, improvement_rate)"),
        ("", "expenses_assumptions_us_placeholders.csv"),
        ("", "sp500_scenario_projection_monthly.csv (month, sp500_level for months 0..N)"),
        ("", ""),
        ("Outputs", "spia_projection_model.xlsx, illustrations/*.png"),
        ("", ""),
        ("Note", "This workbook is built for a fixed row count; change issue/horizon in the launcher and re-download."),
    ]

    row = 3
    for k, v in lines:
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        if k in {"Purpose", "Core commands", "Typical CSV inputs", "Outputs", "Note"}:
            ws[f"A{row}"].font = Font(bold=True)
        row += 1

    ws["A28"] = "spia_projection.py (excerpt)"
    ws["A28"].font = Font(bold=True)
    spia_text = _read_text(BASE_DIR / "spia_projection.py", fallback="spia_projection.py not found")
    excerpt = "\n".join(spia_text.splitlines()[:140])
    ws["A29"] = excerpt

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120


def build_workbook_from_spec(
    spec: ExcelBuildSpec,
    out_path: Path | None = None,
    *,
    python_snapshot: ExcelPythonSnapshot | None = None,
    mc_snapshot: MCExcelSnapshot | None = None,
    alm_snapshot: ALMExcelSnapshot | None = None,
) -> Path | bytes:
    """
    Write a workbook to `out_path`. If `out_path` is None, return the file as bytes (for downloads).

    Optional ``python_snapshot`` embeds the launcher run on the ModelCheck sheet for formula validation.
    Optional ``alm_snapshot`` adds ``ALM_Projection`` plus Dashboard ALM summaries from the latest ALM run.
    """
    yc = spec.yield_curve_df.copy()
    if "maturity_years" not in yc.columns or "zero_rate" not in yc.columns:
        raise ValueError("yield_curve_df must have columns maturity_years, zero_rate")

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    ws_inputs = wb.active
    _write_inputs(ws_inputs, spec)

    ws_yc = wb.create_sheet("YieldCurve")
    _write_simple_table(ws_yc, "Zero curve nodes (continuously compounded)", yc)

    ws_bq = wb.create_sheet("BaseQx")
    _write_simple_table(ws_bq, "Base annual q_x by integer age", spec.base_qx_df)

    ws_mp = wb.create_sheet("MP2016_Long")
    if spec.mp_improvement_long_df is not None and not spec.mp_improvement_long_df.empty:
        _write_simple_table(ws_mp, "MP-2016 improvement rates (long)", spec.mp_improvement_long_df)
    else:
        ws_mp["A1"] = "MP-2016 (not used for static mortality)"
        ws_mp["A1"].font = Font(bold=True, size=12)
        ws_mp.cell(row=3, column=1, value="age").font = Font(bold=True)
        ws_mp.cell(row=3, column=2, value="year").font = Font(bold=True)
        ws_mp.cell(row=3, column=3, value="improvement_rate").font = Font(bold=True)

    n_y = len(yc)
    y_last_row = 3 + n_y
    n_months = spec.n_months
    ws_monthly = wb.create_sheet("MonthlyCurve")
    _write_monthly_curve_logdf(ws_monthly, n_months=n_months, y_last_row=y_last_row)

    ws_idx = wb.create_sheet("IndexScenario")
    _write_simple_table(
        ws_idx,
        "Index scenario: S&P 500 proxy levels by payment month (illustrative unless replaced)",
        spec.index_scenario_df,
    )
    idx_last_row = 3 + len(spec.index_scenario_df)

    ws_proj = wb.create_sheet("Projection")
    _write_projection(ws_proj, n_months=n_months, y_last_row=y_last_row, idx_last_row=idx_last_row)

    alm_layout: ALMDashboardLayout | None = None
    if alm_snapshot is not None:
        alm_layout = _write_alm_projection_sheet(wb, alm_snapshot)

    if python_snapshot is not None:
        _write_model_check_sheet(
            wb,
            python_snapshot,
            alm_layout=alm_layout,
            alm_snapshot=alm_snapshot,
        )

    if mc_snapshot is not None:
        _write_mc_summary_sheet(wb, mc_snapshot)

    _write_dashboard(wb, n_months=n_months, alm_layout=alm_layout)
    _write_python_tab(wb)

    if out_path is None:
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    out_path = Path(out_path)
    wb.save(out_path)
    return out_path


def _load_default_csv_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    yc = pd.read_csv(BASE_DIR / "treasury_zero_rate_curve_latest.csv")
    rp = pd.read_csv(BASE_DIR / "rp2014_male_healthy_annuitant_qx_2014.csv")
    mp = pd.read_csv(BASE_DIR / "mp2016_male_improvement_rates.csv")
    exp = pd.read_csv(BASE_DIR / "expenses_assumptions_us_placeholders.csv")
    return yc, rp, mp, exp


def _expenses_from_placeholder_csv(exp: pd.DataFrame) -> tuple[float, float, float]:
    key_to_val = {str(k): float(v) for k, v in zip(exp["key"], exp["value"])}
    key_to_unit = {str(k): str(u).lower().strip() for k, u in zip(exp["key"], exp["unit"])}
    pol = float(key_to_val.get("policy_expense_dollars", 0.0))
    monthly_ex = float(key_to_val.get("monthly_expense_dollars", 0.0))
    praw = float(key_to_val.get("premium_expense_rate", 0.0))
    u = key_to_unit.get("premium_expense_rate", "")
    if u in {"percent", "pct", "%"}:
        prem = praw / 100.0 if praw > 1.0 else praw
    else:
        prem = praw
    return pol, prem, monthly_ex


def build_workbook(out_path: Path | None = None) -> Path:
    """CLI/default build using on-disk CSVs (legacy entry point)."""
    yc, rp, mp, exp = _load_default_csv_data()
    pol, prem, monthly_ex = _expenses_from_placeholder_csv(exp)
    n_m = _n_months_for_contract(65, 110)
    scen_path = BASE_DIR / "sp500_scenario_projection_monthly.csv"
    try:
        s0, lv = sp.load_index_scenario_monthly_csv(str(scen_path), n_months=n_m)
    except FileNotFoundError:
        s0, lv = sp.flat_index_scenario(n_m)
    idx_rows = [{"month": 0, "sp500_level": float(s0)}]
    for k in range(1, n_m + 1):
        idx_rows.append({"month": int(k), "sp500_level": float(lv[k - 1])})
    idx_df = pd.DataFrame(idx_rows)

    spec = ExcelBuildSpec(
        issue_age=65,
        sex="male",
        benefit_annual=100_000.0,
        payment_freq_per_year=12,
        valuation_year=2025,
        horizon_age=110,
        spread=0.0,
        yield_curve_df=yc,
        mortality_excel_mode="rp_mp",
        base_qx_df=rp.rename(columns={c: c.strip() for c in rp.columns}),
        mp_improvement_long_df=mp,
        policy_expense_dollars=pol,
        premium_expense_rate=prem,
        monthly_expense_dollars=monthly_ex,
        yield_mode_label="zero_csv (CLI default)",
        mortality_mode_label="rp2014_mp2016 (CLI default)",
        expense_mode_label="csv (CLI default)",
        n_months=n_m,
        expense_annual_inflation=0.025,
        index_scenario_df=idx_df,
    )
    target = Path(out_path) if out_path is not None else OUT_XLSX
    result = build_workbook_from_spec(spec, out_path=target)
    assert isinstance(result, Path)
    return result


if __name__ == "__main__":
    out = build_workbook()
    print(f"Created workbook: {out.resolve()}")
