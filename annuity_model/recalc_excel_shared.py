"""
Shared Excel pieces for product **recalculation** workbooks (formula grids auditors can edit).

New products should reuse these sheet names and patterns so tooling, docs, and parity tests stay aligned:
``Inputs``; ``YieldCurve`` + ``MonthlyCurve`` for discount factors; ``Liabilities`` for the monthly
cashflow grid; optional ``ALM_Engine`` / ``ALM_Projection``; ``ModelCheck`` for Python snapshot vs formulas.
"""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

# Canonical sheet names (match SPIA / Term recalc exports).
RECALC_INPUTS_SHEET = "Inputs"
RECALC_YIELD_CURVE_SHEET = "YieldCurve"
RECALC_MONTHLY_CURVE_SHEET = "MonthlyCurve"
RECALC_LIABILITIES_SHEET = "Liabilities"


def yield_curve_sheet_name() -> str:
    """Canonical yield node sheet name (ALM ladder formulas reference this tab)."""
    return RECALC_YIELD_CURVE_SHEET


def write_simple_curve_df(ws, title: str, df: pd.DataFrame) -> None:
    """Same layout as SPIA ``YieldCurve``: title row 1, header row 3, data row 4+."""
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    for c, col in enumerate(df.columns, start=1):
        ws.cell(row=3, column=c, value=col).font = Font(bold=True)
    for r, row in enumerate(df.itertuples(index=False), start=4):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=float(val))


def write_yield_curve_sheet(wb: Workbook, df: pd.DataFrame) -> tuple[str, int]:
    """
    Append the yield curve sheet. Returns ``(sheet_name, y_last_row)`` where ``y_last_row``
    is the last row index containing a curve node (for ``MonthlyCurve`` formulas).
    """
    name = yield_curve_sheet_name()
    ws = wb.create_sheet(name)
    write_simple_curve_df(ws, "Zero curve nodes (continuously compounded)", df)
    y_last_row = 3 + len(df)
    return name, int(y_last_row)


def write_monthly_curve_logdf(ws, n_months: int, y_last_row: int) -> None:
    """
    Monthly discount factors consistent with Python ``YieldCurve.discount_factors``:
    log-linear interpolation on DF between curve nodes; flat zero-rate extrapolation beyond endpoints.

    Expects ``Inputs!$B$6`` = payments per year and ``Inputs!$B$9`` = spread (same convention as SPIA ALM).
    """
    ws.title = RECALC_MONTHLY_CURVE_SHEET
    ws["A1"] = "Monthly Discount Factors (log-linear on DF)"
    ws["A1"].font = Font(bold=True, size=12)

    headers = [
        "Month",
        "t_years",
        "BracketIndex",
        "LowerMat",
        "UpperMat",
        "LowerZero",
        "UpperZero",
        "InterpWeight",
        "LogDF_lower_node",
        "LogDF_upper_node",
        "LogDF_t",
        "DiscountFactor",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)

    first = 4
    last = first + n_months - 1
    yc = yield_curve_sheet_name()
    y_rng = f"{yc}!$A$4:$A${y_last_row}"
    z_rng = f"{yc}!$B$4:$B${y_last_row}"

    for r in range(first, last + 1):
        ws[f"A{r}"] = r - first + 1
        ws[f"B{r}"] = f"=A{r}/Inputs!$B$6"
        ws[f"C{r}"] = (
            f"=IF(B{r}<=INDEX({y_rng},1),1,"
            f"IF(B{r}>=INDEX({y_rng},ROWS({y_rng})),ROWS({y_rng})-1,"
            f"MATCH(B{r},{y_rng},1)))"
        )
        ws[f"D{r}"] = f"=INDEX({y_rng},C{r})"
        ws[f"E{r}"] = f"=INDEX({y_rng},C{r}+1)"
        ws[f"F{r}"] = f"=INDEX({z_rng},C{r})"
        ws[f"G{r}"] = f"=INDEX({z_rng},C{r}+1)"
        ws[f"H{r}"] = f"=IF(E{r}=D{r},0,(B{r}-D{r})/(E{r}-D{r}))"
        ws[f"I{r}"] = f"=-(F{r}+Inputs!$B$9)*D{r}"
        ws[f"J{r}"] = f"=-(G{r}+Inputs!$B$9)*E{r}"
        ws[f"K{r}"] = (
            f"=IF(B{r}<=INDEX({y_rng},1),-(INDEX({z_rng},1)+Inputs!$B$9)*B{r},"
            f"IF(B{r}>=INDEX({y_rng},ROWS({y_rng})),-(INDEX({z_rng},ROWS({y_rng}))+Inputs!$B$9)*B{r},"
            f"I{r}+H{r}*(J{r}-I{r})))"
        )
        ws[f"L{r}"] = f"=EXP(K{r})"
