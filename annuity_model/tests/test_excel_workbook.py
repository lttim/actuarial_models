"""Smoke tests for Excel workbook builder."""

from __future__ import annotations

import io

import numpy as np
import pytest
from openpyxl import load_workbook

import pricing_projection as sp
from alm_excel_ladder import ALM_ENGINE_SHEET

from build_pricing_excel_workbook import (
    ALM_ENGINE_FIELD_GUIDE_SHEET,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_ENGINE_STEP_MONTHS,
    ALM_SHEET_NAME,
    ExcelPythonSnapshot,
    LIABILITY_SHEET_NAME,
    alm_excel_downsample_snapshot,
    alm_excel_period_end_indices,
    alm_excel_snapshot_from_result,
    alm_excel_truncate_snapshot,
    build_workbook_from_spec,
    excel_spec_from_launcher,
)


def test_alm_excel_period_end_indices():
    assert alm_excel_period_end_indices(10, 3) == [2, 5, 8, 9]
    assert alm_excel_period_end_indices(1, 3) == [0]
    assert alm_excel_period_end_indices(6, 1) == [0, 1, 2, 3, 4, 5]


def test_model_check_sheet_embeds_python_snapshot():
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=100_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.02, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_scenario_csv_path=None,
        expense_annual_inflation=0.0,
    )
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    snap = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    raw = build_workbook_from_spec(spec, out_path=None, python_snapshot=snap)
    wb = load_workbook(io.BytesIO(raw))
    assert "ModelCheck" in wb.sheetnames
    mc = wb["ModelCheck"]
    assert mc["B5"].value == pytest.approx(res.pv_benefit, rel=1e-9)
    assert mc["B9"].value == pytest.approx(res.annuity_factor, rel=1e-9)
    assert mc["C5"].value == f"={LIABILITY_SHEET_NAME}!X4"


def test_alm_projection_sheet_and_dashboard_links():
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=100_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.02, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_scenario_csv_path=None,
        expense_annual_inflation=0.0,
    )
    asm = sp.ALMAssumptions(
        allocation=sp.alm_default_allocation_spec(),
        rebalance_band=0.10,
        rebalance_frequency_months=1,
        reinvest_rule="hold_cash",
        disinvest_rule="shortest_first",
        rebalance_policy="liquidity_only",
        liquidity_near_liquid_years=0.25,
    )
    alm = sp.run_alm_projection(
        pricing=res,
        yield_curve=yc,
        spread=0.0,
        assumptions=asm,
        initial_asset_market_value=float(res.single_premium),
    )
    alm_snap = alm_excel_snapshot_from_result(
        alm, asm, initial_asset_market_value=float(res.single_premium)
    )
    alm_snap_ds = alm_excel_truncate_snapshot(
        alm_excel_downsample_snapshot(alm_snap, ALM_ENGINE_STEP_MONTHS),
        ALM_EXCEL_PATH_MONTH_CAP,
    )
    n_m = int(res.months.size)
    for m in (0, min(5, n_m - 1), n_m - 1):
        L_alt = sp.liability_pv_after_paid_months(res, yc, 0.0, m)
        assert L_alt == pytest.approx(float(alm.liability_pv[m]), rel=1e-9, abs=1e-6)
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    snap = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    raw = build_workbook_from_spec(
        spec, out_path=None, python_snapshot=snap, alm_snapshot=alm_snap, alm_assumptions=asm
    )
    wb = load_workbook(io.BytesIO(raw))
    max_formula_len = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    max_formula_len = max(max_formula_len, len(v))
    assert max_formula_len <= 8192, (
        f"Excel rejects formulas over 8192 chars (longest was {max_formula_len})"
    )
    wb_cached = load_workbook(io.BytesIO(raw), data_only=True)
    wac = wb_cached[ALM_SHEET_NAME]
    assert wac["C13"].value is not None
    assert wac["D13"].value is not None
    assert wac["E13"].value is not None
    assert wac["F13"].value is not None
    assert wac["H13"].value is not None
    assert float(wac["C13"].value) == pytest.approx(
        sum(float(wac[f"{chr(ord('H')+k)}13"].value) for k in range(6)),
        rel=1e-9,
        abs=1e-2,
    )
    assert float(wac["F13"].value) == pytest.approx(
        float(alm_snap_ds.asset_market_value[0])
        - float(alm_snap_ds.liability_pv[0])
        - float(alm_snap_ds.borrowing_balance[0]),
        rel=1e-9,
        abs=1e-3,
    )
    assert ALM_SHEET_NAME in wb.sheetnames
    assert ALM_ENGINE_SHEET in wb.sheetnames
    assert ALM_ENGINE_FIELD_GUIDE_SHEET in wb.sheetnames
    assert LIABILITY_SHEET_NAME in wb.sheetnames
    ws_alm = wb[ALM_SHEET_NAME]
    dr = 13
    assert int(ws_alm[f"A{dr}"].value) == int(alm_snap_ds.month_index[0]) + 1
    n_full = int(alm_snap.asset_market_value.shape[0])
    last_row = dr + int(alm_snap_ds.asset_market_value.shape[0]) - 1
    assert int(ws_alm[f"A{last_row}"].value) == int(alm_snap_ds.month_index[-1]) + 1
    assert int(alm_snap_ds.asset_market_value.shape[0]) == min(ALM_EXCEL_PATH_MONTH_CAP, n_full)
    assert isinstance(ws_alm[f"C{dr}"].value, str) and str(ws_alm[f"C{dr}"].value).startswith("=SUM(")
    assert isinstance(ws_alm[f"H{dr}"].value, str) and str(ws_alm[f"H{dr}"].value).startswith(f"={ALM_ENGINE_SHEET}!")
    assert isinstance(ws_alm[f"D{dr}"].value, str) and str(ws_alm[f"D{dr}"].value).startswith("=")
    assert isinstance(ws_alm[f"F{dr}"].value, str) and str(ws_alm[f"F{dr}"].value).startswith("=")
    assert isinstance(ws_alm[f"G{dr}"].value, str) and str(ws_alm[f"G{dr}"].value).startswith("=")
    dash = wb["Dashboard"]
    assert dash["B67"].value == f"={ALM_SHEET_NAME}!B3"

    mc = wb["ModelCheck"]
    assert mc["A10"].value == "ALM checks (ALM_Projection sheet)"
    assert mc["B11"].value == pytest.approx(float(alm_snap.initial_asset_market_value), rel=1e-9)
    assert mc["B12"].value == pytest.approx(float(alm_snap_ds.asset_market_value[0]), rel=1e-9)
    assert mc["C12"].value == f"={ALM_SHEET_NAME}!C13"
    assert mc["D12"].value == "=C12-B12"
