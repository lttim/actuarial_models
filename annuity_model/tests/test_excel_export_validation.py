"""
End-to-end Excel export validation gate.

These tests build a complete workbook for **every implemented product** (SPIA,
Term Life, RILA), open it back with openpyxl, and run the static formula
validator from :mod:`excel_workbook_validator`. They guarantee that no exported
workbook ships with a formula Excel would flag with
"Removed Records: Formula from /xl/worksheets/sheetN.xml part".

Whenever a workbook builder is changed (new sheet, new formula pattern), these
tests re-validate every cell in the produced workbook against the function
arity table in :mod:`excel_workbook_validator`. If the change introduces a
malformed formula the test fails with the offending sheet, cell and formula.

Targeted unit tests for the validator itself live further below.
"""

from __future__ import annotations

import io

import numpy as np
import pytest
from openpyxl import load_workbook

import iul_projection as iul
import pricing_projection as sp
import rila_projection as rp
import term_projection as tp
from build_pricing_excel_workbook import (
    ExcelPythonSnapshot,
    alm_excel_snapshot_from_result,
    build_workbook_from_spec,
    excel_spec_from_launcher,
)
from build_rila_excel_workbook import (
    build_rila_workbook_from_spec,
    rila_excel_spec_from_launcher,
)
from build_iul_excel_workbook import (
    build_iul_workbook_from_spec,
    iul_excel_spec_from_launcher,
)
from build_term_excel_workbook import (
    build_term_workbook_from_spec,
    term_excel_spec_from_launcher,
)
from excel_workbook_validator import (
    ExcelWorkbookValidationError,
    FormulaIssue,
    validate_formula,
    validate_workbook,
    validate_workbook_or_raise,
)


def _validate_xlsx_bytes(raw: bytes) -> None:
    """End-to-end gate: workbook must pass strict-mode validation.

    Strict mode (added 2026-04) additionally rejects calls to Excel
    functions that are neither registered in
    ``excel_workbook_validator.FUNCTION_ARITIES`` nor explicitly allow-listed
    in ``_STRICT_MODE_ALLOWED_UNREGISTERED``. This catches typos like
    ``AVERGE`` (Excel will repair to ``#NAME?`` silently) and unintentional
    new built-in dependencies before they reach a builder. The non-strict
    code path remains available for partial / WIP workbooks where the
    caller knowingly uses a built-in we have not enumerated yet.
    """
    wb = load_workbook(io.BytesIO(raw), data_only=False)
    issues = validate_workbook(wb, strict=True)
    assert issues == [], (
        "Workbook failed strict-mode validation. Excel may flag with "
        "'Removed Records: Formula' on load, OR you used a function not yet "
        "registered in excel_workbook_validator.FUNCTION_ARITIES. "
        "If the function is intentional, register its arity (or add to "
        "_STRICT_MODE_ALLOWED_UNREGISTERED if it doesn't need an arity check). "
        "Issues:\n"
        + "\n".join(f"  - {iss}" for iss in issues[:25])
        + ("" if len(issues) <= 25 else f"\n  ... and {len(issues) - 25} more.")
    )


def test_spia_workbook_passes_excel_formula_validation():
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=100_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.02, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_scenario_csv_path=None,
        expense_annual_inflation=0.0,
    )
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    snap = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    raw = build_workbook_from_spec(spec, out_path=None, python_snapshot=snap)
    _validate_xlsx_bytes(raw)


def test_term_workbook_passes_excel_formula_validation():
    contract = tp.TermLifeContract(
        issue_age=40,
        sex="male",
        death_benefit=250_000.0,
        monthly_premium=200.0,
        term_years=20,
        premium_mode="level_monthly",
        benefit_timing="eoy_death",
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=60,
        spread=0.0,
        valuation_year=None,
    )
    spec = term_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=60,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
    )
    _ = res  # Consume to silence linter for value not used post-build
    raw = build_term_workbook_from_spec(spec)
    _validate_xlsx_bytes(raw)


def test_rila_workbook_passes_excel_formula_validation():
    contract = rp.RILAContract(
        issue_age=55,
        sex="male",
        participation=0.85,
        cap=0.09,
        floor=-0.02,
        rider_fee_annual=0.008,
    )
    yc = sp.YieldCurve.from_flat_rate(0.035)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    n_months = int(round((90 - contract.issue_age) * 12))
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.004, 0.02, size=n_months))
    res = rp.price_rila_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.01,
    )
    spec = rila_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.01,
    )
    raw = build_rila_workbook_from_spec(spec)
    _validate_xlsx_bytes(raw)


def test_rila_long_horizon_workbook_passes_excel_formula_validation():
    """The original repro: issue 65 -> horizon 110 (540 months); was producing repaired files."""
    contract = rp.RILAContract(
        issue_age=65,
        sex="male",
        participation=1.0,
        cap=0.10,
        floor=0.0,
        rider_fee_annual=0.01,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.01, 25.0)
    n_months = int(round((110 - contract.issue_age) * 12))
    levels = np.full(n_months, 100.0, dtype=float)
    res = rp.price_rila_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=110,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = rila_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=110,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    raw = build_rila_workbook_from_spec(spec)
    _validate_xlsx_bytes(raw)


def test_iul_workbook_passes_excel_formula_validation():
    contract = iul.IULContract(
        issue_age=45,
        sex="male",
        face_amount=250_000.0,
        single_premium=25_000.0,
        premium_load_pct=0.06,
        monthly_expense_charge=7.50,
        participation=1.0,
        cap=0.10,
        floor=0.0,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    n_months = (80 - contract.issue_age) * 12
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    res = iul.price_iul_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = iul_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    raw = build_iul_workbook_from_spec(spec)
    _validate_xlsx_bytes(raw)


# -----------------------------------------------------------------------------
# Unit tests for the validator itself
# -----------------------------------------------------------------------------


def test_validator_flags_if_with_two_args():
    issues = validate_formula("Sheet1", "A1", '=IF(A1="",IFERROR(B1,0))')
    assert any("IF called with 2 arg(s)" in i.message for i in issues)


def test_validator_accepts_proper_if_with_three_args():
    assert validate_formula("Sheet1", "A1", '=IF(A1="","",B1)') == []


def test_validator_flags_unbalanced_parens():
    issues = validate_formula("Sheet1", "A1", "=SUM(B1:B5")
    assert any("unbalanced" in i.message for i in issues)


def test_validator_flags_unbalanced_quotes():
    issues = validate_formula("Sheet1", "A1", '=IF(A1="oops,B1,C1)')
    assert any("unterminated string" in i.message or "unbalanced" in i.message for i in issues)


def test_validator_flags_excel_error_literal_inside_formula():
    issues = validate_formula("Sheet1", "A1", "=IF(A1=#REF!,1,0)")
    assert any("#REF!" in i.message for i in issues)


def test_validator_ignores_nonformula_strings():
    assert validate_formula("Sheet1", "A1", "Hello, world") == []
    assert validate_formula("Sheet1", "A1", "") == []


def test_validator_ignores_unknown_function_names():
    """Unknown function names must not produce false positives (whitelist behaviour)."""
    assert validate_formula("Sheet1", "A1", "=MY_CUSTOM(A1, B1, C1, D1, E1, F1)") == []


def test_validator_handles_strings_with_commas():
    """Commas inside string literals must not be counted as argument separators."""
    assert validate_formula("Sheet1", "A1", '=IF(A1=", , ,",1,0)') == []


def test_validator_flags_iferror_with_one_arg():
    issues = validate_formula("Sheet1", "A1", "=IFERROR(VLOOKUP(A1,B:C,2,FALSE))")
    assert any("IFERROR called with 1 arg" in i.message for i in issues)


def test_validator_workbook_raises_on_bad_formulas():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = '=IF(B1="",C1)'  # missing else
    with pytest.raises(ExcelWorkbookValidationError) as excinfo:
        validate_workbook_or_raise(wb)
    assert "Removed Records" in str(excinfo.value)
    assert any(isinstance(i, FormulaIssue) for i in excinfo.value.issues)


def test_validator_workbook_passes_on_clean_workbook():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = '=IF(A1=0,"zero","nonzero")'
    ws["C1"] = "=SUM(A1:A10)"
    ws["D1"] = "=IFERROR(VLOOKUP(A1,A:B,2,FALSE),0)"
    validate_workbook_or_raise(wb)


def test_strict_mode_flags_unknown_excel_function():
    """Strict mode rejects calls to unregistered functions (typo guard)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    ws["B1"] = "=AVERGE(A1:A3)"

    issues_lax = validate_workbook(wb, strict=False)
    assert issues_lax == [], (
        "non-strict mode must remain quiet on unknown functions; got: " + repr(issues_lax)
    )
    issues_strict = validate_workbook(wb, strict=True)
    assert any("AVERGE" in i.message for i in issues_strict), (
        "strict mode must flag the typo'd function name; got: " + repr(issues_strict)
    )


def test_strict_mode_or_raise_propagates_unknown_function():
    """``validate_workbook_or_raise(..., strict=True)`` raises on unknown funcs."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=BOGUSFN(1, 2)"
    validate_workbook_or_raise(wb, strict=False)
    with pytest.raises(ExcelWorkbookValidationError) as excinfo:
        validate_workbook_or_raise(wb, strict=True)
    assert "BOGUSFN" in str(excinfo.value)


def test_strict_mode_accepts_all_registered_functions():
    """Strict mode is not allowed to flag a function that *is* registered."""
    import openpyxl

    from excel_workbook_validator import FUNCTION_ARITIES

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = '=IF(A1=0,"zero","nonzero")'
    ws["C1"] = "=SUM(A1:A10)"
    ws["D1"] = "=IFERROR(VLOOKUP(A1,A:B,2,FALSE),0)"
    ws["E1"] = "=SUMPRODUCT(A1:A3,A1:A3)"
    issues = validate_workbook(wb, strict=True)
    assert issues == [], (
        "strict mode flagged a registered function; FUNCTION_ARITIES has "
        f"{len(FUNCTION_ARITIES)} entries. Issues: {issues!r}"
    )


def test_strict_mode_template_cache_isolated_from_lax():
    """Strict and lax results must not bleed through the formula-template cache."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=NEWFANGLED(1, 2)"
    lax = validate_workbook(wb, strict=False)
    strict = validate_workbook(wb, strict=True)
    assert lax == []
    assert any("NEWFANGLED" in i.message for i in strict)
    lax_again = validate_workbook(wb, strict=False)
    assert lax_again == [], (
        "lax mode must stay clean even after strict mode populated the cache; "
        "got: " + repr(lax_again)
    )


def test_validator_flags_trailing_empty_arg():
    """The 'f-string lost its substitution' bug pattern: ``IF(a, b, )``."""
    issues = validate_formula("Sheet1", "A1", "=IF(A1=1,B1,)")
    assert any("implicit empty argument" in i.message for i in issues)
    issues2 = validate_formula("Sheet1", "A1", "=IFERROR(B1,)")
    assert any("implicit empty argument" in i.message for i in issues2)


def test_validator_accepts_explicit_empty_string_arg():
    """``IFERROR(x, "")`` is intentional and must NOT be flagged."""
    assert validate_formula("Sheet1", "A1", '=IFERROR(VLOOKUP(A1,B:C,2,FALSE),"")') == []
    assert validate_formula("Sheet1", "A1", '=IF(B1>0,B1,"")') == []


def test_validator_flags_cross_sheet_reference_to_empty_column():
    """
    Regression for the RILA ALM 'not reconciling' bug.

    A formula that references ``Other!S:S`` while sheet ``Other`` has no data in
    column S must be flagged. Excel silently coerces the missing column to zero
    in SUMPRODUCT/INDEX, so without this check ModelCheck columns C and D drift
    from the Python snapshot only after the user opens the file.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    main = wb.active
    main.title = "Main"
    other = wb.create_sheet("Other")
    other["A1"] = "Month"
    other["A2"] = 1
    other["O1"] = "Discount"
    other["O2"] = 0.99
    main["A1"] = "=SUMPRODUCT(Other!$S:$S,Other!$O:$O)"
    issues = validate_workbook(wb)
    assert any("Other" in i.message and "S" in i.message for i in issues), (
        "validator must flag the empty cross-sheet column reference; got: " + repr(issues)
    )


def test_validator_flags_cross_sheet_indirect_reference_to_empty_column():
    """The same check must see through string literals consumed by INDIRECT(...)."""
    import openpyxl

    wb = openpyxl.Workbook()
    main = wb.active
    main.title = "Main"
    other = wb.create_sheet("Other")
    other["A1"] = "Month"
    other["A2"] = 1
    other["M1"] = "ExpTotalCF"
    other["M2"] = 100.0
    main["A1"] = '=SUMPRODUCT(INDIRECT("Other!S2:S5"),INDIRECT("Other!O2:O5"))'
    issues = validate_workbook(wb)
    msgs = " | ".join(i.message for i in issues)
    assert "S" in msgs and "Other" in msgs, (
        "validator must catch INDIRECT-built references to a missing column; got: " + repr(issues)
    )


def test_validator_accepts_cross_sheet_reference_to_populated_column():
    """No false positive when the referenced column actually has data."""
    import openpyxl

    wb = openpyxl.Workbook()
    main = wb.active
    main.title = "Main"
    other = wb.create_sheet("Other")
    other["A1"] = "Month"
    other["A2"] = 1
    other["S1"] = "ExpTotalCF"
    other["S2"] = 100.0
    main["A1"] = "=SUMPRODUCT(Other!$S:$S,Other!$A:$A)"
    main["A2"] = '=SUMPRODUCT(INDIRECT("Other!S2:S5"),INDIRECT("Other!A2:A5"))'
    assert validate_workbook(wb) == []


def test_rila_alm_projection_uses_liability_column_m_not_s():
    """
    Regression for the original user-reported defect.

    RILA's Liabilities sheet places ``ExpTotalCF`` in column M, not S. If any
    ALM_Projection or ALM_Engine formula references ``Liabilities!S`` (the SPIA
    layout) the workbook will open in Excel but reconcile to zero / produce
    large differences in ModelCheck. This test rebuilds the RILA workbook with
    ALM and asserts that no formula references ``Liabilities!S`` while at least
    one references ``Liabilities!M``.
    """
    contract = rp.RILAContract(
        issue_age=55,
        sex="male",
        participation=0.85,
        cap=0.09,
        floor=-0.02,
        rider_fee_annual=0.008,
    )
    yc = sp.YieldCurve.from_flat_rate(0.035)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    n_months = int(round((85 - contract.issue_age) * 12))
    rng = np.random.default_rng(0)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.004, 0.02, size=n_months))
    res = rp.price_rila_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=85,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.01,
    )
    asm = sp.ALMAssumptions(
        allocation=sp.alm_default_allocation_spec(),
        rebalance_band=0.05,
        rebalance_frequency_months=12,
        reinvest_rule="pro_rata",
        disinvest_rule="shortest_first",
        rebalance_policy="liquidity_only",
    )
    alm_result = sp.run_alm_projection_from_pricing_result(
        pricing=res,
        yield_curve=yc,
        spread=0.0,
        assumptions=asm,
        initial_asset_market_value=float(res.single_premium),
    )
    alm_snap = alm_excel_snapshot_from_result(
        alm_result, asm, initial_asset_market_value=float(res.single_premium)
    )
    spec = rila_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=85,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.01,
    )
    raw = build_rila_workbook_from_spec(
        spec,
        alm_assumptions=asm,
        alm_snapshot=alm_snap,
    )
    wb = load_workbook(io.BytesIO(raw), data_only=False)
    found_m = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                assert "Liabilities!S" not in v and "Liabilities!$S" not in v, (
                    f"RILA ALM formula still references Liabilities!S — would reconcile to "
                    f"zero. Sheet={ws.title!r} cell={c.coordinate} formula={v!r}"
                )
                if "Liabilities!M" in v or 'Liabilities!M"' in v:
                    found_m = True
    assert found_m, (
        "Expected at least one RILA ALM formula to reference Liabilities!M (ExpTotalCF) "
        "after the column-letter parameterization fix."
    )


def test_portfolio_workbook_passes_strict_validation() -> None:
    from pathlib import Path

    from build_portfolio_excel_workbook import build_portfolio_workbook_bytes
    from inforce_io import load_policy_inputs_from_csv
    from portfolio import Portfolio
    from portfolio_runner import run_portfolio

    csv_path = Path(__file__).resolve().parent / "data" / "inforce" / "example_v1" / "inforce.csv"
    policies = load_policy_inputs_from_csv(csv_path)
    from pricing_scenario_materialize import ANN_MODEL_ROOT, run_scenario_for_portfolio_policies

    sex = "female" if str(policies[0].contract.sex).lower() == "female" else "male"
    scen = run_scenario_for_portfolio_policies({}, policies, sex=sex, repo_root=ANN_MODEL_ROOT)
    res = run_portfolio(portfolio=Portfolio(policies=policies), scenario=scen)
    raw = build_portfolio_workbook_bytes(res)
    _validate_xlsx_bytes(raw)
