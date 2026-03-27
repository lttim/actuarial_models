from __future__ import annotations

import io

import numpy as np
import pytest
from openpyxl import load_workbook

import pricing_projection as sp
import term_projection as tp
from build_term_excel_workbook import build_term_workbook_from_spec, term_excel_spec_from_launcher
from build_pricing_excel_workbook import LIABILITY_SHEET_NAME


pytestmark = [pytest.mark.parity, pytest.mark.product_term]


def _setup_case() -> tuple[tp.TermLifeContract, sp.YieldCurve, sp.MortalityTableQx]:
    contract = tp.TermLifeContract(issue_age=45, sex="male", death_benefit=250_000.0, monthly_premium=250.0, term_years=20)
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.01, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    return contract, yc, mort


def test_term_eoy_claims_only_on_policy_year_end_months():
    contract, yc, mort = _setup_case()
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=None,
    )
    non_zero_months = res.months[np.asarray(res.expected_claim_cashflows) > 0.0]
    assert non_zero_months.size > 0
    assert np.all((non_zero_months % 12) == 0)


def test_term_liability_path_drives_alm_without_shape_drift():
    contract, yc, mort = _setup_case()
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=None,
    )
    asm = sp.ALMAssumptions(
        allocation=sp.alm_default_allocation_spec(),
        rebalance_band=0.10,
        rebalance_frequency_months=1,
        reinvest_rule="hold_cash",
        disinvest_rule="shortest_first",
        rebalance_policy="liquidity_only",
        liquidity_near_liquid_years=0.25,
    )
    path = tp.liability_path_from_term_projection(res)
    alm = sp.run_alm_projection_from_liability_path(
        liability_path=path,
        yield_curve=yc,
        spread=0.0,
        assumptions=asm,
        initial_asset_market_value=500_000.0,
    )
    assert alm.asset_market_value.shape == res.expected_total_cashflows.shape
    assert alm.liability_pv.shape == res.expected_total_cashflows.shape
    assert np.isfinite(alm.surplus).all()


def test_term_workbook_modelcheck_reconciles_zero_difference():
    contract, yc, mort = _setup_case()
    spec = term_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
    )
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=None,
    )
    xlsx = build_term_workbook_from_spec(spec)
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    ws_mc = wb["ModelCheck"]
    ws_liab = wb[LIABILITY_SHEET_NAME]
    assert ws_mc["C5"].value == f"={LIABILITY_SHEET_NAME}!X4"
    assert ws_mc["C6"].value == f"={LIABILITY_SHEET_NAME}!X5"
    assert ws_mc["C7"].value == f"={LIABILITY_SHEET_NAME}!X7"
    assert ws_mc["D5"].value == "=C5-B5"
    assert ws_mc["D6"].value == "=C6-B6"
    assert ws_mc["D7"].value == "=C7-B7"
    for coord, needle in (
        ("A4", "=IF(ROW()-3>"),
        ("D4", "=IF(A4="),
        ("G4", "=IF(A4="),
        ("O4", "=IF(A4="),
        ("T4", "=IF(A4="),
    ):
        v = ws_liab[coord].value
        assert isinstance(v, str) and v.startswith("="), coord
        assert needle in v, (coord, v)
    ex_claims = float(np.sum(res.expected_claim_cashflows * res.discount_factors))
    ex_prem = float(np.sum(res.expected_premium_cashflows * res.discount_factors))
    ex_net = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    np.testing.assert_allclose(float(ws_mc["B5"].value), ex_claims, rtol=0.0, atol=1e-9)
    np.testing.assert_allclose(float(ws_mc["B6"].value), ex_prem, rtol=0.0, atol=1e-9)
    np.testing.assert_allclose(float(ws_mc["B7"].value), ex_net, rtol=0.0, atol=1e-9)
    np.testing.assert_allclose(ex_net, ex_claims - ex_prem, rtol=0.0, atol=1e-9)


def test_term_workbook_includes_alm_sheets_when_snapshot_passed():
    contract, yc, mort = _setup_case()
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=None,
    )
    asm = sp.ALMAssumptions(
        allocation=sp.alm_default_allocation_spec(),
        rebalance_band=0.10,
        rebalance_frequency_months=1,
        reinvest_rule="hold_cash",
        disinvest_rule="shortest_first",
        rebalance_policy="liquidity_only",
        liquidity_near_liquid_years=0.25,
    )
    from alm_excel_ladder import ALM_ENGINE_SHEET

    from build_pricing_excel_workbook import (
        ALM_ENGINE_FIELD_GUIDE_SHEET,
        ALM_SHEET_NAME,
        alm_excel_snapshot_from_result,
        alm_excel_downsample_snapshot,
        alm_excel_truncate_snapshot,
        ALM_ENGINE_STEP_MONTHS,
        ALM_EXCEL_PATH_MONTH_CAP,
    )

    aum0 = float(res.single_premium) + 100_000.0
    alm = sp.run_alm_projection_from_pricing_result(
        pricing=res,
        yield_curve=yc,
        spread=0.0,
        assumptions=asm,
        initial_asset_market_value=aum0,
    )
    alm_snap = alm_excel_snapshot_from_result(alm, asm, initial_asset_market_value=aum0)
    alm_snap = alm_excel_truncate_snapshot(
        alm_excel_downsample_snapshot(alm_snap, ALM_ENGINE_STEP_MONTHS),
        ALM_EXCEL_PATH_MONTH_CAP,
    )
    spec = term_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=90,
        spread=0.0,
        valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
    )
    raw = build_term_workbook_from_spec(spec, alm_snapshot=alm_snap, alm_assumptions=asm)
    wb = load_workbook(io.BytesIO(raw), data_only=False)
    assert ALM_SHEET_NAME in wb.sheetnames
    assert ALM_ENGINE_SHEET in wb.sheetnames
    assert ALM_ENGINE_FIELD_GUIDE_SHEET in wb.sheetnames
    mc = wb["ModelCheck"]
    assert "ALM checks" in str(mc["A10"].value)
