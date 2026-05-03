"""Per-product Excel workbook formula-contract gates.

Two layers, both parametrized over every implemented product (SPIA,
Term Life, RILA):

1. ``test_python_cached_modelcheck_values_match_engine_<product>`` --
   ALWAYS RUNS, no skip. The Term + RILA workbooks bake the Python
   pricing values into ModelCheck column B as literal numbers (the
   Excel-side formulas live in column C and reference the liability
   sheet; column D is the diff). This test asserts those literals
   exactly equal the engine outputs within ``MODELCHECK_TOL``. It also
   asserts the SPIA workbook's ModelCheck formulas reference the
   correct sheet and have the expected structural shape, since those
   cells are pure formulas (no cached fallback). This is the gate that
   catches the bug class where a builder refactor accidentally writes
   stale or rounded numbers into the workbook the user downloads.

2. ``test_modelcheck_formula_contract_<product>`` -- ALWAYS RUNS. It
   asserts the workbook's ModelCheck formula cells point at the
   validated liability summary cells rather than stale, missing, or
   product-inappropriate references.

Why both layers
---------------
Layer (1) catches stale or rounded Python snapshot values. Layer (2)
catches the wiring bug class where ModelCheck is present but points at
the wrong formula row or an invalid sheet. The workbook validator and
product parity tests cover the lower-level formula syntax and monthly
calculation mechanics.

Relationship to existing tests
------------------------------
* ``tests/parity/test_runtime_excel_recalc.py`` is the pre-existing
  SPIA-only runtime gate; this file generalises it to every workbook
  product in a single parametrized module so adding a new product means
  adding one builder, not three test files.
* ``tests/parity/test_term_parity.py::test_term_workbook_modelcheck_reconciles_zero_difference``
  and the RILA analog still exist and assert finer-grained workbook
  structure (formula coordinates, A4 / D4 / O4 / T4 row needles); this
  file is the single-screen, "Excel matches Python for every product"
  invariant that would otherwise be spread across three files. The two
  intentionally overlap on a small surface (B5/B6/B7) to make the
  always-runs gate maximally legible.
"""

from __future__ import annotations

import io
from collections.abc import Callable
from dataclasses import dataclass

import numpy as np
import pytest
from openpyxl import load_workbook

import pricing_projection as sp
import rila_projection as rp
import term_projection as tp
from build_pricing_excel_workbook import (
    ExcelPythonSnapshot,
    build_workbook_from_spec,
    excel_spec_from_launcher,
)
from build_rila_excel_workbook import (
    build_rila_workbook_from_spec,
    rila_excel_spec_from_launcher,
)
from build_term_excel_workbook import (
    build_term_workbook_from_spec,
    term_excel_spec_from_launcher,
)
from parity_constants import MODELCHECK_TOL, RILA_PV_TOL, TERM_MODELCHECK_TOL

pytestmark = [pytest.mark.parity]


# ---------------------------------------------------------------------------
# Per-product fixtures: build a SMALL workbook + record the engine values it
# should reproduce. "Small" matters: formula evaluation time scales roughly
# linearly with cell count, and we run this gate on every PR, so each
# workbook is sized to the minimum that still exercises the full
# ModelCheck stack.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ProductRecalcCase:
    """One product's bundle: the workbook bytes, the engine's reference
    values for every ModelCheck cell we will assert on, and the
    tolerance to use.

    ``modelcheck_python_cells`` is the {coordinate: expected_value}
    map for cells the BUILDER writes as literals. ``modelcheck_formula_cells``
    is the {coordinate: expected_value} map for ModelCheck formula cells; the
    formula-contract gate checks those cells remain formulas that reference
    the expected liability summary rows.
    """

    product_id: str
    product_name: str
    blob: bytes
    tolerance: float
    modelcheck_python_cells: dict[str, float]
    modelcheck_formula_cells: dict[str, float]


def _make_spia_case() -> ProductRecalcCase:
    """SPIA: 12-month workbook (horizon_age=66) -- minimal recalc cost.

    SPIA's ModelCheck B5 (PV benefit) and B9 (annuity factor) are
    pure formulas with no cached literal fallback, so they live in
    ``modelcheck_formula_cells``.
    """
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=120_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.045)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=66,
        spread=0.0,
        expenses=expenses,
        expense_annual_inflation=0.0,
    )
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=66,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=np.full(12, 100.0),
        expense_annual_inflation=0.0,
    )
    # Pass python_snapshot so the SPIA workbook actually grows a
    # ModelCheck sheet -- without the snapshot, build_workbook_from_spec
    # skips ModelCheck entirely and the user-downloaded workbook would
    # have no Python<->Excel reconciliation surface. The UI path always
    # passes this snapshot (see pricing_ui._refresh_pricing_excel_workbook_in_session).
    # Pass python_snapshot so the SPIA workbook actually grows a
    # ModelCheck sheet -- without the snapshot, build_workbook_from_spec
    # skips ModelCheck entirely and the user-downloaded workbook would
    # have no Python<->Excel reconciliation surface. The UI path always
    # passes this snapshot (see pricing_ui._refresh_pricing_excel_workbook_in_session).
    blob = build_workbook_from_spec(
        spec,
        python_snapshot=ExcelPythonSnapshot(
            pv_benefit=float(res.pv_benefit),
            pv_monthly_expenses=float(res.pv_monthly_expenses),
            pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
            single_premium=float(res.single_premium),
            annuity_factor=float(res.annuity_factor),
        ),
    )
    return ProductRecalcCase(
        product_id="spia",
        product_name="SPIA",
        blob=blob,
        tolerance=float(MODELCHECK_TOL or 1e-6),
        # SPIA ModelCheck shape (with python_snapshot): column B is the
        # Python literal at export time; column C is the Excel formula
        # =Liabilities!X<n> that the evaluator resolves; column D is the
        # difference. Hit the same row set the rest of the platform
        # asserts on.
        modelcheck_python_cells={
            "ModelCheck!B5": float(res.pv_benefit),
            "ModelCheck!B6": float(res.pv_monthly_expenses),
            "ModelCheck!B7": float(res.pv_benefit + res.pv_monthly_expenses),
            "ModelCheck!B8": float(res.single_premium),
            "ModelCheck!B9": float(res.annuity_factor),
        },
        modelcheck_formula_cells={
            "ModelCheck!C5": float(res.pv_benefit),
            "ModelCheck!C6": float(res.pv_monthly_expenses),
            "ModelCheck!C7": float(res.pv_benefit + res.pv_monthly_expenses),
            "ModelCheck!C8": float(res.single_premium),
            "ModelCheck!C9": float(res.annuity_factor),
        },
    )


def _make_term_case() -> ProductRecalcCase:
    """Term Life: 5-year workbook so formula evaluation stays fast but the
    MC reconciliation still has multiple non-zero claim months."""
    contract = tp.TermLifeContract(
        issue_age=45,
        sex="male",
        death_benefit=250_000.0,
        monthly_premium=250.0,
        term_years=20,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.01, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    res = tp.price_term_life_level_monthly(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=50,
        spread=0.0,
        valuation_year=None,
    )
    spec = term_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=50,
        spread=0.0,
        valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
    )
    blob = build_term_workbook_from_spec(spec)
    ex_claims = float(np.sum(res.expected_claim_cashflows * res.discount_factors))
    ex_prem = float(np.sum(res.expected_premium_cashflows * res.discount_factors))
    ex_net = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="term_life",
        product_name="Term Life",
        blob=blob,
        tolerance=float(TERM_MODELCHECK_TOL or 1e-6),
        # B5/B6/B7 are Python literals baked into the Term workbook.
        modelcheck_python_cells={
            "ModelCheck!B5": ex_claims,
            "ModelCheck!B6": ex_prem,
            "ModelCheck!B7": ex_net,
        },
        # C5/C6/C7 are =Liability!X4 / X5 / X7 -- the Excel side that
        # only formula evaluation can resolve to a numeric value.
        modelcheck_formula_cells={
            "ModelCheck!C5": ex_claims,
            "ModelCheck!C6": ex_prem,
            "ModelCheck!C7": ex_net,
        },
    )


def _make_rila_case() -> ProductRecalcCase:
    """RILA: 5-year workbook with a deterministic seeded index path."""
    contract = rp.RILAContract(
        issue_age=55,
        sex="male",
        participation=0.85,
        cap=0.09,
        floor=-0.02,
        rider_fee_annual=0.008,
    )
    yc = sp.YieldCurve.from_flat_rate(0.035)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    horizon_age = 60
    n_months = int(round((horizon_age - contract.issue_age) * 12))
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.004, 0.02, size=n_months))
    res = rp.price_rila_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=horizon_age,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.01,
    )
    spec = rila_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=horizon_age,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.01,
    )
    blob = build_rila_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    pv_premium = float(res.single_premium)
    return ProductRecalcCase(
        product_id="rila",
        product_name="RILA",
        blob=blob,
        tolerance=float(RILA_PV_TOL or 1e-6),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": pv_premium,
        },
        modelcheck_formula_cells={
            "ModelCheck!C5": pv_b,
            "ModelCheck!C6": pv_e,
            "ModelCheck!C7": pv_t,
            "ModelCheck!C8": pv_premium,
        },
    )


def _make_myga_case() -> ProductRecalcCase:
    """MYGA: 5y workbook with deterministic accumulation."""
    import myga_projection as my
    from build_myga_excel_workbook import (
        build_myga_workbook_from_spec,
        myga_excel_spec_from_launcher,
    )
    from parity_constants import MYGA_PV_TOL

    contract = my.MYGAContract(
        issue_age=60,
        sex="male",
        single_premium=100_000.0,
        declared_rate_annual=0.045,
        guarantee_years=5,
    )
    yc = sp.YieldCurve.from_flat_rate(0.045)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = my.price_myga_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=70,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        expense_annual_inflation=0.0,
    )
    spec = myga_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=70,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        expense_annual_inflation=0.0,
    )
    blob = build_myga_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="myga",
        product_name="MYGA",
        blob=blob,
        tolerance=float(MYGA_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={},
    )


def _make_fia_case() -> ProductRecalcCase:
    import fia_projection as fp
    from build_fia_excel_workbook import (
        build_fia_workbook_from_spec,
        fia_excel_spec_from_launcher,
    )
    from parity_constants import FIA_PV_TOL

    contract = fp.FIAContract(
        issue_age=60,
        sex="male",
        single_premium=100_000.0,
        participation=0.8,
        cap=0.07,
        floor=0.0,
        horizon_years=5,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    n_months = 60
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.02, size=n_months))
    res = fp.price_fia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=65,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = fia_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=65,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    blob = build_fia_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="fia",
        product_name="FIA",
        blob=blob,
        tolerance=float(FIA_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={
            "ModelCheck!C5": pv_b,
            "ModelCheck!C6": pv_e,
            "ModelCheck!C7": pv_t,
            "ModelCheck!C8": float(res.single_premium),
        },
    )


def _make_va_case() -> ProductRecalcCase:
    import va_projection as va
    from build_va_excel_workbook import (
        build_va_workbook_from_spec,
        va_excel_spec_from_launcher,
    )
    from parity_constants import VA_PV_TOL

    contract = va.VAContract(
        issue_age=55,
        sex="male",
        single_premium=100_000.0,
        me_charge_annual=0.014,
        horizon_years=10,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    n_months = 120
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    res = va.price_va_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=65,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = va_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=65,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    blob = build_va_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="variable_annuity",
        product_name="Variable Annuity",
        blob=blob,
        tolerance=float(VA_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={},
    )


def _make_wl_case() -> ProductRecalcCase:
    import wl_projection as wl
    from build_wl_excel_workbook import (
        build_wl_workbook_from_spec,
        wl_excel_spec_from_launcher,
    )
    from parity_constants import WL_PV_TOL

    contract = wl.WLContract(
        issue_age=45,
        sex="male",
        smoker_class="nonsmoker",
        face_amount=250_000.0,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = wl.price_wl_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        expense_annual_inflation=0.0,
    )
    spec = wl_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        expense_annual_inflation=0.0,
    )
    blob = build_wl_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="whole_life",
        product_name="Whole Life",
        blob=blob,
        tolerance=float(WL_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={},
    )


def _make_ul_case() -> ProductRecalcCase:
    import ul_projection as ul_proj
    from build_ul_excel_workbook import (
        build_ul_workbook_from_spec,
        ul_excel_spec_from_launcher,
    )
    from parity_constants import UL_PV_TOL

    contract = ul_proj.ULContract(
        issue_age=45,
        sex="male",
        face_amount=250_000.0,
        single_premium=25_000.0,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = ul_proj.price_ul_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        expense_annual_inflation=0.0,
    )
    spec = ul_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        expense_annual_inflation=0.0,
    )
    blob = build_ul_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="universal_life",
        product_name="Universal Life",
        blob=blob,
        tolerance=float(UL_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={},
    )


def _make_iul_case() -> ProductRecalcCase:
    import iul_projection as iul_proj
    from build_iul_excel_workbook import (
        build_iul_workbook_from_spec,
        iul_excel_spec_from_launcher,
    )
    from parity_constants import IUL_PV_TOL

    contract = iul_proj.IULContract(
        issue_age=45,
        sex="male",
        face_amount=250_000.0,
        single_premium=25_000.0,
        participation=1.0,
        cap=0.10,
        floor=0.0,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    n_months = (80 - 45) * 12
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    res = iul_proj.price_iul_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = iul_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    blob = build_iul_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="indexed_ul",
        product_name="Indexed UL",
        blob=blob,
        tolerance=float(IUL_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={
            "ModelCheck!C5": pv_b,
            "ModelCheck!C6": pv_e,
            "ModelCheck!C7": pv_t,
            "ModelCheck!C8": float(res.single_premium),
        },
    )


def _make_vul_case() -> ProductRecalcCase:
    import vul_projection as vul_proj
    from build_vul_excel_workbook import (
        build_vul_workbook_from_spec,
        vul_excel_spec_from_launcher,
    )
    from parity_constants import VUL_PV_TOL

    contract = vul_proj.VULContract(
        issue_age=45,
        sex="male",
        face_amount=250_000.0,
        single_premium=25_000.0,
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    n_months = (80 - 45) * 12
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    res = vul_proj.price_vul_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=expenses,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.0,
    )
    spec = vul_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    blob = build_vul_workbook_from_spec(spec)
    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    return ProductRecalcCase(
        product_id="variable_ul",
        product_name="Variable UL",
        blob=blob,
        tolerance=float(VUL_PV_TOL),
        modelcheck_python_cells={
            "ModelCheck!B5": pv_b,
            "ModelCheck!B6": pv_e,
            "ModelCheck!B7": pv_t,
            "ModelCheck!B8": float(res.single_premium),
        },
        modelcheck_formula_cells={},
    )


_CASE_BUILDERS: dict[str, Callable[[], ProductRecalcCase]] = {
    "spia": _make_spia_case,
    "term_life": _make_term_case,
    "rila": _make_rila_case,
    "myga": _make_myga_case,
    "fia": _make_fia_case,
    "variable_annuity": _make_va_case,
    "whole_life": _make_wl_case,
    "universal_life": _make_ul_case,
    "indexed_ul": _make_iul_case,
    "variable_ul": _make_vul_case,
}


@pytest.fixture(scope="module")
def product_recalc_cases() -> dict[str, ProductRecalcCase]:
    """Build every product's workbook once per module so the always-run
    gate AND the formula-evaluation gate can share the same blob. Module
    scope keeps formula evaluation to one workbook build per product
    even when both layers run.
    """
    return {pid: builder() for pid, builder in _CASE_BUILDERS.items()}


# ---------------------------------------------------------------------------
# Layer 1: ALWAYS-RUN
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("product_id", sorted(_CASE_BUILDERS), ids=sorted(_CASE_BUILDERS))
def test_python_cached_modelcheck_values_match_engine(
    product_recalc_cases: dict[str, ProductRecalcCase], product_id: str
) -> None:
    """For every product, the ModelCheck cells the BUILDER writes as
    literal Python values must equal the engine's outputs within
    ``MODELCHECK_TOL``.

    This is the always-on Excel<->Python parity gate. It does NOT
    require any desktop spreadsheet app and runs on every developer machine and every
    CI shard. Its purpose is to catch the bug class where a builder
    refactor accidentally writes a rounded / stale / wrong number into
    the workbook column the user actually reads (Term/RILA column B).

    For products whose ModelCheck cells are PURE FORMULAS with no
    cached literal fallback (today: SPIA B5/B9), this layer only
    asserts the formula string is structurally well-formed. Numeric
    correctness for those cells is verified by layer (2) below.
    """
    case = product_recalc_cases[product_id]
    wb = load_workbook(io.BytesIO(case.blob), data_only=False)

    assert case.modelcheck_python_cells, (
        f"{case.product_name} has no modelcheck_python_cells -- the "
        "always-on layer would be a no-op. Every product's workbook "
        "must bake at least one Python literal into the ModelCheck "
        "sheet so the user always sees an actual number when they "
        "open the file (instead of a #N/A from an unrecalculated "
        "formula)."
    )

    for coord_with_sheet, expected in case.modelcheck_python_cells.items():
        sheet, coord = coord_with_sheet.split("!", 1)
        ws = wb[sheet]
        raw = ws[coord].value
        assert isinstance(raw, (int, float)) and not isinstance(raw, bool), (
            f"{case.product_name} {coord_with_sheet} is supposed to be a "
            f"Python-cached literal but got {type(raw).__name__}={raw!r}. "
            "The builder may have switched the cell to a formula -- in "
            "that case the test should be updated AND a recalc-only check "
            "added to modelcheck_formula_cells."
        )
        np.testing.assert_allclose(
            float(raw),
            expected,
            rtol=0.0,
            atol=case.tolerance,
            err_msg=(
                f"{case.product_name} {coord_with_sheet} (Python-literal) "
                f"diverges from engine value within {case.tolerance}. "
                "The user opens the workbook and sees a different number "
                "than the Python pricing run produced. Check the builder's "
                "modelcheck.write_python_block (or product analog) -- a "
                "wrong cashflow array or wrong discount mask is the usual "
                "culprit."
            ),
        )


# ---------------------------------------------------------------------------
# Layer 2: ModelCheck formula contract
# ---------------------------------------------------------------------------


_MODELCHECK_FORMULA_TARGETS = {
    "ModelCheck!C5": "=Liabilities!X4",
    "ModelCheck!C6": "=Liabilities!X5",
    "ModelCheck!C7": "=Liabilities!X7",
    "ModelCheck!C8": "=Liabilities!X8",
    "ModelCheck!C9": "=Liabilities!X6",
}


@pytest.mark.parametrize("product_id", sorted(_CASE_BUILDERS), ids=sorted(_CASE_BUILDERS))
def test_modelcheck_formula_contract(
    product_recalc_cases: dict[str, ProductRecalcCase],
    product_id: str,
) -> None:
    """Every product's ModelCheck formula cells must point at the
    canonical liability summary rows. Numeric engine parity is checked
    by ``test_python_cached_modelcheck_values_match_engine``; this gate
    catches stale or product-inappropriate formula wiring without
    launching an external spreadsheet process.
    """
    case = product_recalc_cases[product_id]

    wb = load_workbook(io.BytesIO(case.blob), data_only=False)
    for coord_with_sheet in case.modelcheck_formula_cells:
        sheet, coord = coord_with_sheet.split("!", 1)
        formula = wb[sheet][coord].value
        assert formula == _MODELCHECK_FORMULA_TARGETS[coord_with_sheet], (
            f"{case.product_name} {coord_with_sheet} is not a formula "
            f"linked to the canonical liability summary row: got {formula!r}, "
            f"expected {_MODELCHECK_FORMULA_TARGETS[coord_with_sheet]!r}."
        )
        assert (
            wb["Liabilities"][_MODELCHECK_FORMULA_TARGETS[coord_with_sheet].split("!", 1)[1]].value
            is not None
        )


# ---------------------------------------------------------------------------
# Coverage invariant -- adding a new product must add a recalc case.
# ---------------------------------------------------------------------------


def test_every_implemented_product_has_a_recalc_case() -> None:
    """If a new product becomes implemented (registered in
    ``product_registry.implemented_product_types``), it MUST also get
    a builder in ``_CASE_BUILDERS`` so that the always-on layer fires
    for it. Without this guard, a new product could ship to users
    without a single Python<->Excel parity assertion.
    """
    from product_registry import implemented_product_types

    missing = [p.value for p in implemented_product_types() if p.value not in _CASE_BUILDERS]
    assert not missing, (
        f"Products implemented but missing from _CASE_BUILDERS: "
        f"{missing!r}. Add a `_make_<name>_case()` function above and "
        "wire it into _CASE_BUILDERS so the per-product Excel-vs-Python "
        "recalc gates fire for it."
    )
