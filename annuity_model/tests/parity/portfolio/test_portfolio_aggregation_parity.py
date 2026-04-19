"""Portfolio liability aggregation invariants (parity-adjacent)."""

from __future__ import annotations

import io
from pathlib import Path

import numpy as np
import pytest

from build_portfolio_excel_workbook import build_portfolio_workbook_bytes
from excel_workbook_validator import validate_workbook_or_raise
from inforce_io import load_policy_inputs_from_csv
from openpyxl import load_workbook
from portfolio import Portfolio
from portfolio_runner import run_portfolio
from portfolio_scenario import default_run_scenario

pytestmark = pytest.mark.parity


def test_inforce_csv_portfolio_rollups_sum_to_total() -> None:
    root = Path(__file__).resolve().parents[3]
    csv_path = root / "tests/data/inforce/example_v1/inforce.csv"
    policies = load_policy_inputs_from_csv(csv_path)
    res = run_portfolio(portfolio=Portfolio(policies=policies), scenario=default_run_scenario())
    n = len(res.liability_path_total.expected_total_cashflows)
    summed = np.zeros(n)
    for _pt, path in sorted(res.rollups_by_product_type.items(), key=lambda x: x[0].value):
        cf = path.expected_total_cashflows
        summed[: len(cf)] += cf
    np.testing.assert_allclose(
        summed,
        res.liability_path_total.expected_total_cashflows,
        atol=1e-9,
    )


def test_portfolio_workbook_passes_validator_and_modelcheck_formula_shape() -> None:
    root = Path(__file__).resolve().parents[3]
    policies = load_policy_inputs_from_csv(root / "tests/data/inforce/example_v1/inforce.csv")
    res = run_portfolio(portfolio=Portfolio(policies=policies), scenario=default_run_scenario())
    raw = build_portfolio_workbook_bytes(res)
    wb = load_workbook(io.BytesIO(raw), data_only=False)
    validate_workbook_or_raise(wb)
    ws = wb["ModelCheck"]
    n = len(res.liability_path_total.expected_total_cashflows)
    for r in range(2, 2 + n):
        cell = ws.cell(row=r, column=2)
        assert isinstance(cell.value, str) and cell.value.startswith("=SUM(")
