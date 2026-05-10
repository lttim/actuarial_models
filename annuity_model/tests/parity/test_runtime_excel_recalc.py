"""SPIA workbook ModelCheck formula contract gate.

The classic parity tests compare Python to a *Python simulation* of the Excel
formulas (``tests/parity/excel_formula_sim.py``). That guards against logic
drift, but not against a builder bug that emits a *different* formula string
than the simulation expects -- the workbook is broken in Excel even though
parity is green.

This test closes the gap that previously lived in a subprocess-backed
workbook recalc gate: it builds a small SPIA workbook end-to-end and
asserts ModelCheck has both the Python literal snapshot users see on
open and the formula links into the validated liability summary.
"""

from __future__ import annotations

import io

import numpy as np
import pytest
from openpyxl import load_workbook

from annuity_model import pricing_projection as sp
from annuity_model.build_pricing_excel_workbook import (
    ExcelPythonSnapshot,
    build_workbook_from_spec,
    excel_spec_from_launcher,
)
from annuity_model.parity_constants import MODELCHECK_TOL

pytestmark = [pytest.mark.parity, pytest.mark.product_spia, pytest.mark.slow]


def _build_small_spia_workbook() -> tuple[bytes, sp.SPIAProjectionResult]:
    """Build the smallest realistic SPIA workbook for a quick contract check.

    horizon_age=66 gives 12 projection months (one year), which is the
    minimum that exercises the full ModelCheck stack without the heavyweight
    60-month ALM ladder dominating recalc time.
    """
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=120_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.045)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    expenses = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=66,
        spread=0.0,
        expenses=expenses,
        expense_annual_inflation=0.0,
    )
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=66,
        spread=0.0,
        valuation_year=2025,
        expenses=expenses,
        yield_mode_label="flat",
        mortality_mode_label="qx",
        expense_mode_label="manual",
        index_s0=100.0,
        index_levels_at_payment=np.full(12, 100.0),
        expense_annual_inflation=0.0,
    )
    blob = build_workbook_from_spec(
        spec,
        python_snapshot=ExcelPythonSnapshot(
            pv_benefit=float(res.pv_benefit),
            pv_monthly_expenses=float(res.pv_monthly_expenses),
            pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
            single_premium=float(res.single_premium),
            annuity_factor=float(res.annuity_factor),
        ),
    )
    return blob, res


def test_modelcheck_python_literals_and_formula_links_are_wired() -> None:
    """ModelCheck must pair Python literals with canonical formula links."""
    blob, res = _build_small_spia_workbook()

    wb = load_workbook(io.BytesIO(blob), data_only=False)
    ws = wb["ModelCheck"]

    np.testing.assert_allclose(
        float(ws["B5"].value),
        float(res.pv_benefit),
        rtol=0.0,
        atol=MODELCHECK_TOL or 1e-6,
    )
    np.testing.assert_allclose(
        float(ws["B9"].value),
        float(res.annuity_factor),
        rtol=0.0,
        atol=MODELCHECK_TOL or 1e-6,
    )
    assert ws["C5"].value == "=Liabilities!X4"
    assert ws["C9"].value == "=Liabilities!X6"
    assert wb["Liabilities"]["X4"].value is not None
    assert wb["Liabilities"]["X6"].value is not None
