"""Deep Excel assessment gates for every implemented product.

These tests are the mandatory replacement layer for the removed
LibreOffice-backed runtime recalc gate: every product must build a real
workbook, pass strict static validation, expose a ModelCheck surface, and
survive a save/reopen round trip. Experimental formula evaluators are kept
advisory through ``tests/workbook_recalc_backends.py``.
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from test_regression_matrix import _FIXTURE_BUILDERS, _build_excel_workbook
from workbook_recalc_backends import candidate_backends

from annuity_model.excel_workbook_validator import validate_workbook
from annuity_model.product_excel import registered_builders
from annuity_model.product_registry import ProductType, implemented_product_types

pytestmark = [pytest.mark.regression]

ERROR_LITERALS = ("#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#NUM!", "#N/A", "#NULL!")
DIRECT_REF_RE = re.compile(r"^='?([^'!]+)'?!([A-Z]{1,3}[1-9][0-9]*)$")

PARITY_COVERAGE_FILES: dict[ProductType, tuple[str, ...]] = {
    ProductType.SPIA: (
        "tests/parity/test_alm_parity.py",
        "tests/parity/test_runtime_excel_recalc.py",
        "tests/parity/test_excel_recalc_per_product.py",
    ),
    ProductType.TERM_LIFE: (
        "tests/parity/test_term_parity.py",
        "tests/parity/test_excel_recalc_per_product.py",
    ),
    ProductType.RILA: (
        "tests/parity/test_rila_parity.py",
        "tests/parity/test_excel_recalc_per_product.py",
    ),
    ProductType.MYGA: ("tests/parity/test_myga_actuarial.py",),
    ProductType.FIA: ("tests/parity/test_fia_actuarial.py",),
    ProductType.VARIABLE_ANNUITY: ("tests/parity/test_va_actuarial.py",),
    ProductType.WHOLE_LIFE: ("tests/parity/test_wl_actuarial.py",),
    ProductType.UNIVERSAL_LIFE: ("tests/parity/test_ul_actuarial.py",),
    ProductType.INDEXED_UL: ("tests/parity/test_iul_actuarial.py",),
    ProductType.VARIABLE_UL: ("tests/parity/test_vul_actuarial.py",),
}


def _formulas(ws: Any) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                found.append((cell.coordinate, cell.value))
    return found


def _error_literals(wb: Any) -> list[str]:
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(token in value for token in ERROR_LITERALS):
                    errors.append(f"{ws.title}!{cell.coordinate}={value!r}")
    return errors


def _formula_count(wb: Any) -> int:
    return sum(len(_formulas(ws)) for ws in wb.worksheets)


@pytest.mark.parametrize("product_type", implemented_product_types(), ids=lambda p: p.value)
def test_deep_excel_product_assessment_matrix(product_type: ProductType) -> None:
    """Every product prices, exports, validates, exposes ModelCheck, and round-trips."""

    assert product_type in registered_builders(), (
        f"{product_type.value} is implemented but has no registered Excel builder."
    )
    assert product_type in _FIXTURE_BUILDERS, (
        f"{product_type.value} is missing from the regression-matrix fixture builders."
    )

    fixture = _FIXTURE_BUILDERS[product_type]("baseline")
    raw = _build_excel_workbook(fixture)
    assert raw.startswith(b"PK"), f"{product_type.value} builder did not return .xlsx bytes"

    wb = load_workbook(io.BytesIO(raw), data_only=False)
    issues = validate_workbook(wb, strict=True)
    assert issues == [], f"{product_type.value} workbook failed strict validation:\n" + "\n".join(
        f"  - {issue}" for issue in issues[:25]
    )
    assert not _error_literals(wb), f"{product_type.value} workbook contains Excel error literals"

    assert "ModelCheck" in wb.sheetnames, f"{product_type.value} workbook has no ModelCheck sheet"
    modelcheck = wb["ModelCheck"]
    modelcheck_formulas = _formulas(modelcheck)
    assert modelcheck_formulas, f"{product_type.value} ModelCheck has no formula cells"
    assert any(
        isinstance(cell.value, (int, float)) for row in modelcheck.iter_rows() for cell in row
    ), f"{product_type.value} ModelCheck has no Python numeric snapshot values"

    direct_refs: list[tuple[str, str, str]] = []
    for coord, formula in modelcheck_formulas:
        match = DIRECT_REF_RE.match(formula)
        if match:
            direct_refs.append((coord, match.group(1), match.group(2)))
    assert direct_refs, f"{product_type.value} ModelCheck has no direct cross-sheet formula links"
    for source_coord, target_sheet, target_coord in direct_refs:
        assert target_sheet in wb.sheetnames, (
            f"{product_type.value} ModelCheck!{source_coord} references missing sheet "
            f"{target_sheet!r}"
        )
        assert wb[target_sheet][target_coord].value is not None, (
            f"{product_type.value} ModelCheck!{source_coord} references blank "
            f"{target_sheet}!{target_coord}"
        )

    formula_count_before = _formula_count(wb)
    roundtrip = io.BytesIO()
    wb.save(roundtrip)
    roundtrip.seek(0)
    reopened = load_workbook(roundtrip, data_only=False)
    assert reopened.sheetnames == wb.sheetnames
    assert _formula_count(reopened) == formula_count_before
    assert validate_workbook(reopened, strict=True) == []


def test_every_product_has_parity_and_excel_assessment_coverage() -> None:
    """Coverage invariant for the product/parity/Excel validation matrix."""

    products = set(implemented_product_types())
    assert set(registered_builders()) == products
    assert set(PARITY_COVERAGE_FILES) == products

    root = Path(__file__).resolve().parents[1]
    missing = [
        f"{product.value}: {path}"
        for product, paths in PARITY_COVERAGE_FILES.items()
        for path in paths
        if not (root / path).is_file()
    ]
    assert not missing, "Missing product parity coverage files:\n" + "\n".join(missing)


def test_formula_corpus_static_contract_accepts_required_patterns() -> None:
    """Hand-authored formula corpus for constructs emitted by product workbooks."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["x", "y"])
    ws.append([1.0, 10.0])
    ws.append([2.0, 0.0])
    ws.append([3.0, None])

    mc = wb.create_sheet("ModelCheck")
    mc["A1"] = "Formula corpus"
    mc["B2"] = "=Data!A2"
    mc["B3"] = "=IF(Data!B3=0,0,Data!A3/Data!B3)"
    mc["B4"] = '=IFERROR(Data!A4/Data!B4,"")'
    mc["B5"] = "=SUMPRODUCT(Data!A2:A4,Data!B2:B4)"
    mc["B6"] = "=INDEX(Data!A2:A4,2)"
    mc["B7"] = '=IF(Data!B4="",0,Data!B4)'

    assert validate_workbook(wb, strict=True) == []


@pytest.mark.excel_recalc_candidate
def test_candidate_recalc_backends_are_advisory_against_formula_corpus() -> None:
    """Future recalc backends must start advisory and explicit."""

    backends = candidate_backends()
    if not backends:
        pytest.skip("No advisory workbook recalc backend is installed or enabled.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 2.0
    ws["A2"] = 3.0
    mc = wb.create_sheet("ModelCheck")
    mc["B2"] = "=SUM(Data!A1:A2)"
    raw = io.BytesIO()
    wb.save(raw)

    for backend in backends:
        values = backend.recalc(raw.getvalue(), ["ModelCheck!B2"], timeout=30.0)
        assert values["ModelCheck!B2"] == 5.0, backend.name
