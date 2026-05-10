"""End-to-end Streamlit UI smoke gate (runs on every pytest invocation).

Why this file exists
--------------------
The per-product files (``test_apptest_spia.py`` / ``term`` / ``rila``)
prove the **Pricing Run** form renders for each product. That is
necessary but not sufficient. This module pins the broader contract the
user expects of the app:

1. **App boots clean from both entry points** -- ``pricing_ui.py`` (the
   in-package launcher used by ``run_pricing_ui.{sh,bat}``) AND
   ``streamlit_app.py`` (the Streamlit-Cloud entry point). A regression
   in only the cloud entry point used to ship silently because no test
   exercised it.
2. **Every sidebar section renders** without raising. Catches the
   "page X breaks because session_state key Y was renamed" bug class
   that the new ``scripts/audit_session_state.py`` is designed to
   prevent at static-analysis level -- this is the runtime gate that
   complements it.
3. **Clicking "Run pricing" actually produces a result** for each
   implemented product (SPIA / Term / RILA). This is the
   functionality test the user asked for: the core action of the app
   must work end-to-end on every run.
4. **The post-run Excel download surface** stays valid -- the workbook
   bytes attached to the download_button must pass strict-mode
   ``excel_workbook_validator``. Catches a builder regression that
   would otherwise only surface when a user opens the file in Excel.

Performance budget
------------------
Streamlit AppTest is slow (~1-3 s per ``.run()``); each per-product
end-to-end test costs ~5-10 s. We accept that cost because these are
the *only* runtime UI tests in the suite and the alternative (skipping
on every non-CI machine) is exactly the silent-rot pattern the audit
script was meant to break. The whole module finishes well under 60 s
on a 2024-era laptop.

Skip semantics
--------------
The module skips ONLY when ``streamlit.testing.v1`` is unimportable
(streamlit < 1.28). On modern installs (we pin newer in
``requirements.lock``), no test in this file is allowed to skip.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from .conftest import (
    assert_no_exceptions,
    load_pricing_ui,
    navigate_to_pricing_run,
    select_product,
)


def _session_get(at: Any, key: str) -> Any:
    """Read from AppTest session_state without crashing on missing keys.

    Streamlit's ``SafeSessionState`` only supports attribute access and
    raises ``AttributeError`` for missing keys -- ``.get(...)`` is NOT
    available, despite mimicking a dict-like surface in many other
    spots. Wrap that here so test-side code reads naturally.
    """
    try:
        return getattr(at.session_state, key)
    except (AttributeError, KeyError):
        return None


def _find_section_radio(at: Any) -> Any:
    """Locate the sidebar 'Section' radio by its label.

    The radio is constructed in ``pricing_ui.main()`` as
    ``st.radio("Section", options=SECTION_ORDER, format_func=...)``.
    The label "Section" is unique across the app, which makes this
    lookup robust both on first paint AND after the Pricing Run page
    has rendered its own product/yield-curve/mortality radios further
    down the tree (so ``at.radio[0]`` no longer refers to the section
    radio).
    """
    for r in at.radio:
        if getattr(r, "label", None) == "Section":
            return r
    raise AssertionError(
        "Could not find the sidebar 'Section' radio (label='Section'). "
        f"Saw radios with labels {[getattr(r, 'label', None) for r in at.radio]!r}."
    )


def _navigate_to_section(at: Any, section: str) -> None:
    """Click the sidebar Section radio for *section* and re-run.

    *section* is the underlying option value from ``SECTION_ORDER``
    (``"alm"``, ``"run"``, ...) -- AppTest's ``set_value`` on a
    radio with a ``format_func`` accepts the underlying value, not the
    formatted label.
    """
    _find_section_radio(at).set_value(section).run()


ROOT = Path(__file__).resolve().parent.parent.parent
REPO_ROOT = ROOT.parent
SRC_ROOT = ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from annuity_model.build_portfolio_excel_workbook import (
    build_portfolio_workbook_bytes,  # noqa: E402
)
from annuity_model.excel_workbook_validator import validate_workbook  # noqa: E402
from annuity_model.pricing_run_form_state import PORTFOLIO_KEY, RUN_KEY  # noqa: E402

# ---------------------------------------------------------------------------
# Boot tests
# ---------------------------------------------------------------------------


@pytest.mark.ui
def test_pricing_ui_boots_without_exception(streamlit_apptest_module) -> None:
    """``pricing_ui.py`` -- the in-package launcher path -- must render
    on first paint with zero script-level exceptions. This is the
    minimum-viable contract: a regression here means every developer
    using ``run_pricing_ui.sh`` opens to a broken page.
    """
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    assert at.sidebar, (
        "pricing_ui.py rendered without a sidebar -- the section radio "
        "and run controls live there. Something killed the sidebar layer."
    )


@pytest.mark.ui
def test_overview_renders_product_readiness_badges(streamlit_apptest_module) -> None:
    """Overview must expose product maturity and assumption status."""
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    assert_no_exceptions(at, context="render overview product readiness badges")
    rendered_markdown = "\n".join(str(getattr(item, "value", "")) for item in at.markdown)
    assert "Mechanics-production" in rendered_markdown
    assert "Assumptions: demo-safe-with-waiver" in rendered_markdown


@pytest.mark.ui
def test_diagnostics_export_empty_state_is_actionable(streamlit_apptest_module) -> None:
    """Preparing diagnostics before pricing should guide the user, not crash."""
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    buttons = [button for button in at.button if button.label == "Prepare diagnostics JSON"]
    assert buttons, "Diagnostics export button missing from the sidebar."
    buttons[0].click().run()
    assert_no_exceptions(at, context="prepare diagnostics before pricing run")
    warnings = [warning.value for warning in at.warning]
    assert "Run Pricing Run first to populate diagnostics." in warnings


@pytest.mark.ui
def test_streamlit_cloud_entry_boots_without_exception(streamlit_apptest_module) -> None:
    """``streamlit_app.py`` at the repo root is the entry point
    Streamlit Community Cloud uses. It is a thin launcher that injects
    ``annuity_model/src`` onto ``sys.path`` and calls ``pricing_ui.main()``.
    A regression here ships silently to Cloud users until someone
    actually opens the deployment.
    """
    AppTest = streamlit_apptest_module
    cloud_entry = REPO_ROOT / "streamlit_app.py"
    assert cloud_entry.is_file(), (
        f"missing {cloud_entry}; the Streamlit Cloud entry point was "
        "deleted or moved -- update streamlit_cloud config + this test."
    )
    at = AppTest.from_file(str(cloud_entry), default_timeout=60)
    at.run()
    if at.exception:
        raise AssertionError(
            "streamlit_app.py raised on initial render:\n"
            + "\n---\n".join(str(e.value) for e in at.exception)
        )


# ---------------------------------------------------------------------------
# Per-section render tests (parametrized)
# ---------------------------------------------------------------------------
#
# Base SECTION_ORDER from pricing_ui.py omits "portfolio"; the live sidebar
# inserts ``"portfolio"`` after ``"run"`` when portfolio v1 is enabled. We test
# every navigable section EXCEPT "tests" -- that section, when rendered,
# executes the project's own pytest suite live inside the page, which would
# recursively re-enter the test runner and obliterate the timing budget.

_SECTIONS_TO_RENDER: list[str] = [
    "overview",
    "run",
    "portfolio",
    "alm",
    "what_if",
    "excel_replicator",
]


@pytest.mark.ui
@pytest.mark.parametrize("section", _SECTIONS_TO_RENDER)
def test_each_sidebar_section_renders_without_exception(
    streamlit_apptest_module,
    section: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Selecting each sidebar section must complete without raising.

    The per-section pages each do their own session_state lookups, and
    several of them gate on ``pricing_res`` being None vs populated.
    Both branches must render cleanly: this test exercises the "no
    pricing yet" branch (the harder one -- it has to render an
    informative empty state, not crash on a missing key).
    """
    if section == "portfolio":
        # Sidebar lists Portfolio only when ``portfolio_v1_enabled()``; pin on
        # so this parametrized gate matches CI / launcher defaults.
        monkeypatch.setenv("ANNUITY_MODEL_PORTFOLIO_V1", "1")
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    _navigate_to_section(at, section)
    assert_no_exceptions(at, context=f"render '{section}' section without prior pricing run")
    if section == "excel_replicator":
        empty_state_messages = [info.value for info in at.info]
        assert (
            "Run pricing first in the Pricing Run section to populate the Excel Replicator."
            in empty_state_messages
        )


# ---------------------------------------------------------------------------
# End-to-end functionality: click "Run pricing" and verify result
# ---------------------------------------------------------------------------


def _click_run_pricing(at: Any) -> Any:
    """Find and click the primary 'Run pricing' button, then re-run.

    The Pricing Run page renders ``st.button("Run pricing", type="primary")``
    once per render. AppTest exposes buttons via ``at.button``; we
    match by label rather than index so a future button-reorder doesn't
    silently click the wrong button.
    """
    candidates = [b for b in at.button if b.label == "Run pricing"]
    assert candidates, (
        "Pricing Run page is missing the 'Run pricing' button. The user "
        "cannot price anything. Either the button label was renamed (in "
        "which case update this test AND the docs) or the button stopped "
        "rendering for the selected product (check "
        "_is_implemented_product gating in pricing_ui.py)."
    )
    candidates[0].click().run()
    return at


def _set_deterministic_pricing_inputs(at: Any) -> None:
    """Pin the random-ish inputs so the pricing run is deterministic.

    We don't need to set every widget -- the form has sensible defaults
    seeded by ``build_run_form_seed_defaults``. We nail down the shared
    age/horizon controls to values that are identical across products
    and small enough for AppTest's 60-second rerun budget. This still
    exercises the full form -> adapter -> engine -> workbook path, but
    avoids using the production default horizon_age=110 for long RILA
    workbook builds inside the UI harness.
    """
    # Match against RUN_KEY.ISSUE_AGE rather than the raw literal -- the
    # tests/test_run_state_key_drift.py ratchet forbids new files from
    # introducing raw "run_*" session-state literals.
    for inp in at.number_input:
        if inp.key == RUN_KEY.ISSUE_AGE:
            inp.set_value(55)
        elif inp.key == RUN_KEY.HORIZON_AGE:
            inp.set_value(65)


_PRODUCT_RUN_CASES: list[tuple[str, str, str]] = [
    # (product_value, friendly_name, product-specific result descriptor)
    ("spia", "SPIA", "single_premium"),
    ("term_life", "Term Life", "single_premium"),
    ("rila", "RILA", "single_premium"),
    ("myga", "MYGA", "single_premium"),
    ("fia", "FIA", "single_premium"),
    ("variable_annuity", "Variable Annuity", "single_premium"),
    ("whole_life", "Whole Life", "single_premium"),
    ("universal_life", "Universal Life", "single_premium"),
    ("indexed_ul", "Indexed UL", "single_premium"),
    ("variable_ul", "Variable UL", "single_premium"),
]


@pytest.mark.ui
@pytest.mark.parametrize(
    "product_value,product_label,result_attr",
    _PRODUCT_RUN_CASES,
    ids=[c[0] for c in _PRODUCT_RUN_CASES],
)
def test_run_pricing_button_populates_session_result(
    streamlit_apptest_module,
    product_value: str,
    product_label: str,
    result_attr: str,
) -> None:
    """End-to-end: select product, click 'Run pricing', verify a real
    pricing result lands in ``st.session_state['pricing_res']``.

    This is THE functionality test for the user-facing app. A green run
    here means the full chain works: form -> adapter dispatch -> engine
    -> session state -> render. A red run typically means one of:
      * adapter dispatch broke (product_registry / get_product_adapter)
      * engine raised on the seed inputs (look at the AppTest exception
        message for the engine ValueError)
      * session_state plumbing broke (the result was computed but never
        stored, so the rest of the app sees None)
    """
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    navigate_to_pricing_run(at)
    select_product(at, product_value)
    assert_no_exceptions(at, context=f"select {product_label}")
    _set_deterministic_pricing_inputs(at)
    at.run()  # Re-render with the pinned inputs.
    assert_no_exceptions(at, context=f"set deterministic inputs for {product_label}")

    _click_run_pricing(at)
    assert_no_exceptions(at, context=f"click 'Run pricing' for {product_label}")

    res = _session_get(at, "pricing_res")
    assert res is not None, (
        f"Clicking 'Run pricing' for {product_label} did not populate "
        "st.session_state['pricing_res']. Either the engine raised "
        "(check stderr / AppTest exception above), or the success path "
        "in pricing_ui.py stopped writing the result key."
    )
    assert hasattr(res, result_attr), (
        f"{product_label} pricing result is missing required attribute "
        f"{result_attr!r}; got {type(res).__name__} with attrs "
        f"{sorted(a for a in dir(res) if not a.startswith('_'))[:10]}..."
    )

    contract = _session_get(at, "pricing_contract")
    assert contract is not None, (
        f"{product_label} run produced a pricing_res but no "
        "pricing_contract; downstream pages (ALM / Excel Replicator / "
        "What-If) will all break because they look up the contract."
    )

    meta = _session_get(at, "pricing_meta")
    assert meta is not None, (
        f"{product_label} run did not write pricing_meta; the metadata "
        "panel and the Excel Replicator both depend on it."
    )


# ---------------------------------------------------------------------------
# Post-run downstream pages must render
# ---------------------------------------------------------------------------


@pytest.mark.ui
@pytest.mark.parametrize(
    "product_value,product_label",
    [(c[0], c[1]) for c in _PRODUCT_RUN_CASES],
    ids=[c[0] for c in _PRODUCT_RUN_CASES],
)
def test_alm_section_renders_after_pricing_run(
    streamlit_apptest_module, product_value: str, product_label: str
) -> None:
    """After a successful pricing run, navigating to the ALM section
    must render without raising. This catches the bug class where
    the ALM page assumes a session_state key that the pricing-run
    success path forgot to write for a particular product.
    """
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    navigate_to_pricing_run(at)
    select_product(at, product_value)
    _set_deterministic_pricing_inputs(at)
    at.run()
    _click_run_pricing(at)
    assert _session_get(at, "pricing_res") is not None, (
        f"Pricing run for {product_label} did not produce a result -- "
        "the downstream ALM render test cannot proceed. See "
        "test_run_pricing_button_populates_session_result above for the "
        "primary failure."
    )

    _navigate_to_section(at, "alm")
    assert_no_exceptions(at, context=f"render ALM section after {product_label} pricing run")


@pytest.mark.ui
@pytest.mark.parametrize(
    "product_value,product_label",
    [(c[0], c[1]) for c in _PRODUCT_RUN_CASES],
    ids=[c[0] for c in _PRODUCT_RUN_CASES],
)
def test_excel_replicator_section_renders_after_pricing_run(
    streamlit_apptest_module, product_value: str, product_label: str
) -> None:
    """Same as the ALM test but for the Excel Replicator page. This
    page builds and exposes the workbook, so a regression here means
    the user cannot download the Excel artifact -- the entire reason
    the app exists for many of its users.
    """
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    navigate_to_pricing_run(at)
    select_product(at, product_value)
    _set_deterministic_pricing_inputs(at)
    at.run()
    _click_run_pricing(at)
    assert _session_get(at, "pricing_res") is not None

    _navigate_to_section(at, "excel_replicator")
    assert_no_exceptions(
        at,
        context=f"render Excel Replicator after {product_label} pricing run",
    )


# ---------------------------------------------------------------------------
# Excel download surface must produce a workbook that passes strict validation
# ---------------------------------------------------------------------------


@pytest.mark.ui
@pytest.mark.parametrize(
    "product_value,product_label",
    [(c[0], c[1]) for c in _PRODUCT_RUN_CASES],
    ids=[c[0] for c in _PRODUCT_RUN_CASES],
)
def test_excel_download_workbook_passes_strict_validation(
    streamlit_apptest_module, product_value: str, product_label: str
) -> None:
    """The Excel workbook the UI builds for the download_button must
    pass the strict-mode workbook validator.

    AppTest does not expose ``st.download_button`` as a typed widget
    (it surfaces as ``UnknownElement``), so introspecting the button
    payload directly isn't possible. Instead we assert against the
    bytes that the UI stores in session_state under
    ``pricing_xlsx_bytes`` -- this is *exactly* what the
    ``download_button(data=...)`` argument receives. If those bytes
    pass validation, the user gets a workbook Excel can open; if they
    don't, the user downloads a file Excel rejects with "Removed
    Records: Formula" or similar.

    This is the UI-side complement to
    ``tests/test_excel_export_validation.py`` -- that file verifies
    the engine path; this verifies the path the *user* actually
    triggers.
    """
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    navigate_to_pricing_run(at)
    select_product(at, product_value)
    _set_deterministic_pricing_inputs(at)
    at.run()
    _click_run_pricing(at)
    assert _session_get(at, "pricing_res") is not None

    _navigate_to_section(at, "excel_replicator")
    assert_no_exceptions(at, context=f"render Excel Replicator for {product_label} download")

    build_err = _session_get(at, "pricing_xlsx_built_error")
    assert build_err is None, (
        f"Excel build for {product_label} failed with: {build_err!r}. "
        "The user would see an st.error on the Excel Replicator page "
        "instead of a download_button."
    )

    raw = _session_get(at, "pricing_xlsx_bytes")
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 1024, (
        f"{product_label} session_state['pricing_xlsx_bytes'] is "
        f"empty/missing ({type(raw).__name__}, "
        f"len={len(raw) if raw else 0}). The download_button would "
        "have nothing to serve."
    )

    wb = load_workbook(io.BytesIO(bytes(raw)), data_only=False)
    issues = validate_workbook(wb, strict=True)
    assert issues == [], (
        f"{product_label} workbook bytes destined for the UI "
        "download_button failed strict validation. The user would "
        "download a workbook Excel rejects with 'Removed Records: "
        "Formula' on open. Issues:\n"
        + "\n".join(f"  - {iss}" for iss in issues[:25])
        + ("" if len(issues) <= 25 else f"\n  ... {len(issues) - 25} more.")
    )


@pytest.mark.ui
def test_portfolio_manual_product_type_change_renders_without_exception(
    streamlit_apptest_module,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Changing a manual row's product must not assign session keys after widgets bind.

    Regression: ``_portfolio_push_defaults_to_session`` wrote
    ``portfolio_row_<id>_product_type`` after the product ``st.selectbox``
    for that key was already instantiated, which Streamlit rejects with
    ``StreamlitAPIException``. CSV-only portfolio tests never exercised
    this path.
    """
    monkeypatch.setenv("ANNUITY_MODEL_PORTFOLIO_V1", "1")
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    _navigate_to_section(at, "portfolio")
    assert_no_exceptions(at, context="render portfolio section")

    add_btn = [b for b in at.button if b.key == "portfolio_manual_add_button"]
    assert add_btn, "Add policy button missing"
    add_btn[0].click().run()
    assert_no_exceptions(at, context="add portfolio manual row")

    pt_boxes = [
        s
        for s in at.selectbox
        if getattr(s, "key", None)
        and str(s.key).startswith("portfolio_row_")
        and str(s.key).endswith("_product_type")
    ]
    assert pt_boxes, "manual portfolio product_type selectbox missing"
    pt_boxes[0].set_value("term_life").run()
    assert_no_exceptions(
        at,
        context="change portfolio manual row product from SPIA to term_life",
    )


@pytest.mark.ui
def test_portfolio_section_upload_run_and_workbook(
    streamlit_apptest_module,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Portfolio section: upload canonical inforce, run, workbook validates."""
    monkeypatch.setenv("ANNUITY_MODEL_PORTFOLIO_V1", "1")
    AppTest = streamlit_apptest_module
    at = load_pricing_ui(AppTest)
    _navigate_to_section(at, "portfolio")
    assert_no_exceptions(at, context="render portfolio section")

    csv_path = ROOT / "tests" / "data" / "inforce" / "example_v1" / "inforce.csv"
    raw_csv = csv_path.read_bytes()
    matched_fu = [fu for fu in at.file_uploader if fu.key == "portfolio_inforce_uploader"]
    assert matched_fu, "portfolio file uploader missing"
    matched_fu[0].upload("inforce.csv", raw_csv, mime_type="text/csv").run()

    run_btn = [b for b in at.button if b.key == "portfolio_run_button"]
    assert run_btn, "Run portfolio button missing"
    run_btn[0].click().run()
    assert_no_exceptions(at, context="portfolio pricing run")

    pres = _session_get(at, PORTFOLIO_KEY.RESULT)
    assert pres is not None, "portfolio_res not populated after Run portfolio"
    assert getattr(pres, "alm_result", None) is not None, (
        "baseline portfolio ALM should populate for canonical inforce (positive aggregate premium)"
    )
    proj_ms = [w for w in at.multiselect if str(w.key).startswith("portfolio_proj_series_")]
    assert proj_ms, "portfolio liability projection multiselect missing"
    wf_ms = [w for w in at.multiselect if str(w.key).startswith("portfolio_wf_series_")]
    assert wf_ms, "portfolio waterfall multiselect missing"

    xlsx = build_portfolio_workbook_bytes(pres)
    wb = load_workbook(io.BytesIO(bytes(xlsx)), data_only=False)
    issues = validate_workbook(wb, strict=True)
    assert issues == [], f"portfolio workbook failed validation: {issues[:5]}"
