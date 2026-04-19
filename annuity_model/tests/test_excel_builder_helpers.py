"""Unit tests for the declarative Excel-builder helpers (P1, 2026-04).

The helpers (:func:`write_inputs_sheet`, :func:`write_liability_summary_block`,
:func:`write_kv_rows`) collapse the three near-identical chunks of code
that previously lived inline in build_pricing/term/rila. These tests pin
their behavior so a future refactor can confidently rely on them.

We assert raw cell content (no formula recalculation) -- the helpers
operate purely on openpyxl Worksheet APIs and do not invoke Excel.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from excel_builder_helpers import (
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_kv_rows,
    write_liability_summary_block,
)

pytestmark = [pytest.mark.invariant]


# ---------------------------------------------------------------------------
# write_inputs_sheet
# ---------------------------------------------------------------------------


def test_write_inputs_sheet_writes_title_and_rows() -> None:
    wb = Workbook()
    ws = wb.active
    spec = InputsSheetSpec(
        title="My Title",
        rows=(("Issue Age", 65), ("Sex", "male"), ("Spread", 0.05)),
    )
    write_inputs_sheet(ws, spec)

    assert ws["A1"].value == "My Title"
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.size == 12.0

    assert ws["A3"].value == "Issue Age"
    assert ws["B3"].value == 65
    assert ws["A4"].value == "Sex"
    assert ws["B4"].value == "male"
    assert ws["A5"].value == "Spread"
    assert ws["B5"].value == 0.05


def test_write_inputs_sheet_writes_notes_when_provided() -> None:
    wb = Workbook()
    ws = wb.active
    spec = InputsSheetSpec(
        title="Inputs",
        rows=(("k1", 1),),
        notes=("First note.", "Second note."),
    )
    write_inputs_sheet(ws, spec)

    assert ws["D3"].value == "Notes"
    assert ws["D4"].value == "First note."
    assert ws["D5"].value == "Second note."


def test_write_inputs_sheet_skips_notes_when_empty() -> None:
    """Empty notes tuple must not write a "Notes" header (otherwise
    the SPIA refactor would no longer match its previous output)."""
    wb = Workbook()
    ws = wb.active
    spec = InputsSheetSpec(title="t", rows=(("k", 1),))
    write_inputs_sheet(ws, spec)
    assert ws["D3"].value is None


def test_write_inputs_sheet_accepts_formula_values() -> None:
    """Excel formula strings (starting with '=') must round-trip."""
    wb = Workbook()
    ws = wb.active
    spec = InputsSheetSpec(
        title="t",
        rows=(("Derived", "=B3*12"),),
    )
    write_inputs_sheet(ws, spec)
    assert ws["B3"].value == "=B3*12"


# ---------------------------------------------------------------------------
# write_liability_summary_block
# ---------------------------------------------------------------------------


def test_write_liability_summary_block_writes_header_and_rows() -> None:
    wb = Workbook()
    ws = wb.active
    spec = LiabilitySummaryBlockSpec(
        rows=(
            (4, "PV claims", "=SUM(T4:T100)"),
            (5, "PV premiums", "=SUM(U4:U100)"),
            (9, "Reserve at t=0", "=X7"),
        ),
    )
    write_liability_summary_block(ws, spec)

    assert ws["W3"].value == "Summary"
    assert ws["W3"].font.bold is True

    assert ws["W4"].value == "PV claims"
    assert ws["X4"].value == "=SUM(T4:T100)"
    assert ws["W5"].value == "PV premiums"
    assert ws["X5"].value == "=SUM(U4:U100)"
    assert ws["W9"].value == "Reserve at t=0"
    assert ws["X9"].value == "=X7"


def test_write_liability_summary_block_supports_alternate_columns() -> None:
    """Future products may want the block in different columns; the helper
    must honor label_column / value_column / header_row."""
    wb = Workbook()
    ws = wb.active
    spec = LiabilitySummaryBlockSpec(
        rows=((11, "label", "=1+1"),),
        label_column="AA",
        value_column="AB",
        header_row=10,
    )
    write_liability_summary_block(ws, spec)
    assert ws["AA10"].value == "Summary"
    assert ws["AA11"].value == "label"
    assert ws["AB11"].value == "=1+1"


# ---------------------------------------------------------------------------
# write_kv_rows
# ---------------------------------------------------------------------------


def test_write_kv_rows_returns_next_row_index() -> None:
    wb = Workbook()
    ws = wb.active
    next_row = write_kv_rows(
        ws,
        (("a", 1), ("b", 2), ("c", 3)),
        start_row=3,
    )
    assert next_row == 6
    assert ws["A3"].value == "a"
    assert ws["B5"].value == 3


def test_write_kv_rows_handles_empty_iterable() -> None:
    wb = Workbook()
    ws = wb.active
    next_row = write_kv_rows(ws, (), start_row=10)
    # No rows written -> "next" stays at start.
    assert next_row == 10
    assert ws["A10"].value is None
