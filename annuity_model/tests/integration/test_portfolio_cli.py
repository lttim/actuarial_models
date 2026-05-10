"""CLI portfolio-run integration vs golden summary JSON."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
INFORCE = ROOT / "tests/data/inforce/example_v1/inforce.csv"
EXPECTED = ROOT / "tests/data/inforce/example_v1/expected_summary.json"


@pytest.mark.integration
def test_cli_portfolio_run_matches_expected_summary(tmp_path: Path) -> None:
    out = tmp_path / "out"
    cmd = [
        sys.executable,
        "-m",
        "annuity_model.cli",
        "portfolio-run",
        "--inforce",
        str(INFORCE),
        "--out",
        str(out),
    ]
    env = {
        **__import__("os").environ,
        "PYTHONPATH": str(ROOT / "src"),
        "ANNUITY_MODEL_PORTFOLIO_V1": "1",
    }
    subprocess.run(cmd, check=True, cwd=str(ROOT), env=env)
    got = json.loads((out / "portfolio_summary.json").read_text(encoding="utf-8"))
    exp = json.loads(EXPECTED.read_text(encoding="utf-8"))
    assert got == exp
    assert (out / "run_ledger.sqlite3").is_file()
    wb = load_workbook(out / "portfolio.xlsx", data_only=False)
    assert "RunLedger" in wb.sheetnames
    assert "AssumptionEvidence" in wb.sheetnames
