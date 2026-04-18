from __future__ import annotations

import math
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import pricing_projection as sp
import rila_projection as rp
from excel_builder_helpers import (
    ALM_ENGINE_STEP_MONTHS,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_PROJECTION_FIRST_DATA_ROW,
    LIABILITY_SHEET_NAME,
    ALMExcelSnapshot,
    ExcelPythonSnapshot,
    alm_excel_downsample_snapshot,
    alm_excel_truncate_snapshot,
    inject_alm_projection_formula_cached_values,
    liability_layout_for,
    write_alm_projection_sheet,
    write_model_check_sheet,
)
from recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

RILA_PROJ_MAX_ROWS = 600
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_MTH_QX = "MortalMonthly"
SHEET_INDEX = "IndexScenario"

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_PART = 4
_IN_ROW_CAP = 5
_IN_ROW_FLOOR = 6
_IN_ROW_RIDER = 7
_IN_ROW_FREQ = 8
_IN_ROW_VAL = 9
_IN_ROW_HORIZON = 10
_IN_ROW_SPREAD = 11
_IN_ROW_PREMIUM = 12
_IN_ROW_EXP_MTH = 13
_IN_ROW_EXP_INF = 14
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


def _n_months_cell() -> str:
    return _in_addr("B", _IN_ROW_NMONTHS)


def _fill_mortal_monthly_rpmp(
    ws_m,
    *,
    mort: sp.MortalityTableRP2014MP2016,
    issue_age: int,
    valuation_year: int,
    n_months: int,
) -> None:
    dt = 1.0 / 12.0
    for k in range(1, n_months + 1):
        r = 1 + k
        m_index = k - 1
        age_start = issue_age + m_index * dt
        age_int = int(math.floor(age_start))
        calendar_year_start = valuation_year + 1 + (m_index // 12)
        qxv = float(
            mort.qx_at_int_age_and_calendar_year(age_int=age_int, calendar_year=calendar_year_start)
        )
        ws_m.cell(row=r, column=1, value=f'=IF(ROW()-1>{_n_months_cell()},"",ROW()-1)')
        ws_m.cell(row=r, column=2, value=int(age_int))
        ws_m.cell(row=r, column=3, value=int(calendar_year_start))
        ws_m.cell(row=r, column=4, value=float(qxv))
    last_data = 1 + n_months
    last_cap = 1 + RILA_PROJ_MAX_ROWS
    for r in range(last_data + 1, last_cap + 1):
        ws_m.cell(row=r, column=1, value=f'=IF(ROW()-1>{_n_months_cell()},"",ROW()-1)')
        for c in (2, 3, 4):
            ws_m.cell(row=r, column=c, value="")


def _qx_lookup_expr(acell: str, mode: Literal["qx_table", "mortal_monthly"]) -> str:
    clamp_inner = "MIN(MAX({inner},0),0.999999)"
    if mode == "qx_table":
        inner = (
            f"INDEX({SHEET_QX}!$B$2:$B$50000,"
            f"MATCH(INT({_in_addr('B', _IN_ROW_ISSUE_AGE)}+({acell}-1)/12),"
            f"{SHEET_QX}!$A$2:$A$50000,0))"
        )
        return clamp_inner.format(inner=inner)
    inner = f"INDEX({SHEET_MTH_QX}!$D$2:$D$50000,MATCH({acell},{SHEET_MTH_QX}!$A$2:$A$50000,0))"
    return clamp_inner.format(inner=inner)


def _survival_end_formula(
    r: int, acell: str, mode: Literal["qx_table", "mortal_monthly"], *, first_row: int
) -> str:
    qx_e = _qx_lookup_expr(acell, mode)
    p_m = f"EXP(-(-LN(1-{qx_e}))/12)"
    if r == first_row:
        return f'=IF({acell}="","",{p_m})'
    prev = f"D{r-1}"
    return f'=IF({acell}="","",{prev}*{p_m})'


@dataclass(frozen=True)
class RILAExcelBuildSpec:
    contract: rp.RILAContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float
    index_s0: float
    index_levels_payment: np.ndarray


def rila_excel_spec_from_launcher(
    *,
    contract: rp.RILAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
) -> RILAExcelBuildSpec:
    return RILAExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
        index_s0=float(index_s0),
        index_levels_payment=np.asarray(index_levels_at_payment, dtype=float),
    )


def build_rila_workbook_from_spec(
    spec: RILAExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
    alm_snapshot: ALMExcelSnapshot | None = None,
    alm_assumptions: sp.ALMAssumptions | None = None,
) -> bytes:
    s0_py = float(spec.index_s0)
    levels_py = np.asarray(spec.index_levels_payment, dtype=float)

    res = rp.price_rila_single_premium(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=(
            spec.valuation_year if not isinstance(spec.mortality, sp.MortalityTableQx) else None
        ),
        expenses=spec.expenses,
        index_s0=s0_py,
        index_levels_payment=levels_py,
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    ws_in["A1"] = "RILA Inputs (matches model launcher / Python)"
    ws_in["A1"].font = Font(bold=True, size=12)
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Participation", spec.contract.participation),
        ("Cap (decimal / yr segment)", spec.contract.cap),
        ("Floor (decimal / yr segment)", spec.contract.floor),
        ("Rider fee (annual, on AV)", spec.contract.rider_fee_annual),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Single premium (Python export)", float(res.single_premium)),
        ("Monthly expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws_in[f"A{i}"] = k
        ws_in[f"B{i}"] = v

    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{RILA_PROJ_MAX_ROWS})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(
        ws_mc_curve,
        n_months=RILA_PROJ_MAX_ROWS,
        y_last_row=y_last_row,
        payments_per_year_ref=_in_addr("B", _IN_ROW_FREQ),
        spread_ref=_in_addr("B", _IN_ROW_SPREAD),
    )

    ws_ix = wb.create_sheet(SHEET_INDEX)
    ws_ix["A1"] = "month"
    ws_ix["B1"] = "index_level"
    L = rp.levels_end_by_policy_month(s0=float(s0_py), levels_payment=levels_py)
    for j in range(L.shape[0]):
        ws_ix.cell(row=2 + j, column=1, value=int(j))
        ws_ix.cell(row=2 + j, column=2, value=float(L[j]))

    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx[i]))
        mort_mode: Literal["qx_table", "mortal_monthly"] = "qx_table"
    else:
        mort_mode = "mortal_monthly"
        ws_m = wb.create_sheet(SHEET_MTH_QX)
        ws_m["A1"] = "month"
        ws_m["B1"] = "age_int"
        ws_m["C1"] = "calendar_year_start"
        ws_m["D1"] = "qx_annual"
        _fill_mortal_monthly_rpmp(
            ws_m,
            mort=spec.mortality,
            issue_age=spec.contract.issue_age,
            valuation_year=int(spec.valuation_year),
            n_months=min(int(res.months.shape[0]), RILA_PROJ_MAX_ROWS),
        )

    nm_ref = _n_months_cell()
    last_cap_row = 3 + RILA_PROJ_MAX_ROWS
    first = 4
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"
    idx_b = f"{SHEET_INDEX}!$B$2:$B$10000"
    idx_a = f"{SHEET_INDEX}!$A$2:$A$10000"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "RILA liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={_in_addr('B', _IN_ROW_ISSUE_AGE)}"
    ws_pr["V2"] = "=X9"

    hdr = (
        "Month",
        "t_years",
        "AttainedAge",
        "SurvivalEnd",
        "SurvivalStart",
        "MonthDeathProb",
        "IndexLevel_m",
        "RawSegReturn",
        "CreditedSeg",
        "AV_end",
        "ExpBenefitCF",
        "ExpExpenseCF",
        "ExpTotalCF",
        "",
        "DiscountFactor",
        "PVBenefitCF",
        "PVExpenseCF",
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    ws_pr.cell(row=3, column=10, value=0.0)

    part = _in_addr("B", _IN_ROW_PART)
    cap = _in_addr("B", _IN_ROW_CAP)
    fl = _in_addr("B", _IN_ROW_FLOOR)
    rider = _in_addr("B", _IN_ROW_RIDER)
    prem = _in_addr("B", _IN_ROW_PREMIUM)
    exp_m = _in_addr("B", _IN_ROW_EXP_MTH)
    exp_if = _in_addr("B", _IN_ROW_EXP_INF)

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f"=IF({a}=\"\",\"\",{a}/{_in_addr('B', _IN_ROW_FREQ)})")
        ws_pr.cell(
            row=r,
            column=3,
            value=f"=IF({a}=\"\",\"\",{_in_addr('B', _IN_ROW_ISSUE_AGE)}+({a}-1)/{_in_addr('B', _IN_ROW_FREQ)})",
        )
        ws_pr.cell(row=r, column=4, value=_survival_end_formula(r, a, mort_mode, first_row=first))
        if r == first:
            d_surv_start = f'=IF({a}="","",1)'
        else:
            d_surv_start = f'=IF({a}="","",D{r-1})'
        ws_pr.cell(row=r, column=5, value=d_surv_start)
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(
            row=r,
            column=7,
            value=(f'=IF({a}="","",IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0)),""))'),
        )
        ws_pr.cell(
            row=r,
            column=8,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0))/INDEX({idx_b},MATCH({a}-12,{idx_a},0))-1,0),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=9,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"MAX({fl},MIN({cap},{part}*H{r})),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=10,
            value=(f'=IF({a}="","",' f"(IF({a}=1,{prem},J{r-1})*(1+I{r}))*(1-{rider}/12))"),
        )
        ws_pr.cell(row=r, column=11, value=f'=IF({a}="",0,F{r}*J{r})')
        ws_pr.cell(
            row=r,
            column=12,
            value=(f'=IF({a}="",0,{exp_m}*' f"POWER(1+((1+{exp_if})^(1/12)-1),{a}-1)*D{r})"),
        )
        ws_pr.cell(row=r, column=13, value=f'=IF({a}="",0,K{r}+L{r})')
        ws_pr.cell(
            row=r,
            column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=16, value=f'=IF({a}="",0,K{r}*O{r})')
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,L{r}*O{r})')

    money_cols = (7, 8, 9, 10, 11, 12, 13, 16, 17)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 15):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    ws_pr["W3"] = "Summary"
    ws_pr["W3"].font = Font(bold=True)
    ws_pr["W4"] = "PV benefits (claims)"
    ws_pr["X4"] = f"=SUM(P{first}:P{last_cap_row})"
    ws_pr["W5"] = "PV expenses"
    ws_pr["X5"] = f"=SUM(Q{first}:Q{last_cap_row})"
    ws_pr["W6"] = "Σ l_end · v (annuity-style factor)"
    ws_pr["X6"] = f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})"
    ws_pr["W7"] = "PV total (ben+exp)"
    ws_pr["X7"] = "=X4+X5"
    ws_pr["W8"] = "Single premium (export)"
    ws_pr["X8"] = f"={SHEET_INPUTS}!$B${_IN_ROW_PREMIUM}"
    ws_pr["W9"] = "Reserve at t=0"
    ws_pr["X9"] = "=X7"

    alm_layout = None
    alm_snap_for_book = None
    if alm_snapshot is not None:
        if alm_assumptions is None:
            raise ValueError("alm_assumptions is required when alm_snapshot is provided.")
        alm_snap_for_book = alm_excel_downsample_snapshot(alm_snapshot, int(ALM_ENGINE_STEP_MONTHS))
        alm_snap_for_book = alm_excel_truncate_snapshot(alm_snap_for_book, ALM_EXCEL_PATH_MONTH_CAP)
        # Liability column letters live in liability_layouts.LIABILITY_LAYOUTS
        # so the validator and parity tests can cross-check them. We pass the
        # string code rather than ProductType to avoid a registry-builder cycle.
        _rila_layout = liability_layout_for("rila")
        alm_layout = write_alm_projection_sheet(
            wb,
            alm_snap_for_book,
            alm_assumptions,
            n_months=int(res.months.shape[0]),
            y_last_row=int(y_last_row),
            engine_step_months=int(ALM_ENGINE_STEP_MONTHS),
            yield_curve_spread_ref=_in_addr("B", _IN_ROW_SPREAD),
            liability_total_col=_rila_layout.total_cf_col,
            liability_discount_col=_rila_layout.discount_col,
        )

    snap_py = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    rila_rows: list[tuple[str, float, str, str]] = [
        ("PV benefits", float(res.pv_benefit), f"={LIABILITY_SHEET_NAME}!X4", "money"),
        (
            "PV monthly expenses",
            float(res.pv_monthly_expenses),
            f"={LIABILITY_SHEET_NAME}!X5",
            "money",
        ),
        (
            "PV monthly total (ben+exp)",
            float(res.pv_benefit + res.pv_monthly_expenses),
            f"={LIABILITY_SHEET_NAME}!X7",
            "money",
        ),
        ("Single premium", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Annuity factor", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=alm_layout,
        alm_snapshot=alm_snap_for_book,
        pricing_rows=rila_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME}; optional ALM_Projection)",
        subtitle=(
            "RILA: column B is Python at export; column C references Liabilities summary. "
            "IndexScenario holds L[month] for end-of-month index levels. "
            "Premium in Inputs is the priced single premium from Python."
        ),
    )

    from excel_workbook_validator import validate_workbook_or_raise

    validate_workbook_or_raise(wb)

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if alm_snap_for_book is not None:
        data = inject_alm_projection_formula_cached_values(
            data,
            first_data_row=int(ALM_PROJECTION_FIRST_DATA_ROW),
            snap=alm_snap_for_book,
        )

    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
