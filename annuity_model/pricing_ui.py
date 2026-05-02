"""
Unified Streamlit workspace for the SPIA model: overview, configurable pricing run,
interactive charts, and embedded unit-test dashboard.

Run from the annuity_model folder:
    streamlit run pricing_ui.py
Or: run_pricing_ui.bat (Windows) / ./run_pricing_ui.sh (macOS, Linux).

New Pricing Run numeric inputs and ``run_*`` session keys: extend
``pricing_run_form_state.build_run_form_seed_defaults`` and use
``run_number_input`` / ``ensure_session_choice`` so Streamlit does not default
widgets to ``min_value`` on first paint.
"""

from __future__ import annotations

import dataclasses
import datetime as _dt
import io
import json
import os
import sys
import time
import uuid
from collections.abc import MutableMapping
from pathlib import Path
from typing import Any, Literal

os.environ.setdefault("MPLBACKEND", "Agg")

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pricing_projection as sp
import rila_projection as rp
import term_projection as tp
from alm_excel_ladder import ALM_ENGINE_SHEET
from build_pricing_excel_workbook import (
    ALM_ENGINE_FIELD_GUIDE_SHEET,
    ALM_ENGINE_STEP_MONTHS,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_PROJECTION_FIRST_DATA_ROW,
    ALM_SHEET_NAME,
    LIABILITY_SHEET_NAME,
    ALMExcelSnapshot,
    ExcelPythonSnapshot,
    MCExcelSnapshot,
    alm_excel_downsample_snapshot,
    alm_excel_snapshot_from_result,
    alm_excel_truncate_snapshot,
    mc_excel_snapshot_from_result,
)
from build_portfolio_excel_workbook import build_portfolio_workbook_bytes
from inforce_io import load_policy_inputs_from_csv
from liability_aggregation import padded_cashflows_on_portfolio_grid
from portfolio import PolicyInput, Portfolio, PortfolioResult, RunScenario
from portfolio_config import (
    portfolio_disabled_explanation_markdown,
    portfolio_sidebar_visible,
    portfolio_v1_enabled,
)
from portfolio_runner import run_portfolio
from portfolio_summary import portfolio_result_to_summary_dict
from pricing_scenario_materialize import (
    build_mortality_from_seeds,
    build_yield_curve_from_seeds,
    run_scenario_for_portfolio_policies,
)
from pricing_run_form_state import (
    PORTFOLIO_INFORCE_SCRATCH_COLUMNS,
    PORTFOLIO_KEY,
    PRICING_RUN_NUMBER_INPUT_KEYS,
    RUN_KEY,
    build_run_form_seed_defaults,
    default_inforce_scratch_row,
    ensure_session_choice,
    run_number_input,
)
from product_excel import build_product_workbook
from product_registry import (
    ProductType,
    get_mortality_mode_label,
    get_pricing_metrics,
    get_product_adapter,
    get_product_capabilities,
    get_product_default_mortality_mode,
    get_product_mortality_mode_options,
    get_product_ui_config,
    get_term_contract_ui_config,
    parse_term_benefit_timing_label,
    parse_term_length_label_to_years,
    parse_term_premium_mode_label,
    product_label,
    product_options_for_ui,
)
from products.indexed_ul.ui import (
    build_indexed_ul_contract_from_session,
    render_indexed_ul_pricing_controls,
)
from products.rila.ui import build_rila_contract_from_session, render_rila_pricing_controls
from test_dashboard import render_unit_tests_page


def _maybe_alm_excel_snapshot_for_workbook() -> ALMExcelSnapshot | None:
    alm = st.session_state.get("alm_last")
    asm = st.session_state.get("alm_last_assumptions")
    if not isinstance(alm, sp.ALMResult) or not isinstance(asm, sp.ALMAssumptions):
        return None
    if st.session_state.get("alm_last_pricing_run_id") != st.session_state.get("pricing_run_id"):
        return None
    aum_tag = st.session_state.get("alm_last_initial_asset_market_value")
    return alm_excel_snapshot_from_result(
        alm,
        asm,
        initial_asset_market_value=float(aum_tag) if aum_tag is not None else None,
    )


def _refresh_pricing_excel_workbook_in_session() -> None:
    """Rebuild `pricing_xlsx_bytes` from the current pricing result and optional MC/ALM session state."""
    res = st.session_state.get("pricing_res")
    contract = st.session_state.get("pricing_contract")
    ctx = st.session_state.get("pricing_excel_context") or {}
    if res is None or contract is None:
        return
    yc = ctx.get("yield_curve")
    mort = ctx.get("mortality")
    if not isinstance(yc, sp.YieldCurve) or not isinstance(
        mort, (sp.MortalityTableQx, sp.MortalityTableRP2014MP2016)
    ):
        return
    expenses = ctx.get("expenses")
    if not isinstance(expenses, sp.ExpenseAssumptions):
        return
    meta = st.session_state.get("pricing_meta") or {}
    product_raw = st.session_state.get("pricing_product_type", ProductType.SPIA.value)
    try:
        product_type = ProductType(str(product_raw))
    except ValueError:
        product_type = ProductType.SPIA
    adapter = get_product_adapter(product_type)
    vy_raw = ctx.get("valuation_year")
    vy = int(vy_raw) if vy_raw is not None else 2025
    mc_snap: MCExcelSnapshot | None = None
    mc = st.session_state.get("pricing_mc")
    mc_params = st.session_state.get("pricing_mc_params") or {}
    if mc is not None and hasattr(mc, "single_premium") and hasattr(mc, "annuity_factor"):
        # SPIA / RILA-shaped MC results only; new-product MC results omit
        # those scalar fields (they expose pv_benefit + per-path arrays
        # only in v1).
        mc_snap = mc_excel_snapshot_from_result(
            mc,
            annual_drift=float(mc_params.get("annual_drift", 0.06)),
            annual_vol=float(mc_params.get("annual_vol", 0.15)),
            s0=float(mc_params.get("s0", 100.0)),
        )
    alm_snap = _maybe_alm_excel_snapshot_for_workbook()
    alm_asm = st.session_state.get("alm_last_assumptions")
    try:
        spec = adapter.excel_spec_from_run(
            contract=contract,
            yield_curve=yc,
            mortality=mort,
            horizon_age=int(ctx.get("horizon_age", 110)),
            spread=float(ctx.get("spread", 0.0)),
            valuation_year=vy,
            expenses=expenses,
            yield_mode_label=str(meta.get("yield_mode", "")),
            mortality_mode_label=str(meta.get("mortality_mode", "")),
            expense_mode_label=str(meta.get("expense_mode", "")),
            index_s0=float(res.index_s0),
            index_levels_at_payment=res.index_level_at_payment,
            expense_annual_inflation=float(res.expense_annual_inflation),
        )
        st.session_state["pricing_xlsx_bytes"] = build_product_workbook(
            product_type=product_type,
            spec=spec,
            out_path=None,
            python_snapshot=ExcelPythonSnapshot(
                pv_benefit=float(res.pv_benefit),
                pv_monthly_expenses=float(res.pv_monthly_expenses),
                pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
                single_premium=float(res.single_premium),
                annuity_factor=float(res.annuity_factor),
            ),
            mc_snapshot=mc_snap,
            alm_snapshot=alm_snap,
            alm_assumptions=alm_asm if isinstance(alm_asm, sp.ALMAssumptions) else None,
        )
        st.session_state["pricing_xlsx_has_mc"] = mc_snap is not None
        st.session_state["pricing_xlsx_has_alm"] = alm_snap is not None
        st.session_state.pop("pricing_xlsx_built_error", None)
    except Exception as ex:
        st.session_state["pricing_xlsx_bytes"] = None
        st.session_state.pop("pricing_xlsx_has_mc", None)
        st.session_state.pop("pricing_xlsx_has_alm", None)
        st.session_state["pricing_xlsx_built_error"] = repr(ex)


def _ensure_excel_workbook_includes_current_alm() -> None:
    """If ALM completed after the last Excel build, regenerate the workbook so download includes ALM_Projection."""
    if not isinstance(st.session_state.get("pricing_xlsx_bytes"), bytes):
        return
    want_alm = _maybe_alm_excel_snapshot_for_workbook() is not None
    has_alm = bool(st.session_state.get("pricing_xlsx_has_alm", False))
    if want_alm != has_alm:
        _refresh_pricing_excel_workbook_in_session()


def _resolve_path(p: str) -> Path:
    path = Path(p.strip())
    if path.is_absolute():
        return path
    return (ROOT / path).resolve()


def _round_for_visuals(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    numeric_cols = out.select_dtypes(include=[np.number]).columns
    out.loc[:, numeric_cols] = out.loc[:, numeric_cols].round(0)
    return out


def _alm_surplus_chart(ages: np.ndarray | pd.Series, surplus: np.ndarray | pd.Series) -> None:
    """Surplus vs attained age with a y = 0 reference line (above / below zero)."""
    df = pd.DataFrame(
        {
            "Attained age": np.asarray(ages, dtype=float),
            "Surplus": np.asarray(surplus, dtype=float),
        }
    )
    line = (
        alt.Chart(df)
        .mark_line()
        .encode(
            x=alt.X("Attained age:Q", title="Attained age"),
            y=alt.Y("Surplus:Q", title="Surplus ($)"),
        )
    )
    rule = (
        alt.Chart(pd.DataFrame({"y": [0.0]}))
        .mark_rule(color="#888", strokeDash=[4, 4])
        .encode(y="y:Q")
    )
    layered = (
        (line + rule)
        .properties(
            title="Surplus (asset market value minus liability PV)",
            height=320,
        )
        .resolve_scale(y="shared")
    )
    st.altair_chart(layered.interactive(), use_container_width=True)


def _number_cols_no_decimals(df: pd.DataFrame) -> dict[str, st.column_config.NumberColumn]:
    numeric_cols = list(df.select_dtypes(include=[np.number]).columns)
    return {c: st.column_config.NumberColumn(format="%,.0f") for c in numeric_cols}


def _pop_session_keys(keys: list[str]) -> None:
    for k in keys:
        st.session_state.pop(k, None)


def _invalidate_diagnostics_export() -> None:
    _pop_session_keys(["diagnostics_json_bytes", "diagnostics_json_filename"])


def _clear_dependent_state_on_pricing_change() -> None:
    # ALM and What-if artifacts are pricing-baseline dependent; clear on pricing changes.
    _pop_session_keys(
        [
            "alm_last",
            "alm_last_assumptions",
            "alm_last_initial_asset_market_value",
            "alm_current_assumptions",
            "alm_current_initial_asset_market_value",
            "whatif_last_params",
            "whatif_last_base_res",
            "whatif_last_shocked_res",
            "whatif_last_baseline_mc",
            "whatif_last_shocked_mc",
            "whatif_last_shocked_curve",
            "whatif_last_shocked_mortality",
            "whatif_last_alm_base",
            "whatif_last_alm_after",
            "whatif_last_alm_assumptions",
            "whatif_last_pricing_run_id",
            "alm_last_pricing_run_id",
        ]
    )
    st.session_state.pop("what_if_mc_cache", None)
    _invalidate_diagnostics_export()


def _serialize_array(arr: Any, *, include_full: bool, max_points: int = 250) -> Any:
    """Serialize numpy-like arrays for JSON with optional truncation."""
    if arr is None:
        return None
    a = np.asarray(arr)
    if include_full or a.size <= max_points:
        return a.tolist()
    # Keep file sizes manageable while still giving a shape + endpoints.
    head_n = min(10, a.size)
    tail_n = min(10, a.size)
    return {
        "truncated": True,
        "len": int(a.size),
        "shape": list(a.shape),
        "head": a[:head_n].tolist(),
        "tail": a[-tail_n:].tolist(),
    }


def _contract_to_dict(
    contract: sp.SPIAContract | tp.TermLifeContract | rp.RILAContract,
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "issue_age": int(contract.issue_age),
        "sex": str(contract.sex),
        "benefit_annual": float(contract.benefit_annual),
        "benefit_timing": str(getattr(contract, "benefit_timing", "")),
        "payment_freq_per_year": int(getattr(contract, "payment_freq_per_year", 1)),
        "payment_cessation": str(getattr(contract, "payment_cessation", "")),
    }
    if isinstance(contract, tp.TermLifeContract):
        out["death_benefit"] = float(contract.death_benefit)
        out["monthly_premium"] = float(contract.monthly_premium)
        out["term_years"] = int(contract.term_years)
        out["premium_mode"] = str(contract.premium_mode)
    if isinstance(contract, rp.RILAContract):
        out["rila_participation"] = float(contract.participation)
        out["rila_cap"] = float(contract.cap)
        out["rila_floor"] = float(contract.floor)
        out["rila_rider_fee_annual"] = float(contract.rider_fee_annual)
    return out


def _yield_curve_to_dict(yc: sp.YieldCurve) -> dict[str, Any]:
    return {
        "maturities_years": _serialize_array(yc.maturities_years, include_full=True),
        "zero_rates": _serialize_array(yc.zero_rates, include_full=True),
    }


def _mortality_to_dict(mort: Any) -> dict[str, Any]:
    out: dict[str, Any] = {}
    if hasattr(mort, "ages"):
        out["ages"] = _serialize_array(mort.ages, include_full=True)
    if hasattr(mort, "qx"):
        out["qx"] = _serialize_array(mort.qx, include_full=True)
    if hasattr(mort, "qx_at_int_age"):
        out["qx_at_int_age"] = _serialize_array(mort.qx_at_int_age, include_full=True)
    out["type"] = type(mort).__name__
    return out


def _pricing_result_to_dict(
    res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
    contract_state: sp.SPIAContract | tp.TermLifeContract | rp.RILAContract,
    *,
    include_full: bool,
) -> dict[str, Any]:
    return {
        "contract": _contract_to_dict(contract_state),
        "single_premium": float(res.single_premium),
        "pv_benefit": float(res.pv_benefit),
        "pv_monthly_expenses": float(res.pv_monthly_expenses),
        "annuity_factor": float(res.annuity_factor),
        "times_years": _serialize_array(res.times_years, include_full=include_full),
        "months": _serialize_array(res.months, include_full=include_full),
        "expected_total_cashflows": _serialize_array(
            res.expected_total_cashflows, include_full=include_full
        ),
        "economic_reserve": _serialize_array(res.economic_reserve, include_full=include_full),
        "survival_to_payment": _serialize_array(res.survival_to_payment, include_full=include_full),
        # Index / inflation scaffolding (needed for full Before/After diagnostics).
        "index_s0": float(res.index_s0),
        "index_level_at_payment": _serialize_array(
            res.index_level_at_payment, include_full=include_full
        ),
        "index_simple_return": _serialize_array(res.index_simple_return, include_full=include_full),
        "index_log_return": _serialize_array(res.index_log_return, include_full=include_full),
        "index_cumulative_return": _serialize_array(
            res.index_cumulative_return, include_full=include_full
        ),
    }


def _alm_result_to_dict(
    alm: sp.ALMResult, asm: sp.ALMAssumptions | None, *, include_buckets: bool, include_full: bool
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "assumptions": None,
        "month_index": _serialize_array(alm.month_index, include_full=True),
        "times_years": _serialize_array(alm.times_years, include_full=include_full),
        "asset_market_value": _serialize_array(alm.asset_market_value, include_full=include_full),
        "liability_pv": _serialize_array(alm.liability_pv, include_full=include_full),
        "surplus": _serialize_array(alm.surplus, include_full=include_full),
        "funding_ratio": _serialize_array(alm.funding_ratio, include_full=include_full),
        "liquidity_buffer_months": _serialize_array(
            alm.liquidity_buffer_months, include_full=include_full
        ),
        "borrowing_balance": _serialize_array(alm.borrowing_balance, include_full=include_full),
        "pv01_assets": float(alm.pv01_assets),
        "pv01_liabilities": float(alm.pv01_liabilities),
        "pv01_net": float(alm.pv01_net),
        "duration_assets_mac": float(alm.duration_assets_mac),
        "duration_liabilities_mac": float(alm.duration_liabilities_mac),
        "duration_gap": float(alm.duration_gap),
    }
    if asm is not None:
        out["assumptions"] = _alm_assumptions_to_dict(asm)
    if include_buckets:
        out["bucket_asset_mv"] = _serialize_array(alm.bucket_asset_mv, include_full=True)
    else:
        out["bucket_asset_mv"] = {
            "shape": list(alm.bucket_asset_mv.shape),
        }
    return out


def _alm_assumptions_to_dict(asm: sp.ALMAssumptions) -> dict[str, Any]:
    return {
        "rebalance_band": float(asm.rebalance_band),
        "rebalance_frequency_months": int(asm.rebalance_frequency_months),
        "reinvest_rule": str(asm.reinvest_rule),
        "disinvest_rule": str(asm.disinvest_rule),
        "rebalance_policy": str(asm.rebalance_policy),
        "borrowing_policy": str(asm.borrowing_policy),
        "borrowing_rate_mode": str(asm.borrowing_rate_mode),
        "borrowing_rate_tenor_years": float(asm.borrowing_rate_tenor_years),
        "borrowing_spread_annual": float(asm.borrowing_spread_annual),
        "borrowing_rate_annual": float(asm.borrowing_rate_annual),
        "liquidity_near_liquid_years": float(asm.liquidity_near_liquid_years),
        "allocation": {
            "buckets": [
                {"name": b.name, "tenor_years": float(b.tenor_years)}
                for b in asm.allocation.buckets
            ],
            "weights": _serialize_array(asm.allocation.weights, include_full=True),
        },
    }


def _whatif_result_to_dict(
    *,
    base_res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult,
    shocked_res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult,
    baseline_mc: Any,
    shocked_mc: Any,
    whatif_params: dict[str, Any],
    alm_base: sp.ALMResult | None,
    alm_after: sp.ALMResult | None,
    asm: sp.ALMAssumptions | None,
    include_full: bool,
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "whatif_params": whatif_params,
        "base": {
            "single_premium": float(base_res.single_premium),
            "pv_benefit": float(base_res.pv_benefit),
            "pv_monthly_expenses": float(base_res.pv_monthly_expenses),
            "economic_reserve_issue": (
                float(base_res.economic_reserve[0]) if base_res.economic_reserve.size else None
            ),
            "times_years": _serialize_array(base_res.times_years, include_full=include_full),
            "economic_reserve": _serialize_array(
                base_res.economic_reserve, include_full=include_full
            ),
            "index_s0": float(base_res.index_s0),
            "index_level_at_payment": _serialize_array(
                base_res.index_level_at_payment, include_full=include_full
            ),
            "index_simple_return": _serialize_array(
                base_res.index_simple_return, include_full=include_full
            ),
            "index_log_return": _serialize_array(
                base_res.index_log_return, include_full=include_full
            ),
            "index_cumulative_return": _serialize_array(
                base_res.index_cumulative_return, include_full=include_full
            ),
        },
        "after": {
            "single_premium": float(shocked_res.single_premium),
            "pv_benefit": float(shocked_res.pv_benefit),
            "pv_monthly_expenses": float(shocked_res.pv_monthly_expenses),
            "economic_reserve_issue": (
                float(shocked_res.economic_reserve[0])
                if shocked_res.economic_reserve.size
                else None
            ),
            "times_years": _serialize_array(shocked_res.times_years, include_full=include_full),
            "economic_reserve": _serialize_array(
                shocked_res.economic_reserve, include_full=include_full
            ),
            "index_s0": float(shocked_res.index_s0),
            "index_level_at_payment": _serialize_array(
                shocked_res.index_level_at_payment, include_full=include_full
            ),
            "index_simple_return": _serialize_array(
                shocked_res.index_simple_return, include_full=include_full
            ),
            "index_log_return": _serialize_array(
                shocked_res.index_log_return, include_full=include_full
            ),
            "index_cumulative_return": _serialize_array(
                shocked_res.index_cumulative_return, include_full=include_full
            ),
        },
        "tail_risk_mc": {
            "baseline": {
                "n_sims": int(getattr(baseline_mc, "n_sims", 0)),
                "premium_mean": float(getattr(baseline_mc, "premium_mean", float("nan"))),
                "premium_median": float(getattr(baseline_mc, "premium_median", float("nan"))),
                "premium_p05": float(getattr(baseline_mc, "premium_p05", float("nan"))),
                "premium_p95": float(getattr(baseline_mc, "premium_p95", float("nan"))),
            },
            "after": {
                "n_sims": int(getattr(shocked_mc, "n_sims", 0)),
                "premium_mean": float(getattr(shocked_mc, "premium_mean", float("nan"))),
                "premium_median": float(getattr(shocked_mc, "premium_median", float("nan"))),
                "premium_p05": float(getattr(shocked_mc, "premium_p05", float("nan"))),
                "premium_p95": float(getattr(shocked_mc, "premium_p95", float("nan"))),
            },
        },
    }
    if alm_base is not None:
        # Always include bucket time series for diagnostics completeness.
        out["alm_base"] = _alm_result_to_dict(
            alm_base, asm, include_buckets=True, include_full=include_full
        )
    else:
        out["alm_base"] = None
    if alm_after is not None:
        # Always include bucket time series for diagnostics completeness.
        out["alm_after"] = _alm_result_to_dict(
            alm_after, asm, include_buckets=True, include_full=include_full
        )
    else:
        out["alm_after"] = None
    return out


MortalityMode = Literal[
    "synthetic",
    "qx_csv",
    "rp2014_mp2016",
    "us_ssa_2015_period",
    "cso_2017_ult",
]
YieldMode = Literal["flat", "zero_csv", "par_bootstrap"]
ExpenseMode = Literal["csv", "manual"]

SECTION_LABELS: dict[str, str] = {
    "overview": "Overview",
    "run": "Pricing Run",
    "portfolio": "Portfolio (multi-policy)",
    "alm": "ALM",
    "what_if": "What-if Analysis",
    "excel_replicator": "Excel Replicator",
    "tests": "Unit Tests",
}
SECTION_ORDER: list[str] = [
    "overview",
    "run",
    "alm",
    "what_if",
    "excel_replicator",
    "tests",
]


def _dynamic_overview_features() -> list[str]:
    options = list(product_options_for_ui())
    available_products = ", ".join(product_label(p) for p in options) if options else "None"
    mc_products = [
        product_label(p) for p in options if get_product_capabilities(p).supports_monte_carlo
    ]
    econ_products = [
        product_label(p) for p in options if get_product_capabilities(p).supports_economic_scenario
    ]
    return [
        f"Supported product run types: {available_products}.",
        "Run-time pricing dispatch is centralized in the product registry adapters.",
        f"Economic scenario controls enabled for: {', '.join(econ_products) if econ_products else 'None'}.",
        f"Monte Carlo pricing enabled for: {', '.join(mc_products) if mc_products else 'None'}.",
        "Yield curve sources: flat rate, zero-curve CSV, or par-yield CSV bootstrapped to zeros.",
        "Mortality sources are product-scoped and configured by registry defaults/options.",
        "ALM tab supports Treasury ladder projection, reinvestment/disinvestment policy controls, and KPI output tied to the active pricing run.",
        "What-if analysis provides before/after/impact views across pricing and ALM dimensions.",
        "Excel replicator export includes parity-oriented workbook output with optional MC and ALM snapshots.",
        "Embedded unit-test dashboard is available from the Unit Tests section.",
    ]


def _seed_run_form_state_from_last_inputs() -> None:
    meta = st.session_state.get("pricing_meta") or {}
    saved_inputs = st.session_state.get("pricing_run_inputs") or {}
    product_default = str(
        st.session_state.get(
            "pricing_product_type", meta.get("product_type", ProductType.SPIA.value)
        )
    )
    try:
        default_product_type = ProductType(product_default)
    except ValueError:
        default_product_type = ProductType.SPIA

    defaults = build_run_form_seed_defaults(
        product_default=product_default,
        saved_inputs=saved_inputs,
        meta=meta,
        default_product_type=default_product_type,
    )
    st.session_state["_pricing_run_numeric_seeds"] = {
        k: defaults[k] for k in PRICING_RUN_NUMBER_INPUT_KEYS if k in defaults
    }
    for k, v in defaults.items():
        if k in PRICING_RUN_NUMBER_INPUT_KEYS:
            continue
        st.session_state.setdefault(k, v)


def _normalize_run_state_for_selected_product(
    state: MutableMapping[str, Any],
    *,
    selected_product: ProductType,
    switched_product: bool,
) -> None:
    """Normalize run-form state so UI values remain valid across product switches and reruns."""
    capabilities = get_product_capabilities(selected_product)

    # Keep enumerated controls valid for current product.
    mortality_options = list(get_product_mortality_mode_options(selected_product))
    default_mortality_mode = get_product_default_mortality_mode(selected_product)
    if switched_product and selected_product in (
        ProductType.SPIA,
        ProductType.RILA,
        ProductType.MYGA,
        ProductType.FIA,
        ProductType.VARIABLE_ANNUITY,
        ProductType.WHOLE_LIFE,
        ProductType.UNIVERSAL_LIFE,
        ProductType.INDEXED_UL,
        ProductType.VARIABLE_UL,
    ):
        state["run_m_mode"] = default_mortality_mode
    # Force run_use_index ON for products that need it for meaningful crediting
    if switched_product and selected_product in (
        ProductType.RILA,
        ProductType.FIA,
        ProductType.INDEXED_UL,
        ProductType.VARIABLE_UL,
        ProductType.VARIABLE_ANNUITY,
    ):
        state["run_use_index"] = True
    # Force run_use_index OFF for products that don't use index
    if switched_product and selected_product in (
        ProductType.MYGA,
        ProductType.WHOLE_LIFE,
        ProductType.UNIVERSAL_LIFE,
    ):
        state["run_use_index"] = False
    if str(state.get("run_m_mode", "")) not in mortality_options:
        state["run_m_mode"] = default_mortality_mode

    y_mode = str(state.get("run_y_mode", "par_bootstrap"))
    if y_mode not in ("flat", "zero_csv", "par_bootstrap"):
        state["run_y_mode"] = "par_bootstrap"
    expense_mode = str(state.get("run_expense_mode", "csv"))
    if expense_mode not in ("csv", "manual"):
        state["run_expense_mode"] = "csv"
    if str(state.get("run_sex", "male")) not in ("male", "female"):
        state["run_sex"] = "male"

    # Keep path-like inputs nonblank; blank state often appears after product switching.
    if not str(state.get("run_index_csv", "")).strip():
        state["run_index_csv"] = sp.DEFAULT_SP500_SCENARIO_CSV
    if not str(state.get("run_qx_csv", "")).strip():
        state["run_qx_csv"] = sp.DEFAULT_MORTALITY_QX_CSV
    if not str(state.get("run_rp_xlsx", "")).strip():
        state["run_rp_xlsx"] = sp.DEFAULT_RP2014_XLSX
    if not str(state.get("run_rp_out", "")).strip():
        state["run_rp_out"] = sp.DEFAULT_RP2014_MALE_HEALTHY_QX_CSV
    if not str(state.get("run_mp_xlsx", "")).strip():
        state["run_mp_xlsx"] = sp.DEFAULT_MP2016_XLSX
    if not str(state.get("run_mp_out", "")).strip():
        state["run_mp_out"] = sp.DEFAULT_MP2016_MALE_IMPROVEMENT_CSV

    # Product capabilities govern whether these toggles should remain enabled.
    if not capabilities.supports_economic_scenario:
        state["run_use_index"] = False
    if not capabilities.supports_monte_carlo:
        state["run_mc_enable"] = False

    # Term premium defaults can get seeded as 0.0 from prior SPIA runs.
    if selected_product == ProductType.TERM_LIFE:
        default_term_premium = float(get_term_contract_ui_config().default_monthly_premium)
        current_term_premium_raw = state.get("run_term_monthly_premium")
        if current_term_premium_raw is None:
            state["run_term_monthly_premium"] = default_term_premium
        else:
            current_term_premium = float(current_term_premium_raw)
            if current_term_premium <= 0.0:
                state["run_term_monthly_premium"] = default_term_premium


def _build_yield_curve(
    mode: YieldMode,
    *,
    flat_rate: float,
    zero_csv: str,
    par_csv: str,
    coupon_freq: int,
) -> sp.YieldCurve:
    seeds = {
        RUN_KEY.Y_MODE: mode,
        RUN_KEY.FLAT_RATE: flat_rate,
        RUN_KEY.ZERO_CSV: zero_csv,
        RUN_KEY.PAR_CSV: par_csv,
        RUN_KEY.COUPON_FREQ: int(coupon_freq),
    }
    return build_yield_curve_from_seeds(seeds, repo_root=ROOT)


def _build_mortality(
    mode: MortalityMode,
    *,
    product_type: ProductType,
    sex: Literal["male", "female"],
    qx_csv: str,
    rp_xlsx: str,
    rp_out_csv: str,
    mp_xlsx: str,
    mp_out_csv: str,
) -> tuple[sp.MortalityTableQx | sp.MortalityTableRP2014MP2016, bool]:
    """Returns (mortality, needs_valuation_year). Delegates to :mod:`pricing_scenario_materialize`."""
    seeds = {
        RUN_KEY.M_MODE: mode,
        RUN_KEY.QX_CSV: qx_csv,
        RUN_KEY.RP_XLSX: rp_xlsx,
        RUN_KEY.RP_OUT: rp_out_csv,
        RUN_KEY.MP_XLSX: mp_xlsx,
        RUN_KEY.MP_OUT: mp_out_csv,
    }
    return build_mortality_from_seeds(seeds, product_type=product_type, sex=sex, repo_root=ROOT)


def _render_overview() -> None:
    st.header("Model overview")
    st.markdown(
        "This workspace runs the pricing and projection engine with product adapters, "
        "scenario analysis, and Excel parity checks."
    )
    st.caption(
        "Overview content is generated from the product registry and shared section metadata "
        "to reduce documentation drift after model updates."
    )

    st.subheader("Current feature set")
    for i, feat in enumerate(_dynamic_overview_features(), start=1):
        st.markdown(f"{i}. {feat}")

    st.subheader("Workspace sections")
    section_labels = [SECTION_LABELS[k] for k in SECTION_ORDER if k != "overview"]
    st.markdown(
        "Use the sidebar to navigate: " + " | ".join(f"**{name}**" for name in section_labels) + "."
    )


def _result_dataframe(
    res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
) -> pd.DataFrame:
    expected_payment_pv = res.expected_benefit_cashflows * res.discount_factors
    cumulative_pv = np.cumsum(expected_payment_pv)
    return pd.DataFrame(
        {
            "month": res.months,
            "time_years": res.times_years,
            "age_at_payment": res.ages_at_payment,
            "survival": res.survival_to_payment,
            "discount_factor": res.discount_factors,
            "index_level": res.index_level_at_payment,
            "index_simple_return": res.index_simple_return,
            "index_log_return": res.index_log_return,
            "index_cumulative_return": res.index_cumulative_return,
            "benefit_nominal": res.benefit_nominal_scheduled,
            "expense_nominal": res.expense_nominal_scheduled,
            "expected_benefit": res.expected_benefit_cashflows,
            "expected_expense": res.expected_expense_cashflows,
            "expected_total": res.expected_total_cashflows,
            "expected_payment_pv": expected_payment_pv,
            "cumulative_benefit_pv": cumulative_pv,
        }
    )


def _render_pricing_run_charts(
    res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
    contract: sp.SPIAContract | tp.TermLifeContract | rp.RILAContract,
    expenses: sp.ExpenseAssumptions | None,
    product_type: ProductType,
) -> None:
    expected_payment_pv = res.expected_benefit_cashflows * res.discount_factors
    cumulative_expected_payment_pv = np.cumsum(expected_payment_pv)
    ages_r = contract.issue_age + res.reserve_times_years

    st.subheader("Run charts")
    st.markdown("**Cumulative PV benefits**")
    st.line_chart(
        pd.DataFrame(
            {
                "age": res.ages_at_payment,
                "cumulative_pv_benefits": np.rint(cumulative_expected_payment_pv),
            }
        ).set_index("age")
    )

    st.markdown("**Economic reserve** (benefit + monthly expense, PV roll-forward)")
    st.line_chart(
        pd.DataFrame({"age": ages_r, "reserve": np.rint(res.economic_reserve)}).set_index("age")
    )

    if product_type != ProductType.TERM_LIFE and not isinstance(expenses, sp.ExpenseAssumptions):
        st.warning(
            "Profit decomposition unavailable: pricing expense assumptions were not found in session state."
        )
    else:
        _render_profit_decomposition_chart(
            res, contract, expenses=expenses, product_type=product_type
        )


def _build_profit_waterfall_chart_df(rows: list[tuple[str, float, bool]]) -> pd.DataFrame:
    """Walking waterfall data for Altair: each *change* bar spans cumulative *start*→*end* (signed *delta*).

    * First *is_total*: pillar from 0 to the starting anchor (SPIA: undiscounted level benefit; Term: undiscounted claims).
    * *Change* rows: floating bar from running level to running + table amount (``delta`` may be negative).
    * Last *is_total*: reconciliation pillar from 0 to modeled premium / net PV.

    Row order in *rows* is preserved left-to-right on the x-axis.
    """
    records: list[dict[str, float | str]] = []
    running = 0.0
    n = len(rows)
    for i, (label, val, is_total) in enumerate(rows):
        v = float(val)
        if is_total and i == 0 or is_total and i == n - 1:
            records.append(
                {
                    "Step": label,
                    "start": 0.0,
                    "end": v,
                    "delta": v,
                    "lo": 0.0,
                    "hi": v,
                    "bar_color": "Total",
                }
            )
            running = v
        elif is_total:
            records.append(
                {
                    "Step": label,
                    "start": 0.0,
                    "end": v,
                    "delta": v,
                    "lo": min(0.0, v),
                    "hi": max(0.0, v),
                    "bar_color": "Total",
                }
            )
            running = v
        else:
            s = running
            e = running + v
            records.append(
                {
                    "Step": label,
                    "start": s,
                    "end": e,
                    "delta": v,
                    "lo": min(s, e),
                    "hi": max(s, e),
                    "bar_color": "Increase" if v >= 0.0 else "Decrease",
                }
            )
            running = e
    return pd.DataFrame(records)


def _altair_profit_waterfall_chart(df: pd.DataFrame) -> alt.Chart:
    color_scale = alt.Scale(
        domain=["Total", "Increase", "Decrease"], range=["#1f77b4", "#2ca02c", "#d62728"]
    )
    bars = (
        alt.Chart(df)
        .mark_bar(size=36)
        .encode(
            x=alt.X("Step:N", sort=None, axis=alt.Axis(labelAngle=-30, labelLimit=280, title=None)),
            y=alt.Y("hi:Q", title="Amount ($)"),
            y2="lo:Q",
            color=alt.Color(
                "bar_color:N",
                scale=color_scale,
                legend=alt.Legend(orient="top", title=None, labelFontSize=11),
            ),
            tooltip=[
                alt.Tooltip("Step:N", title="Component"),
                alt.Tooltip("delta:Q", format=",.2f", title="Step ($)"),
                alt.Tooltip("start:Q", format=",.2f", title="From ($)"),
                alt.Tooltip("end:Q", format=",.2f", title="To ($)"),
            ],
        )
    )
    rule = (
        alt.Chart(pd.DataFrame({"y": [0.0]}))
        .mark_rule(color="#888888", strokeDash=[4, 3])
        .encode(y=alt.Y("y:Q"))
    )
    return (bars + rule).properties(height=440)


def _render_profit_decomposition_chart(
    res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
    contract: sp.SPIAContract | tp.TermLifeContract | rp.RILAContract,
    expenses: sp.ExpenseAssumptions | None,
    product_type: ProductType,
) -> None:
    st.subheader("Profit decomposition waterfall")
    rows, caption = _build_profit_decomposition_rows(
        res=res,
        contract=contract,
        expenses=expenses,
        product_type=product_type,
    )

    wf_df = _build_profit_waterfall_chart_df(rows)
    st.altair_chart(_altair_profit_waterfall_chart(wf_df), use_container_width=True)
    st.caption(
        "Blue = subtotal / reconciliation pillars from zero; green = upward step; red = downward step (table amount)."
    )

    table = pd.DataFrame([{"Component": label, "Amount ($)": val} for label, val, _ in rows])
    table_display = _round_for_visuals(table)
    st.dataframe(
        table_display,
        use_container_width=True,
        hide_index=True,
        column_config=_number_cols_no_decimals(table_display),
    )
    st.caption(caption)


def _build_profit_decomposition_rows(
    *,
    res: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
    contract: sp.SPIAContract | tp.TermLifeContract | rp.RILAContract,
    expenses: sp.ExpenseAssumptions | None,
    product_type: ProductType,
) -> tuple[list[tuple[str, float, bool]], str]:
    if product_type == ProductType.TERM_LIFE:
        undiscounted_expected_claims = float(np.sum(res.expected_benefit_cashflows))
        discounting_effect = float(res.pv_benefit - undiscounted_expected_claims)
        pv_premiums = float(-float(res.pv_monthly_expenses))
        rows = [
            ("Undiscounted expected claims", undiscounted_expected_claims, True),
            ("Discounting effect", discounting_effect, False),
            ("Policyholder premium PV (funding)", -pv_premiums, False),
            ("Net PV (claims - premiums)", float(res.single_premium), True),
        ]
        caption = (
            "Interpretation: this view starts from undiscounted expected claims, applies discounting to get "
            "PV claims, then subtracts PV policyholder premiums to arrive at net PV."
        )
        return rows, caption

    if product_type == ProductType.SPIA and isinstance(contract, sp.SPIAContract):
        b_month = float(contract.benefit_annual) / 12.0
        n_months = int(res.months.size)
        level_benefit_certain_undisc = float(b_month * n_months)
        level_benefit_mort_undisc = float(np.sum(b_month * res.survival_to_payment))
        level_benefit_mort_disc = float(
            np.sum(b_month * res.survival_to_payment * res.discount_factors)
        )

        mortality_effect = level_benefit_mort_undisc - level_benefit_certain_undisc
        discounting_effect = level_benefit_mort_disc - level_benefit_mort_undisc
        benefit_design_effect = float(res.pv_benefit - level_benefit_mort_disc)
        issue_expense = (
            float(expenses.policy_expense_dollars)
            if isinstance(expenses, sp.ExpenseAssumptions)
            else 0.0
        )
        expense_component = issue_expense + float(res.pv_monthly_expenses)
        margin_component = float(
            res.single_premium - (res.pv_benefit + res.pv_monthly_expenses + issue_expense)
        )
        rows = [
            ("Undiscounted level benefits (certain life)", level_benefit_certain_undisc, True),
            ("Mortality effect", mortality_effect, False),
            ("Discounting effect", discounting_effect, False),
            ("Benefit design effect (e.g., indexation)", benefit_design_effect, False),
            ("Expenses (issue + monthly PV)", expense_component, False),
            ("Margin / premium load", margin_component, False),
            ("Single premium", float(res.single_premium), True),
        ]
        caption = (
            "Interpretation: start from level benefits paid with certainty and no discounting, then layer in "
            "mortality, discounting, product design effects (such as indexation), expenses, and premium load/margin "
            "to reconcile to modeled single premium."
        )
        return rows, caption

    # Generic fallback for future products where detailed decomposition may differ.
    issue_expense = (
        float(expenses.policy_expense_dollars)
        if isinstance(expenses, sp.ExpenseAssumptions)
        else 0.0
    )
    monthly_component = float(res.pv_monthly_expenses)
    rows = [
        ("PV benefits", float(res.pv_benefit), True),
        ("PV monthly cashflow component", monthly_component, False),
        ("Issue expense", issue_expense, False),
        ("Modeled net premium / value", float(res.single_premium), True),
    ]
    caption = (
        "Interpretation: generic product-level decomposition showing the core PV building blocks. "
        "Product-specific decomposition should replace this as each product is implemented."
    )
    return rows, caption


def _merge_profit_waterfall_row_sets(
    row_sets: list[list[tuple[str, float, bool]]],
) -> list[tuple[str, float, bool]]:
    if not row_sets:
        raise ValueError("row_sets must be non-empty")
    ref = row_sets[0]
    for rs in row_sets[1:]:
        if len(rs) != len(ref):
            raise ValueError("incompatible waterfall row counts")
        for i, (lab, _, is_tot) in enumerate(ref):
            lab2, _, is2 = rs[i]
            if lab != lab2 or is_tot != is2:
                raise ValueError("incompatible waterfall row labels or total flags")
    return [
        (ref[i][0], float(sum(float(rs[i][1]) for rs in row_sets)), ref[i][2])
        for i in range(len(ref))
    ]


def _build_portfolio_generic_pv_bridge_rows(
    policy_results: tuple[Any, ...] | list[Any],
    expenses: sp.ExpenseAssumptions | None,
) -> tuple[list[tuple[str, float, bool]], str]:
    sum_pv_b = 0.0
    sum_pv_m = 0.0
    sum_sp = 0.0
    n_pol = 0
    for pr in policy_results:
        p = pr.pricing
        if hasattr(p, "pv_benefit"):
            sum_pv_b += float(getattr(p, "pv_benefit"))
        if hasattr(p, "pv_monthly_expenses"):
            sum_pv_m += float(getattr(p, "pv_monthly_expenses"))
        if hasattr(p, "single_premium"):
            sum_sp += float(getattr(p, "single_premium"))
        n_pol += 1
    issue_total = (
        float(expenses.policy_expense_dollars) * n_pol
        if isinstance(expenses, sp.ExpenseAssumptions)
        else 0.0
    )
    rows = [
        ("PV benefits (portfolio sum)", sum_pv_b, True),
        ("PV monthly cashflow component (portfolio sum)", sum_pv_m, False),
        ("Issue expense (portfolio sum)", issue_total, False),
        ("Modeled net premium / value (portfolio sum)", sum_sp, True),
    ]
    caption = (
        "Portfolio-level PV bridge: per-policy PV components and issue expense summed. "
        "For mixed product books, intermediate steps are not a single-product story."
    )
    return rows, caption


def _build_portfolio_profit_decomposition_rows(
    res: PortfolioResult,
    expenses: sp.ExpenseAssumptions | None,
) -> tuple[list[tuple[str, float, bool]], str]:
    return _build_portfolio_profit_decomposition_rows_for_policy_results(
        tuple(res.policy_results), expenses
    )


def _build_portfolio_profit_decomposition_rows_for_policy_results(
    policy_results: tuple[Any, ...] | list[Any],
    expenses: sp.ExpenseAssumptions | None,
) -> tuple[list[tuple[str, float, bool]], str]:
    prs = list(policy_results)
    if not prs:
        raise ValueError("portfolio waterfall requires at least one policy result")
    types = {pr.product_type for pr in prs}
    if len(types) == 1:
        pt = next(iter(types))
        try:
            row_sets: list[list[tuple[str, float, bool]]] = []
            for pr in prs:
                rows_one, _c = _build_profit_decomposition_rows(
                    res=pr.pricing,  # type: ignore[arg-type]
                    contract=pr.contract,
                    expenses=expenses,
                    product_type=pt,
                )
                row_sets.append(rows_one)
            merged = _merge_profit_waterfall_row_sets(row_sets)
            if pt == ProductType.TERM_LIFE:
                cap = (
                    "Portfolio Term book: same waterfall interpretation as Pricing Run; table amounts sum policies."
                )
            elif pt == ProductType.SPIA:
                cap = "Portfolio SPIA book: same ladder as Pricing Run; table amounts sum policies."
            else:
                cap = "Homogeneous portfolio: PV bridge rows summed across policies."
            return merged, cap
        except (ValueError, TypeError, AttributeError):
            pass
    return _build_portfolio_generic_pv_bridge_rows(tuple(prs), expenses)


def _render_portfolio_profit_waterfall(
    res: PortfolioResult, expenses: sp.ExpenseAssumptions | None
) -> None:
    st.subheader("Portfolio profit decomposition")
    types_sorted = sorted({pr.product_type for pr in res.policy_results}, key=lambda x: x.value)
    options = ["Aggregate"] + [pt.value for pt in types_sorted]
    run_tag = int(st.session_state.get(PORTFOLIO_KEY.RUN_ID, 0))
    selected = st.multiselect(
        "Waterfall series (aggregate plus product types)",
        options=options,
        default=options,
        key=f"portfolio_wf_series_{run_tag}",
    )
    if not selected:
        st.info("Select at least one waterfall series.")
        return
    for series in selected:
        if series == "Aggregate":
            prs = tuple(res.policy_results)
            title = "Aggregate"
        else:
            pt = ProductType(series)
            prs = tuple(pr for pr in res.policy_results if pr.product_type == pt)
            title = f"Product type: {series}"
            if not prs:
                continue
        rows, caption = _build_portfolio_profit_decomposition_rows_for_policy_results(prs, expenses)
        wf_df = _build_profit_waterfall_chart_df(rows)
        st.markdown(f"**{title}**")
        net_prem = sum(float(getattr(pr.pricing, "single_premium", 0.0) or 0.0) for pr in prs)
        if net_prem < 0:
            st.caption(
                "Note: modeled premium / value is negative (e.g. net premium outflow to policyholders)."
            )
        st.altair_chart(_altair_profit_waterfall_chart(wf_df), use_container_width=True)
        st.caption(
            "Blue = subtotal / reconciliation pillars from zero; green = upward step; red = downward step (table amount)."
        )
        table = pd.DataFrame([{"Component": label, "Amount ($)": val} for label, val, _ in rows])
        table_display = _round_for_visuals(table)
        st.dataframe(
            table_display,
            use_container_width=True,
            hide_index=True,
            column_config=_number_cols_no_decimals(table_display),
        )
        st.caption(caption)


def _render_portfolio_liability_projection_chart(res: PortfolioResult) -> None:
    st.subheader("Liability cashflow projection")
    st.caption(
        "Nominal expected outflows from priced liability paths (the same series fed to ALM). "
        "Per-product paths are zero-padded after the last month for that type."
    )
    n = len(res.liability_path_total.expected_total_cashflows)
    ty = np.asarray(res.liability_path_total.times_years, dtype=float)
    types_sorted = sorted(res.rollups_by_product_type.keys(), key=lambda x: x.value)
    options = ["Aggregate"] + [pt.value for pt in types_sorted]
    run_tag = int(st.session_state.get(PORTFOLIO_KEY.RUN_ID, 0))
    selected = st.multiselect(
        "Series to plot (aggregate plus product types)",
        options=options,
        default=options,
        key=f"portfolio_proj_series_{run_tag}",
    )
    cumulative = st.toggle("Cumulative cashflows", value=False, key="portfolio_projection_cumulative")
    if not selected:
        st.info("Select at least one series to plot.")
        return
    frames: list[pd.DataFrame] = []
    if "Aggregate" in selected:
        cf = np.asarray(res.liability_path_total.expected_total_cashflows, dtype=float)
        y = np.cumsum(cf) if cumulative else cf
        frames.append(pd.DataFrame({"time_years": ty, "series": "Aggregate", "value": y}))
    for pt in types_sorted:
        if pt.value not in selected:
            continue
        path = res.rollups_by_product_type[pt]
        cf = padded_cashflows_on_portfolio_grid(path, n)
        y = np.cumsum(cf) if cumulative else cf
        frames.append(pd.DataFrame({"time_years": ty, "series": pt.value, "value": y}))
    plot_df = pd.concat(frames, ignore_index=True)
    y_title = "Cumulative cashflow ($)" if cumulative else "Expected cashflow ($)"
    chart = (
        alt.Chart(plot_df)
        .mark_line()
        .encode(
            x=alt.X("time_years:Q", title="Time (years)"),
            y=alt.Y("value:Q", title=y_title),
            color=alt.Color("series:N", title="Series"),
            tooltip=[
                alt.Tooltip("time_years:Q", format=".4f", title="Time (y)"),
                alt.Tooltip("series:N", title="Series"),
                alt.Tooltip("value:Q", format=",.2f", title="Amount"),
            ],
        )
        .properties(height=420)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_portfolio_alm_baseline_section(res: PortfolioResult) -> None:
    alm = res.alm_result
    if alm is None:
        return
    st.subheader("Portfolio ALM (baseline)")
    st.caption(
        "Deterministic Treasury ladder versus **aggregated** liability outflows, using the same yield curve and "
        "credit spread as this portfolio run. Rules match `alm_engine_baseline_assumptions` (Pricing Run ALM "
        "defaults), not the interactive optimizer."
    )
    fr = np.asarray(alm.funding_ratio, dtype=float)
    fr0 = float(fr[0]) if fr.size else float("nan")
    min_fr = float(np.nanmin(fr)) if fr.size else float("nan")
    min_surp = float(np.min(np.asarray(alm.surplus, dtype=float))) if alm.surplus.size else float("nan")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Initial funding ratio", f"{fr0:.4f}")
    with c2:
        st.metric("Min funding ratio", f"{min_fr:.4f}")
    with c3:
        st.metric("Min surplus ($)", f"{min_surp:,.0f}")
    with c4:
        st.metric("Duration gap (years)", f"{alm.duration_gap:.4f}")
    st.caption(
        f"PV01 net: {alm.pv01_net:,.2f} · PV01 assets: {alm.pv01_assets:,.2f} · "
        f"PV01 liabilities: {alm.pv01_liabilities:,.2f}"
    )
    ty_vis = np.asarray(alm.times_years, dtype=float)
    sur = np.asarray(alm.surplus, dtype=float)
    df_s = pd.DataFrame({"Time (years)": ty_vis, "Surplus": sur})
    line = (
        alt.Chart(df_s)
        .mark_line()
        .encode(
            x=alt.X("Time (years):Q", title="Time (years)"),
            y=alt.Y("Surplus:Q", title="Surplus ($)"),
        )
    )
    rule = (
        alt.Chart(pd.DataFrame({"y": [0.0]}))
        .mark_rule(color="#888", strokeDash=[4, 4])
        .encode(y="y:Q")
    )
    st.altair_chart((line + rule).properties(height=320, title="Surplus path"), use_container_width=True)
    fr_df = pd.DataFrame({"Time (years)": ty_vis, "Funding ratio": fr})
    fr_chart = (
        alt.Chart(fr_df)
        .mark_line(color="#1f77b4")
        .encode(
            x=alt.X("Time (years):Q"),
            y=alt.Y("Funding ratio:Q", scale=alt.Scale(zero=False)),
        )
        .properties(height=260, title="Funding ratio")
    )
    st.altair_chart(fr_chart, use_container_width=True)


def _execute_portfolio_pricing(
    policies: tuple[PolicyInput, ...],
    scen: RunScenario,
) -> tuple[PortfolioResult, str | None]:
    """Price portfolio with baseline ALM when aggregate premium supports it."""
    alm_asm = sp.alm_engine_baseline_assumptions()
    try:
        res = run_portfolio(
            portfolio=Portfolio(policies=policies),
            scenario=scen,
            alm_assumptions=alm_asm,
        )
        return res, None
    except ValueError as exc:
        msg = str(exc)
        if "initial_asset_market_value" in msg or "single_premium" in msg:
            res = run_portfolio(
                portfolio=Portfolio(policies=policies),
                scenario=scen,
                alm_assumptions=None,
            )
            return res, (
                "Portfolio ALM was skipped: aggregate assets inferred from single premiums are not positive. "
                "Pricing and liability aggregation are shown without ALM."
            )
        raise


def _shock_yield_curve(curve: sp.YieldCurve, rate_shift_bps: float) -> sp.YieldCurve:
    shift = float(rate_shift_bps) / 10000.0
    return sp.YieldCurve(
        maturities_years=np.asarray(curve.maturities_years, dtype=float).copy(),
        zero_rates=np.asarray(curve.zero_rates, dtype=float).copy() + shift,
    )


def _key_rate_bump_curve(
    curve: sp.YieldCurve,
    *,
    key_tenor_years: float,
    key_tenors_years: np.ndarray,
    bump_bps: float = 1.0,
) -> sp.YieldCurve:
    return sp.yield_curve_key_rate_bump(
        curve,
        key_tenor_years=key_tenor_years,
        key_tenors_years=key_tenors_years,
        bump_bps=bump_bps,
    )


def _shock_mortality(
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    longevity_improvement_pct: float,
) -> sp.MortalityTableQx | sp.MortalityTableRP2014MP2016:
    # Positive longevity improvement means lower qx.
    factor = max(0.01, 1.0 - float(longevity_improvement_pct) / 100.0)
    if isinstance(mortality, sp.MortalityTableQx):
        return sp.MortalityTableQx(
            ages=np.asarray(mortality.ages, dtype=int).copy(),
            qx=np.clip(np.asarray(mortality.qx, dtype=float) * factor, 0.0, 0.999999),
        )
    shocked_base = sp.MortalityTableQx(
        ages=np.asarray(mortality.base_qx_2014.ages, dtype=int).copy(),
        qx=np.clip(np.asarray(mortality.base_qx_2014.qx, dtype=float) * factor, 0.0, 0.999999),
    )
    return sp.MortalityTableRP2014MP2016(
        base_qx_2014=shocked_base,
        mp2016_ages=np.asarray(mortality.mp2016_ages, dtype=int).copy(),
        mp2016_years=np.asarray(mortality.mp2016_years, dtype=int).copy(),
        mp2016_i_matrix=np.asarray(mortality.mp2016_i_matrix, dtype=float).copy(),
        base_year=int(mortality.base_year),
    )


def _equity_regime_params(regime: str) -> tuple[float, float]:
    mapping = {
        "defensive": (0.03, 0.10),
        "base": (0.06, 0.15),
        "bullish": (0.09, 0.20),
        "stressed": (-0.02, 0.35),
    }
    return mapping.get(regime, mapping["base"])


def _deterministic_index_levels_from_regime(
    *, s0: float, annual_drift: float, n_months: int
) -> np.ndarray:
    dt = 1.0 / 12.0
    months = np.arange(1, n_months + 1, dtype=float)
    return float(s0) * np.exp(float(annual_drift) * months * dt)


def _render_impact_metric(
    label: str, before_val: float, after_val: float, money: bool = True
) -> None:
    delta = float(after_val - before_val)
    if money:
        st.metric(label, f"${after_val:,.0f}", delta=f"${delta:,.0f}")
    else:
        st.metric(label, f"{after_val:,.0f}", delta=f"{delta:+,.0f}")


def _mc_cache_get_or_compute(
    key: tuple[object, ...],
    *,
    contract: sp.SPIAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int | None,
    expenses: sp.ExpenseAssumptions,
    expense_annual_inflation: float,
    n_sims: int,
    annual_drift: float,
    annual_vol: float,
    seed: int,
    s0: float,
) -> sp.SPIAMonteCarloResult:
    cache = st.session_state.setdefault("what_if_mc_cache", {})
    if key in cache:
        return cache[key]
    out = sp.price_spia_single_premium_monte_carlo(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=horizon_age,
        spread=spread,
        valuation_year=valuation_year,
        expenses=expenses,
        expense_annual_inflation=expense_annual_inflation,
        n_sims=n_sims,
        annual_drift=annual_drift,
        annual_vol=annual_vol,
        seed=seed,
        s0=s0,
    )
    cache[key] = out
    return out


def compute_what_if_term_shocked_pricing(
    *,
    base_contract: tp.TermLifeContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int | None,
    term_monthly_premium_mult: float = 1.0,
) -> tp.TermLifeProjectionResult:
    """Reprice Term for what-if; kept pure for unit tests (regression vs SPIA horizon mismatch)."""
    shocked_contract = dataclasses.replace(
        base_contract,
        monthly_premium=float(base_contract.monthly_premium) * float(term_monthly_premium_mult),
    )
    return tp.price_term_life_level_monthly(
        contract=shocked_contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=horizon_age,
        spread=spread,
        valuation_year=valuation_year,
    )


def build_alm_pricing_for_mc_scenario(
    *,
    product_type: ProductType,
    scenario_source: str,
    baseline_pricing: Any,
    contract: Any,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int | None,
    expenses: sp.ExpenseAssumptions,
    expense_annual_inflation: float,
    mc_n_sims: int,
    mc_seed: int,
    mc_scenario_idx: int,
    mc_params: dict[str, Any],
) -> Any:
    """
    Single-path MC repricing for ALM liability PV (SPIA or RILA); other products return baseline.
    """
    if scenario_source != "MC simulation (single path)":
        return baseline_pricing
    n_months = int(baseline_pricing.months.size)
    idx_paths = sp.simulate_index_levels_gbm(
        n_sims=mc_n_sims,
        n_months=n_months,
        s0=float(mc_params.get("s0", 100.0) or 100.0),
        annual_drift=float(mc_params.get("annual_drift", 0.06) or 0.06),
        annual_vol=float(mc_params.get("annual_vol", 0.15) or 0.15),
        seed=mc_seed,
    )
    idx_one = idx_paths[int(mc_scenario_idx)]
    idx_levels_payment = np.asarray(idx_one[1:], dtype=float)
    idx_s0 = float(idx_one[0])
    if product_type == ProductType.SPIA and isinstance(contract, sp.SPIAContract):
        return sp.price_spia_single_premium(
            contract=contract,
            yield_curve=yield_curve,
            mortality=mortality,
            horizon_age=int(horizon_age),
            spread=spread,
            valuation_year=int(valuation_year) if valuation_year is not None else None,
            expenses=expenses,
            expense_annual_inflation=float(expense_annual_inflation),
            index_s0=idx_s0,
            index_levels_payment=idx_levels_payment,
        )
    if product_type == ProductType.RILA and isinstance(contract, rp.RILAContract):
        return rp.price_rila_single_premium(
            contract=contract,
            yield_curve=yield_curve,
            mortality=mortality,
            horizon_age=int(horizon_age),
            spread=spread,
            valuation_year=int(valuation_year) if valuation_year is not None else None,
            expenses=expenses,
            expense_annual_inflation=float(expense_annual_inflation),
            index_s0=idx_s0,
            index_levels_payment=idx_levels_payment,
        )
    return baseline_pricing


def _run_alm_from_session_pricing(
    *,
    pricing: sp.SPIAProjectionResult | tp.TermLifeProjectionResult | rp.RILAProjectionResult,
    yield_curve: sp.YieldCurve,
    spread: float,
    assumptions: sp.ALMAssumptions,
    initial_asset_market_value: float,
    asset_curve: sp.YieldCurve | None = None,
    liability_cashflows: np.ndarray | None = None,
) -> sp.ALMResult:
    return sp.run_alm_projection_from_pricing_result(
        pricing=pricing,
        yield_curve=yield_curve,
        spread=spread,
        assumptions=assumptions,
        initial_asset_market_value=float(initial_asset_market_value),
        asset_curve=asset_curve,
        liability_cashflows=liability_cashflows,
    )


def _render_what_if_studio() -> None:
    st.header("What-if Analysis")
    st.caption("Live scenario shocks relative to the latest baseline run in Pricing Run.")

    base_res = st.session_state.get("pricing_res")
    base_contract = st.session_state.get("pricing_contract")
    product_raw = st.session_state.get("pricing_product_type", ProductType.SPIA.value)
    try:
        product_type = ProductType(str(product_raw))
    except ValueError:
        product_type = ProductType.SPIA
    ctx = st.session_state.get("pricing_excel_context") or {}
    base_curve = ctx.get("yield_curve")
    base_mort = ctx.get("mortality")
    base_expenses = ctx.get("expenses")
    if (
        not isinstance(base_expenses, sp.ExpenseAssumptions)
        and product_type == ProductType.TERM_LIFE
    ):
        base_expenses = sp.ExpenseAssumptions(0.0, 0.0, 0.0)

    if (
        base_res is None
        or base_contract is None
        or not isinstance(base_curve, sp.YieldCurve)
        or not isinstance(base_expenses, sp.ExpenseAssumptions)
        or not isinstance(base_mort, (sp.MortalityTableQx, sp.MortalityTableRP2014MP2016))
    ):
        st.info("Run pricing first in Pricing Run to set a baseline for What-if analysis.")
        return

    if product_type == ProductType.RILA:
        st.info(
            "What-if Analysis is not yet wired for RILA in this release; use Pricing Run and ALM with deterministic scenarios."
        )
        return

    c1, c2, c3 = st.columns(3)
    with c1:
        rate_shift_bps = st.slider(
            "Rates shift (bps)", min_value=-300, max_value=300, value=0, step=5
        )
        spread_shift_bps = st.slider(
            "Credit spread shift (bps)", min_value=-300, max_value=300, value=0, step=5
        )
    with c2:
        longevity_improvement_pct = st.slider(
            "Longevity improvement shock (%)",
            min_value=-20.0,
            max_value=20.0,
            value=0.0,
            step=0.5,
            help="Positive values reduce mortality rates (longer lives).",
        )
        inflation_shift_pct = 0.0
        if product_type == ProductType.SPIA:
            inflation_shift_pct = st.slider(
                "Expense inflation shift (%)", min_value=-5.0, max_value=10.0, value=0.0, step=0.1
            )
    with c3:
        expense_ratio_mult = 1.0
        equity_regime = "base"
        mc_sims = 800
        term_monthly_premium_mult = 1.0
        if product_type == ProductType.SPIA:
            expense_ratio_mult = st.slider(
                "Expense ratio multiplier", min_value=0.50, max_value=2.00, value=1.00, step=0.05
            )
            equity_regime = st.selectbox(
                "Equity regime",
                options=["defensive", "base", "bullish", "stressed"],
                index=1,
                format_func=lambda x: x.capitalize(),
            )
            mc_sims = st.slider(
                "Tail-risk MC simulations", min_value=200, max_value=5000, value=800, step=200
            )
        elif product_type == ProductType.TERM_LIFE:
            term_monthly_premium_mult = st.slider(
                "Monthly premium multiplier",
                min_value=0.50,
                max_value=1.50,
                value=1.00,
                step=0.01,
                help="Applies to Term monthly premium to test premium adequacy sensitivity.",
            )

    alm_asset_parallel_bps = 0
    alm_twist_short_bps = 0
    alm_twist_long_bps = 0
    alm_liability_cf_pct = 0.0
    if product_type == ProductType.SPIA:
        st.markdown("**ALM add-on shocks**")
        st.caption(
            "These apply on top of the main What-if scenario. Asset curve shocks tilt **mark-to-market** on the Treasury ladder; "
            "liability stress scales **After** SPIA outflows in the ALM engine. "
            "Uses assumptions from the **ALM** tab if you ran them there; otherwise built-in defaults."
        )
        wa1, wa2, wa3, wa4 = st.columns(4)
        with wa1:
            alm_asset_parallel_bps = st.slider(
                "Asset earned-rate parallel shift (bps)",
                min_value=-200,
                max_value=200,
                value=0,
                step=5,
                help="Extra parallel shift on the **After** zero curve for **asset** discounting only.",
            )
        with wa2:
            alm_twist_short_bps = st.slider("Twist: short-end add-on (bps)", -75, 75, 0, 5)
        with wa3:
            alm_twist_long_bps = st.slider("Twist: long-end add-on (bps)", -75, 75, 0, 5)
        with wa4:
            alm_liability_cf_pct = st.slider(
                "Liability outflow stress (%)",
                -40.0,
                40.0,
                0.0,
                0.5,
                help="Scales **After** SPIA cash outflows in the ALM projection (stress on liquidity / disinvestment).",
            )

    alm_whatif_base: sp.ALMResult | None = None
    alm_whatif_after: sp.ALMResult | None = None
    asm_whatif_used: sp.ALMAssumptions | None = None
    try:
        horizon_age = int(ctx.get("horizon_age", 110))
        base_spread = float(ctx.get("spread", 0.0))
        valuation_year = ctx.get("valuation_year")
        base_infl = float(base_res.expense_annual_inflation)
        s0 = float(base_res.index_s0)
        n_months = int(base_res.months.size)

        shocked_curve = _shock_yield_curve(base_curve, float(rate_shift_bps))
        shocked_mort = _shock_mortality(base_mort, float(longevity_improvement_pct))
        shocked_expenses = sp.ExpenseAssumptions(
            policy_expense_dollars=float(base_expenses.policy_expense_dollars)
            * float(expense_ratio_mult),
            premium_expense_rate=min(
                0.99, float(base_expenses.premium_expense_rate) * float(expense_ratio_mult)
            ),
            monthly_expense_dollars=float(base_expenses.monthly_expense_dollars)
            * float(expense_ratio_mult),
        )
        shocked_infl = max(-0.99, base_infl + float(inflation_shift_pct) / 100.0)
        shocked_spread = base_spread + float(spread_shift_bps) / 10000.0

        baseline_mc = None
        shocked_mc = None
        if product_type == ProductType.SPIA:
            # Equity regime controls how the (deterministic) index levels used for "After" evolve.
            #
            # Key requirement: when What-if dials are at identity (all 0 / multipliers at 1) and
            # equity_regime == "base", "After" must reproduce the Pricing Run deterministic result.
            # We do that by applying a regime-specific multiplicative tilt to the Pricing Run's
            # actual baseline index_level_at_payment.
            drift_map, vol_map = _equity_regime_params(equity_regime)
            drift_base_map, vol_base_map = _equity_regime_params("base")

            base_is_identity = (
                equity_regime == "base"
                and abs(float(rate_shift_bps)) < 1e-12
                and abs(float(spread_shift_bps)) < 1e-12
                and abs(float(inflation_shift_pct)) < 1e-12
                and abs(float(longevity_improvement_pct)) < 1e-9
                and abs(float(expense_ratio_mult) - 1.0) < 1e-9
            )

            # Monte Carlo drift/vol are anchored to the Pricing Run's MC parameters so that
            # equity_regime=="base" gives an identity for tail-risk stats too.
            base_mc_params = st.session_state.get("pricing_mc_params") or {}
            base_drift = float(base_mc_params.get("annual_drift", 0.06))
            base_vol = float(base_mc_params.get("annual_vol", 0.15))

            if base_is_identity:
                idx_levels = np.asarray(base_res.index_level_at_payment, dtype=float)
            else:
                idx_regime_det = _deterministic_index_levels_from_regime(
                    s0=s0, annual_drift=drift_map, n_months=n_months
                )
                idx_base_det = _deterministic_index_levels_from_regime(
                    s0=s0, annual_drift=drift_base_map, n_months=n_months
                )
                idx_base_det = np.asarray(idx_base_det, dtype=float)
                idx_regime_det = np.asarray(idx_regime_det, dtype=float)
                scale = idx_regime_det / idx_base_det
                idx_levels = np.asarray(base_res.index_level_at_payment, dtype=float) * scale

            if equity_regime == "base":
                drift_mc = base_drift
                vol_mc = base_vol
            else:
                # Scale regime drifts/vols relative to the regime mapping's "base" so the meaning
                # of defensive/bullish/stressed stays consistent even if the Pricing Run used different MC inputs.
                drift_mc = (
                    base_drift * (drift_map / drift_base_map)
                    if abs(drift_base_map) > 1e-15
                    else base_drift
                )
                vol_mc = (
                    base_vol * (vol_map / vol_base_map) if abs(vol_base_map) > 1e-15 else base_vol
                )

            shocked_res = sp.price_spia_single_premium(
                contract=base_contract,
                yield_curve=shocked_curve,
                mortality=shocked_mort,
                horizon_age=horizon_age,
                spread=shocked_spread,
                valuation_year=int(valuation_year) if valuation_year is not None else None,
                expenses=shocked_expenses,
                index_s0=s0,
                index_levels_payment=idx_levels,
                expense_annual_inflation=shocked_infl,
            )
        elif product_type == ProductType.TERM_LIFE and isinstance(
            base_contract, tp.TermLifeContract
        ):
            shocked_res = compute_what_if_term_shocked_pricing(
                base_contract=base_contract,
                yield_curve=shocked_curve,
                mortality=shocked_mort,
                horizon_age=horizon_age,
                spread=shocked_spread,
                valuation_year=int(valuation_year) if valuation_year is not None else None,
                term_monthly_premium_mult=float(term_monthly_premium_mult),
            )
        else:
            raise ValueError(
                f"What-if analysis is not implemented for product type: {product_type.value}"
            )
        vy = int(valuation_year) if valuation_year is not None else None
        if product_type == ProductType.SPIA:
            baseline_key = (
                "baseline",
                int(mc_sims),
                int(horizon_age),
                float(base_spread),
                float(base_infl),
                float(base_drift),
                float(base_vol),
                float(s0),
                int(base_contract.issue_age),
                float(base_contract.benefit_annual),
            )
            baseline_mc = _mc_cache_get_or_compute(
                baseline_key,
                contract=base_contract,
                yield_curve=base_curve,
                mortality=base_mort,
                horizon_age=horizon_age,
                spread=base_spread,
                valuation_year=vy,
                expenses=base_expenses,
                expense_annual_inflation=base_infl,
                n_sims=int(mc_sims),
                annual_drift=float(base_drift),
                annual_vol=float(base_vol),
                seed=42,
                s0=float(s0),
            )
            shocked_key = (
                "shocked",
                int(mc_sims),
                int(horizon_age),
                float(shocked_spread),
                float(shocked_infl),
                float(drift_mc),
                float(vol_mc),
                float(s0),
                float(rate_shift_bps),
                float(spread_shift_bps),
                float(longevity_improvement_pct),
                float(expense_ratio_mult),
                str(equity_regime),
                int(base_contract.issue_age),
                float(base_contract.benefit_annual),
            )
            shocked_mc = _mc_cache_get_or_compute(
                shocked_key,
                contract=base_contract,
                yield_curve=shocked_curve,
                mortality=shocked_mort,
                horizon_age=horizon_age,
                spread=shocked_spread,
                valuation_year=vy,
                expenses=shocked_expenses,
                expense_annual_inflation=shocked_infl,
                n_sims=int(mc_sims),
                annual_drift=float(drift_mc),
                annual_vol=float(vol_mc),
                seed=42,
                s0=float(s0),
            )

            try:
                asm_wf = st.session_state.get("alm_last_assumptions")
                if not isinstance(asm_wf, sp.ALMAssumptions):
                    asm_wf = st.session_state.get("alm_current_assumptions")
                if not isinstance(asm_wf, sp.ALMAssumptions):
                    asm_wf = sp.ALMAssumptions(
                        allocation=sp.alm_default_allocation_spec(),
                        rebalance_band=0.05,
                        rebalance_frequency_months=1,
                        reinvest_rule="pro_rata",
                        disinvest_rule="shortest_first",
                        rebalance_policy="liquidity_only",
                        borrowing_policy="borrow_after_assets_insufficient",
                        borrowing_rate_mode="scenario_linked",
                        borrowing_rate_tenor_years=1.0,
                        borrowing_spread_annual=0.01,
                        borrowing_rate_annual=0.05,
                        liquidity_near_liquid_years=0.25,
                    )
                asm_whatif_used = asm_wf
                aum_wf = st.session_state.get("alm_last_initial_asset_market_value")
                if aum_wf is None:
                    aum_wf = st.session_state.get("alm_current_initial_asset_market_value")
                aum_wf_use = (
                    float(aum_wf)
                    if isinstance(aum_wf, (int, float, np.floating))
                    else float(base_res.single_premium)
                )
                alm_whatif_base = _run_alm_from_session_pricing(
                    pricing=base_res,
                    yield_curve=base_curve,
                    spread=base_spread,
                    assumptions=asm_wf,
                    initial_asset_market_value=aum_wf_use,
                )
                yc_alm_asset = sp.yield_curve_twist_linear_bps(
                    sp.yield_curve_parallel_bps(shocked_curve, float(alm_asset_parallel_bps)),
                    bps_short=float(alm_twist_short_bps),
                    bps_long=float(alm_twist_long_bps),
                )
                cf_alm = np.asarray(shocked_res.expected_total_cashflows, dtype=float) * (
                    1.0 + float(alm_liability_cf_pct) / 100.0
                )
                alm_whatif_after = _run_alm_from_session_pricing(
                    pricing=shocked_res,
                    yield_curve=shocked_curve,
                    spread=shocked_spread,
                    assumptions=asm_wf,
                    initial_asset_market_value=aum_wf_use,
                    asset_curve=yc_alm_asset,
                    liability_cashflows=cf_alm,
                )
            except Exception as alm_ex:
                alm_whatif_base = None
                alm_whatif_after = None
                asm_whatif_used = None
                st.warning(f"ALM what-if layer skipped: {alm_ex!r}")
    except Exception as ex:
        st.error(f"What-if scenario failed: {ex!r}")
        return

    st.session_state["whatif_last_params"] = {
        "rates_shift_bps": float(rate_shift_bps),
        "spread_shift_bps": float(spread_shift_bps),
        "inflation_shift_pct": float(inflation_shift_pct),
        "longevity_improvement_pct": float(longevity_improvement_pct),
        "expense_ratio_mult": float(expense_ratio_mult),
        "equity_regime": str(equity_regime),
        "mc_sims": int(mc_sims),
        "term_monthly_premium_mult": float(term_monthly_premium_mult),
        "alm_asset_parallel_bps": float(alm_asset_parallel_bps),
        "alm_twist_short_bps": float(alm_twist_short_bps),
        "alm_twist_long_bps": float(alm_twist_long_bps),
        "alm_liability_cf_pct": float(alm_liability_cf_pct),
    }
    st.session_state["whatif_last_base_res"] = base_res
    st.session_state["whatif_last_shocked_res"] = shocked_res
    st.session_state["whatif_last_baseline_mc"] = baseline_mc
    st.session_state["whatif_last_shocked_mc"] = shocked_mc
    st.session_state["whatif_last_shocked_curve"] = shocked_curve
    st.session_state["whatif_last_shocked_mortality"] = shocked_mort
    st.session_state["whatif_last_alm_base"] = alm_whatif_base
    st.session_state["whatif_last_alm_after"] = alm_whatif_after
    st.session_state["whatif_last_alm_assumptions"] = asm_whatif_used
    st.session_state["whatif_last_pricing_run_id"] = st.session_state.get("pricing_run_id")
    _invalidate_diagnostics_export()

    st.subheader("Before vs after vs impact")
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        _render_impact_metric(
            "Single premium",
            float(base_res.single_premium),
            float(shocked_res.single_premium),
            money=True,
        )
    with m2:
        base_margin = float(
            base_res.single_premium - (base_res.pv_benefit + base_res.pv_monthly_expenses)
        )
        shocked_margin = float(
            shocked_res.single_premium - (shocked_res.pv_benefit + shocked_res.pv_monthly_expenses)
        )
        _render_impact_metric("Margin", base_margin, shocked_margin, money=True)
    with m3:
        _render_impact_metric(
            "Reserve at issue",
            float(base_res.economic_reserve[0]),
            float(shocked_res.economic_reserve[0]),
            money=True,
        )
    with m4:
        if baseline_mc is not None and shocked_mc is not None:
            _render_impact_metric(
                "Tail risk (P95 premium)",
                float(baseline_mc.premium_p95),
                float(shocked_mc.premium_p95),
                money=True,
            )
        else:
            _render_impact_metric(
                "PV benefit", float(base_res.pv_benefit), float(shocked_res.pv_benefit), money=True
            )

    compare_df = pd.DataFrame(
        {
            "Metric": ["Single premium", "Margin", "Reserve at issue"],
            "Before": [
                float(base_res.single_premium),
                float(
                    base_res.single_premium - (base_res.pv_benefit + base_res.pv_monthly_expenses)
                ),
                float(base_res.economic_reserve[0]),
            ],
            "After": [
                float(shocked_res.single_premium),
                float(
                    shocked_res.single_premium
                    - (shocked_res.pv_benefit + shocked_res.pv_monthly_expenses)
                ),
                float(shocked_res.economic_reserve[0]),
            ],
        }
    )
    if baseline_mc is not None and shocked_mc is not None:
        compare_df = pd.concat(
            [
                compare_df,
                pd.DataFrame(
                    {
                        "Metric": ["Tail risk (P95 premium)"],
                        "Before": [float(baseline_mc.premium_p95)],
                        "After": [float(shocked_mc.premium_p95)],
                    }
                ),
            ],
            ignore_index=True,
        )
    compare_df["Impact"] = compare_df["After"] - compare_df["Before"]
    compare_display = _round_for_visuals(compare_df)
    st.dataframe(
        compare_display,
        use_container_width=True,
        hide_index=True,
        column_config=_number_cols_no_decimals(compare_display),
    )

    st.markdown("**Reserve path impact**")
    reserve_df = pd.DataFrame(
        {
            "age": base_contract.issue_age + base_res.reserve_times_years,
            "Before reserve": base_res.economic_reserve,
            "After reserve": shocked_res.economic_reserve,
            "Impact": shocked_res.economic_reserve - base_res.economic_reserve,
        }
    ).set_index("age")
    reserve_display = _round_for_visuals(reserve_df)
    # Clean up x-axis labels and monetary formatting.
    reserve_display.index = np.round(reserve_display.index.values.astype(float), 2)
    for col in ["Before reserve", "After reserve", "Impact"]:
        reserve_display[col] = reserve_display[col].astype(int)
    st.line_chart(reserve_display[["Before reserve", "After reserve"]])
    st.bar_chart(reserve_display[["Impact"]])

    if baseline_mc is not None and shocked_mc is not None:
        st.markdown("**Tail-risk distribution impact (single premium)**")
        c1, c2 = st.columns(2)
        with c1:
            counts_b, edges_b = np.histogram(baseline_mc.single_premium, bins=35)
            mids_b = 0.5 * (edges_b[:-1] + edges_b[1:])
            mids_b_disp = np.rint(mids_b).astype(int)
            bin_labels_b = [f"{int(v):,}" for v in mids_b_disp]
            df_b = pd.DataFrame(
                {"bin": bin_labels_b, "count_before": counts_b.astype(int)}
            ).set_index("bin")
            st.bar_chart(_round_for_visuals(df_b))
        with c2:
            counts_a, edges_a = np.histogram(shocked_mc.single_premium, bins=35)
            mids_a = 0.5 * (edges_a[:-1] + edges_a[1:])
            mids_a_disp = np.rint(mids_a).astype(int)
            bin_labels_a = [f"{int(v):,}" for v in mids_a_disp]
            df_a = pd.DataFrame(
                {"bin": bin_labels_a, "count_after": counts_a.astype(int)}
            ).set_index("bin")
            st.bar_chart(_round_for_visuals(df_a))

    if alm_whatif_base is not None and alm_whatif_after is not None:
        _alm_b = alm_whatif_base
        _alm_a = alm_whatif_after
        st.subheader("ALM KPI impact")
        st.caption(
            "Before = ALM on the **Pricing Run** baseline; After = ALM on the shocked liability pricing with **After** curve "
            "for liability PV, optional extra asset **earned-rate** shifts, and scaled outflows. "
            "**Liquidity buffer** = (cash + bonds within near-liquid residual maturity) divided by mean expected monthly "
            "outflow over the next 12 months."
        )

        def _alm_snap(r: sp.ALMResult) -> dict[str, float]:
            return {
                "fr_m1": float(r.funding_ratio[0]) if r.funding_ratio.size else float("nan"),
                "surp_m1": float(r.surplus[0]) if r.surplus.size else float("nan"),
                "liq_m1": (
                    float(r.liquidity_buffer_months[0])
                    if r.liquidity_buffer_months.size
                    else float("nan")
                ),
                "pv01_net": float(r.pv01_net),
                "dur_gap": float(r.duration_gap),
            }

        sb = _alm_snap(_alm_b)
        sa = _alm_snap(_alm_a)
        alm_cmp = pd.DataFrame(
            {
                "Metric": [
                    "Funding ratio (month-end 1)",
                    "Surplus ($)",
                    "Liquidity buffer (months)",
                    "PV01 net ($ per 1bp)",
                    "Duration gap (years)",
                ],
                "Before": [
                    sb["fr_m1"],
                    sb["surp_m1"],
                    sb["liq_m1"],
                    sb["pv01_net"],
                    sb["dur_gap"],
                ],
                "After": [
                    sa["fr_m1"],
                    sa["surp_m1"],
                    sa["liq_m1"],
                    sa["pv01_net"],
                    sa["dur_gap"],
                ],
            }
        )
        alm_cmp["Impact"] = alm_cmp["After"] - alm_cmp["Before"]
        alm_show = alm_cmp.copy()
        alm_show[["Before", "After", "Impact"]] = alm_show[["Before", "After", "Impact"]].round(4)
        money_rows = alm_show["Metric"].isin(["Surplus ($)", "PV01 net ($ per 1bp)"])
        if bool(money_rows.any()):
            alm_show.loc[money_rows, ["Before", "After", "Impact"]] = (
                alm_show.loc[money_rows, ["Before", "After", "Impact"]].round(0).astype(int)
            )
        st.dataframe(alm_show, use_container_width=True, hide_index=True)

        age_alm = np.round((base_contract.issue_age + _alm_b.times_years).astype(float), 2)
        path_cmp = pd.DataFrame(
            {
                "Funding ratio (before)": _alm_b.funding_ratio,
                "Funding ratio (after)": _alm_a.funding_ratio,
            },
            index=age_alm,
        )
        st.markdown("**Funding ratio path**")
        st.line_chart(path_cmp)
        sur_cmp = pd.DataFrame(
            {"Surplus before": _alm_b.surplus, "Surplus after": _alm_a.surplus},
            index=age_alm,
        )
        st.markdown("**Surplus path**")
        sur_disp = _round_for_visuals(sur_cmp)
        sur_disp[["Surplus before", "Surplus after"]] = sur_disp[
            ["Surplus before", "Surplus after"]
        ].astype(int)
        st.line_chart(sur_disp)

        st.markdown("**PV assets and liabilities**")
        pv_cmp = pd.DataFrame(
            {
                "PV assets (before)": _alm_b.asset_market_value,
                "PV assets (after)": _alm_a.asset_market_value,
                "PV liabilities (before)": _alm_b.liability_pv,
                "PV liabilities (after)": _alm_a.liability_pv,
            },
            index=age_alm,
        )
        pv_disp = _round_for_visuals(pv_cmp)
        for c in [
            "PV assets (before)",
            "PV assets (after)",
            "PV liabilities (before)",
            "PV liabilities (after)",
        ]:
            pv_disp[c] = pv_disp[c].astype(int)
        st.line_chart(pv_disp)

        st.markdown("**ALM key rate duration (before vs after)**")
        try:
            asm_krd_wf = asm_whatif_used if isinstance(asm_whatif_used, sp.ALMAssumptions) else None
            if asm_krd_wf is not None:
                key_tenors = np.array(
                    [
                        float(b.tenor_years)
                        for b in asm_krd_wf.allocation.buckets
                        if float(b.tenor_years) > 1e-12
                    ],
                    dtype=float,
                )
                if key_tenors.size > 0:
                    a0_wf = st.session_state.get("alm_last_initial_asset_market_value")
                    if not isinstance(a0_wf, (int, float, np.floating)):
                        a0_wf = st.session_state.get("alm_current_initial_asset_market_value")
                    a0 = (
                        float(a0_wf)
                        if isinstance(a0_wf, (int, float, np.floating))
                        else float(base_res.single_premium)
                    )

                    def _compute_krd_set(
                        *,
                        curve_liab: sp.YieldCurve,
                        curve_asset: sp.YieldCurve,
                        spread_use: float,
                        cashflows_use: np.ndarray,
                        scenario_label: str,
                    ) -> list[dict[str, float | str]]:
                        w_use = np.asarray(asm_krd_wf.allocation.weights, dtype=float)
                        bond_tenors = np.array(
                            [float(b.tenor_years) for b in asm_krd_wf.allocation.buckets[1:]],
                            dtype=float,
                        )
                        df0_asset = curve_asset.discount_factors(bond_tenors, spread=spread_use)
                        target_mv_bonds = w_use[1:] * a0
                        bond_faces = np.where(df0_asset > 1e-15, target_mv_bonds / df0_asset, 0.0)
                        l0 = float(
                            np.sum(
                                cashflows_use
                                * curve_liab.discount_factors(
                                    base_res.times_years, spread=spread_use
                                )
                            )
                        )
                        net0 = max(1e-9, a0 - l0)
                        out_rows: list[dict[str, float | str]] = []
                        for kt in key_tenors:
                            cl_b = _key_rate_bump_curve(
                                curve_liab,
                                key_tenor_years=float(kt),
                                key_tenors_years=key_tenors,
                                bump_bps=1.0,
                            )
                            ca_b = _key_rate_bump_curve(
                                curve_asset,
                                key_tenor_years=float(kt),
                                key_tenors_years=key_tenors,
                                bump_bps=1.0,
                            )
                            a_b = float(
                                w_use[0] * a0
                                + np.sum(
                                    bond_faces
                                    * ca_b.discount_factors(bond_tenors, spread=spread_use)
                                )
                            )
                            l_b = float(
                                np.sum(
                                    cashflows_use
                                    * cl_b.discount_factors(base_res.times_years, spread=spread_use)
                                )
                            )
                            out_rows.extend(
                                [
                                    {
                                        "Tenor": f"{kt:g}Y",
                                        "Tenor years": float(kt),
                                        "Scenario": scenario_label,
                                        "Series": "Assets KRD",
                                        "KRD": -((a_b - a0) / (max(1e-9, a0) * 1e-4)),
                                    },
                                    {
                                        "Tenor": f"{kt:g}Y",
                                        "Tenor years": float(kt),
                                        "Scenario": scenario_label,
                                        "Series": "Liabilities KRD",
                                        "KRD": -((l_b - l0) / (max(1e-9, l0) * 1e-4)),
                                    },
                                    {
                                        "Tenor": f"{kt:g}Y",
                                        "Tenor years": float(kt),
                                        "Scenario": scenario_label,
                                        "Series": "Surplus KRD",
                                        "KRD": -(((a_b - l_b) - (a0 - l0)) / (net0 * 1e-4)),
                                    },
                                ]
                            )
                        return out_rows

                    krd_rows = []
                    krd_rows.extend(
                        _compute_krd_set(
                            curve_liab=base_curve,
                            curve_asset=base_curve,
                            spread_use=float(base_spread),
                            cashflows_use=np.asarray(
                                base_res.expected_total_cashflows, dtype=float
                            ),
                            scenario_label="Before",
                        )
                    )
                    krd_rows.extend(
                        _compute_krd_set(
                            curve_liab=shocked_curve,
                            curve_asset=yc_alm_asset,
                            spread_use=float(shocked_spread),
                            cashflows_use=np.asarray(cf_alm, dtype=float),
                            scenario_label="After",
                        )
                    )
                    krd_wf_df = pd.DataFrame(krd_rows).sort_values(
                        ["Series", "Tenor years", "Scenario"]
                    )
                    tenor_order = [f"{float(t):g}Y" for t in np.sort(np.unique(key_tenors))]
                    series_order = ["Assets KRD", "Liabilities KRD", "Surplus KRD"]
                    # Faceted specs often render with a narrow default plot width in Streamlit (squished left,
                    # empty space right). Use stacked panels with container width so plots fill the column.
                    enc_x = alt.X(
                        "Tenor:N",
                        sort=tenor_order,
                        title="Key tenor",
                        axis=alt.Axis(labelAngle=0, labelPadding=4),
                    )

                    def _wf_krd_panel(
                        subtitle: str, df_sub: pd.DataFrame, *, show_legend: bool
                    ) -> alt.Chart:
                        color_enc = alt.Color(
                            "Scenario:N",
                            sort=["Before", "After"],
                            scale=alt.Scale(
                                domain=["Before", "After"], range=["#4c78a8", "#f58518"]
                            ),
                            legend=(
                                alt.Legend(orient="top", direction="horizontal")
                                if show_legend
                                else None
                            ),
                        )
                        return (
                            alt.Chart(df_sub)
                            .mark_line(point=True, strokeWidth=2.5)
                            .encode(
                                x=enc_x,
                                y=alt.Y("KRD:Q", title="KRD (years)"),
                                color=color_enc,
                                tooltip=[
                                    alt.Tooltip("Series:N"),
                                    alt.Tooltip("Tenor:N"),
                                    alt.Tooltip("Scenario:N"),
                                    alt.Tooltip("KRD:Q", format=".4f"),
                                ],
                            )
                            .properties(width="container", height=115, title=subtitle)
                        )

                    panels = [
                        _wf_krd_panel(
                            s,
                            krd_wf_df[krd_wf_df["Series"] == s],
                            show_legend=(i == 0),
                        )
                        for i, s in enumerate(series_order)
                    ]
                    krd_wf_chart = (
                        alt.vconcat(*panels, spacing=8)
                        .resolve_scale(y="independent")
                        .configure_view(strokeWidth=0)
                    )
                    st.altair_chart(krd_wf_chart, use_container_width=True)
                    st.caption(
                        "Each panel compares Before vs After KRD at key tenors for one series, with independent y-scales to keep "
                        "the view readable when Surplus KRD magnitudes are much larger."
                    )
                else:
                    st.info("No positive tenors available for What-if ALM KRD chart.")
            else:
                st.info("What-if ALM KRD chart unavailable: ALM assumptions were not available.")
        except Exception as ex:
            st.info(f"What-if ALM KRD chart unavailable for current inputs: {ex!r}")

    st.caption(
        "Impact shown as After - Before. Tail risk uses the 95th percentile of simulated premiums under the selected equity regime."
    )


def _render_shared_pricing_economics_controls(
    economics_product: ProductType, *, include_monte_carlo: bool
) -> None:
    """Yield / mortality / expenses / horizon / spread / economic scenario; same ``run_*`` keys as Pricing Run."""
    with st.expander("Yield curve", expanded=True):
        _y_opts = ("flat", "zero_csv", "par_bootstrap")
        ensure_session_choice(st.session_state, RUN_KEY.Y_MODE, _y_opts, "par_bootstrap")
        st.radio(
            "Source",
            options=list(_y_opts),
            format_func=lambda x: {
                "flat": "Flat zero rate",
                "zero_csv": "Zero curve CSV",
                "par_bootstrap": "Par yields CSV → bootstrap zeros",
            }[x],
            horizontal=True,
            key=RUN_KEY.Y_MODE,
        )
        y_mode = str(st.session_state.get(RUN_KEY.Y_MODE, "par_bootstrap"))
        flat_rate = float(st.session_state.get(RUN_KEY.FLAT_RATE, 0.04))
        zero_csv = str(st.session_state.get(RUN_KEY.ZERO_CSV, sp.DEFAULT_ZERO_CURVE_CSV))
        par_csv = str(st.session_state.get(RUN_KEY.PAR_CSV, sp.DEFAULT_PAR_CURVE_CSV))
        coupon_freq = int(st.session_state.get(RUN_KEY.COUPON_FREQ, 2))
        if y_mode == "flat":
            run_number_input(
                "Flat continuously compounded zero rate",
                RUN_KEY.FLAT_RATE,
                default=0.04,
                format="%.4f",
            )
        elif y_mode == "zero_csv":
            st.text_input("Zero curve CSV path", key=RUN_KEY.ZERO_CSV)
        else:
            st.text_input("Par yield CSV path", key=RUN_KEY.PAR_CSV)
            run_number_input(
                "Coupon payments per year", RUN_KEY.COUPON_FREQ, default=2, min_value=1, step=1
            )

    with st.expander("Mortality", expanded=True):
        mortality_options = list(get_product_mortality_mode_options(economics_product))
        ensure_session_choice(
            st.session_state,
            RUN_KEY.M_MODE,
            mortality_options,
            get_product_default_mortality_mode(economics_product),
        )
        st.radio(
            "Table",
            options=mortality_options,
            format_func=lambda x: get_mortality_mode_label(str(x)),
            horizontal=True,
            key=RUN_KEY.M_MODE,
        )
        m_mode = str(st.session_state.get(RUN_KEY.M_MODE, ""))
        if m_mode == "qx_csv":
            st.text_input("q_x CSV (columns age, qx)", key=RUN_KEY.QX_CSV)
        elif m_mode == "rp2014_mp2016":
            st.caption(
                "SOA workbooks are optional if matching CSV extracts already exist beside the xlsx paths."
            )
            if not str(st.session_state.get(RUN_KEY.RP_XLSX, "")).strip():
                st.session_state[RUN_KEY.RP_XLSX] = sp.DEFAULT_RP2014_XLSX
            if not str(st.session_state.get(RUN_KEY.RP_OUT, "")).strip():
                st.session_state[RUN_KEY.RP_OUT] = sp.DEFAULT_RP2014_MALE_HEALTHY_QX_CSV
            if not str(st.session_state.get(RUN_KEY.MP_XLSX, "")).strip():
                st.session_state[RUN_KEY.MP_XLSX] = sp.DEFAULT_MP2016_XLSX
            if not str(st.session_state.get(RUN_KEY.MP_OUT, "")).strip():
                st.session_state[RUN_KEY.MP_OUT] = sp.DEFAULT_MP2016_MALE_IMPROVEMENT_CSV
            st.text_input("RP-2014 xlsx", key=RUN_KEY.RP_XLSX)
            st.text_input("RP-2014 healthy male qx cache CSV", key=RUN_KEY.RP_OUT)
            st.text_input("MP-2016 xlsx", key=RUN_KEY.MP_XLSX)
            st.text_input("MP-2016 improvement cache CSV", key=RUN_KEY.MP_OUT)
        elif m_mode == "us_ssa_2015_period":
            st.caption(
                "Source: SSA actuarial life table (US Social Security area population), period year 2015."
            )

    with st.expander("Expenses & valuation", expanded=True):
        _exp_opts = ("csv", "manual")
        ensure_session_choice(st.session_state, RUN_KEY.EXPENSE_MODE, _exp_opts, "csv")
        st.radio(
            "Expenses",
            options=list(_exp_opts),
            format_func=lambda x: "Load from CSV" if x == "csv" else "Enter manually",
            horizontal=True,
            key=RUN_KEY.EXPENSE_MODE,
        )
        expense_mode = str(st.session_state.get(RUN_KEY.EXPENSE_MODE, "csv"))
        if expense_mode == "csv":
            st.text_input("Expenses CSV path", key=RUN_KEY.EXPENSES_CSV)
        else:
            run_number_input("Policy expense at issue ($)", RUN_KEY.POLICY_EXPENSE, default=0.0)
            run_number_input(
                "Premium expense (% of single premium)",
                RUN_KEY.PREMIUM_EXPENSE_PCT,
                default=0.0,
                min_value=0.0,
                max_value=99.99,
                help="Enter 2 for 2%. Must stay below 100%.",
            )
            run_number_input(
                "Monthly expense while alive ($)", RUN_KEY.MONTHLY_EXPENSE, default=0.0
            )
        run_number_input(
            "Valuation year (calendar)",
            RUN_KEY.VALUATION_YEAR,
            default=2025,
            min_value=1950,
            max_value=2100,
            help="Used for RP+MP calendar-year mortality; ignored for static/synthetic q_x.",
        )
        run_number_input(
            "Horizon age (stop monthly grid)",
            RUN_KEY.HORIZON_AGE,
            default=110,
            min_value=1,
            max_value=130,
        )
        run_number_input(
            "Credit spread added to zero rate", RUN_KEY.SPREAD, default=0.0, format="%.4f"
        )

    product_caps = get_product_capabilities(economics_product)
    can_use_economic_scenario = bool(product_caps.supports_economic_scenario)
    can_use_monte_carlo = bool(product_caps.supports_monte_carlo)
    if can_use_economic_scenario:
        _econ_title = (
            "Economic scenario (RILA segment crediting & SPIA benefit indexation)"
            if economics_product == ProductType.RILA
            else "Economic scenario (benefit indexation & expense inflation)"
        )
        with st.expander(_econ_title, expanded=True):
            st.checkbox(
                "Use S&P 500 proxy CSV for index levels (month, sp500_level)",
                help="If off, index is flat (zero equity returns). Required for meaningful RILA crediting.",
                key=RUN_KEY.USE_INDEX,
            )
            st.text_input(
                "Index scenario CSV (columns: month, sp500_level for months 0..N)",
                key=RUN_KEY.INDEX_CSV,
            )
            run_number_input(
                "Expense annual inflation (%, not tied to S&P)",
                RUN_KEY.EXPENSE_INFLATION_PCT,
                default=2.5,
                min_value=0.0,
                max_value=25.0,
                help="Applied monthly as (1 + annual)^(1/12) to maintenance expenses only.",
            )

    if include_monte_carlo and can_use_monte_carlo:
        with st.expander("Monte Carlo (stochastic index assumption)", expanded=True):
            st.checkbox(
                "Enable Monte Carlo on index returns",
                help="Simulates index paths and reprices for each path. Mortality, curve, and expense inflation remain deterministic.",
                key=RUN_KEY.MC_ENABLE,
            )
            run_number_input(
                "Number of simulations",
                RUN_KEY.MC_N_SIMS,
                default=100,
                min_value=100,
                max_value=20000,
                step=100,
            )
            run_number_input(
                "Random seed",
                RUN_KEY.MC_SEED,
                default=42,
                min_value=0,
                max_value=2_147_483_647,
                step=1,
            )
            run_number_input(
                "Annual drift (%)",
                RUN_KEY.MC_DRIFT_PCT,
                default=6.0,
                min_value=-50.0,
                max_value=50.0,
                step=0.1,
                help=(
                    "Real-world equity drift used for index simulation. For RILA pricing, "
                    "drift well above the risk-free rate combined with a high cap/participation can "
                    "push some paths into infeasibility (PV death benefits per $1 premium ≥ 1). "
                    "Reduce drift, cap, or participation if too many paths are skipped."
                ),
            )
            run_number_input(
                "Annual volatility (%)",
                RUN_KEY.MC_VOL_PCT,
                default=15.0,
                min_value=0.0,
                max_value=200.0,
                step=0.1,
            )
            run_number_input(
                "Initial index level (S0)", RUN_KEY.MC_S0, default=100.0, min_value=0.01, step=1.0
            )


def _snapshot_pricing_economics_from_session(
    selected_product: ProductType,
) -> tuple[Any, ...]:
    """Read economics widgets from session after expanders ran (Pricing Run ``if run`` path)."""
    y_mode = str(st.session_state.get(RUN_KEY.Y_MODE, "par_bootstrap"))
    flat_rate = float(st.session_state.get(RUN_KEY.FLAT_RATE, 0.04))
    zero_csv = str(st.session_state.get(RUN_KEY.ZERO_CSV, sp.DEFAULT_ZERO_CURVE_CSV))
    par_csv = str(st.session_state.get(RUN_KEY.PAR_CSV, sp.DEFAULT_PAR_CURVE_CSV))
    coupon_freq = int(st.session_state.get(RUN_KEY.COUPON_FREQ, 2))
    m_mode = str(st.session_state.get(RUN_KEY.M_MODE, "rp2014_mp2016"))
    qx_csv = str(st.session_state.get(RUN_KEY.QX_CSV, sp.DEFAULT_MORTALITY_QX_CSV))
    rp_xlsx = str(st.session_state.get(RUN_KEY.RP_XLSX, sp.DEFAULT_RP2014_XLSX))
    rp_out = str(st.session_state.get(RUN_KEY.RP_OUT, sp.DEFAULT_RP2014_MALE_HEALTHY_QX_CSV))
    mp_xlsx = str(st.session_state.get(RUN_KEY.MP_XLSX, sp.DEFAULT_MP2016_XLSX))
    mp_out = str(st.session_state.get(RUN_KEY.MP_OUT, sp.DEFAULT_MP2016_MALE_IMPROVEMENT_CSV))
    expense_mode = str(st.session_state.get(RUN_KEY.EXPENSE_MODE, "csv"))
    pol = float(st.session_state.get(RUN_KEY.POLICY_EXPENSE, 0.0))
    prem_pct = float(st.session_state.get(RUN_KEY.PREMIUM_EXPENSE_PCT, 0.0))
    monthly_ex = float(st.session_state.get(RUN_KEY.MONTHLY_EXPENSE, 0.0))
    valuation_year = int(st.session_state.get(RUN_KEY.VALUATION_YEAR, 2025))
    horizon_age = int(st.session_state.get(RUN_KEY.HORIZON_AGE, 110))
    spread = float(st.session_state.get(RUN_KEY.SPREAD, 0.0))
    product_caps = get_product_capabilities(selected_product)
    can_use_economic_scenario = bool(product_caps.supports_economic_scenario)
    use_index = bool(st.session_state.get(RUN_KEY.USE_INDEX, True)) if can_use_economic_scenario else False
    index_csv = str(st.session_state.get(RUN_KEY.INDEX_CSV, sp.DEFAULT_SP500_SCENARIO_CSV))
    expense_inflation_pct = float(st.session_state.get(RUN_KEY.EXPENSE_INFLATION_PCT, 2.5))
    can_use_monte_carlo = bool(product_caps.supports_monte_carlo)
    mc_enable = bool(st.session_state.get(RUN_KEY.MC_ENABLE, False)) if can_use_monte_carlo else False
    mc_n_sims = int(st.session_state.get(RUN_KEY.MC_N_SIMS, 100))
    mc_seed = int(st.session_state.get(RUN_KEY.MC_SEED, 42))
    mc_drift_pct = float(st.session_state.get(RUN_KEY.MC_DRIFT_PCT, 6.0))
    mc_vol_pct = float(st.session_state.get(RUN_KEY.MC_VOL_PCT, 15.0))
    mc_s0 = float(st.session_state.get(RUN_KEY.MC_S0, 100.0))
    expenses_csv = str(st.session_state.get(RUN_KEY.EXPENSES_CSV, sp.DEFAULT_EXPENSES_CSV))
    return (
        y_mode,  # type: ignore[return-value]
        flat_rate,
        zero_csv,
        par_csv,
        coupon_freq,
        m_mode,  # type: ignore[return-value]
        qx_csv,
        rp_xlsx,
        rp_out,
        mp_xlsx,
        mp_out,
        expense_mode,
        pol,
        prem_pct,
        monthly_ex,
        valuation_year,
        horizon_age,
        spread,
        use_index,
        index_csv,
        expense_inflation_pct,
        mc_enable,
        mc_n_sims,
        mc_seed,
        mc_drift_pct,
        mc_vol_pct,
        mc_s0,
        expenses_csv,
    )


def _render_run_and_results() -> None:
    st.header("Pricing Run")
    _seed_run_form_state_from_last_inputs()
    st.markdown(
        """
        <style>
            .product-type-callout {
                border: 2px solid #1f77b4;
                border-radius: 10px;
                padding: 10px 12px;
                background: rgba(31, 119, 180, 0.08);
                margin-bottom: 10px;
            }
            .product-type-callout strong {
                font-size: 1.05rem;
            }
        </style>
        <div class="product-type-callout">
            <strong>Primary input: Product Type</strong><br/>
            This selection controls which pricing engine, assumptions, and downstream outputs are active for this run.
        </div>
        """,
        unsafe_allow_html=True,
    )
    product_options = list(product_options_for_ui())
    product_values = [p.value for p in product_options]
    if st.session_state.get("run_product_type") not in product_values and product_values:
        st.session_state["run_product_type"] = product_values[0]
    selected_product = st.selectbox(
        "Product type",
        options=product_values,
        format_func=lambda raw: product_label(ProductType(raw)),
        help="Run exactly one product per execution.",
        key="run_product_type",
    )
    selected_product = ProductType(selected_product)
    last_product_raw = st.session_state.get("_run_last_product_type")
    switched_product = (
        last_product_raw is not None and str(last_product_raw) != selected_product.value
    )
    _normalize_run_state_for_selected_product(
        st.session_state,
        selected_product=selected_product,
        switched_product=switched_product,
    )
    st.session_state["_run_last_product_type"] = selected_product.value
    product_ui_cfg = get_product_ui_config(selected_product)
    if product_ui_cfg.selected_info_message:
        st.info(product_ui_cfg.selected_info_message)

    with st.expander("Contract", expanded=True):
        c1, c2, c3 = st.columns(3)
        issue_age = run_number_input(
            "Issue age", "run_issue_age", default=65, min_value=0, max_value=120, step=1
        )
        sex = c2.selectbox("Sex (metadata)", options=["male", "female"], key="run_sex")
        if selected_product == ProductType.TERM_LIFE:
            term_ui = get_term_contract_ui_config()
            benefit_annual = run_number_input(
                term_ui.death_benefit_label,
                "run_term_benefit_annual",
                default=float(term_ui.default_death_benefit),
                min_value=1.0,
                step=10_000.0,
            )
            t1, t2, t3 = st.columns(3)
            term_choice = t1.selectbox(
                "Term length", options=list(term_ui.term_length_options), key="run_term_length"
            )
            premium_mode_choice = t2.selectbox(
                "Premium mode",
                options=list(term_ui.premium_mode_options),
                key="run_term_premium_mode",
            )
            benefit_timing_choice = t3.selectbox(
                "Benefit timing",
                options=list(term_ui.benefit_timing_options),
                key="run_term_benefit_timing",
            )
            monthly_premium = run_number_input(
                "Monthly premium ($)",
                "run_term_monthly_premium",
                default=float(term_ui.default_monthly_premium),
                min_value=0.0,
                step=10.0,
                replace_non_positive=True,
            )
        elif selected_product == ProductType.RILA:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            render_rila_pricing_controls(st, run_number_input)
        elif selected_product == ProductType.MYGA:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            m1, m2, m3 = st.columns(3)
            with m1:
                run_number_input(
                    "Single premium ($)", "run_myga_single_premium",
                    default=100_000.0, min_value=1.0, step=1_000.0,
                )
            with m2:
                run_number_input(
                    "Declared rate (annual decimal)", "run_myga_declared_rate",
                    default=0.045, min_value=-0.5, max_value=1.0, format="%.4f",
                )
            with m3:
                run_number_input(
                    "Guarantee years", "run_myga_guarantee_years",
                    default=5, min_value=1, max_value=30, step=1,
                )
        elif selected_product == ProductType.FIA:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            f1, f2, f3, f4, f5 = st.columns(5)
            with f1:
                run_number_input(
                    "Single premium ($)", "run_fia_single_premium",
                    default=100_000.0, min_value=1.0, step=1_000.0,
                )
            with f2:
                run_number_input(
                    "Participation", "run_fia_participation",
                    default=0.80, min_value=0.0, max_value=5.0, format="%.4f",
                )
            with f3:
                run_number_input(
                    "Annual cap", "run_fia_cap",
                    default=0.07, min_value=-1.0, max_value=2.0, format="%.4f",
                )
            with f4:
                run_number_input(
                    "Annual floor", "run_fia_floor",
                    default=0.0, min_value=-1.0, max_value=1.0, format="%.4f",
                )
            with f5:
                run_number_input(
                    "Horizon years", "run_fia_horizon_years",
                    default=10, min_value=1, max_value=40, step=1,
                )
        elif selected_product == ProductType.VARIABLE_ANNUITY:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            v1, v2, v3 = st.columns(3)
            with v1:
                run_number_input(
                    "Single premium ($)", "run_va_single_premium",
                    default=100_000.0, min_value=1.0, step=1_000.0,
                )
            with v2:
                run_number_input(
                    "M&E charge (annual)", "run_va_me_charge",
                    default=0.014, min_value=0.0, max_value=0.05,
                    format="%.4f", help="Industry typical 100-200 bps.",
                )
            with v3:
                run_number_input(
                    "Horizon years", "run_va_horizon_years",
                    default=20, min_value=1, max_value=40, step=1,
                )
        elif selected_product == ProductType.WHOLE_LIFE:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            w1, w2 = st.columns(2)
            with w1:
                run_number_input(
                    "Face amount ($)", "run_wl_face_amount",
                    default=250_000.0, min_value=1.0, step=10_000.0,
                )
            with w2:
                st.selectbox(
                    "Smoker class", options=["nonsmoker", "smoker"],
                    key="run_wl_smoker_class",
                )
        elif selected_product == ProductType.UNIVERSAL_LIFE:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            u1, u2, u3 = st.columns(3)
            with u1:
                run_number_input(
                    "Face amount ($)", "run_ul_face_amount",
                    default=250_000.0, min_value=1.0, step=10_000.0,
                )
                st.selectbox(
                    "Smoker class", options=["nonsmoker", "smoker"],
                    key="run_ul_smoker_class",
                )
            with u2:
                run_number_input(
                    "Single premium ($)", "run_ul_single_premium",
                    default=25_000.0, min_value=1.0, step=1_000.0,
                )
                run_number_input(
                    "Premium load (decimal)", "run_ul_premium_load",
                    default=0.06, min_value=0.0, max_value=0.5, format="%.4f",
                )
            with u3:
                run_number_input(
                    "Monthly expense charge ($)", "run_ul_monthly_expense",
                    default=7.50, min_value=0.0, step=0.50,
                )
                run_number_input(
                    "Declared rate (annual)", "run_ul_declared_rate",
                    default=0.04, min_value=-0.5, max_value=1.0, format="%.4f",
                )
        elif selected_product == ProductType.INDEXED_UL:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            render_indexed_ul_pricing_controls(st, run_number_input)
        elif selected_product == ProductType.VARIABLE_UL:
            benefit_annual = 0.0
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0
            x1, x2, x3 = st.columns(3)
            with x1:
                run_number_input(
                    "Face amount ($)", "run_vul_face_amount",
                    default=250_000.0, min_value=1.0, step=10_000.0,
                )
                st.selectbox(
                    "Smoker class", options=["nonsmoker", "smoker"],
                    key="run_vul_smoker_class",
                )
            with x2:
                run_number_input(
                    "Single premium ($)", "run_vul_single_premium",
                    default=25_000.0, min_value=1.0, step=1_000.0,
                )
                run_number_input(
                    "Premium load", "run_vul_premium_load",
                    default=0.06, min_value=0.0, max_value=0.5, format="%.4f",
                )
            with x3:
                run_number_input(
                    "Monthly expense", "run_vul_monthly_expense",
                    default=7.50, min_value=0.0, step=0.50,
                )
        else:
            benefit_annual = run_number_input(
                "Annual benefit ($)",
                "run_spia_benefit_annual",
                default=100_000.0,
                min_value=0.0,
                step=1_000.0,
            )
            term_choice = "n/a"
            premium_mode_choice = "n/a"
            benefit_timing_choice = "n/a"
            monthly_premium = 0.0

    _render_shared_pricing_economics_controls(selected_product, include_monte_carlo=True)

    product_caps = get_product_capabilities(selected_product)
    can_use_monte_carlo = bool(product_caps.supports_monte_carlo)

    _is_implemented_product = selected_product in (
        ProductType.SPIA,
        ProductType.TERM_LIFE,
        ProductType.RILA,
        ProductType.MYGA,
        ProductType.FIA,
        ProductType.VARIABLE_ANNUITY,
        ProductType.WHOLE_LIFE,
        ProductType.UNIVERSAL_LIFE,
        ProductType.INDEXED_UL,
        ProductType.VARIABLE_UL,
    )
    run = st.button(
        "Run pricing",
        type="primary",
        disabled=not _is_implemented_product,
        help=(
            None
            if _is_implemented_product
            else f"{product_label(selected_product)} is not implemented in this release."
        ),
    )

    if run:
        try:
            snap = _snapshot_pricing_economics_from_session(selected_product)
            (
                y_mode,
                flat_rate,
                zero_csv,
                par_csv,
                coupon_freq,
                m_mode,
                qx_csv,
                rp_xlsx,
                rp_out,
                mp_xlsx,
                mp_out,
                expense_mode,
                pol,
                prem_pct,
                monthly_ex,
                valuation_year,
                horizon_age,
                spread,
                use_index,
                index_csv,
                expense_inflation_pct,
                mc_enable,
                mc_n_sims,
                mc_seed,
                mc_drift_pct,
                mc_vol_pct,
                mc_s0,
                expenses_csv,
            ) = snap
            adapter = get_product_adapter(selected_product)
            yc = _build_yield_curve(
                y_mode,  # type: ignore[arg-type]
                flat_rate=flat_rate,
                zero_csv=zero_csv,
                par_csv=par_csv,
                coupon_freq=coupon_freq,
            )
            mort, needs_vy = _build_mortality(
                m_mode,  # type: ignore[arg-type]
                product_type=selected_product,
                sex="male" if sex == "male" else "female",
                qx_csv=qx_csv,
                rp_xlsx=rp_xlsx,
                rp_out_csv=rp_out,
                mp_xlsx=mp_xlsx,
                mp_out_csv=mp_out,
            )
            vy: int | None = int(valuation_year) if needs_vy else None
            vy_inputs = int(valuation_year)
            idx_path = str(_resolve_path(index_csv)) if use_index else None
            expense_annual_inflation = float(expense_inflation_pct) / 100.0

            if selected_product == ProductType.TERM_LIFE:
                # Parse Term selectbox labels via the canonical maps in
                # product_registry. Hard-coding term_years/premium_mode/
                # benefit_timing here is a regression -- see
                # tests/test_pricing_ui_term_config.py.
                term_years_value = parse_term_length_label_to_years(str(term_choice))
                premium_mode_value = parse_term_premium_mode_label(str(premium_mode_choice))
                benefit_timing_value = parse_term_benefit_timing_label(
                    str(benefit_timing_choice)
                )
                contract = tp.TermLifeContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    death_benefit=float(benefit_annual),
                    monthly_premium=float(monthly_premium),
                    term_years=int(term_years_value),
                    premium_mode=premium_mode_value,  # type: ignore[arg-type]
                    benefit_timing=benefit_timing_value,  # type: ignore[arg-type]
                )
            elif selected_product == ProductType.RILA:
                contract = build_rila_contract_from_session(
                    st.session_state,
                    issue_age=int(issue_age),
                    sex=str(sex),
                )
            elif selected_product == ProductType.MYGA:
                import myga_projection as my_proj
                contract = my_proj.MYGAContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    single_premium=float(st.session_state.get("run_myga_single_premium", 100_000.0)),
                    declared_rate_annual=float(st.session_state.get("run_myga_declared_rate", 0.045)),
                    guarantee_years=int(st.session_state.get("run_myga_guarantee_years", 5)),
                )
            elif selected_product == ProductType.FIA:
                import fia_projection as fp_proj
                contract = fp_proj.FIAContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    single_premium=float(st.session_state.get("run_fia_single_premium", 100_000.0)),
                    participation=float(st.session_state.get("run_fia_participation", 0.80)),
                    cap=float(st.session_state.get("run_fia_cap", 0.07)),
                    floor=float(st.session_state.get("run_fia_floor", 0.0)),
                    horizon_years=int(st.session_state.get("run_fia_horizon_years", 10)),
                )
            elif selected_product == ProductType.VARIABLE_ANNUITY:
                import va_projection as va_proj
                contract = va_proj.VAContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    single_premium=float(st.session_state.get("run_va_single_premium", 100_000.0)),
                    me_charge_annual=float(st.session_state.get("run_va_me_charge", 0.014)),
                    horizon_years=int(st.session_state.get("run_va_horizon_years", 20)),
                )
            elif selected_product == ProductType.WHOLE_LIFE:
                import wl_projection as wl_proj
                contract = wl_proj.WLContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    smoker_class=str(st.session_state.get("run_wl_smoker_class", "nonsmoker")),  # type: ignore[arg-type]
                    face_amount=float(st.session_state.get("run_wl_face_amount", 250_000.0)),
                )
            elif selected_product == ProductType.UNIVERSAL_LIFE:
                import ul_projection as ul_proj
                contract = ul_proj.ULContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    smoker_class=str(st.session_state.get("run_ul_smoker_class", "nonsmoker")),  # type: ignore[arg-type]
                    face_amount=float(st.session_state.get("run_ul_face_amount", 250_000.0)),
                    single_premium=float(st.session_state.get("run_ul_single_premium", 25_000.0)),
                    premium_load_pct=float(st.session_state.get("run_ul_premium_load", 0.06)),
                    monthly_expense_charge=float(st.session_state.get("run_ul_monthly_expense", 7.50)),
                    declared_rate_annual=float(st.session_state.get("run_ul_declared_rate", 0.04)),
                )
            elif selected_product == ProductType.INDEXED_UL:
                contract = build_indexed_ul_contract_from_session(
                    st.session_state,
                    issue_age=int(issue_age),
                    sex=str(sex),
                )
            elif selected_product == ProductType.VARIABLE_UL:
                import vul_projection as vul_proj
                contract = vul_proj.VULContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    smoker_class=str(st.session_state.get("run_vul_smoker_class", "nonsmoker")),  # type: ignore[arg-type]
                    face_amount=float(st.session_state.get("run_vul_face_amount", 250_000.0)),
                    single_premium=float(st.session_state.get("run_vul_single_premium", 25_000.0)),
                    premium_load_pct=float(st.session_state.get("run_vul_premium_load", 0.06)),
                    monthly_expense_charge=float(st.session_state.get("run_vul_monthly_expense", 7.50)),
                )
            else:
                contract = sp.SPIAContract(
                    issue_age=int(issue_age),
                    sex="male" if sex == "male" else "female",
                    benefit_annual=float(benefit_annual),
                )

            expenses_arg: sp.ExpenseAssumptions | None = None
            if expense_mode == "manual":
                expenses_arg = sp.ExpenseAssumptions(
                    policy_expense_dollars=pol,
                    premium_expense_rate=prem_pct / 100.0,
                    monthly_expense_dollars=monthly_ex,
                )
                expenses_used = expenses_arg
            else:
                try:
                    expenses_used = sp.ExpenseAssumptions.load_from_csv(
                        str(_resolve_path(expenses_csv))
                    )
                except (FileNotFoundError, ValueError, KeyError):
                    expenses_used = sp.ExpenseAssumptions(0.0, 0.0, 0.0)

            res = adapter.price(
                contract=contract,
                yield_curve=yc,
                mortality=mort,
                horizon_age=int(horizon_age),
                spread=float(spread),
                valuation_year=vy,
                expenses=expenses_arg,
                expenses_csv_path=str(_resolve_path(expenses_csv)),
                index_scenario_csv_path=idx_path,
                expense_annual_inflation=expense_annual_inflation,
            )
            _clear_dependent_state_on_pricing_change()
            st.session_state["pricing_res"] = res
            st.session_state["pricing_contract"] = contract
            st.session_state["pricing_product_type"] = selected_product.value
            st.session_state["pricing_run_id"] = int(st.session_state.get("pricing_run_id", 0)) + 1
            st.session_state["pricing_err"] = None
            st.session_state["pricing_meta"] = {
                "product_type": selected_product.value,
                "yield_mode": y_mode,
                "mortality_mode": m_mode,
                "expense_mode": expense_mode,
                "mc_enabled": bool(mc_enable and can_use_monte_carlo),
                "use_index": bool(use_index),
                "index_scenario_csv_path": idx_path,
            }
            st.session_state["pricing_run_inputs"] = {
                "sex": "male" if sex == "male" else "female",
                "issue_age": int(issue_age),
                "benefit_annual": float(benefit_annual),
                "horizon_age": int(horizon_age),
                "valuation_year": vy_inputs,
                "spread": float(spread),
                "expense_annual_inflation": float(expense_annual_inflation),
                "use_index": bool(use_index),
                "index_scenario_csv_path": idx_path,
                "mc_enabled": bool(mc_enable and can_use_monte_carlo),
                "mc_n_sims": int(mc_n_sims),
                "mc_seed": int(mc_seed),
                "mc_annual_drift": float(mc_drift_pct) / 100.0,
                "mc_annual_vol": float(mc_vol_pct) / 100.0,
                "mc_s0": float(mc_s0),
                "mc_base_settings_for_tail_risk": {
                    "annual_drift": float(mc_drift_pct) / 100.0,
                    "annual_vol": float(mc_vol_pct) / 100.0,
                    "seed": int(mc_seed),
                    "s0": float(mc_s0),
                },
                "term_length": term_choice,
                "term_premium_mode": premium_mode_choice,
                "term_benefit_timing": benefit_timing_choice,
                "term_monthly_premium": float(monthly_premium),
                "rila_participation": float(st.session_state.get("run_rila_participation", 1.0)),
                "rila_cap": float(st.session_state.get("run_rila_cap", 0.10)),
                "rila_floor": float(st.session_state.get("run_rila_floor", 0.0)),
                "rila_rider_fee_annual": float(st.session_state.get("run_rila_rider_fee", 0.01)),
                "mortality_qx_csv": qx_csv,
                "mortality_rp_xlsx": rp_xlsx,
                "mortality_rp_out_csv": rp_out,
                "mortality_mp_xlsx": mp_xlsx,
                "mortality_mp_out_csv": mp_out,
            }
            st.session_state["pricing_excel_context"] = {
                "contract": contract,
                "yield_curve": yc,
                "mortality": mort,
                "horizon_age": int(horizon_age),
                "spread": float(spread),
                "valuation_year": vy_inputs,
                "expenses": expenses_used,
                "yield_mode": y_mode,
                "mortality_mode": m_mode,
                "expense_mode": expense_mode,
            }

            # --- Monte Carlo (run before Excel so MC sheet can be embedded) ---
            mc_snap_for_excel: MCExcelSnapshot | None = None
            if mc_enable and can_use_monte_carlo:
                mc = adapter.price_monte_carlo(
                    contract=contract,
                    yield_curve=yc,
                    mortality=mort,
                    horizon_age=int(horizon_age),
                    spread=float(spread),
                    valuation_year=vy,
                    expenses=expenses_arg,
                    expenses_csv_path=str(_resolve_path(expenses_csv)),
                    expense_annual_inflation=expense_annual_inflation,
                    n_sims=int(mc_n_sims),
                    annual_drift=float(mc_drift_pct) / 100.0,
                    annual_vol=float(mc_vol_pct) / 100.0,
                    seed=int(mc_seed),
                    s0=float(mc_s0),
                )
                st.session_state["pricing_mc"] = mc
                st.session_state["pricing_mc_params"] = {
                    "annual_drift": float(mc_drift_pct) / 100.0,
                    "annual_vol": float(mc_vol_pct) / 100.0,
                    "s0": float(mc_s0),
                    "n_sims": int(mc_n_sims),
                    "seed": int(mc_seed),
                }
                # mc_excel_snapshot_from_result currently expects the
                # SPIA / RILA MC result shape (single_premium etc.). The
                # new product MC results omit those scalar fields. We
                # only embed an MC snapshot in the Excel for products
                # whose MC result is shape-compatible.
                if hasattr(mc, "single_premium") and hasattr(mc, "annuity_factor"):
                    mc_snap_for_excel = mc_excel_snapshot_from_result(
                        mc,
                        annual_drift=float(mc_drift_pct) / 100.0,
                        annual_vol=float(mc_vol_pct) / 100.0,
                        s0=float(mc_s0),
                    )
            else:
                st.session_state.pop("pricing_mc", None)
                st.session_state.pop("pricing_mc_params", None)

            # --- Excel workbook (built after MC so MC_Summary sheet can be included) ---
            _refresh_pricing_excel_workbook_in_session()
        except Exception as e:
            _clear_dependent_state_on_pricing_change()
            st.session_state["pricing_err"] = f"{type(e).__name__}: {e}"
            st.session_state["pricing_res"] = None
            st.session_state.pop("pricing_product_type", None)
            st.session_state.pop("pricing_run_inputs", None)
            st.session_state.pop("pricing_excel_context", None)
            st.session_state.pop("pricing_xlsx_bytes", None)
            st.session_state.pop("pricing_xlsx_built_error", None)
            st.session_state.pop("pricing_mc", None)
            st.session_state.pop("pricing_mc_params", None)
            st.session_state.pop("pricing_xlsx_has_mc", None)
            st.session_state.pop("pricing_xlsx_has_alm", None)

    err = st.session_state.get("pricing_err")
    res = st.session_state.get("pricing_res")
    contract_state = st.session_state.get("pricing_contract")

    if err:
        st.error(err)
    if res is not None and contract_state is not None:
        st.success("Pricing completed.")
        meta = st.session_state.get("pricing_meta") or {}

        product_raw = str(meta.get("product_type", ProductType.SPIA.value))
        product_type = (
            ProductType(product_raw)
            if product_raw in {p.value for p in ProductType}
            else ProductType.SPIA
        )
        m1, m2, m3, m4 = st.columns(4)
        metrics = get_pricing_metrics(product_type, res)
        for col, metric in zip((m1, m2, m3, m4), metrics):
            formatted = f"${metric.value:,.0f}" if metric.is_money else f"{metric.value:,.0f}"
            col.metric(metric.label, formatted)

        st.caption(
            f"Yield: {meta.get('yield_mode')}; mortality: {meta.get('mortality_mode')}; "
            f"expenses: {meta.get('expense_mode')}."
        )
        mc_res = st.session_state.get("pricing_mc")
        if mc_res is not None:
            st.subheader("Monte Carlo summary (index-path uncertainty)")
            # The full SPIA / RILA-shaped MC summary block uses the
            # ``premium_*`` aggregate fields. The new product MC results
            # (FIA / VA / IUL / VUL) only expose pv_benefit and the
            # per-path arrays in v1; render a compact summary for them
            # and skip the histogram premium chart.
            has_premium_summary = all(
                hasattr(mc_res, name)
                for name in (
                    "premium_mean", "premium_median", "premium_p05", "premium_p95",
                )
            )
            if has_premium_summary:
                n_infeasible_mc = int(getattr(mc_res, "n_infeasible", 0))
                n_feasible_mc = int(getattr(mc_res, "n_feasible", mc_res.n_sims))
                if n_infeasible_mc > 0:
                    pct = 100.0 * n_infeasible_mc / max(int(mc_res.n_sims), 1)
                    worst = float(getattr(mc_res, "infeasible_max_loading", 0.0))
                    st.warning(
                        f"{n_infeasible_mc:,} of {mc_res.n_sims:,} Monte Carlo paths ({pct:.1f}%) "
                        f"were economically infeasible (PV death benefits per $1 premium + premium "
                        f"expense rate exceeded 1; worst observed loading {worst:.4f}). Statistics "
                        "below use only the feasible paths. Reduce participation, cap, MC drift, or "
                        "horizon to lower the share."
                    )
                a1, a2, a3, a4 = st.columns(4)
                a1.metric("Mean premium", f"${mc_res.premium_mean:,.0f}")
                a2.metric("Median premium", f"${mc_res.premium_median:,.0f}")
                a3.metric("P5 premium", f"${mc_res.premium_p05:,.0f}")
                a4.metric("P95 premium", f"${mc_res.premium_p95:,.0f}")
                st.caption(
                    f"Simulations: {mc_res.n_sims:,} (feasible {n_feasible_mc:,}; infeasible {n_infeasible_mc:,})"
                )
                prem_arr = np.asarray(mc_res.single_premium, dtype=float)
                prem_arr = prem_arr[np.isfinite(prem_arr)]
                if prem_arr.size > 0:
                    hist_counts, hist_edges = np.histogram(prem_arr, bins=40)
                    hist_df = pd.DataFrame(
                        {
                            "premium_bin_mid": 0.5 * (hist_edges[:-1] + hist_edges[1:]),
                            "count": hist_counts,
                        }
                    ).set_index("premium_bin_mid")
                    st.line_chart(_round_for_visuals(hist_df))
            else:
                # Compact summary for new-product MC results: just show
                # mean PV(benefit) and (if available) E[AV(T)].
                pv_b_arr = np.asarray(getattr(mc_res, "pv_benefit", np.array([])), dtype=float)
                pv_b_arr = pv_b_arr[np.isfinite(pv_b_arr)]
                cols = st.columns(3)
                cols[0].metric("MC paths", f"{int(mc_res.n_sims):,}")
                if pv_b_arr.size > 0:
                    cols[1].metric("Mean PV(benefit)", f"${pv_b_arr.mean():,.0f}")
                if hasattr(mc_res, "av_end_mean"):
                    cols[2].metric("Mean AV(T)", f"${float(mc_res.av_end_mean):,.0f}")
                st.caption(
                    "Compact MC summary -- the new product engines expose pv_benefit + "
                    "per-path arrays only in v1. Future releases may add the full "
                    "premium-distribution surface used by SPIA / RILA."
                )

        df = _result_dataframe(res)
        st.subheader("Month-by-month projection")
        df_display = _round_for_visuals(df)
        st.dataframe(
            df_display,
            use_container_width=True,
            height=360,
            column_config=_number_cols_no_decimals(df_display),
        )
        csv_bytes = df_display.to_csv(index=False).encode("utf-8")
        c_dl1, c_dl2 = st.columns(2)
        with c_dl1:
            st.download_button(
                "Download projection CSV",
                data=csv_bytes,
                file_name=get_product_ui_config(product_type).projection_csv_filename,
                mime="text/csv",
            )
        with c_dl2:
            st.caption("Excel download moved to the Excel Replicator section.")
        ctx = st.session_state.get("pricing_excel_context") or {}
        expenses = ctx.get("expenses")
        _render_pricing_run_charts(res, contract_state, expenses, product_type=product_type)


def _read_workbook_cell_float(ws: Any, coord: str) -> float | None:
    v = ws[coord].value
    if v is None:
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(x):
        return None
    return x


def _alm_workbook_mirror_snapshot(
    alm: sp.ALMResult,
    asm: sp.ALMAssumptions | None,
    *,
    initial_asset_market_value: float | None,
) -> ALMExcelSnapshot | None:
    """Same truncation/downsample as ``build_workbook_from_spec`` (must match embedded ALM path)."""
    if asm is None:
        return None
    try:
        raw = alm_excel_snapshot_from_result(
            alm,
            asm,
            initial_asset_market_value=initial_asset_market_value,
        )
        ds = alm_excel_downsample_snapshot(raw, int(ALM_ENGINE_STEP_MONTHS))
        return alm_excel_truncate_snapshot(ds, int(ALM_EXCEL_PATH_MONTH_CAP))
    except Exception:
        return None


def _alm_modelcheck_key_assets_surplus_df(
    *,
    alm: sp.ALMResult,
    xlsx_bytes: bytes | None,
    dr: int = ALM_PROJECTION_FIRST_DATA_ROW,
    mirror_snap: ALMExcelSnapshot | None = None,
) -> pd.DataFrame:
    """
    ModelCheck-style table: Python path vs workbook (ALM_Projection).

    Uses the same truncated ALM snapshot as the Excel export when ``mirror_snap`` is provided so the
    Python column matches ModelCheck column B and the first ``ALM_EXCEL_PATH_MONTH_CAP`` rows on the sheet.

    Surplus from Excel is read as **C−D−E** when cached values exist (then **F** is not used). If the
    workbook has no cached results for those cells (typical before a full Excel recalc), the expected
    Excel column is **NaN** — do not treat that as a match. New downloads embed snapshot caches on
    **ALM_Projection** C–F and per-bucket columns **H+** so ``=SUM(bucket…)`` matches **C** under
    ``data_only`` until Excel recalculates (recalc may still refresh values from **ALM_Engine** formulas).
    """
    if mirror_snap is not None:
        a_mv = np.asarray(mirror_snap.asset_market_value, dtype=float)
        l_pv = np.asarray(mirror_snap.liability_pv, dtype=float)
        debt_b = np.asarray(mirror_snap.borrowing_balance, dtype=float)
        surp = a_mv - l_pv - debt_b
    else:
        a_mv = np.asarray(alm.asset_market_value, dtype=float)
        surp = np.asarray(alm.surplus, dtype=float)
    n = int(a_mv.size)
    if n < 1:
        return pd.DataFrame()

    n_rows = int(min(ALM_EXCEL_PATH_MONTH_CAP, n))
    if n_rows < 1:
        return pd.DataFrame()
    last_lab = f"ALM asset MV (month {n_rows} on sheet)"
    last_s_lab = f"ALM surplus (month {n_rows} on sheet)"
    if n_rows == 1:
        asset_specs: list[tuple[int, int, str]] = [(0, 0, "ALM asset MV (month 1 on sheet)")]
        surp_specs: list[tuple[int, int, str]] = [(0, 0, "ALM surplus (month 1 on sheet)")]
    else:
        asset_specs = [
            (0, 0, "ALM asset MV (month 1 on sheet)"),
            (n_rows - 1, n_rows - 1, last_lab),
        ]
        surp_specs = [
            (0, 0, "ALM surplus (month 1 on sheet)"),
            (n_rows - 1, n_rows - 1, last_s_lab),
        ]

    ws = None
    if isinstance(xlsx_bytes, bytes) and xlsx_bytes:
        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            if ALM_SHEET_NAME in wb.sheetnames:
                ws = wb[ALM_SHEET_NAME]
        except Exception:
            ws = None

    rows: list[dict[str, Any]] = []
    for excel_off, py_idx, lab in asset_specs:
        py = float(a_mv[py_idx])
        r = dr + excel_off
        ex: float | None = None
        if ws is not None:
            ex = _read_workbook_cell_float(ws, f"C{r}")
        if ex is None:
            # Do not substitute Python: empty formula caches would fake a match (openpyxl data_only).
            ex = float("nan") if ws is not None else py
        rows.append(
            {
                "Metric": lab,
                "Python snapshot": py,
                "Expected Excel value (after recalc)": ex,
                "Difference (Excel - Python)": float(ex - py) if np.isfinite(ex) else float("nan"),
            }
        )

    for excel_off, py_idx, lab in surp_specs:
        py = float(surp[py_idx])
        r = dr + excel_off
        ex: float | None = None
        if ws is not None:
            c = _read_workbook_cell_float(ws, f"C{r}")
            d = _read_workbook_cell_float(ws, f"D{r}")
            e_b = _read_workbook_cell_float(ws, f"E{r}")
            if c is not None and d is not None and e_b is not None:
                ex = float(c - d - e_b)
            else:
                ex = _read_workbook_cell_float(ws, f"F{r}")
        if ex is None:
            ex = float("nan") if ws is not None else py
        rows.append(
            {
                "Metric": lab,
                "Python snapshot": py,
                "Expected Excel value (after recalc)": ex,
                "Difference (Excel - Python)": float(ex - py) if np.isfinite(ex) else float("nan"),
            }
        )

    return pd.DataFrame(rows)


def _render_excel_replicator() -> None:
    st.header("Excel Replicator")
    st.caption(
        "Download the formula workbook and review parity metrics aligned with the workbook ModelCheck sheet."
    )

    res = st.session_state.get("pricing_res")
    contract_state = st.session_state.get("pricing_contract")
    if res is None or contract_state is None:
        st.info("Run pricing first in the Pricing Run section to populate the Excel Replicator.")
        return

    meta = st.session_state.get("pricing_meta") or {}
    product_raw = str(meta.get("product_type", ProductType.SPIA.value))
    product_type = (
        ProductType(product_raw)
        if product_raw in {p.value for p in ProductType}
        else ProductType.SPIA
    )

    _ensure_excel_workbook_includes_current_alm()

    m1, m2, m3, m4 = st.columns(4)
    metrics = get_pricing_metrics(product_type, res)
    for col, metric in zip((m1, m2, m3, m4), metrics):
        formatted = f"${metric.value:,.0f}" if metric.is_money else f"{metric.value:,.0f}"
        col.metric(f"Python {metric.label.lower()}", formatted)

    modelcheck = pd.DataFrame(
        [
            {
                "Metric": "PV benefits",
                "Python snapshot": float(res.pv_benefit),
                "Expected Excel value (after recalc)": float(res.pv_benefit),
                "Difference (Excel - Python)": 0.0,
            },
            {
                "Metric": "PV monthly expenses",
                "Python snapshot": float(res.pv_monthly_expenses),
                "Expected Excel value (after recalc)": float(res.pv_monthly_expenses),
                "Difference (Excel - Python)": 0.0,
            },
            {
                "Metric": "PV monthly total (ben+exp)",
                "Python snapshot": float(res.pv_benefit + res.pv_monthly_expenses),
                "Expected Excel value (after recalc)": float(
                    res.pv_benefit + res.pv_monthly_expenses
                ),
                "Difference (Excel - Python)": 0.0,
            },
            {
                "Metric": "Single premium",
                "Python snapshot": float(res.single_premium),
                "Expected Excel value (after recalc)": float(res.single_premium),
                "Difference (Excel - Python)": 0.0,
            },
            {
                "Metric": "Annuity factor",
                "Python snapshot": float(res.annuity_factor),
                "Expected Excel value (after recalc)": float(res.annuity_factor),
                "Difference (Excel - Python)": 0.0,
            },
        ]
    )

    st.subheader("ModelCheck parity dashboard")
    modelcheck_display = _round_for_visuals(modelcheck)
    st.dataframe(
        modelcheck_display,
        use_container_width=True,
        hide_index=True,
        column_config=_number_cols_no_decimals(modelcheck_display),
    )
    st.caption(
        f"Workbook references: PV benefits `{LIABILITY_SHEET_NAME}!X4`, PV monthly expenses `{LIABILITY_SHEET_NAME}!X5`, "
        f"PV monthly total `{LIABILITY_SHEET_NAME}!X7`, single premium `{LIABILITY_SHEET_NAME}!X8`, annuity factor `{LIABILITY_SHEET_NAME}!X6`."
    )
    st.caption(
        "After opening the workbook and recalculating, the ModelCheck tab differences should be near zero "
        "if Inputs match this run (especially spread B9 and valuation year)."
    )

    alm_chk = st.session_state.get("alm_last")
    alm_chk_rid = st.session_state.get("alm_last_pricing_run_id")
    pr_chk_rid = st.session_state.get("pricing_run_id")
    if isinstance(alm_chk, sp.ALMResult) and alm_chk_rid == pr_chk_rid:
        st.subheader("ModelCheck — ALM (assets & surplus)")
        xb_mc = st.session_state.get("pricing_xlsx_bytes")
        asm_chk = st.session_state.get("alm_last_assumptions")
        aum_chk = st.session_state.get("alm_last_initial_asset_market_value")
        aum_opt = float(aum_chk) if aum_chk is not None else None
        mirror = (
            _alm_workbook_mirror_snapshot(
                alm_chk,
                asm_chk if isinstance(asm_chk, sp.ALMAssumptions) else None,
                initial_asset_market_value=aum_opt,
            )
            if isinstance(asm_chk, sp.ALMAssumptions)
            else None
        )
        alm_mc_df = _alm_modelcheck_key_assets_surplus_df(
            alm=alm_chk,
            xlsx_bytes=xb_mc if isinstance(xb_mc, bytes) else None,
            mirror_snap=mirror,
        )
        if not alm_mc_df.empty:
            alm_mc_disp = _round_for_visuals(alm_mc_df)
            st.dataframe(
                alm_mc_disp,
                use_container_width=True,
                hide_index=True,
                column_config=_number_cols_no_decimals(alm_mc_disp),
            )
            n_mon = int(np.asarray(alm_chk.asset_market_value).size)
            n_on_sheet = int(min(ALM_EXCEL_PATH_MONTH_CAP, n_mon))
            lr = ALM_PROJECTION_FIRST_DATA_ROW + n_on_sheet - 1
            st.caption(
                f"Workbook **{ALM_SHEET_NAME}** / **{ALM_ENGINE_SHEET}** show the **first {n_on_sheet}** monthly ALM steps "
                f"(cap {ALM_EXCEL_PATH_MONTH_CAP}; Python may have more months). Rows **{ALM_PROJECTION_FIRST_DATA_ROW}**–**{lr}**. "
                f"**C** = SUM buckets; **D** from **{LIABILITY_SHEET_NAME}**; **F** = C−D−E. "
                "Parity uses cached **C−D−E** (embedded on export). After a full recalc in Excel, "
                "saved values may differ if formulas diverge from Python; re-download to reset caches."
            )

    # --- Monte Carlo distribution dashboard ---
    mc_res = st.session_state.get("pricing_mc")
    mc_params = st.session_state.get("pricing_mc_params") or {}
    # The full MC distribution dashboard expects SPIA / RILA-shaped MC
    # results. New product MC results omit the per-metric arrays this
    # block iterates over.
    has_full_mc_surface = mc_res is not None and all(
        hasattr(mc_res, name)
        for name in (
            "single_premium",
            "pv_benefit",
            "pv_monthly_expenses",
            "pv_monthly_total",
            "annuity_factor",
        )
    )
    if has_full_mc_surface:
        st.divider()
        st.subheader("Monte Carlo distribution statistics")
        n_sims_disp = mc_params.get("n_sims", mc_res.n_sims)
        drift_disp = mc_params.get("annual_drift", 0.0)
        vol_disp = mc_params.get("annual_vol", 0.0)
        s0_disp = mc_params.get("s0", 100.0)
        st.caption(
            f"{n_sims_disp:,} simulations | GBM drift {drift_disp * 100:.1f}% | "
            f"vol {vol_disp * 100:.1f}% | S\u2080 {s0_disp:.2f} | "
            "Mortality, yield curve, and expense inflation are deterministic across paths."
        )

        n_infeasible_mc2 = int(getattr(mc_res, "n_infeasible", 0))
        n_feasible_mc2 = int(getattr(mc_res, "n_feasible", mc_res.n_sims))
        if n_infeasible_mc2 > 0:
            pct2 = 100.0 * n_infeasible_mc2 / max(int(mc_res.n_sims), 1)
            st.info(
                f"Distribution stats below are over the {n_feasible_mc2:,} feasible paths "
                f"({n_infeasible_mc2:,} infeasible / {pct2:.1f}% omitted)."
            )

        _mc_metrics: list[tuple[str, np.ndarray]] = [
            ("Single Premium ($)", mc_res.single_premium),
            ("PV Benefit ($)", mc_res.pv_benefit),
            ("PV Monthly Expenses ($)", mc_res.pv_monthly_expenses),
            ("PV Monthly Total ($)", mc_res.pv_monthly_total),
            ("Annuity Factor", mc_res.annuity_factor),
        ]
        stat_rows = []
        for name, arr in _mc_metrics:
            a = np.asarray(arr, dtype=float)
            finite = a[np.isfinite(a)]
            if finite.size == 0:
                stat_rows.append(
                    {
                        "Metric": name,
                        "Mean": float("nan"),
                        "Std Dev": float("nan"),
                        "P5": float("nan"),
                        "P25": float("nan"),
                        "Median": float("nan"),
                        "P75": float("nan"),
                        "P95": float("nan"),
                    }
                )
                continue
            stat_rows.append(
                {
                    "Metric": name,
                    "Mean": float(np.mean(finite)),
                    "Std Dev": float(np.std(finite)),
                    "P5": float(np.percentile(finite, 5)),
                    "P25": float(np.percentile(finite, 25)),
                    "Median": float(np.median(finite)),
                    "P75": float(np.percentile(finite, 75)),
                    "P95": float(np.percentile(finite, 95)),
                }
            )
        stats_df = pd.DataFrame(stat_rows)
        stats_display = _round_for_visuals(stats_df)
        st.dataframe(
            stats_display,
            use_container_width=True,
            hide_index=True,
            column_config=_number_cols_no_decimals(stats_display),
        )

        st.markdown("**Premium & key metric distributions**")
        ch1, ch2 = st.columns(2)

        def _hist_df(arr: np.ndarray, n_bins: int = 35) -> pd.DataFrame | None:
            a = np.asarray(arr, dtype=float)
            a = a[np.isfinite(a)]
            if a.size == 0:
                return None
            counts, edges = np.histogram(a, bins=n_bins)
            mids = 0.5 * (edges[:-1] + edges[1:])
            return pd.DataFrame({"bin": np.rint(mids), "count": counts}).set_index("bin")

        def _maybe_chart(col, title: str, arr: np.ndarray) -> None:
            with col:
                st.markdown(title)
                df_h = _hist_df(arr)
                if df_h is None:
                    st.caption("No feasible paths to plot.")
                else:
                    st.bar_chart(df_h)

        _maybe_chart(ch1, "Single premium", mc_res.single_premium)
        _maybe_chart(ch2, "PV benefit", mc_res.pv_benefit)
        ch3, ch4 = st.columns(2)
        _maybe_chart(ch3, "Annuity factor", mc_res.annuity_factor)
        _maybe_chart(ch4, "PV monthly total", mc_res.pv_monthly_total)

        st.caption(
            "The MC_Summary sheet in the downloaded workbook contains the same statistics table "
            "and a premium distribution chart embedded as an Excel bar chart."
        )
    else:
        st.info(
            "Monte Carlo was not enabled for this run. Enable it in the Pricing Run section "
            "and re-run to see distribution statistics here and in the Excel workbook."
        )

    st.divider()
    xb = st.session_state.get("pricing_xlsx_bytes")
    xlsx_has_mc: bool = st.session_state.get("pricing_xlsx_has_mc", False)
    xlsx_has_alm: bool = st.session_state.get("pricing_xlsx_has_alm", False)
    if isinstance(xb, bytes) and xb:
        if xlsx_has_mc:
            st.success(
                "Workbook includes **MC_Summary** sheet with distribution statistics table and premium histogram chart.",
                icon="✅",
            )
        else:
            st.warning(
                "Workbook does **not** include MC_Summary — MC was disabled or not run when this workbook was built. "
                "Enable Monte Carlo in Pricing Run and click **Run pricing** again to regenerate.",
                icon="⚠️",
            )
        if xlsx_has_alm:
            st.success(
                f"Workbook includes **{LIABILITY_SHEET_NAME}**, **ALM_Projection** / **ALM_Engine** / **{ALM_ENGINE_FIELD_GUIDE_SHEET}** "
                f"(first {ALM_EXCEL_PATH_MONTH_CAP} months of the ALM path).",
                icon="✅",
            )
        else:
            st.info(
                "This workbook file does not yet include **ALM_Projection** — run ALM, then return here to refresh the download."
            )
        suffix_parts: list[str] = []
        if xlsx_has_mc:
            suffix_parts.append("MC_Summary")
        if xlsx_has_alm:
            suffix_parts.append("ALM")
        mc_label = (" + " + " + ".join(suffix_parts)) if suffix_parts else ""
        help_bits = ["ModelCheck parity vs Python pricing"]
        if xlsx_has_mc:
            help_bits.append("MC statistics chart")
        if xlsx_has_alm:
            help_bits.append("ALM path sheets")
        st.download_button(
            f"Download Excel recalculation workbook{mc_label}",
            data=xb,
            file_name=get_product_ui_config(product_type).recalc_workbook_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Workbook includes " + ", ".join(help_bits) + ".",
            type="primary",
        )
    elif st.session_state.get("pricing_xlsx_built_error"):
        st.error(f"Excel export unavailable: {st.session_state['pricing_xlsx_built_error']}")
    else:
        st.warning("Excel workbook not available yet for this run.")


def _render_alm_section() -> None:
    st.header("Asset–liability management (ALM)")

    res = st.session_state.get("pricing_res")
    contract_state = st.session_state.get("pricing_contract")
    ctx = st.session_state.get("pricing_excel_context") or {}
    yc = ctx.get("yield_curve")
    spr = float(ctx.get("spread", 0.0))
    meta_alm = st.session_state.get("pricing_meta") or {}
    pt_raw = str(meta_alm.get("product_type", ProductType.SPIA.value))
    try:
        alm_product_type = ProductType(pt_raw)
    except ValueError:
        alm_product_type = ProductType.SPIA
    if alm_product_type == ProductType.TERM_LIFE:
        liab_label = "Term"
    elif alm_product_type == ProductType.RILA:
        liab_label = "RILA"
    else:
        liab_label = "SPIA"
    st.caption(
        f"Dynamic Treasury ladder + cash versus priced {liab_label} outflows. Earned rate on assets and liability discounting use "
        "the same zero curve + credit spread as **Pricing Run** for consistency. "
        "Rebalancing uses a drift band versus target weights on the review months you choose."
    )

    if res is None or contract_state is None or not isinstance(yc, sp.YieldCurve):
        st.info("Run **Pricing Run** first. ALM anchors on that liability path, curve, and spread.")
        return

    # Allow the ALM engine to run against either the Pricing Run deterministic baseline,
    # or against a single Monte Carlo index path scenario (liability PV + discounting).
    scenario_source: str = "Base (Pricing Run deterministic)"
    mc_params = st.session_state.get("pricing_mc_params") or {}
    mc_n_sims = int(mc_params.get("n_sims", 0) or 0)
    mc_seed = int(mc_params.get("seed", 42) or 42)
    mc_scenario_idx: int = 0
    if mc_n_sims > 0:
        scenario_source = st.selectbox(
            "ALM pricing scenario (for liability PV and discounting)",
            options=["Base (Pricing Run deterministic)", "MC simulation (single path)"],
            index=0,
        )
        if scenario_source == "MC simulation (single path)":
            mc_scenario_idx = st.number_input(
                "MC simulation index (0-based)",
                min_value=0,
                max_value=max(0, mc_n_sims - 1),
                value=0,
                step=1,
            )
    else:
        st.caption(
            "MC scenario selection is unavailable because Pricing Run MC inputs are missing."
        )

    base_spec = sp.alm_default_allocation_spec()
    n_bk = len(base_spec.buckets)
    # Initialize allocation widget state once; keyed widgets then read from session state only.
    for i in range(n_bk):
        k = f"alm_alloc_{i}"
        if k not in st.session_state:
            st.session_state[k] = float(round(base_spec.weights[i] * 100.0, 2))
    # Apply optimized weights safely on next rerun (avoid mutating active widget keys mid-run).
    pending_alloc = st.session_state.pop("alm_alloc_pending", None)
    if isinstance(pending_alloc, (list, tuple, np.ndarray)) and len(pending_alloc) == n_bk:
        try:
            for i, wi in enumerate(np.asarray(pending_alloc, dtype=float)):
                st.session_state[f"alm_alloc_{i}"] = float(wi * 100.0)
        except Exception:
            pass
    opt_notice = st.session_state.pop("alm_opt_notice", None)
    if isinstance(opt_notice, dict):
        msg = str(opt_notice.get("message", ""))
        level = str(opt_notice.get("level", "info"))
        if level == "success":
            st.success(msg)
        elif level == "warning":
            st.warning(msg)
        else:
            st.info(msg)

    with st.expander("Target allocation (% weights, must sum to 100%)", expanded=True):
        cols = st.columns(min(n_bk, 6))
        raw: list[float] = []
        for i, b in enumerate(base_spec.buckets):
            with cols[i % len(cols)]:
                raw.append(
                    float(
                        st.number_input(
                            f"{b.name} %",
                            min_value=0.0,
                            max_value=100.0,
                            step=0.5,
                            key=f"alm_alloc_{i}",
                        )
                    )
                )
        norm_run = st.checkbox("Normalize percentages to 100% on run", value=True)
        ws = np.array(raw, dtype=float) / 100.0
        s = float(np.sum(ws))
        if s <= 0.0:
            st.error("Allocation must include positive weights.")
        elif abs(s - 1.0) > 1e-3 and not norm_run:
            st.warning(
                f"Weights currently sum to {s * 100:.2f}%. Enable normalization or adjust inputs."
            )
        elif abs(s - 1.0) > 1e-3 and norm_run:
            st.info(f"Weights sum to {s * 100:.2f}% — will scale to 100% on run.")

    with st.expander("Rebalancing, flows, liquidity definition", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            band_pct = st.slider("Drift band vs targets (± share of AUM)", 0.5, 20.0, 5.0, 0.5)
            freq_m = st.number_input("Check rebalance every N months", min_value=1, value=1, step=1)
            near_liq_y = st.number_input(
                "Near-liquid residual maturity (years)",
                min_value=0.0,
                max_value=3.0,
                value=0.25,
                step=0.05,
                help="Liquidity buffer counts cash plus bond market value in buckets with residual maturity here or below.",
            )
            borrow_policy = st.selectbox(
                "Borrowing policy",
                options=["borrow_before_selling", "borrow_after_assets_insufficient"],
                index=1,
                format_func=lambda x: {
                    "borrow_before_selling": "Always borrow before selling assets",
                    "borrow_after_assets_insufficient": "Borrow only when asset portfolio is insufficient",
                }[x],
            )
            borrow_rate_mode = st.selectbox(
                "Borrowing rate basis",
                options=["scenario_linked", "fixed"],
                index=0,
                format_func=lambda x: {
                    "scenario_linked": "Scenario-linked (selected tenor rate + spread)",
                    "fixed": "Fixed annual borrowing rate",
                }[x],
            )
            is_scenario_linked = borrow_rate_mode == "scenario_linked"
            borrow_rate_tenor = st.selectbox(
                "Scenario-linked borrowing tenor",
                options=[0.25, 0.5, 1.0, 2.0, 3.0, 5.0],
                index=2,
                format_func=lambda x: f"{x:g}Y",
                help="Curve tenor used to derive borrowing base rate in scenario-linked mode.",
                disabled=not is_scenario_linked,
            )
            # Logical default: 1Y curve+spread plus 100 bps floor at 3%.
            df_t = float(
                yc.discount_factors(np.array([float(borrow_rate_tenor)], dtype=float), spread=spr)[
                    0
                ]
            )
            base_t = -np.log(max(df_t, 1e-15)) / float(borrow_rate_tenor)
            borrow_rate_default_pct = float(max(0.03, base_t + 0.01) * 100.0)
            borrow_spread_bps = st.number_input(
                "Borrowing spread over selected tenor scenario rate (bps)",
                min_value=0.0,
                max_value=2000.0,
                value=100.0,
                step=5.0,
                help="Used when borrowing rate basis is scenario-linked.",
                disabled=not is_scenario_linked,
            )
            borrow_rate_pct = st.number_input(
                "Fixed borrowing rate (annual, %)",
                min_value=0.0,
                max_value=50.0,
                value=round(borrow_rate_default_pct, 2),
                step=0.1,
                help="Used only when borrowing rate basis is fixed.",
                disabled=is_scenario_linked,
            )
        with c2:
            rebalance_policy = st.selectbox(
                "Rebalance policy",
                options=["liquidity_only", "full_target"],
                index=0,
                format_func=lambda x: {
                    "liquidity_only": "Hold to maturity bias (sell only for liquidity shortfall)",
                    "full_target": "Full target rebalance (trade back to weights when drift breaches band)",
                }[x],
            )
            reinvest = st.selectbox(
                "Reinvest matured principal",
                options=["hold_cash", "pro_rata"],
                index=1,
                format_func=lambda x: {
                    "hold_cash": "Keep in cash until band rebalance (or next review)",
                    "pro_rata": "Re-deploy into bonds pro-rata to bond targets (excess over cash target)",
                }[x],
            )
            disinvest = st.selectbox(
                "Disinvest if cash is short after outflows",
                options=["shortest_first", "pro_rata"],
                index=0,
                format_func=lambda x: {
                    "shortest_first": "Shortest residual maturity first",
                    "pro_rata": "Pro-rata across bond buckets",
                }[x],
            )

    aum0 = st.number_input(
        "Initial asset market value ($)",
        min_value=0.0,
        value=float(res.single_premium),
        step=10_000.0,
        help="Usually the priced single premium invested at issue.",
    )

    # Persist the current ALM selection so What-if and diagnostics can reflect the user's latest inputs
    # even before they click "Run ALM projection".
    try:
        ws_run_current = ws.copy()
        if norm_run and float(np.sum(ws_run_current)) > 0.0:
            ws_run_current = ws_run_current / float(np.sum(ws_run_current))
        alloc_current = sp.ALMAllocationSpec(buckets=base_spec.buckets, weights=ws_run_current)
        asm_current = sp.ALMAssumptions(
            allocation=alloc_current,
            rebalance_band=float(band_pct) / 100.0,
            rebalance_frequency_months=int(freq_m),
            reinvest_rule=reinvest,  # type: ignore[arg-type]
            disinvest_rule=disinvest,  # type: ignore[arg-type]
            rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
            borrowing_policy=borrow_policy,  # type: ignore[arg-type]
            borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
            borrowing_rate_tenor_years=float(borrow_rate_tenor),
            borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
            borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
            liquidity_near_liquid_years=float(near_liq_y),
        )
        st.session_state["alm_current_assumptions"] = asm_current
        st.session_state["alm_current_initial_asset_market_value"] = float(aum0)
    except Exception:
        # Don't block UI if current weights/band inputs are temporarily invalid while user is editing.
        st.session_state.pop("alm_current_assumptions", None)
        st.session_state.pop("alm_current_initial_asset_market_value", None)

    run_alm = st.button("Run ALM projection", type="primary")
    opt_col1, opt_col2 = st.columns([2, 1])
    with opt_col1:
        opt_surplus_constraint = st.selectbox(
            "Optimization surplus constraint",
            options=["Path never negative", "Ending surplus non-negative"],
            index=1,
            help="Select whether optimization enforces surplus >= 0 at every month, or only at the ending month.",
        )
    with opt_col2:
        opt_samples = st.number_input(
            "Optimization samples",
            min_value=100,
            max_value=5000,
            value=1200,
            step=100,
            help="Requested random candidates; runtime controls may cap evaluated candidates for responsiveness.",
        )
    opt_objective = st.selectbox(
        "Optimization objective",
        options=[
            "Balanced mix (diversified weights)",
            "Match liability KRD by tenor (fast screen + ALM)",
        ],
        index=0,
        help=(
            "Balanced mix scores weights for diversification vs targets, then runs ALM on each candidate until caps. "
            "KRD match uses a larger random pool plus simplex coordinate refinement on the analytical hedge score, "
            "then runs ALM on more top candidates (higher time budget) for surplus feasibility."
        ),
    )
    run_alm_opt = st.button("Optimize allocation and run ALM")

    def _build_pricing_for_selected_scenario() -> Any:
        mort = ctx.get("mortality")
        expenses = ctx.get("expenses")
        valuation_year = ctx.get("valuation_year")
        horizon_age = ctx.get("horizon_age")
        if not isinstance(mort, (sp.MortalityTableQx, sp.MortalityTableRP2014MP2016)):
            raise ValueError("Pricing Run mortality missing from session state.")
        if not isinstance(expenses, sp.ExpenseAssumptions):
            raise ValueError("Pricing Run expenses missing from session state.")
        return build_alm_pricing_for_mc_scenario(
            product_type=alm_product_type,
            scenario_source=scenario_source,
            baseline_pricing=res,
            contract=contract_state,
            yield_curve=yc,
            mortality=mort,
            horizon_age=int(horizon_age),
            spread=spr,
            valuation_year=int(valuation_year) if valuation_year is not None else None,
            expenses=expenses,
            expense_annual_inflation=float(res.expense_annual_inflation),
            mc_n_sims=mc_n_sims,
            mc_seed=mc_seed,
            mc_scenario_idx=int(mc_scenario_idx),
            mc_params=mc_params,
        )

    if run_alm:
        if aum0 <= 0.0:
            st.error("Initial assets must be positive.")
        elif float(np.sum(ws)) <= 0.0:
            st.error("Invalid allocation.")
        else:
            ws_run = ws.copy()
            if norm_run and float(np.sum(ws_run)) > 0:
                ws_run = ws_run / float(np.sum(ws_run))
            try:
                alloc = sp.ALMAllocationSpec(buckets=base_spec.buckets, weights=ws_run)
                asm = sp.ALMAssumptions(
                    allocation=alloc,
                    rebalance_band=float(band_pct) / 100.0,
                    rebalance_frequency_months=int(freq_m),
                    reinvest_rule=reinvest,  # type: ignore[arg-type]
                    disinvest_rule=disinvest,  # type: ignore[arg-type]
                    rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
                    borrowing_policy=borrow_policy,  # type: ignore[arg-type]
                    borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
                    borrowing_rate_tenor_years=float(borrow_rate_tenor),
                    borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
                    borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
                    liquidity_near_liquid_years=float(near_liq_y),
                )
                pricing_for_alm = _build_pricing_for_selected_scenario()
                out = _run_alm_from_session_pricing(
                    pricing=pricing_for_alm,
                    yield_curve=yc,
                    spread=spr,
                    assumptions=asm,
                    initial_asset_market_value=float(aum0),
                )
                st.session_state["alm_last"] = out
                st.session_state["alm_last_assumptions"] = asm
                st.session_state["alm_last_initial_asset_market_value"] = float(aum0)
                st.session_state["alm_last_pricing_run_id"] = st.session_state.get("pricing_run_id")
                _invalidate_diagnostics_export()
                _refresh_pricing_excel_workbook_in_session()
                st.success("ALM projection complete.")
            except Exception as ex:
                st.error(f"ALM run failed: {ex!r}")

    if run_alm_opt:
        if aum0 <= 0.0:
            st.error("Initial assets must be positive.")
        else:
            try:
                pricing_for_alm = _build_pricing_for_selected_scenario()
                rng = np.random.default_rng(42)
                n_assets = len(base_spec.buckets)
                current_w = ws.copy()
                if float(np.sum(current_w)) > 0:
                    current_w = current_w / float(np.sum(current_w))
                candidates = [current_w, np.asarray(base_spec.weights, dtype=float)]
                # Add a conservative anchor.
                w_cons = np.zeros(n_assets, dtype=float)
                w_cons[0] = 0.20
                rem = 0.80 / max(1, n_assets - 1)
                w_cons[1:] = rem
                candidates.append(w_cons)
                # Structured tenor candidates around diversified bond ladders.
                tenors = np.array([float(b.tenor_years) for b in base_spec.buckets], dtype=float)
                bond_ten = np.clip(tenors[1:], 1e-9, None)
                for cash_w in [0.0, 0.05, 0.10, 0.20]:
                    for tilt in [-1.0, -0.5, 0.0, 0.5, 1.0]:
                        wb = bond_ten**tilt
                        wb = wb / float(np.sum(wb))
                        w_try = np.concatenate(([cash_w], (1.0 - cash_w) * wb))
                        candidates.append(w_try)
                # Explicitly include near-even spreads to accelerate convergence to diversified mixes.
                for cash_w in [0.0, 0.05, 0.10]:
                    wb_even = np.full(
                        n_assets - 1, (1.0 - cash_w) / float(max(1, n_assets - 1)), dtype=float
                    )
                    candidates.append(np.concatenate(([cash_w], wb_even)))

                # Runtime guardrails to avoid very long runs on large horizons.
                if opt_surplus_constraint == "Ending surplus non-negative":
                    # Default mode is easier to satisfy; use tighter limits for better responsiveness.
                    max_eval = min(220, max(50, int(opt_samples)))
                    time_budget_sec = 6.0
                else:
                    max_eval = min(400, max(80, int(opt_samples)))
                    time_budget_sec = 12.0
                opt_krd_match = str(opt_objective).startswith("Match")
                if opt_krd_match:
                    # Allow more full ALM checks and wall-clock time for KRD matching.
                    time_budget_sec = min(32.0, float(time_budget_sec) + 14.0)
                    max_eval = min(520, int(max_eval) + 140)

                max_random = max(0, max_eval - len(candidates))
                # Random simplex samples (bounded by max_eval).
                alpha = np.ones(n_assets, dtype=float)
                for _ in range(min(int(opt_samples), max_random)):
                    candidates.append(rng.dirichlet(alpha))
                if opt_krd_match:
                    extra_draws = min(950, max(320, int(opt_samples) * 2 + 150))
                    for _ in range(extra_draws):
                        candidates.append(rng.dirichlet(alpha))

                key_tenors_opt = np.array(
                    [
                        float(b.tenor_years)
                        for b in base_spec.buckets[1:]
                        if float(b.tenor_years) > 1e-12
                    ],
                    dtype=float,
                )
                bond_tenors_opt = key_tenors_opt.copy()

                tenor_axis = np.array(
                    [float(b.tenor_years) for b in base_spec.buckets], dtype=float
                )
                target_tenor = float(np.median(tenor_axis[1:])) if tenor_axis.size > 1 else 0.0
                best_score = float("inf")
                best_end_surplus = -float("inf")
                best_w: np.ndarray | None = None
                best_out: sp.ALMResult | None = None
                best_min_surplus = -float("inf")
                best_fallback_w: np.ndarray | None = None
                best_fallback_out: sp.ALMResult | None = None
                best_krd_mismatch = float("nan")

                start_t = time.perf_counter()
                eval_count = 0

                def _run_one_alm_candidate(
                    w_try: np.ndarray, *, objective_score: float | None
                ) -> None:
                    nonlocal best_score, best_end_surplus, best_w, best_out, eval_count
                    nonlocal best_min_surplus, best_fallback_w, best_fallback_out, best_krd_mismatch
                    w_norm = np.asarray(w_try, dtype=float)
                    s_wm = float(np.sum(w_norm))
                    if s_wm <= 1e-15:
                        return
                    w_norm = w_norm / s_wm
                    alloc_try = sp.ALMAllocationSpec(buckets=base_spec.buckets, weights=w_norm)
                    asm_try = sp.ALMAssumptions(
                        allocation=alloc_try,
                        rebalance_band=float(band_pct) / 100.0,
                        rebalance_frequency_months=int(freq_m),
                        reinvest_rule=reinvest,  # type: ignore[arg-type]
                        disinvest_rule=disinvest,  # type: ignore[arg-type]
                        rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
                        borrowing_policy=borrow_policy,  # type: ignore[arg-type]
                        borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
                        borrowing_rate_tenor_years=float(borrow_rate_tenor),
                        borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
                        borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
                        liquidity_near_liquid_years=float(near_liq_y),
                    )
                    out_try = _run_alm_from_session_pricing(
                        pricing=pricing_for_alm,
                        yield_curve=yc,
                        spread=spr,
                        assumptions=asm_try,
                        initial_asset_market_value=float(aum0),
                    )
                    eval_count += 1

                    min_surp = float(np.min(np.asarray(out_try.surplus, dtype=float)))
                    if min_surp > best_min_surplus:
                        best_min_surplus = min_surp
                        best_fallback_w = w_norm.copy()
                        best_fallback_out = out_try

                    if opt_surplus_constraint == "Path never negative":
                        feasible = bool(np.all(np.asarray(out_try.surplus, dtype=float) >= -1e-6))
                    else:
                        feasible = bool(float(out_try.surplus[-1]) >= -1e-6)
                    if not feasible:
                        return

                    end_surplus = float(out_try.surplus[-1])
                    if opt_krd_match:
                        sc = float(objective_score) if objective_score is not None else float("inf")
                        if sc < best_score - 1e-15 or (
                            abs(sc - best_score) <= 1e-15 and end_surplus > best_end_surplus
                        ):
                            best_score = sc
                            best_end_surplus = end_surplus
                            best_w = w_norm.copy()
                            best_out = out_try
                            best_krd_mismatch = sc
                        return

                    w_eval = w_norm
                    w_bond = w_eval[1:] if w_eval.size > 1 else np.asarray([], dtype=float)
                    if w_bond.size > 0:
                        bond_sum = float(np.sum(w_bond))
                        if bond_sum > 1e-12:
                            w_bond_norm = w_bond / bond_sum
                            even_penalty = float(np.std(w_bond_norm))
                        else:
                            even_penalty = 1.0
                    else:
                        even_penalty = 0.0
                    tenor_score = float(np.dot(w_eval, tenor_axis))
                    tenor_dev_penalty = abs(tenor_score - target_tenor) / max(1.0, target_tenor)
                    long_penalty = float(w_eval[-1]) if w_eval.size > 1 else 0.0
                    concentration_penalty = float(np.max(w_eval))
                    score = (
                        1.00 * even_penalty
                        + 0.40 * tenor_dev_penalty
                        + 0.35 * long_penalty
                        + 0.25 * concentration_penalty
                    )
                    if score < best_score - 1e-12 or (
                        abs(score - best_score) <= 1e-12 and end_surplus > best_end_surplus
                    ):
                        best_score = score
                        best_end_surplus = end_surplus
                        best_w = w_norm.copy()
                        best_out = out_try

                if opt_krd_match:
                    if key_tenors_opt.size == 0:
                        st.error("KRD matching requires bond buckets with positive tenor.")
                    else:
                        liab_krd_vec = sp.liability_key_rate_durations_years(
                            yc,
                            float(spr),
                            np.asarray(pricing_for_alm.expected_total_cashflows, dtype=float),
                            np.asarray(pricing_for_alm.times_years, dtype=float),
                            key_tenors_opt,
                        )

                        def _analytical_krd_mismatch(wv: np.ndarray) -> float:
                            wv = np.maximum(np.asarray(wv, dtype=float), 0.0)
                            s = float(np.sum(wv))
                            if s <= 1e-15:
                                return float("inf")
                            wv = wv / s
                            ak = sp.initial_ladder_asset_key_rate_durations_years(
                                yc,
                                float(spr),
                                float(aum0),
                                wv,
                                bond_tenors_opt,
                                key_tenors_opt,
                            )
                            return float(
                                sp.key_rate_duration_hedge_mismatch_score(ak, liab_krd_vec)
                            )

                        stage1_n = min(len(candidates), max(520, int(opt_samples) * 2 + 120))
                        krd_scored: list[tuple[float, np.ndarray]] = []
                        for w_try in candidates[:stage1_n]:
                            w_arr = np.asarray(w_try, dtype=float)
                            if float(np.sum(w_arr)) <= 1e-15:
                                continue
                            w_arr = w_arr / float(np.sum(w_arr))
                            sc = _analytical_krd_mismatch(w_arr)
                            krd_scored.append((sc, w_arr))
                        krd_scored.sort(key=lambda t: t[0])

                        refined_seen: set[tuple[float, ...]] = set()
                        seeds_to_refine: list[np.ndarray] = []
                        for _sc, wv in krd_scored[: min(18, len(krd_scored))]:
                            key = tuple(np.round(wv, 5).tolist())
                            if key in refined_seen:
                                continue
                            refined_seen.add(key)
                            seeds_to_refine.append(wv)
                            if len(seeds_to_refine) >= 10:
                                break
                        for w_seed in seeds_to_refine:
                            w_ref, sc_ref = sp.refine_weights_on_probability_simplex(
                                w_seed,
                                _analytical_krd_mismatch,
                                max_rounds=32,
                                transfer_fracs=(0.08, 0.05, 0.03, 0.02, 0.01, 0.006),
                            )
                            krd_scored.append((sc_ref, w_ref))

                        for rank in range(min(4, len(krd_scored))):
                            w_anchor = krd_scored[rank][1]
                            conc = 35.0 + 12.0 * float(rank)
                            for _ in range(72):
                                alpha_loc = (
                                    np.maximum(np.asarray(w_anchor, dtype=float), 1e-4) * conc
                                    + 0.06
                                )
                                w_loc = rng.dirichlet(alpha_loc)
                                krd_scored.append((_analytical_krd_mismatch(w_loc), w_loc))

                        krd_scored.sort(key=lambda t: t[0])
                        top_m = min(58, max(38, max_eval // 2 + 8), len(krd_scored))
                        alm_picked: list[tuple[float, np.ndarray]] = []
                        seen_alm_weights: set[tuple[float, ...]] = set()
                        for sc_i, w_i in krd_scored:
                            keyw = tuple(np.round(np.asarray(w_i, dtype=float), 5).tolist())
                            if keyw in seen_alm_weights:
                                continue
                            seen_alm_weights.add(keyw)
                            alm_picked.append((sc_i, w_i))
                            if len(alm_picked) >= top_m:
                                break
                        for sc_i, w_i in alm_picked:
                            if (time.perf_counter() - start_t) >= time_budget_sec:
                                break
                            _run_one_alm_candidate(w_i, objective_score=sc_i)
                else:
                    for w_try in candidates:
                        if eval_count >= max_eval:
                            break
                        if (time.perf_counter() - start_t) >= time_budget_sec:
                            break
                        _run_one_alm_candidate(np.asarray(w_try, dtype=float), objective_score=None)

                if best_w is None or best_out is None:
                    if best_fallback_w is None or best_fallback_out is None:
                        st.warning(
                            "No feasible allocation found under the selected surplus constraint. "
                            "This can happen when constraints are too strict for current assumptions "
                            "(cashflows, borrowing policy/rate, rebalance policy, and curve)."
                        )
                    else:
                        st.session_state["alm_alloc_pending"] = np.asarray(
                            best_fallback_w, dtype=float
                        ).tolist()
                        asm_best = sp.ALMAssumptions(
                            allocation=sp.ALMAllocationSpec(
                                buckets=base_spec.buckets, weights=best_fallback_w
                            ),
                            rebalance_band=float(band_pct) / 100.0,
                            rebalance_frequency_months=int(freq_m),
                            reinvest_rule=reinvest,  # type: ignore[arg-type]
                            disinvest_rule=disinvest,  # type: ignore[arg-type]
                            rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
                            borrowing_policy=borrow_policy,  # type: ignore[arg-type]
                            borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
                            borrowing_rate_tenor_years=float(borrow_rate_tenor),
                            borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
                            borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
                            liquidity_near_liquid_years=float(near_liq_y),
                        )
                        st.session_state["alm_last"] = best_fallback_out
                        st.session_state["alm_last_assumptions"] = asm_best
                        st.session_state["alm_last_initial_asset_market_value"] = float(aum0)
                        st.session_state["alm_last_pricing_run_id"] = st.session_state.get(
                            "pricing_run_id"
                        )
                        _invalidate_diagnostics_export()
                        _refresh_pricing_excel_workbook_in_session()
                        st.session_state["alm_opt_notice"] = {
                            "level": "warning",
                            "message": (
                                "No feasible allocation found within runtime limits; showing nearest candidate "
                                "(highest minimum surplus). Target allocation inputs updated."
                            ),
                        }
                        st.rerun()
                else:
                    st.session_state["alm_alloc_pending"] = np.asarray(best_w, dtype=float).tolist()
                    asm_best = sp.ALMAssumptions(
                        allocation=sp.ALMAllocationSpec(buckets=base_spec.buckets, weights=best_w),
                        rebalance_band=float(band_pct) / 100.0,
                        rebalance_frequency_months=int(freq_m),
                        reinvest_rule=reinvest,  # type: ignore[arg-type]
                        disinvest_rule=disinvest,  # type: ignore[arg-type]
                        rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
                        borrowing_policy=borrow_policy,  # type: ignore[arg-type]
                        borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
                        borrowing_rate_tenor_years=float(borrow_rate_tenor),
                        borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
                        borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
                        liquidity_near_liquid_years=float(near_liq_y),
                    )
                    st.session_state["alm_last"] = best_out
                    st.session_state["alm_last_assumptions"] = asm_best
                    st.session_state["alm_last_initial_asset_market_value"] = float(aum0)
                    st.session_state["alm_last_pricing_run_id"] = st.session_state.get(
                        "pricing_run_id"
                    )
                    _invalidate_diagnostics_export()
                    _refresh_pricing_excel_workbook_in_session()
                    if opt_krd_match:
                        krd_msg = (
                            "Optimized allocation found (KRD screen: match asset key-rate sensitivities to liability "
                            f"by tenor; mean sq. rel. error {best_krd_mismatch:.4f}) and ALM projection completed. "
                            f"Weighted tenor: {float(np.dot(np.asarray(best_w, dtype=float), tenor_axis)):.2f}Y; "
                            f"ending surplus: ${float(best_out.surplus[-1]):,.0f}. "
                            "Target allocation inputs updated."
                        )
                    else:
                        krd_msg = (
                            "Optimized allocation found (balanced tenor spread with anti-concentration bias) "
                            "and ALM projection completed. "
                            f"Weighted tenor: {float(np.dot(np.asarray(best_w, dtype=float), tenor_axis)):.2f}Y; "
                            f"ending surplus: ${float(best_out.surplus[-1]):,.0f}. "
                            "Target allocation inputs updated."
                        )
                    st.session_state["alm_opt_notice"] = {"level": "success", "message": krd_msg}
                    st.rerun()
                st.caption(
                    f"Optimization evaluated {eval_count} ALM projection(s) "
                    f"(cap {max_eval}, time budget {time_budget_sec:.0f}s)."
                    + (
                        " KRD mode ranks weights analytically first, then ALM-checks only the best few."
                        if opt_krd_match
                        else ""
                    )
                )
            except Exception as ex:
                st.error(f"ALM optimization failed: {ex!r}")

    last = st.session_state.get("alm_last")
    if isinstance(last, sp.ALMResult):
        st.subheader("ALM metrics (first month-end)")
        st.caption(
            "Path metrics are recorded after each month’s flows and trades. Scalar PV01 and durations are issue-time (initial portfolio)."
        )
        m1, m2, m3, m4, m5 = st.columns(5)
        with m1:
            fr0 = float(last.funding_ratio[0]) if last.funding_ratio.size else float("nan")
            st.metric("Funding ratio (month 1)", f"{fr0:.3f}")
        with m2:
            st.metric("Surplus ($)", f"${float(last.surplus[0]):,.0f}")
        with m3:
            st.metric("PV01 net ($/bp)", f"{float(last.pv01_net):,.0f}")
        with m4:
            st.metric("Duration gap (y)", f"{float(last.duration_gap):.2f}")
        with m5:
            lb0 = (
                float(last.liquidity_buffer_months[0])
                if last.liquidity_buffer_months.size
                else float("nan")
            )
            st.metric("Liquidity buffer (mo)", f"{lb0:.2f}")

        st.subheader("Paths (attained age)")
        age_ax = contract_state.issue_age + last.times_years
        st.markdown("##### Asset market value and liability present value")
        st.line_chart(
            _round_for_visuals(
                pd.DataFrame(
                    {
                        "Asset market value": last.asset_market_value,
                        "Liability PV": last.liability_pv,
                    },
                    index=age_ax,
                )
            )
        )
        _alm_surplus_chart(age_ax, last.surplus)
        st.markdown("##### Liquidity buffer (months of mean monthly outflows)")
        st.line_chart(
            pd.DataFrame({"Liquidity buffer (months)": last.liquidity_buffer_months}, index=age_ax)
        )
        st.markdown("##### Borrowing balance")
        st.line_chart(
            _round_for_visuals(
                pd.DataFrame({"Borrowing balance": last.borrowing_balance}, index=age_ax)
            )
        )

        asm_vis = st.session_state.get("alm_last_assumptions")
        if isinstance(asm_vis, sp.ALMAssumptions):
            bucket_specs = list(asm_vis.allocation.buckets)
        else:
            bucket_specs = list(base_spec.buckets)
        # Keep all ALM legends/series in logical tenor order: cash, then shortest to longest tenor.
        order_idx = sorted(
            range(len(bucket_specs)), key=lambda i: float(bucket_specs[i].tenor_years)
        )
        ordered_specs = [bucket_specs[i] for i in order_idx]
        ordered_names = [b.name for b in ordered_specs]

        bucket_df_raw = pd.DataFrame(
            last.bucket_asset_mv.T, columns=[b.name for b in bucket_specs], index=age_ax
        )
        bucket_df = bucket_df_raw.reindex(columns=ordered_names)
        st.markdown("**Bucket market values**")
        bucket_mv_long = (
            bucket_df.reset_index()
            .rename(columns={"index": "Attained age"})
            .melt(id_vars=["Attained age"], var_name="Asset type", value_name="Bucket market value")
        )
        bucket_mv_long["Asset type"] = pd.Categorical(
            bucket_mv_long["Asset type"], categories=ordered_names, ordered=True
        )
        bucket_mv_chart = (
            alt.Chart(bucket_mv_long)
            .mark_line()
            .encode(
                x=alt.X("Attained age:Q", title="Attained age"),
                y=alt.Y("Bucket market value:Q", title="Market value ($)"),
                color=alt.Color(
                    "Asset type:N",
                    title="Asset type",
                    sort=ordered_names,
                    legend=alt.Legend(
                        orient="top", direction="horizontal", columns=len(ordered_names)
                    ),
                ),
                order=alt.Order("Asset type:N", sort="ascending"),
                tooltip=[
                    alt.Tooltip("Attained age:Q", format=".2f"),
                    alt.Tooltip("Asset type:N"),
                    alt.Tooltip("Bucket market value:Q", format=",.0f"),
                ],
            )
            .properties(height=320)
        )
        st.altair_chart(bucket_mv_chart.interactive(), use_container_width=True)

        st.markdown("##### Portfolio composition by asset type (%)")
        aum_series = pd.Series(last.asset_market_value, index=age_ax, dtype=float).replace(
            0.0, np.nan
        )
        weight_pct_df = bucket_df.div(aum_series, axis=0).fillna(0.0) * 100.0
        comp_df = (
            weight_pct_df.reset_index()
            .rename(columns={"index": "Attained age"})
            .melt(id_vars=["Attained age"], var_name="Asset type", value_name="Portfolio share (%)")
        )
        comp_df["Asset type"] = pd.Categorical(
            comp_df["Asset type"], categories=ordered_names, ordered=True
        )
        comp_chart = (
            alt.Chart(comp_df)
            .mark_area()
            .encode(
                x=alt.X("Attained age:Q", title="Attained age"),
                y=alt.Y(
                    "Portfolio share (%):Q",
                    stack=True,
                    title="Portfolio share (%)",
                    scale=alt.Scale(domain=[0, 100]),
                ),
                color=alt.Color(
                    "Asset type:N",
                    title="Asset type",
                    sort=ordered_names,
                    legend=alt.Legend(
                        orient="top", direction="horizontal", columns=len(ordered_names)
                    ),
                ),
                order=alt.Order("Asset type:N", sort="ascending"),
            )
            .properties(height=320)
        )
        st.altair_chart(comp_chart.interactive(), use_container_width=True)

        # Yield decomposition: portfolio weighted yield plus per-asset-class contributions.
        # Contributions are weight * bucket annualized zero yield (incl. spread), shown in percentage points.
        tenors = np.array([float(b.tenor_years) for b in ordered_specs], dtype=float)
        bucket_yield = np.zeros_like(tenors, dtype=float)
        for i, T in enumerate(tenors):
            if T <= 1e-12:
                bucket_yield[i] = 0.0
            else:
                dff = float(yc.discount_factors(np.array([T], dtype=float), spread=spr)[0])
                bucket_yield[i] = -np.log(max(dff, 1e-15)) / T

        weight_df = bucket_df.div(aum_series, axis=0).fillna(0.0)
        contrib_pp_df = weight_df.mul(bucket_yield, axis=1) * 100.0
        total_yield_pct = contrib_pp_df.sum(axis=1).rename("Total portfolio yield (%)")
        total_line_df = total_yield_pct.reset_index().rename(columns={"index": "Attained age"})
        st.markdown("##### Total portfolio yield")
        yld_total = (
            alt.Chart(total_line_df)
            .mark_line(color="#1f77b4", strokeWidth=3)
            .encode(
                x=alt.X("Attained age:Q", title="Attained age"),
                y=alt.Y("Total portfolio yield (%):Q", title="Total yield (%)"),
                tooltip=[
                    alt.Tooltip("Attained age:Q", format=".2f"),
                    alt.Tooltip("Total portfolio yield (%):Q", format=".4f"),
                ],
            )
            .properties(height=320)
        )
        st.altair_chart(yld_total.interactive(), use_container_width=True)

        kpi_tbl = pd.DataFrame(
            {
                "PV01 assets": [last.pv01_assets],
                "PV01 liabilities": [last.pv01_liabilities],
                "Mac duration assets": [last.duration_assets_mac],
                "Mac duration liabilities": [last.duration_liabilities_mac],
            }
        )
        kpi_tbl_show = kpi_tbl.copy()
        kpi_tbl_show["PV01 assets"] = kpi_tbl_show["PV01 assets"].map(lambda x: f"{x:,.0f}")
        kpi_tbl_show["PV01 liabilities"] = kpi_tbl_show["PV01 liabilities"].map(
            lambda x: f"{x:,.0f}"
        )
        kpi_tbl_show["Mac duration assets"] = kpi_tbl_show["Mac duration assets"].round(4)
        kpi_tbl_show["Mac duration liabilities"] = kpi_tbl_show["Mac duration liabilities"].round(4)
        st.dataframe(kpi_tbl_show, use_container_width=True, hide_index=True)

        st.markdown("##### Key rate duration by tenor (1 bp localized bump)")
        try:
            pricing_for_krd = _build_pricing_for_selected_scenario()
            key_tenors = np.array(
                [float(b.tenor_years) for b in ordered_specs if float(b.tenor_years) > 1e-12],
                dtype=float,
            )
            if key_tenors.size > 0:
                asm_krd = (
                    asm_vis
                    if isinstance(asm_vis, sp.ALMAssumptions)
                    else sp.ALMAssumptions(
                        allocation=sp.ALMAllocationSpec(buckets=base_spec.buckets, weights=ws),
                        rebalance_band=float(band_pct) / 100.0,
                        rebalance_frequency_months=int(freq_m),
                        reinvest_rule=reinvest,  # type: ignore[arg-type]
                        disinvest_rule=disinvest,  # type: ignore[arg-type]
                        rebalance_policy=rebalance_policy,  # type: ignore[arg-type]
                        borrowing_policy=borrow_policy,  # type: ignore[arg-type]
                        borrowing_rate_mode=borrow_rate_mode,  # type: ignore[arg-type]
                        borrowing_rate_tenor_years=float(borrow_rate_tenor),
                        borrowing_spread_annual=float(borrow_spread_bps) / 10000.0,
                        borrowing_rate_annual=float(borrow_rate_pct) / 100.0,
                        liquidity_near_liquid_years=float(near_liq_y),
                    )
                )
                a0 = float(aum0)
                base_cf = np.asarray(pricing_for_krd.expected_total_cashflows, dtype=float)
                l0 = float(
                    np.sum(base_cf * yc.discount_factors(pricing_for_krd.times_years, spread=spr))
                )
                net0 = max(1e-9, a0 - l0)
                w_krd = np.asarray(asm_krd.allocation.weights, dtype=float)
                bond_tenors = np.array(
                    [float(b.tenor_years) for b in asm_krd.allocation.buckets[1:]], dtype=float
                )
                df0_bonds = yc.discount_factors(bond_tenors, spread=spr)
                target_mv_bonds = w_krd[1:] * a0
                bond_faces = np.where(df0_bonds > 1e-15, target_mv_bonds / df0_bonds, 0.0)
                rows: list[dict[str, float | str]] = []
                for kt in key_tenors:
                    yc_b = _key_rate_bump_curve(
                        yc, key_tenor_years=float(kt), key_tenors_years=key_tenors, bump_bps=1.0
                    )
                    dfb_bonds = yc_b.discount_factors(bond_tenors, spread=spr)
                    a_b = float(w_krd[0] * a0 + np.sum(bond_faces * dfb_bonds))
                    l_b = float(
                        np.sum(
                            base_cf * yc_b.discount_factors(pricing_for_krd.times_years, spread=spr)
                        )
                    )
                    rows.append(
                        {
                            "Tenor": f"{kt:g}Y",
                            "Tenor years": float(kt),
                            "Assets KRD": -((a_b - a0) / (max(1e-9, a0) * 1e-4)),
                            "Liabilities KRD": -((l_b - l0) / (max(1e-9, l0) * 1e-4)),
                            "Surplus KRD": -(((a_b - l_b) - (a0 - l0)) / (net0 * 1e-4)),
                        }
                    )
                krd_df = pd.DataFrame(rows).sort_values("Tenor years")
                krd_long = krd_df.melt(
                    id_vars=["Tenor", "Tenor years"],
                    value_vars=["Assets KRD", "Liabilities KRD", "Surplus KRD"],
                    var_name="Series",
                    value_name="Key rate duration",
                )
                krd_bars_df = krd_long[
                    krd_long["Series"].isin(["Assets KRD", "Liabilities KRD"])
                ].copy()
                krd_surplus_df = krd_long[krd_long["Series"] == "Surplus KRD"].copy()
                tenor_order = krd_df["Tenor"].tolist()

                bars = (
                    alt.Chart(krd_bars_df)
                    .mark_bar()
                    .encode(
                        x=alt.X("Tenor:N", sort=tenor_order, title="Key tenor"),
                        y=alt.Y("Key rate duration:Q", title="Assets/Liabilities KRD (years)"),
                        color=alt.Color(
                            "Series:N",
                            title="Series",
                            sort=["Assets KRD", "Liabilities KRD"],
                            legend=alt.Legend(orient="top", direction="horizontal"),
                        ),
                        xOffset=alt.XOffset("Series:N"),
                        tooltip=[
                            alt.Tooltip("Tenor:N"),
                            alt.Tooltip("Series:N"),
                            alt.Tooltip("Key rate duration:Q", format=".4f"),
                        ],
                    )
                )

                surplus_line = (
                    alt.Chart(krd_surplus_df)
                    .mark_line(color="#d62728", strokeWidth=3, point=True)
                    .encode(
                        x=alt.X("Tenor:N", sort=tenor_order, title="Key tenor"),
                        y=alt.Y("Key rate duration:Q", title="Surplus KRD (years)"),
                        tooltip=[
                            alt.Tooltip("Tenor:N"),
                            alt.Tooltip("Series:N"),
                            alt.Tooltip("Key rate duration:Q", format=".4f"),
                        ],
                    )
                )

                st.altair_chart(
                    alt.layer(bars, surplus_line)
                    .resolve_scale(y="independent")
                    .properties(height=320),
                    use_container_width=True,
                )
                st.caption(
                    "Interpretation: Surplus KRD is the key-rate sensitivity of net surplus (assets minus liabilities), "
                    "normalized by current surplus. Because the denominator is surplus rather than total assets or liabilities, "
                    "Surplus KRD can be much larger in magnitude when surplus is small."
                )
            else:
                st.info("No positive tenors available for key rate duration chart.")
        except Exception as ex:
            st.info(f"Key rate duration chart unavailable for current inputs: {ex!r}")


def _sidebar_section_options() -> list[str]:
    if portfolio_sidebar_visible(st.session_state):
        return [*SECTION_ORDER[:2], "portfolio", *SECTION_ORDER[2:]]
    return list(SECTION_ORDER)


def _portfolio_row_prefix(row_id: str) -> str:
    return f"portfolio_row_{row_id}_"


def _portfolio_reference_product_for_economics(row_ids: list[str]) -> ProductType:
    if row_ids:
        rid = row_ids[0]
        pfx = _portfolio_row_prefix(rid)
        raw = st.session_state.get(pfx + "product_type", ProductType.SPIA.value)
        try:
            return ProductType(str(raw).strip())
        except ValueError:
            return ProductType.SPIA
    try:
        return ProductType(str(st.session_state.get("run_product_type", ProductType.SPIA.value)))
    except ValueError:
        return ProductType.SPIA


def _portfolio_run_scenario_for_policies(policies: tuple[PolicyInput, ...]) -> RunScenario:
    sex_raw = str(getattr(policies[0].contract, "sex", "male")).strip().lower()
    sex: Literal["male", "female"] = "female" if sex_raw == "female" else "male"
    return run_scenario_for_portfolio_policies(dict(st.session_state), policies, sex=sex, repo_root=ROOT)


def _next_portfolio_policy_id(used: list[str]) -> str:
    nums: list[int] = []
    for pid in used:
        ps = str(pid).strip()
        if len(ps) >= 2 and ps[0].upper() == "P" and ps[1:].isdigit():
            nums.append(int(ps[1:]))
    n = max(nums, default=0) + 1
    return f"P{n:04d}"


def _portfolio_collect_used_policy_ids(row_ids: list[str], *, skip_row_id: str | None = None) -> list[str]:
    out: list[str] = []
    for rid in row_ids:
        if skip_row_id is not None and rid == skip_row_id:
            continue
        k = _portfolio_row_prefix(rid) + "policy_id"
        if k in st.session_state:
            s = str(st.session_state[k]).strip()
            if s:
                out.append(s)
    return out


def _portfolio_policy_id_options(row_ids: list[str], current_row_id: str) -> list[str]:
    used = _portfolio_collect_used_policy_ids(row_ids, skip_row_id=current_row_id)
    nxt = _next_portfolio_policy_id(used)
    pool = ["P0001", "P0002", "P0003", "P0004", nxt]
    ordered: list[str] = []
    for p in pool:
        if p not in ordered:
            ordered.append(p)
    return ordered


def _portfolio_wipe_row_keys(row_id: str, *, except_keys: frozenset[str] | None = None) -> None:
    pfx = _portfolio_row_prefix(row_id)
    skip = except_keys or frozenset()
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and k.startswith(pfx) and k not in skip:
            del st.session_state[k]


def _portfolio_push_defaults_to_session(
    row_id: str,
    d: dict[str, Any],
    *,
    skip_session_keys: frozenset[str] | None = None,
) -> None:
    """Write non-None defaults to session; None removes the key so widgets use fresh seeds.

    Keys listed in *skip_session_keys* are not written or popped. Use this when a
    widget for that key was already instantiated in the current run (Streamlit
    forbids assigning ``st.session_state[widget_key]`` after the widget exists).
    """
    pfx = _portfolio_row_prefix(row_id)
    skip = skip_session_keys or frozenset()
    for col in PORTFOLIO_INFORCE_SCRATCH_COLUMNS:
        v = d.get(col)
        key = pfx + col
        if key in skip:
            continue
        if v is None:
            st.session_state.pop(key, None)
        else:
            st.session_state[key] = v


def _portfolio_add_manual_row() -> None:
    rows: list[str] = list(st.session_state.setdefault(PORTFOLIO_KEY.MANUAL_ROWS, []))
    rid = str(uuid.uuid4())
    used = _portfolio_collect_used_policy_ids(rows)
    pid = _next_portfolio_policy_id(used)
    d = default_inforce_scratch_row(ProductType.SPIA)
    d["policy_id"] = pid
    rows.append(rid)
    st.session_state[PORTFOLIO_KEY.MANUAL_ROWS] = rows
    _portfolio_push_defaults_to_session(rid, d)


def _portfolio_remove_last_manual_row() -> None:
    rows: list[str] = list(st.session_state.get(PORTFOLIO_KEY.MANUAL_ROWS) or [])
    if not rows:
        return
    rid = rows.pop()
    _portfolio_wipe_row_keys(rid)
    st.session_state.pop(f"portfolio_meta_{rid}_product_type_prev", None)
    st.session_state[PORTFOLIO_KEY.MANUAL_ROWS] = rows


def _portfolio_row_as_dict_for_dataframe(row_id: str) -> dict[str, Any]:
    pfx = _portfolio_row_prefix(row_id)
    raw_pt = st.session_state.get(pfx + "product_type")
    if not raw_pt:
        raw_pt = ProductType.SPIA.value
    pt = ProductType(str(raw_pt).strip())
    out: dict[str, Any] = dict(default_inforce_scratch_row(pt))
    for col in PORTFOLIO_INFORCE_SCRATCH_COLUMNS:
        key = pfx + col
        if key not in st.session_state:
            continue
        v = st.session_state[key]
        if col == "policy_id":
            out[col] = str(v).strip() if v is not None else ""
            continue
        if col == "sex":
            s = str(v).strip().lower() if v is not None else "male"
            out[col] = s if s in ("male", "female") else "male"
            continue
        if col == "product_type":
            out[col] = str(v).strip()
            continue
        if v is None or v == "":
            out[col] = None
            continue
        if col in ("issue_age", "term_years", "guarantee_years", "horizon_years"):
            try:
                out[col] = int(float(v))
            except (TypeError, ValueError):
                out[col] = None
            continue
        if col == "gmdb_basis":
            out[col] = str(v).strip()
            continue
        try:
            out[col] = float(v)
        except (TypeError, ValueError):
            out[col] = None
    return out


def _portfolio_manual_rows_dataframe() -> pd.DataFrame:
    row_ids: list[str] = list(st.session_state.get(PORTFOLIO_KEY.MANUAL_ROWS) or [])
    recs = [_portfolio_row_as_dict_for_dataframe(rid) for rid in row_ids]
    return pd.DataFrame.from_records(recs, columns=list(PORTFOLIO_INFORCE_SCRATCH_COLUMNS))


def _render_portfolio_contract_fields(row_id: str, pt: ProductType) -> None:
    pfx = _portfolio_row_prefix(row_id)
    if pt == ProductType.SPIA:
        st.number_input(
            "Annual benefit ($)",
            min_value=1.0,
            step=1000.0,
            key=pfx + "benefit_annual",
        )
    elif pt == ProductType.TERM_LIFE:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.number_input(
                "Death benefit ($)",
                min_value=1.0,
                step=10_000.0,
                key=pfx + "death_benefit",
            )
        with c2:
            st.number_input(
                "Monthly premium ($)",
                min_value=0.0,
                step=10.0,
                key=pfx + "monthly_premium",
            )
        with c3:
            st.number_input("Term (years)", min_value=1, max_value=50, step=1, key=pfx + "term_years")
    elif pt == ProductType.RILA:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.number_input("Participation", min_value=0.0, max_value=5.0, format="%.4f", key=pfx + "participation")
        with c2:
            st.number_input("Cap (annual decimal)", min_value=-1.0, max_value=2.0, format="%.4f", key=pfx + "cap")
        with c3:
            st.number_input("Floor (annual decimal)", min_value=-1.0, max_value=1.0, format="%.4f", key=pfx + "floor")
        with c4:
            st.number_input(
                "Rider fee (annual on AV)",
                min_value=0.0,
                max_value=1.0,
                format="%.4f",
                key=pfx + "rider_fee_annual",
            )
    elif pt == ProductType.MYGA:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
        with c2:
            st.number_input(
                "Declared rate (annual decimal)",
                min_value=-0.5,
                max_value=1.0,
                format="%.4f",
                key=pfx + "declared_rate_annual",
            )
        with c3:
            st.number_input("Guarantee years", min_value=1, max_value=30, step=1, key=pfx + "guarantee_years")
    elif pt == ProductType.FIA:
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        with c1:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
        with c2:
            st.number_input("Participation", min_value=0.0, max_value=5.0, format="%.4f", key=pfx + "participation")
        with c3:
            st.number_input("Cap (annual decimal)", min_value=-1.0, max_value=2.0, format="%.4f", key=pfx + "cap")
        with c4:
            st.number_input("Floor (annual decimal)", min_value=-1.0, max_value=1.0, format="%.4f", key=pfx + "floor")
        with c5:
            st.number_input(
                "Rider fee (annual on AV)",
                min_value=0.0,
                max_value=1.0,
                format="%.4f",
                key=pfx + "rider_fee_annual",
            )
        with c6:
            st.number_input("Horizon (years)", min_value=1, max_value=80, step=1, key=pfx + "horizon_years")
    elif pt == ProductType.VARIABLE_ANNUITY:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
        with c2:
            st.number_input(
                "M&E charge (annual decimal)",
                min_value=0.0,
                max_value=1.0,
                format="%.4f",
                key=pfx + "me_charge_annual",
            )
        with c3:
            st.number_input("Horizon (years)", min_value=1, max_value=80, step=1, key=pfx + "horizon_years")
        st.selectbox(
            "GMDB basis",
            options=["return_of_premium", "max_anniversary"],
            key=pfx + "gmdb_basis",
            accept_new_options=True,
        )
    elif pt == ProductType.WHOLE_LIFE:
        st.number_input("Face amount ($)", min_value=1.0, step=10_000.0, key=pfx + "face_amount")
    elif pt == ProductType.UNIVERSAL_LIFE:
        c1, c2 = st.columns(2)
        with c1:
            st.number_input("Face amount ($)", min_value=1.0, step=10_000.0, key=pfx + "face_amount")
        with c2:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
    elif pt == ProductType.INDEXED_UL:
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.number_input("Face amount ($)", min_value=1.0, step=10_000.0, key=pfx + "face_amount")
        with c2:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
        with c3:
            st.number_input("Participation", min_value=0.0, max_value=5.0, format="%.4f", key=pfx + "participation")
        with c4:
            st.number_input("Cap (annual decimal)", min_value=-1.0, max_value=2.0, format="%.4f", key=pfx + "cap")
        with c5:
            st.number_input("Floor (annual decimal)", min_value=-1.0, max_value=1.0, format="%.4f", key=pfx + "floor")
    elif pt == ProductType.VARIABLE_UL:
        c1, c2 = st.columns(2)
        with c1:
            st.number_input("Face amount ($)", min_value=1.0, step=10_000.0, key=pfx + "face_amount")
        with c2:
            st.number_input("Single premium ($)", min_value=1.0, step=1000.0, key=pfx + "single_premium")
    else:
        st.warning(f"No manual fields wired for {pt.value}.")


def _render_portfolio_section() -> None:
    """Multi-policy inforce CSV run (see ``portfolio_config.portfolio_v1_enabled``)."""
    st.subheader("Portfolio (multi-policy)")
    st.caption(
        "Load an inforce CSV (see ``tests/data/inforce/example_v1/inforce.csv`` for column layout). "
        "Shared economics below use the same ``run_*`` keys as **Pricing Run** (deterministic pricing only)."
    )
    st.markdown("##### Manual policies (optional)")
    st.caption(
        "Defaults match the **Pricing Run** tab (from ``build_run_form_seed_defaults``). "
        "Add policies, adjust fields, then **Run from table**."
    )
    st.session_state.setdefault(PORTFOLIO_KEY.MANUAL_ROWS, [])
    b1, b2 = st.columns(2)
    with b1:
        if st.button("Add policy", type="secondary", key="portfolio_manual_add_button"):
            _portfolio_add_manual_row()
    with b2:
        if st.button("Remove last policy", type="secondary", key="portfolio_manual_remove_button"):
            _portfolio_remove_last_manual_row()

    product_options = list(product_options_for_ui())
    product_values = [p.value for p in product_options]
    row_ids = list(st.session_state.get(PORTFOLIO_KEY.MANUAL_ROWS) or [])

    _seed_run_form_state_from_last_inputs()
    ref_pt = _portfolio_reference_product_for_economics(row_ids)
    _normalize_run_state_for_selected_product(
        st.session_state,
        selected_product=ref_pt,
        switched_product=False,
    )
    st.markdown("##### Shared economics (Pricing Run defaults)")
    st.caption(
        "Yield curve, mortality, expenses, horizon, spread, and index scenario match the **Pricing Run** tab; "
        "Monte Carlo is omitted for portfolio batch pricing."
    )
    _render_shared_pricing_economics_controls(ref_pt, include_monte_carlo=False)

    for idx, row_id in enumerate(row_ids):
        pfx = _portfolio_row_prefix(row_id)
        meta_pt_prev = f"portfolio_meta_{row_id}_product_type_prev"
        title_pid = str(st.session_state.get(pfx + "policy_id", "")).strip() or "(new)"
        with st.expander(f"Policy {idx + 1} — {title_pid}", expanded=True):
            sel = st.selectbox(
                "Product type",
                options=product_values,
                format_func=lambda v: product_label(ProductType(v)),
                key=pfx + "product_type",
            )
            if meta_pt_prev in st.session_state and st.session_state[meta_pt_prev] != sel:
                preserve = {
                    "policy_id": str(st.session_state.get(pfx + "policy_id", "")).strip(),
                    "issue_age": st.session_state.get(pfx + "issue_age"),
                    "sex": st.session_state.get(pfx + "sex"),
                }
                newd = default_inforce_scratch_row(ProductType(sel), preserve=preserve)
                _portfolio_wipe_row_keys(row_id, except_keys=frozenset({pfx + "product_type"}))
                _portfolio_push_defaults_to_session(
                    row_id,
                    newd,
                    skip_session_keys=frozenset({pfx + "product_type"}),
                )
            st.session_state[meta_pt_prev] = sel

            opts_pid = _portfolio_policy_id_options(row_ids, row_id)
            cur_pid = str(st.session_state.get(pfx + "policy_id", opts_pid[0] if opts_pid else "")).strip()
            if cur_pid and cur_pid not in opts_pid:
                opts_pid = [cur_pid, *opts_pid]
            st.selectbox(
                "Policy ID",
                options=opts_pid if opts_pid else ["P0001"],
                key=pfx + "policy_id",
                accept_new_options=True,
            )

            c1, c2, c3 = st.columns(3)
            with c1:
                st.number_input("Issue age", min_value=0, max_value=120, step=1, key=pfx + "issue_age")
            with c2:
                st.selectbox(
                    "Sex",
                    options=["male", "female"],
                    key=pfx + "sex",
                    accept_new_options=True,
                )
            with c3:
                st.caption("Product-specific fields below.")
            pt = ProductType(str(st.session_state.get(pfx + "product_type", sel)).strip())
            _render_portfolio_contract_fields(row_id, pt)

    if row_ids:
        st.markdown("##### Assembled inforce preview")
        st.dataframe(
            _portfolio_manual_rows_dataframe(),
            use_container_width=True,
            hide_index=True,
        )

    if st.button("Run from table", type="secondary", key="portfolio_run_table_button"):
        from inforce_io import load_policy_inputs_from_csv_from_dataframe

        try:
            row_ids_run = list(st.session_state.get(PORTFOLIO_KEY.MANUAL_ROWS) or [])
            if not row_ids_run:
                st.error("Add at least one policy (use **Add policy**), or upload a CSV.")
            else:
                df = _portfolio_manual_rows_dataframe()
                if df.empty or "product_type" not in df.columns:
                    st.error("Add at least one row with a product_type.")
                else:
                    pt = df["product_type"].astype(str).str.strip()
                    df = df.loc[pt != ""].loc[pt.str.lower() != "nan"]
                    if df.empty:
                        st.error("Add at least one policy row with a product_type.")
                    else:
                        policies = load_policy_inputs_from_csv_from_dataframe(df)
                        scen = _portfolio_run_scenario_for_policies(tuple(policies))
                        res, alm_skip_msg = _execute_portfolio_pricing(tuple(policies), scen)
                        st.session_state[PORTFOLIO_KEY.RESULT] = res
                        st.session_state[PORTFOLIO_KEY.LAST_SCENARIO] = scen
                        st.session_state[PORTFOLIO_KEY.RUN_ID] = (
                            int(st.session_state.get(PORTFOLIO_KEY.RUN_ID, 0)) + 1
                        )
                        st.session_state[PORTFOLIO_KEY.UPLOAD_NAME] = "scratch_table.csv"
                        st.success("Portfolio pricing completed (from table).")
                        if alm_skip_msg:
                            st.warning(alm_skip_msg)
        except Exception as exc:  # noqa: BLE001 -- surface parse/pricing errors to the user
            st.error(f"{type(exc).__name__}: {exc}")

    up = st.file_uploader("Inforce CSV", type=["csv"], key="portfolio_inforce_uploader")
    if st.button("Run portfolio", type="primary", key="portfolio_run_button"):
        if up is None:
            st.error("Upload an inforce CSV first.")
        else:
            import tempfile

            with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tf:
                tf.write(up.getvalue())
                tmp = Path(tf.name)
            try:
                policies = load_policy_inputs_from_csv(tmp)
                scen = _portfolio_run_scenario_for_policies(tuple(policies))
                res, alm_skip_msg = _execute_portfolio_pricing(tuple(policies), scen)
                st.session_state[PORTFOLIO_KEY.RESULT] = res
                st.session_state[PORTFOLIO_KEY.LAST_SCENARIO] = scen
                st.session_state[PORTFOLIO_KEY.RUN_ID] = (
                    int(st.session_state.get(PORTFOLIO_KEY.RUN_ID, 0)) + 1
                )
                st.session_state[PORTFOLIO_KEY.UPLOAD_NAME] = getattr(up, "name", "inforce.csv")
                st.success("Portfolio pricing completed.")
                if alm_skip_msg:
                    st.warning(alm_skip_msg)
            finally:
                tmp.unlink(missing_ok=True)

    res = st.session_state.get(PORTFOLIO_KEY.RESULT)
    if res is not None:
        summ = portfolio_result_to_summary_dict(res)
        st.json(summ, expanded=False)
        scen_last = st.session_state.get(PORTFOLIO_KEY.LAST_SCENARIO)
        expenses_last = scen_last.expenses if isinstance(scen_last, RunScenario) else None
        _render_portfolio_liability_projection_chart(res)
        _render_portfolio_profit_waterfall(res, expenses_last)
        _render_portfolio_alm_baseline_section(res)
        rows = []
        for pt in sorted(res.rollups_by_product_type, key=lambda x: x.value):
            scal = res.product_type_scalar_rollups[pt]
            rows.append(
                {
                    "product_type": pt.value,
                    "policy_count": scal.policy_count,
                    "sum_single_premium": scal.sum_single_premium,
                    "rollup_cf_sum": float(res.rollups_by_product_type[pt].expected_total_cashflows.sum()),
                }
            )
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        for pr in res.policy_results:
            with st.expander(f"{pr.policy_id} — {pr.product_type.value}"):
                pv = getattr(pr.pricing, "pv_benefit", None)
                sp = getattr(pr.pricing, "single_premium", None)
                st.write(
                    {
                        "pv_benefit": float(pv) if pv is not None else None,
                        "single_premium": float(sp) if sp is not None else None,
                    }
                )
        xlsx = build_portfolio_workbook_bytes(res)
        st.download_button(
            "Download portfolio workbook",
            data=xlsx,
            file_name="portfolio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="portfolio_download_workbook",
        )


def main() -> None:
    st.set_page_config(page_title="Pricing Demo", layout="wide")
    with st.sidebar:
        st.title("Pricing Demo")
        if portfolio_v1_enabled():
            st.session_state.pop(PORTFOLIO_KEY.UI_FORCE_SIDEBAR, None)
            st.caption(
                "Batch / multi-policy: set **Section** (below) to **Portfolio (multi-policy)**."
            )
        else:
            with st.expander("Portfolio section is off — why?", expanded=False):
                st.markdown(portfolio_disabled_explanation_markdown())
                st.caption(
                    "Optional: show the Portfolio page in **Section** for this browser session "
                    "only (Streamlit). CLI `portfolio-run` follows the same enablement rules."
                )
                st.checkbox(
                    "Show Portfolio (multi-policy) in Section list",
                    key=PORTFOLIO_KEY.UI_FORCE_SIDEBAR,
                )
        page = st.radio(
            "Section",
            options=_sidebar_section_options(),
            format_func=lambda x: SECTION_LABELS[x],
        )
        st.divider()
        st.caption(f"Project root: `{ROOT}`")

        st.subheader("Diagnostics export")
        # Diagnostics should be fully self-contained for offline review/debugging.
        # Always include everything (no include/exclude toggles).
        include_full_paths = True
        include_alm_buckets = True
        if st.button("Prepare diagnostics JSON", type="secondary"):
            pricing_res = st.session_state.get("pricing_res")
            pricing_contract = st.session_state.get("pricing_contract")
            pricing_excel_context = st.session_state.get("pricing_excel_context") or {}
            alm_last = st.session_state.get("alm_last")
            alm_last_assumptions = st.session_state.get("alm_last_assumptions")
            alm_current_assumptions = st.session_state.get("alm_current_assumptions")
            alm_current_aum0 = st.session_state.get("alm_current_initial_asset_market_value")

            if pricing_res is None or pricing_contract is None:
                st.warning("Run Pricing Run first to populate diagnostics.")
            else:
                ctx_yc = pricing_excel_context.get("yield_curve")
                ctx_mort = pricing_excel_context.get("mortality")
                ctx_exp = pricing_excel_context.get("expenses")
                payload: dict[str, Any] = {
                    "exported_at_utc": _dt.datetime.utcnow().isoformat() + "Z",
                    "pricing_run_id": st.session_state.get("pricing_run_id"),
                    "pricing_meta": st.session_state.get("pricing_meta") or {},
                    "pricing_run_inputs": st.session_state.get("pricing_run_inputs") or {},
                    "pricing": _pricing_result_to_dict(
                        pricing_res,
                        pricing_contract,
                        include_full=include_full_paths,
                    ),
                    "pricing_inputs": {
                        "horizon_age": pricing_excel_context.get("horizon_age"),
                        "valuation_year": pricing_excel_context.get("valuation_year"),
                        "spread": pricing_excel_context.get("spread"),
                        "yield_curve": (
                            _yield_curve_to_dict(ctx_yc)
                            if isinstance(ctx_yc, sp.YieldCurve)
                            else None
                        ),
                        "mortality": _mortality_to_dict(ctx_mort) if ctx_mort is not None else None,
                        "expenses": (
                            {
                                "policy_expense_dollars": float(
                                    getattr(ctx_exp, "policy_expense_dollars", float("nan"))
                                ),
                                "premium_expense_rate": float(
                                    getattr(ctx_exp, "premium_expense_rate", float("nan"))
                                ),
                                "monthly_expense_dollars": float(
                                    getattr(ctx_exp, "monthly_expense_dollars", float("nan"))
                                ),
                            }
                            if isinstance(ctx_exp, sp.ExpenseAssumptions)
                            else None
                        ),
                        "yield_mode": pricing_excel_context.get("yield_mode"),
                        "mortality_mode": pricing_excel_context.get("mortality_mode"),
                        "expense_mode": pricing_excel_context.get("expense_mode"),
                        "expense_annual_inflation": pricing_excel_context.get(
                            "expense_annual_inflation"
                        ),
                    },
                    "alm": None,
                    "alm_current": None,
                    "what_if": None,
                }

                current_pricing_run_id = st.session_state.get("pricing_run_id")
                alm_run_id = st.session_state.get("alm_last_pricing_run_id")
                whatif_run_id = st.session_state.get("whatif_last_pricing_run_id")

                if isinstance(alm_last, sp.ALMResult) and alm_run_id == current_pricing_run_id:
                    payload["alm"] = _alm_result_to_dict(
                        alm_last,
                        (
                            alm_last_assumptions
                            if isinstance(alm_last_assumptions, sp.ALMAssumptions)
                            else None
                        ),
                        include_buckets=include_alm_buckets,
                        include_full=include_full_paths,
                    )

                if isinstance(alm_current_assumptions, sp.ALMAssumptions):
                    payload["alm_current"] = {
                        "initial_asset_market_value": (
                            float(alm_current_aum0) if alm_current_aum0 is not None else None
                        ),
                        "assumptions": _alm_assumptions_to_dict(alm_current_assumptions),
                    }

                what_if_shocked_res = st.session_state.get("whatif_last_shocked_res")
                what_if_base_res = st.session_state.get("whatif_last_base_res")
                what_if_alm_base = st.session_state.get("whatif_last_alm_base")
                what_if_alm_after = st.session_state.get("whatif_last_alm_after")
                what_if_baseline_mc = st.session_state.get("whatif_last_baseline_mc")
                what_if_shocked_mc = st.session_state.get("whatif_last_shocked_mc")
                what_if_shocked_curve = st.session_state.get("whatif_last_shocked_curve")
                what_if_shocked_mortality = st.session_state.get("whatif_last_shocked_mortality")
                what_if_alm_assumptions = st.session_state.get("whatif_last_alm_assumptions")
                what_if_params = st.session_state.get("whatif_last_params") or {}

                pricing_meta_whatif = st.session_state.get("pricing_meta") or {}
                pt_whatif = str(pricing_meta_whatif.get("product_type", ProductType.SPIA.value))
                what_if_need_mc = pt_whatif != ProductType.TERM_LIFE.value
                if (
                    whatif_run_id == current_pricing_run_id
                    and what_if_shocked_res is not None
                    and what_if_base_res is not None
                    and (
                        not what_if_need_mc
                        or (what_if_baseline_mc is not None and what_if_shocked_mc is not None)
                    )
                ):
                    payload["what_if"] = _whatif_result_to_dict(
                        base_res=what_if_base_res,
                        shocked_res=what_if_shocked_res,
                        baseline_mc=what_if_baseline_mc,
                        shocked_mc=what_if_shocked_mc,
                        whatif_params={
                            **what_if_params,
                            "shocked_curve": (
                                _yield_curve_to_dict(what_if_shocked_curve)
                                if isinstance(what_if_shocked_curve, sp.YieldCurve)
                                else None
                            ),
                            "shocked_mortality": (
                                _mortality_to_dict(what_if_shocked_mortality)
                                if what_if_shocked_mortality is not None
                                else None
                            ),
                        },
                        alm_base=what_if_alm_base,
                        alm_after=what_if_alm_after,
                        asm=(
                            what_if_alm_assumptions
                            if isinstance(what_if_alm_assumptions, sp.ALMAssumptions)
                            else None
                        ),
                        include_full=include_full_paths,
                    )

                st.session_state["diagnostics_json_bytes"] = json.dumps(
                    payload, default=str, ensure_ascii=False, indent=2
                ).encode("utf-8")
                st.session_state["diagnostics_json_filename"] = (
                    f"pricing_diagnostics_{_dt.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
                )
                st.success("Diagnostics JSON prepared. Use Download below.")

        diag_bytes = st.session_state.get("diagnostics_json_bytes")
        diag_name = st.session_state.get("diagnostics_json_filename") or "pricing_diagnostics.json"
        if isinstance(diag_bytes, (bytes, bytearray)) and diag_bytes:
            st.download_button(
                "Download diagnostics JSON",
                data=diag_bytes,
                file_name=diag_name,
                mime="application/json",
                type="primary",
            )

    if page == "overview":
        _render_overview()
    elif page == "run":
        _render_run_and_results()
    elif page == "portfolio":
        _render_portfolio_section()
    elif page == "alm":
        _render_alm_section()
    elif page == "what_if":
        _render_what_if_studio()
    elif page == "excel_replicator":
        _render_excel_replicator()
    else:
        render_unit_tests_page(embedded=True)


if __name__ == "__main__":
    main()
