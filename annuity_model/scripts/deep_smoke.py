"""
Deep smoke test for the 10-product workbook export pipeline.

Builds a real .xlsx for each implemented product on disk, opens it back with
openpyxl, runs the static validator, then inspects ModelCheck for evidence
that the Python <-> Excel parity wiring is intact.

Usage (from annuity_model/, with .venv active):
    python scripts/deep_smoke.py

Output workbooks land in annuity_model/.smoke/out/ (gitignored).
"""

from __future__ import annotations

import os
import sys
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import numpy as np
from openpyxl import load_workbook

from portfolio_config import portfolio_v1_enabled

import fia_projection as fp
import iul_projection as iul
import myga_projection as my
import pricing_projection as sp
import rila_projection as rp
import term_projection as tp
import ul_projection as ul
import va_projection as va
import vul_projection as vul
import wl_projection as wl
from build_fia_excel_workbook import (
    build_fia_workbook_from_spec,
    fia_excel_spec_from_launcher,
)
from build_iul_excel_workbook import (
    build_iul_workbook_from_spec,
    iul_excel_spec_from_launcher,
)
from build_myga_excel_workbook import (
    build_myga_workbook_from_spec,
    myga_excel_spec_from_launcher,
)
from build_pricing_excel_workbook import (
    ExcelPythonSnapshot,
    alm_excel_snapshot_from_result,
    build_workbook_from_spec,
    excel_spec_from_launcher,
)
from build_rila_excel_workbook import (
    build_rila_workbook_from_spec,
    rila_excel_spec_from_launcher,
)
from build_term_excel_workbook import (
    build_term_workbook_from_spec,
    term_excel_spec_from_launcher,
)
from build_ul_excel_workbook import (
    build_ul_workbook_from_spec,
    ul_excel_spec_from_launcher,
)
from build_va_excel_workbook import (
    build_va_workbook_from_spec,
    va_excel_spec_from_launcher,
)
from build_vul_excel_workbook import (
    build_vul_workbook_from_spec,
    vul_excel_spec_from_launcher,
)
from build_wl_excel_workbook import (
    build_wl_workbook_from_spec,
    wl_excel_spec_from_launcher,
)
from excel_workbook_validator import validate_workbook

OUT_DIR = REPO_ROOT / ".smoke" / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def _validate_xlsx_path(path: Path) -> tuple[int, int]:
    """Return (n_sheets, n_formulas) after validating the workbook."""
    wb = load_workbook(path, data_only=False)
    issues = validate_workbook(wb)
    if issues:
        for iss in issues[:10]:
            print(f"      issue: {iss}")
        raise SystemExit(f"VALIDATOR FAILED for {path.name}: {len(issues)} issue(s)")
    n_formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.startswith("="):
                    n_formulas += 1
    return len(wb.worksheets), n_formulas


def _peek_modelcheck(path: Path, sheet: str = "ModelCheck") -> dict[str, object]:
    """Return the first ~10 (label, value) pairs from a ModelCheck sheet."""
    wb = load_workbook(path, data_only=False)
    if sheet not in wb.sheetnames:
        return {"sheet": None}
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        rows.append(tuple(row))
    return {"sheet": sheet, "head": rows}


# ----------------------------- SPIA ---------------------------------------
def build_spia() -> Path:
    contract = sp.SPIAContract(issue_age=65, sex="male", benefit_annual=100_000.0)
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.full_like(ages, 0.02, dtype=float)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    res = sp.price_spia_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_scenario_csv_path=None,
        expense_annual_inflation=0.0,
    )
    spec = excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=80,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.0,
    )
    snap = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    raw = build_workbook_from_spec(spec, out_path=None, python_snapshot=snap)
    out = OUT_DIR / "spia_smoke.xlsx"
    out.write_bytes(raw)
    return out


# ----------------------------- Term ---------------------------------------
def build_term() -> Path:
    contract = tp.TermLifeContract(
        issue_age=40,
        sex="male",
        death_benefit=250_000.0,
        monthly_premium=200.0,
        term_years=20,
        premium_mode="level_monthly",
        benefit_timing="eoy_death",
    )
    yc = sp.YieldCurve.from_flat_rate(0.04)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 0.0)
    spec = term_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=60,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
    )
    raw = build_term_workbook_from_spec(spec)
    out = OUT_DIR / "term_smoke.xlsx"
    out.write_bytes(raw)
    return out


# ----------------------------- RILA ---------------------------------------
def build_rila(*, with_alm: bool = False, label: str = "rila") -> Path:
    contract = rp.RILAContract(
        issue_age=55,
        sex="male",
        participation=0.85,
        cap=0.09,
        floor=-0.02,
        rider_fee_annual=0.008,
    )
    yc = sp.YieldCurve.from_flat_rate(0.035)
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.008 + ages * 2e-5, 1e-6, 0.4)
    mort = sp.MortalityTableQx(ages, qx)
    ex = sp.ExpenseAssumptions(0.0, 0.0, 25.0)
    horizon_age = 85
    n_months = int(round((horizon_age - contract.issue_age) * 12))
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.004, 0.02, size=n_months))
    res = rp.price_rila_single_premium(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=horizon_age,
        spread=0.0,
        valuation_year=None,
        expenses=ex,
        index_s0=100.0,
        index_levels_payment=levels,
        expense_annual_inflation=0.01,
    )
    spec = rila_excel_spec_from_launcher(
        contract=contract,
        yield_curve=yc,
        mortality=mort,
        horizon_age=horizon_age,
        spread=0.0,
        valuation_year=2025,
        expenses=ex,
        yield_mode_label="flat",
        mortality_mode_label="synthetic",
        expense_mode_label="manual",
        index_s0=float(res.index_s0),
        index_levels_at_payment=res.index_level_at_payment,
        expense_annual_inflation=0.01,
    )

    if with_alm:
        asm = sp.ALMAssumptions(
            allocation=sp.alm_default_allocation_spec(),
            rebalance_band=0.05,
            rebalance_frequency_months=12,
            reinvest_rule="pro_rata",
            disinvest_rule="shortest_first",
            rebalance_policy="liquidity_only",
        )
        alm_result = sp.run_alm_projection_from_pricing_result(
            pricing=res,
            yield_curve=yc,
            spread=0.0,
            assumptions=asm,
            initial_asset_market_value=float(res.single_premium),
        )
        alm_snap = alm_excel_snapshot_from_result(
            alm_result, asm, initial_asset_market_value=float(res.single_premium)
        )
        raw = build_rila_workbook_from_spec(spec, alm_assumptions=asm, alm_snapshot=alm_snap)
    else:
        raw = build_rila_workbook_from_spec(spec)

    out = OUT_DIR / f"{label}_smoke.xlsx"
    out.write_bytes(raw)
    return out


def _flat_yc(rate: float = 0.04) -> sp.YieldCurve:
    return sp.YieldCurve.from_flat_rate(rate)


def _synth_mort() -> sp.MortalityTableQx:
    ages = np.arange(0, 121, dtype=int)
    qx = np.clip(0.005 + ages * 1e-5, 1e-6, 0.4)
    return sp.MortalityTableQx(ages, qx)


def build_myga() -> Path:
    contract = my.MYGAContract(
        issue_age=60, sex="male", single_premium=100_000.0,
        declared_rate_annual=0.045, guarantee_years=5,
    )
    spec = myga_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(0.045), mortality=_synth_mort(),
        horizon_age=70, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
    )
    out = OUT_DIR / "myga_smoke.xlsx"
    out.write_bytes(build_myga_workbook_from_spec(spec))
    return out


def build_fia() -> Path:
    contract = fp.FIAContract(
        issue_age=60, sex="male", single_premium=100_000.0,
        participation=0.8, cap=0.07, floor=0.0, horizon_years=10,
    )
    n_months = 120
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.02, size=n_months))
    spec = fia_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=70, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
        index_s0=100.0, index_levels_at_payment=levels,
    )
    out = OUT_DIR / "fia_smoke.xlsx"
    out.write_bytes(build_fia_workbook_from_spec(spec))
    return out


def build_va() -> Path:
    contract = va.VAContract(
        issue_age=55, sex="male", single_premium=100_000.0,
        me_charge_annual=0.014, horizon_years=20,
    )
    n_months = 240
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    spec = va_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=75, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
        index_s0=100.0, index_levels_at_payment=levels,
    )
    out = OUT_DIR / "va_smoke.xlsx"
    out.write_bytes(build_va_workbook_from_spec(spec))
    return out


def build_wl() -> Path:
    contract = wl.WLContract(
        issue_age=45, sex="male", smoker_class="nonsmoker", face_amount=250_000.0,
    )
    spec = wl_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=80, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
    )
    out = OUT_DIR / "wl_smoke.xlsx"
    out.write_bytes(build_wl_workbook_from_spec(spec))
    return out


def build_ul() -> Path:
    contract = ul.ULContract(
        issue_age=45, sex="male", face_amount=250_000.0, single_premium=25_000.0,
    )
    spec = ul_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=80, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
    )
    out = OUT_DIR / "ul_smoke.xlsx"
    out.write_bytes(build_ul_workbook_from_spec(spec))
    return out


def build_iul() -> Path:
    contract = iul.IULContract(
        issue_age=45, sex="male", face_amount=250_000.0, single_premium=25_000.0,
        participation=1.0, cap=0.10, floor=0.0,
    )
    n_months = (80 - 45) * 12
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    spec = iul_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=80, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
        index_s0=100.0, index_levels_at_payment=levels,
    )
    out = OUT_DIR / "iul_smoke.xlsx"
    out.write_bytes(build_iul_workbook_from_spec(spec))
    return out


def build_vul() -> Path:
    contract = vul.VULContract(
        issue_age=45, sex="male", face_amount=250_000.0, single_premium=25_000.0,
    )
    n_months = (80 - 45) * 12
    rng = np.random.default_rng(42)
    levels = 100.0 * np.cumprod(1.0 + rng.normal(0.005, 0.03, size=n_months))
    spec = vul_excel_spec_from_launcher(
        contract=contract, yield_curve=_flat_yc(), mortality=_synth_mort(),
        horizon_age=80, spread=0.0, valuation_year=2025,
        expenses=sp.ExpenseAssumptions(0.0, 0.0, 0.0),
        yield_mode_label="flat", mortality_mode_label="qx",
        expense_mode_label="manual", expense_annual_inflation=0.0,
        index_s0=100.0, index_levels_at_payment=levels,
    )
    out = OUT_DIR / "vul_smoke.xlsx"
    out.write_bytes(build_vul_workbook_from_spec(spec))
    return out


def build_portfolio() -> Path:
    """Mixed-product portfolio workbook (canonical inforce CSV)."""
    from build_portfolio_excel_workbook import build_portfolio_workbook_to_path
    from inforce_io import load_policy_inputs_from_csv
    from portfolio import Portfolio
    from portfolio_runner import run_portfolio
    from pricing_scenario_materialize import run_scenario_for_portfolio_policies

    csv_path = REPO_ROOT / "tests" / "data" / "inforce" / "example_v1" / "inforce.csv"
    policies = load_policy_inputs_from_csv(csv_path)
    pol_t = tuple(policies)
    sex_raw = str(getattr(pol_t[0].contract, "sex", "male")).strip().lower()
    sex = "female" if sex_raw == "female" else "male"
    scen = run_scenario_for_portfolio_policies({}, pol_t, sex=sex)  # type: ignore[arg-type]
    res = run_portfolio(portfolio=Portfolio(policies=pol_t), scenario=scen)
    out = OUT_DIR / "portfolio_smoke.xlsx"
    build_portfolio_workbook_to_path(res, out)
    return out


def main() -> int:
    t0 = time.perf_counter()
    print(f"[smoke] writing workbooks to {OUT_DIR}")
    failures: list[str] = []

    builders: list[tuple[str, object]] = [
        ("SPIA", build_spia),
        ("Term", build_term),
        ("RILA (no ALM)", lambda: build_rila(with_alm=False, label="rila_no_alm")),
        ("RILA + ALM", lambda: build_rila(with_alm=True, label="rila_alm")),
        ("MYGA", build_myga),
        ("FIA", build_fia),
        ("VA", build_va),
        ("WL", build_wl),
        ("UL", build_ul),
        ("IUL", build_iul),
        ("VUL", build_vul),
    ]
    if portfolio_v1_enabled():
        builders.append(("Portfolio", build_portfolio))
    else:
        print("[smoke] skipping Portfolio (see portfolio_config.portfolio_v1_enabled)")

    for name, builder in builders:
        try:
            t = time.perf_counter()
            path = builder()
            n_sheets, n_formulas = _validate_xlsx_path(path)
            elapsed_ms = (time.perf_counter() - t) * 1000
            kb = path.stat().st_size / 1024
            print(
                f"  OK  {name:<14} {path.name:<22} "
                f"{kb:7.1f} KB  {n_sheets} sheets  {n_formulas:>5} formulas  "
                f"{elapsed_ms:6.0f} ms"
            )
            if "rila" in path.name:
                meta = _peek_modelcheck(path)
                if meta.get("sheet"):
                    head = meta.get("head", [])
                    nonblank = [r for r in head if any(c not in (None, "") for c in r)]
                    print(
                        f"      ModelCheck has {len(nonblank)} non-blank head rows (showing first):"
                    )
                    if nonblank:
                        print(f"        {nonblank[0]}")
        except Exception as exc:
            print(f"  FAIL {name}: {exc!r}")
            failures.append(name)

    elapsed_total = time.perf_counter() - t0
    print(f"[smoke] done in {elapsed_total:.2f}s ({len(failures)} failures)")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
