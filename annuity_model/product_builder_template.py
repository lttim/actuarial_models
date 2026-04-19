"""Declarative builder template for the seven new products.

Each new product (MYGA / FIA / VA / WL / UL / IUL / VUL) ships with an
Excel workbook that follows the same skeleton:

* ``Inputs``        — (label, value) rows
* ``YieldCurve``    — copy of the maturity-rate curve
* ``MonthlyCurve``  — derived monthly discount factors (shared helper)
* ``Liabilities``   — per-month grid: month, age, survival, cashflow,
                      discount, PV
* ``ModelCheck``    — Python-vs-Excel reconciliation block

The variation between products is:

* Which input rows go on the Inputs sheet.
* The shape and width of the Liabilities grid (life vs. accumulation).
* The summary block at columns W/X.
* The ModelCheck rows.

This module exposes :func:`build_simple_product_workbook` which takes a
fully populated :class:`SimpleWorkbookSpec` and emits the workbook
bytes. It is intentionally smaller-surface than the SPIA / Term / RILA
hand-rolled builders (those do ALM and have product-specific liability
shapes that don't fit the simple skeleton); the seven new products use
this template.

Section 1.6 of ``docs/seven_product_rollout_plan.md``.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import pricing_projection as sp
from excel_builder_helpers import (
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from excel_workbook_validator import validate_workbook_or_raise
from recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"


@dataclass(frozen=True, slots=True)
class LiabilityGridColumn:
    """One column in the per-month Liabilities grid.

    Attributes
    ----------
    header:
        Column header text (row 3 of the Liabilities sheet).
    column_letter:
        Excel column letter (``A``, ``B``, ..., ``AA``).
    formula_template:
        Excel formula string with ``{a}`` substituted for the row's
        month-cell address (e.g. ``A4``). Use ``{r}`` for the current
        row number when needed (e.g. for self-references).
    number_format:
        openpyxl number_format string (``"#,##0.00"`` for money,
        ``"0.000000"`` for ratios, ``""`` to leave default).
    """

    header: str
    column_letter: str
    formula_template: str
    number_format: str = ""


@dataclass(frozen=True, slots=True)
class ModelCheckRow:
    """One row to insert into the standard ModelCheck block.

    The row format is ``(label, python_literal_value, excel_formula_ref, kind)``
    matching :func:`excel_builder_helpers.write_model_check_sheet`.
    """

    label: str
    python_value: float
    excel_formula: str
    kind: str = "money"  # "money" or "factor"


@dataclass(frozen=True, slots=True)
class SimpleWorkbookSpec:
    """All the data needed to emit one product's workbook via the template."""

    title: str
    """Title cell (A1) of the Inputs sheet."""

    inputs_rows: Sequence[tuple[str, object]]
    """(label, value) rows for the Inputs sheet (rows 3..N)."""

    yield_curve: sp.YieldCurve
    """Yield curve for the YieldCurve / MonthlyCurve sheets."""

    n_months: int
    """Number of liability months for the projection (and grid height)."""

    grid_max_rows: int
    """Maximum rows the Liabilities grid is sized for (>= n_months)."""

    grid_columns: Sequence[LiabilityGridColumn]
    """Column specs for the Liabilities sheet (rows 4..grid_max_rows+3)."""

    liability_summary_rows: Sequence[tuple[int, str, str]]
    """Rows for the W/X summary block on the Liabilities sheet."""

    modelcheck_subtitle: str
    """Subtitle text for the ModelCheck sheet."""

    modelcheck_rows: Sequence[ModelCheckRow]
    """Per-product ModelCheck reconciliation rows."""

    modelcheck_python_snapshot: object
    """ExcelPythonSnapshot-shaped object passed through to write_model_check_sheet."""

    n_months_cell_address: str = "Inputs!$B$18"
    """Where the ``n_months`` formula lives. Per builder convention."""

    n_months_formula: str | None = None
    """If supplied, the formula written into ``n_months_cell_address`` (B18 by default)."""


def build_simple_product_workbook(
    spec: SimpleWorkbookSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    """Emit the standard four-sheet workbook for a simple product.

    The workbook always has:
      * ``Inputs``                — populated from ``spec.inputs_rows``.
      * ``YieldCurve``            — from ``spec.yield_curve``.
      * ``MonthlyCurve_recalc``   — derived (shared helper).
      * ``Liabilities``           — driven by ``spec.grid_columns`` and
        ``spec.liability_summary_rows``.
      * ``ModelCheck``            — driven by ``spec.modelcheck_rows``.
    """
    if spec.grid_max_rows < spec.n_months:
        raise ValueError(
            f"grid_max_rows ({spec.grid_max_rows}) must be >= n_months "
            f"({spec.n_months})."
        )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = "Inputs"
    write_inputs_sheet(
        ws_in,
        InputsSheetSpec(title=spec.title, rows=list(spec.inputs_rows)),
    )
    if spec.n_months_formula is not None:
        # Address looks like "Inputs!$B$18". Strip the sheet prefix so we
        # can write to ws_in directly.
        addr = spec.n_months_cell_address.split("!")[-1].replace("$", "")
        # Add label two columns to the left for clarity (column A).
        col_letter = "".join(c for c in addr if c.isalpha())
        row_num = "".join(c for c in addr if c.isdigit())
        if col_letter == "B":
            ws_in[f"A{row_num}"] = "Model months (formula)"
        ws_in[addr] = spec.n_months_formula
        ws_in[addr].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=spec.grid_max_rows, y_last_row=y_last_row)

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = spec.title
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={spec.n_months_cell_address}"
    nm_ref = spec.n_months_cell_address
    first = 4
    last_cap_row = first - 1 + spec.grid_max_rows

    # Always emit column A with the month index gating formula (so other
    # columns can refer to A{r} as their "is this row in-range" gate).
    ws_pr.cell(row=3, column=1, value="Month").font = Font(bold=True)
    for r in range(first, last_cap_row + 1):
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=1).number_format = "0"

    for col in spec.grid_columns:
        # Header row 3.
        col_idx = _col_letter_to_index(col.column_letter)
        ws_pr.cell(row=3, column=col_idx, value=col.header).font = Font(bold=True)
        for r in range(first, last_cap_row + 1):
            a = f"A{r}"
            value = col.formula_template.format(a=a, r=r)
            cell = ws_pr.cell(row=r, column=col_idx, value=value)
            if col.number_format:
                cell.number_format = col.number_format

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(rows=list(spec.liability_summary_rows)),
    )

    write_model_check_sheet(
        wb,
        spec.modelcheck_python_snapshot,
        alm_layout=None,
        alm_snapshot=None,
        pricing_rows=[
            (row.label, float(row.python_value), row.excel_formula, row.kind)
            for row in spec.modelcheck_rows
        ],
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=spec.modelcheck_subtitle,
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data


def _col_letter_to_index(letter: str) -> int:
    """Convert column letter (``A``..``ZZ``) to 1-based index."""
    s = letter.strip().upper()
    if not s.isalpha():
        raise ValueError(f"invalid column letter {letter!r}")
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


__all__ = [
    "LIABILITY_SHEET_NAME",
    "LiabilityGridColumn",
    "ModelCheckRow",
    "SimpleWorkbookSpec",
    "build_simple_product_workbook",
]
