from __future__ import annotations

import math
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Literal, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from annuity_model import pricing_projection as sp
from annuity_model import rila_projection as rp
from annuity_model.excel_builder_helpers import (
    ALM_ENGINE_STEP_MONTHS,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_PROJECTION_FIRST_DATA_ROW,
    LIABILITY_SHEET_NAME,
    ALMExcelSnapshot,
    ExcelPythonSnapshot,
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    alm_excel_downsample_snapshot,
    alm_excel_truncate_snapshot,
    inject_alm_projection_formula_cached_values,
    liability_layout_for,
    write_alm_projection_sheet,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from annuity_model.policy_features import SegmentAllocation, normalize_segment_allocations
from annuity_model.recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

RILA_PROJ_MAX_ROWS = 600
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_MTH_QX = "MortalMonthly"
SHEET_INDEX = "IndexScenario"
SHEET_SCHEDULES = "PolicySchedules"
SHEET_SEGMENTS = "SegmentAllocations"

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_PART = 4
_IN_ROW_CAP = 5
_IN_ROW_FLOOR = 6
_IN_ROW_RIDER = 7
_IN_ROW_FREQ = 8
_IN_ROW_VAL = 9
_IN_ROW_HORIZON = 10
_IN_ROW_SPREAD = 11
_IN_ROW_PREMIUM = 12
_IN_ROW_EXP_MTH = 13
_IN_ROW_EXP_INF = 14
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


def _n_months_cell() -> str:
    return _in_addr("B", _IN_ROW_NMONTHS)


def _fill_mortal_monthly_rpmp(
    ws_m: Worksheet,
    *,
    mort: sp.MortalityTableRP2014MP2016,
    issue_age: int,
    valuation_year: int,
    n_months: int,
) -> None:
    dt = 1.0 / 12.0
    for k in range(1, n_months + 1):
        r = 1 + k
        m_index = k - 1
        age_start = issue_age + m_index * dt
        age_int = int(math.floor(age_start))
        calendar_year_start = valuation_year + 1 + (m_index // 12)
        qxv = float(
            mort.qx_at_int_age_and_calendar_year(age_int=age_int, calendar_year=calendar_year_start)
        )
        ws_m.cell(row=r, column=1, value=f'=IF(ROW()-1>{_n_months_cell()},"",ROW()-1)')
        ws_m.cell(row=r, column=2, value=int(age_int))
        ws_m.cell(row=r, column=3, value=int(calendar_year_start))
        ws_m.cell(row=r, column=4, value=float(qxv))
    last_data = 1 + n_months
    last_cap = 1 + RILA_PROJ_MAX_ROWS
    for r in range(last_data + 1, last_cap + 1):
        ws_m.cell(row=r, column=1, value=f'=IF(ROW()-1>{_n_months_cell()},"",ROW()-1)')
        for c in (2, 3, 4):
            ws_m.cell(row=r, column=c, value="")


def _qx_lookup_expr(acell: str, mode: Literal["qx_table", "mortal_monthly"]) -> str:
    clamp_inner = "MIN(MAX({inner},0),0.999999)"
    if mode == "qx_table":
        inner = (
            f"INDEX({SHEET_QX}!$B$2:$B$50000,"
            f"MATCH(INT({_in_addr('B', _IN_ROW_ISSUE_AGE)}+({acell}-1)/12),"
            f"{SHEET_QX}!$A$2:$A$50000,0))"
        )
        return clamp_inner.format(inner=inner)
    inner = f"INDEX({SHEET_MTH_QX}!$D$2:$D$50000,MATCH({acell},{SHEET_MTH_QX}!$A$2:$A$50000,0))"
    return clamp_inner.format(inner=inner)


def _survival_end_formula(
    r: int, acell: str, mode: Literal["qx_table", "mortal_monthly"], *, first_row: int
) -> str:
    qx_e = _qx_lookup_expr(acell, mode)
    p_m = f"EXP(-(-LN(1-{qx_e}))/12)"
    if r == first_row:
        return f'=IF({acell}="","",{p_m})'
    prev = f"D{r - 1}"
    return f'=IF({acell}="","",{prev}*{p_m})'


@dataclass(frozen=True)
class RILAExcelBuildSpec:
    contract: rp.RILAContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float
    index_s0: float
    index_levels_payment: np.ndarray


def rila_excel_spec_from_launcher(
    *,
    contract: rp.RILAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
) -> RILAExcelBuildSpec:
    return RILAExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
        index_s0=float(index_s0),
        index_levels_payment=np.asarray(index_levels_at_payment, dtype=float),
    )


def build_rila_workbook_from_spec(
    spec: RILAExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
    alm_snapshot: ALMExcelSnapshot | None = None,
    alm_assumptions: sp.ALMAssumptions | None = None,
) -> bytes:
    s0_py = float(spec.index_s0)
    levels_py = np.asarray(spec.index_levels_payment, dtype=float)

    res = rp.price_rila_single_premium(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=(
            spec.valuation_year if not isinstance(spec.mortality, sp.MortalityTableQx) else None
        ),
        expenses=spec.expenses,
        index_s0=s0_py,
        index_levels_payment=levels_py,
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    wb = Workbook()
    ws_in = cast(Worksheet, wb.active)
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Participation", spec.contract.participation),
        ("Cap (decimal / yr segment)", spec.contract.cap),
        ("Floor (decimal / yr segment)", spec.contract.floor),
        ("Rider fee (annual, on AV)", spec.contract.rider_fee_annual),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Single premium (Python export)", float(res.single_premium)),
        ("Monthly expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(
        ws_in,
        InputsSheetSpec(
            title="RILA Inputs (matches model launcher / Python)",
            rows=rows,
        ),
    )

    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{RILA_PROJ_MAX_ROWS})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(
        ws_mc_curve,
        n_months=RILA_PROJ_MAX_ROWS,
        y_last_row=y_last_row,
        payments_per_year_ref=_in_addr("B", _IN_ROW_FREQ),
        spread_ref=_in_addr("B", _IN_ROW_SPREAD),
    )

    ws_ix = wb.create_sheet(SHEET_INDEX)
    ws_ix["A1"] = "month"
    ws_ix["B1"] = "index_level"
    L = rp.levels_end_by_policy_month(s0=float(s0_py), levels_payment=levels_py)
    for j in range(L.shape[0]):
        ws_ix.cell(row=2 + j, column=1, value=int(j))
        ws_ix.cell(row=2 + j, column=2, value=float(L[j]))

    sched_months = min(int(res.months.shape[0]), RILA_PROJ_MAX_ROWS)
    withdrawal_sched = spec.contract.withdrawals.values(sched_months)
    surrender_rates = spec.contract.surrender_charges.monthly_rates(sched_months)
    ws_sched = wb.create_sheet(SHEET_SCHEDULES)
    ws_sched["A1"] = "month"
    ws_sched["B1"] = "withdrawal"
    ws_sched["C1"] = "surrender_charge_rate"
    ws_sched["G1"] = "death_benefit_type"
    ws_sched["H1"] = str(spec.contract.death_benefit_type)
    ws_sched["G2"] = "glwb_enabled"
    ws_sched["H2"] = bool(spec.contract.glwb.enabled)
    ws_sched["G3"] = "glwb_rollup_monthly"
    ws_sched["H3"] = (
        float((1.0 + float(spec.contract.glwb.rollup_annual)) ** (1.0 / 12.0) - 1.0)
        if spec.contract.glwb.enabled
        else 0.0
    )
    ws_sched["G4"] = "glwb_fee_monthly"
    ws_sched["H4"] = (
        float(spec.contract.glwb.fee_annual) / 12.0 if spec.contract.glwb.enabled else 0.0
    )
    ws_sched["G5"] = "glwb_income_start_month"
    ws_sched["H5"] = int(spec.contract.glwb.income_start_month)
    ws_sched["G6"] = "glwb_withdrawal_monthly"
    ws_sched["H6"] = (
        float(spec.contract.glwb.withdrawal_rate) / 12.0 if spec.contract.glwb.enabled else 0.0
    )
    ws_sched["G7"] = "glwb_ratchet"
    ws_sched["H7"] = bool(spec.contract.glwb.ratchet)
    for c in range(1, 9):
        ws_sched.cell(row=1, column=c).font = Font(bold=True)
    for j in range(RILA_PROJ_MAX_ROWS):
        r = 2 + j
        ws_sched.cell(row=r, column=1, value=int(j + 1))
        if j < sched_months:
            ws_sched.cell(row=r, column=2, value=float(withdrawal_sched[j]))
            ws_sched.cell(row=r, column=3, value=float(surrender_rates[j]))
        else:
            ws_sched.cell(row=r, column=2, value=0.0)
            ws_sched.cell(row=r, column=3, value=0.0)

    allocations = normalize_segment_allocations(
        spec.contract.segment_allocations
        or (
            SegmentAllocation(
                weight=1.0,
                design="cap_floor",
                participation=float(spec.contract.participation),
                cap=float(spec.contract.cap),
                floor=float(spec.contract.floor),
            ),
        )
    )
    ws_seg = wb.create_sheet(SHEET_SEGMENTS)
    seg_hdr = ("weight", "design", "participation", "cap", "floor", "buffer")
    for c, h in enumerate(seg_hdr, start=1):
        ws_seg.cell(row=1, column=c, value=h).font = Font(bold=True)
    for i, alloc in enumerate(allocations, start=2):
        ws_seg.cell(row=i, column=1, value=float(alloc.weight))
        ws_seg.cell(row=i, column=2, value=str(alloc.design))
        ws_seg.cell(row=i, column=3, value=float(alloc.participation))
        ws_seg.cell(row=i, column=4, value=float(alloc.cap))
        ws_seg.cell(row=i, column=5, value=float(alloc.floor))
        ws_seg.cell(row=i, column=6, value=float(alloc.buffer))

    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx[i]))
        mort_mode: Literal["qx_table", "mortal_monthly"] = "qx_table"
    else:
        mort_mode = "mortal_monthly"
        ws_m = wb.create_sheet(SHEET_MTH_QX)
        ws_m["A1"] = "month"
        ws_m["B1"] = "age_int"
        ws_m["C1"] = "calendar_year_start"
        ws_m["D1"] = "qx_annual"
        _fill_mortal_monthly_rpmp(
            ws_m,
            mort=spec.mortality,
            issue_age=spec.contract.issue_age,
            valuation_year=int(spec.valuation_year),
            n_months=min(int(res.months.shape[0]), RILA_PROJ_MAX_ROWS),
        )

    nm_ref = _n_months_cell()
    last_cap_row = 3 + RILA_PROJ_MAX_ROWS
    first = 4
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"
    idx_b = f"{SHEET_INDEX}!$B$2:$B$10000"
    idx_a = f"{SHEET_INDEX}!$A$2:$A$10000"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "RILA liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={_in_addr('B', _IN_ROW_ISSUE_AGE)}"
    ws_pr["V2"] = "=X9"

    hdr = (
        "Month",
        "t_years",
        "AttainedAge",
        "SurvivalEnd",
        "SurvivalStart",
        "MonthDeathProb",
        "IndexLevel_m",
        "RawSegReturn",
        "CreditedSeg",
        "AV_end",
        "ExpBenefitCF",
        "ExpExpenseCF",
        "ExpTotalCF",
        "",
        "DiscountFactor",
        "PVBenefitCF",
        "PVExpenseCF",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "WithdrawalSched_XL",
        "SurrRate_XL",
        "RawSegReturn_XL",
        "CreditedSeg_XL",
        "BenefitBase_XL",
        "GLWBWithdrawal_XL",
        "WithdrawalPaid_XL",
        "RiderFee_XL",
        "AV_end_XL",
        "SurrCharge_XL",
        "SurrValue_XL",
        "DeathBenefit_XL",
        "DeathCF_XL",
        "PolicyAccessCF_XL",
        "ExpBenefitCF_XL",
        "PVBenefitCF_XL",
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    ws_pr.cell(row=3, column=10, value=0.0)

    part = _in_addr("B", _IN_ROW_PART)
    cap = _in_addr("B", _IN_ROW_CAP)
    fl = _in_addr("B", _IN_ROW_FLOOR)
    rider = _in_addr("B", _IN_ROW_RIDER)
    prem = _in_addr("B", _IN_ROW_PREMIUM)
    exp_m = _in_addr("B", _IN_ROW_EXP_MTH)
    exp_if = _in_addr("B", _IN_ROW_EXP_INF)
    sched_month_col = f"{SHEET_SCHEDULES}!$A$2:$A${1 + RILA_PROJ_MAX_ROWS}"
    sched_wd_col = f"{SHEET_SCHEDULES}!$B$2:$B${1 + RILA_PROJ_MAX_ROWS}"
    sched_surr_col = f"{SHEET_SCHEDULES}!$C$2:$C${1 + RILA_PROJ_MAX_ROWS}"
    db_type_ref = f"{SHEET_SCHEDULES}!$H$1"
    glwb_enabled_ref = f"{SHEET_SCHEDULES}!$H$2"
    glwb_roll_ref = f"{SHEET_SCHEDULES}!$H$3"
    glwb_fee_ref = f"{SHEET_SCHEDULES}!$H$4"
    glwb_start_ref = f"{SHEET_SCHEDULES}!$H$5"
    glwb_withdrawal_ref = f"{SHEET_SCHEDULES}!$H$6"
    glwb_ratchet_ref = f"{SHEET_SCHEDULES}!$H$7"

    def _allocation_credit_formula(raw_cell: str) -> str:
        terms: list[str] = []
        for idx in range(len(allocations)):
            row = 2 + idx
            w = f"{SHEET_SEGMENTS}!$A${row}"
            design = f"{SHEET_SEGMENTS}!$B${row}"
            p = f"{SHEET_SEGMENTS}!$C${row}"
            c = f"{SHEET_SEGMENTS}!$D${row}"
            f = f"{SHEET_SEGMENTS}!$E${row}"
            b = f"{SHEET_SEGMENTS}!$F${row}"
            terms.append(
                f'{w}*IF({design}="buffer",'
                f"IF({raw_cell}>=0,MIN({c},{p}*{raw_cell}),MIN(0,{raw_cell}+{b})),"
                f"MAX({f},MIN({c},{p}*{raw_cell})))"
            )
        return "+".join(terms) if terms else "0"

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{_in_addr("B", _IN_ROW_FREQ)})')
        ws_pr.cell(
            row=r,
            column=3,
            value=f'=IF({a}="","",{_in_addr("B", _IN_ROW_ISSUE_AGE)}+({a}-1)/{_in_addr("B", _IN_ROW_FREQ)})',
        )
        ws_pr.cell(row=r, column=4, value=_survival_end_formula(r, a, mort_mode, first_row=first))
        if r == first:
            d_surv_start = f'=IF({a}="","",1)'
        else:
            d_surv_start = f'=IF({a}="","",D{r - 1})'
        ws_pr.cell(row=r, column=5, value=d_surv_start)
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(
            row=r,
            column=7,
            value=(f'=IF({a}="","",IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0)),""))'),
        )
        ws_pr.cell(
            row=r,
            column=8,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0))/INDEX({idx_b},MATCH({a}-12,{idx_a},0))-1,0),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=9,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),MAX({fl},MIN({cap},{part}*H{r})),0))'
            ),
        )
        ws_pr.cell(
            row=r,
            column=10,
            value=(f'=IF({a}="","",(IF({a}=1,{prem},J{r - 1})*(1+I{r}))*(1-{rider}/12))'),
        )
        ws_pr.cell(row=r, column=11, value=f'=IF({a}="",0,F{r}*J{r})')
        ws_pr.cell(
            row=r,
            column=12,
            value=(f'=IF({a}="",0,{exp_m}*POWER(1+((1+{exp_if})^(1/12)-1),{a}-1)*D{r})'),
        )
        ws_pr.cell(row=r, column=13, value=f'=IF({a}="",0,K{r}+L{r})')
        ws_pr.cell(
            row=r,
            column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=16, value=f'=IF({a}="",0,K{r}*O{r})')
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,L{r}*O{r})')
        prev_av = prem if r == first else f"AI{r - 1}"
        prev_base = prem if r == first else f"AE{r - 1}"
        ws_pr.cell(
            row=r,
            column=27,
            value=f'=IF({a}="",0,IFERROR(INDEX({sched_wd_col},MATCH({a},{sched_month_col},0)),0))',
        )
        ws_pr.cell(
            row=r,
            column=28,
            value=f'=IF({a}="",0,IFERROR(INDEX({sched_surr_col},MATCH({a},{sched_month_col},0)),0))',
        )
        ws_pr.cell(
            row=r,
            column=29,
            value=(
                f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),'
                f"IFERROR(INDEX({idx_b},MATCH({a},{idx_a},0))/INDEX({idx_b},MATCH({a}-12,{idx_a},0))-1,0),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=30,
            value=f'=IF({a}="","",IF(AND({a}>=12,MOD({a},12)=0),{_allocation_credit_formula(f"AC{r}")},0))',
        )
        av_after_credit = f"({prev_av}*(1+AD{r}))"
        base_roll = f"IF(AND({glwb_enabled_ref},{a}<{glwb_start_ref}),{prev_base}*(1+{glwb_roll_ref}),{prev_base})"
        base_ratchet = (
            f"IF(AND({glwb_enabled_ref},{glwb_ratchet_ref},{a}<{glwb_start_ref},MOD({a},12)=0),"
            f"MAX({base_roll},{av_after_credit}),{base_roll})"
        )
        ws_pr.cell(row=r, column=31, value=f'=IF({a}="",0,{base_ratchet})')
        ws_pr.cell(
            row=r,
            column=32,
            value=(
                f'=IF({a}="",0,IF(AND({glwb_enabled_ref},{a}>={glwb_start_ref}),'
                f"MIN({av_after_credit},AE{r}*{glwb_withdrawal_ref}),0))"
            ),
        )
        ws_pr.cell(
            row=r,
            column=33,
            value=f'=IF({a}="",0,MIN(MAX(0,{av_after_credit}-AF{r}),AA{r}))',
        )
        ws_pr.cell(
            row=r,
            column=34,
            value=(
                f'=IF({a}="",0,MAX(0,{av_after_credit}-AF{r}-AG{r})*{rider}/12'
                f"+MAX(0,AE{r})*{glwb_fee_ref})"
            ),
        )
        ws_pr.cell(
            row=r,
            column=35,
            value=f'=IF({a}="",0,MAX(0,MAX(0,{av_after_credit}-AF{r}-AG{r})*(1-{rider}/12)-AE{r}*{glwb_fee_ref}))',
        )
        ws_pr.cell(row=r, column=36, value=f'=IF({a}="",0,AI{r}*AB{r})')
        ws_pr.cell(row=r, column=37, value=f'=IF({a}="",0,MAX(0,AI{r}-AJ{r}))')
        ws_pr.cell(
            row=r,
            column=38,
            value=f'=IF({a}="",0,IF({db_type_ref}="return_of_premium",MAX(AI{r},{prem}),AI{r}))',
        )
        ws_pr.cell(row=r, column=39, value=f'=IF({a}="",0,AL{r}*F{r})')
        ws_pr.cell(row=r, column=40, value=f'=IF({a}="",0,(AF{r}+AG{r})*E{r})')
        ws_pr.cell(row=r, column=41, value=f'=IF({a}="",0,AM{r}+AN{r})')
        ws_pr.cell(row=r, column=42, value=f'=IF({a}="",0,AO{r}*O{r})')

    money_cols = (
        7,
        8,
        9,
        10,
        11,
        12,
        13,
        16,
        17,
        27,
        31,
        32,
        33,
        34,
        35,
        36,
        37,
        38,
        39,
        40,
        41,
        42,
    )
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 15, 28, 29, 30):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV benefits (formula mechanics)", f"=SUM(AP{first}:AP{last_cap_row})"),
                (5, "PV expenses", f"=SUM(Q{first}:Q{last_cap_row})"),
                (
                    6,
                    "Σ l_end · v (annuity-style factor)",
                    f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})",
                ),
                (7, "PV total (ben+exp)", "=X4+X5"),
                (8, "Single premium (export)", f"={SHEET_INPUTS}!$B${_IN_ROW_PREMIUM}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    alm_layout = None
    alm_snap_for_book = None
    if alm_snapshot is not None:
        if alm_assumptions is None:
            raise ValueError("alm_assumptions is required when alm_snapshot is provided.")
        alm_snap_for_book = alm_excel_downsample_snapshot(alm_snapshot, int(ALM_ENGINE_STEP_MONTHS))
        alm_snap_for_book = alm_excel_truncate_snapshot(alm_snap_for_book, ALM_EXCEL_PATH_MONTH_CAP)
        # Liability column letters live in liability_layouts.LIABILITY_LAYOUTS
        # so the validator and parity tests can cross-check them. We pass the
        # string code rather than ProductType to avoid a registry-builder cycle.
        _rila_layout = liability_layout_for("rila")
        alm_layout = write_alm_projection_sheet(
            wb,
            alm_snap_for_book,
            alm_assumptions,
            n_months=int(res.months.shape[0]),
            y_last_row=int(y_last_row),
            engine_step_months=int(ALM_ENGINE_STEP_MONTHS),
            yield_curve_spread_ref=_in_addr("B", _IN_ROW_SPREAD),
            liability_total_col=_rila_layout.total_cf_col,
            liability_discount_col=_rila_layout.discount_col,
        )

    snap_py = ExcelPythonSnapshot(
        pv_benefit=float(res.pv_benefit),
        pv_monthly_expenses=float(res.pv_monthly_expenses),
        pv_monthly_total=float(res.pv_benefit + res.pv_monthly_expenses),
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    rila_rows: list[tuple[str, float, str, str]] = [
        ("PV benefits", float(res.pv_benefit), f"={LIABILITY_SHEET_NAME}!X4", "money"),
        (
            "PV monthly expenses",
            float(res.pv_monthly_expenses),
            f"={LIABILITY_SHEET_NAME}!X5",
            "money",
        ),
        (
            "PV monthly total (ben+exp)",
            float(res.pv_benefit + res.pv_monthly_expenses),
            f"={LIABILITY_SHEET_NAME}!X7",
            "money",
        ),
        ("Single premium", float(res.single_premium), f"={LIABILITY_SHEET_NAME}!X8", "money"),
        ("Annuity factor", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=alm_layout,
        alm_snapshot=alm_snap_for_book,
        pricing_rows=rila_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME}; optional ALM_Projection)",
        subtitle=(
            "RILA: column B is Python at export; column C references Liabilities summary. "
            "IndexScenario holds L[month] for end-of-month index levels. "
            "Premium in Inputs is the priced single premium from Python."
        ),
    )

    from annuity_model.excel_workbook_validator import validate_workbook_or_raise

    validate_workbook_or_raise(wb)

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if alm_snap_for_book is not None:
        data = inject_alm_projection_formula_cached_values(
            data,
            first_data_row=int(ALM_PROJECTION_FIRST_DATA_ROW),
            snap=alm_snap_for_book,
        )

    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
