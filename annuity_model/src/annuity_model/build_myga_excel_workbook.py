"""MYGA Excel workbook builder.

Emits a self-contained workbook with the standard skeleton:

* ``Inputs``        — single premium, declared rate, guarantee years, ...
* ``YieldCurve``    — copy of the maturity-rate curve
* ``MonthlyCurve_recalc`` — derived monthly discount factors
* ``Liabilities``   — per-month grid: month, age, survival, AV, claim,
                      expense, total CF, discount, PV columns
* ``ModelCheck``    — Python-vs-Excel reconciliation block

The MYGA cashflow shape (death CF + maturity CF + optional lapse CF)
matches the RILA accumulation pattern; ``total_cf_col=M``,
``discount_col=O`` (per :data:`liability_layouts.LIABILITY_LAYOUTS`).
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from annuity_model import myga_projection as my
from annuity_model import pricing_projection as sp
from annuity_model.excel_builder_helpers import (
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from annuity_model.excel_workbook_validator import validate_workbook_or_raise
from annuity_model.recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_MTH_QX = "MortalMonthly"

MYGA_PROJ_MAX_ROWS = 600

# Inputs row map. Row 6 = payment frequency (matches Term/RILA convention
# so the shared MonthlyCurve helper finds the same cell). Row 9 = spread.
_IN_ROW_ISSUE_AGE = 3
_IN_ROW_SINGLE_PREMIUM = 4
_IN_ROW_DECLARED_RATE = 5
_IN_ROW_FREQ = 6
_IN_ROW_VAL = 7
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_GUAR_YEARS = 10
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


def _n_months_cell() -> str:
    return _in_addr("B", _IN_ROW_NMONTHS)


@dataclass(frozen=True)
class MYGAExcelBuildSpec:
    contract: my.MYGAContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float


def myga_excel_spec_from_launcher(
    *,
    contract: my.MYGAContract,
    yield_curve: sp.YieldCurve,
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    expense_annual_inflation: float,
    **_kwargs: object,
) -> MYGAExcelBuildSpec:
    return MYGAExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
    )


def build_myga_workbook_from_spec(
    spec: MYGAExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    """Build the MYGA workbook bytes (and optionally write to disk)."""
    res = my.price_myga_single_premium(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=(
            spec.valuation_year if not isinstance(spec.mortality, sp.MortalityTableQx) else None
        ),
        expenses=spec.expenses,
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Single premium", float(spec.contract.single_premium)),
        ("Declared rate (annual)", float(spec.contract.declared_rate_annual)),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Guarantee years", int(spec.contract.guarantee_years)),
        ("Monthly expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(
        ws_in,
        InputsSheetSpec(
            title="MYGA Inputs (matches model launcher / Python)",
            rows=rows,
        ),
    )
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{_in_addr('B', _IN_ROW_GUAR_YEARS)}*{_in_addr('B', _IN_ROW_FREQ)})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=MYGA_PROJ_MAX_ROWS, y_last_row=y_last_row)

    # Mortality table sheet (Qx flavour or RP+MP cached annual qx) -- the
    # validator requires that any cross-sheet references resolve.
    if isinstance(spec.mortality, sp.MortalityTableQx):
        ws_q = wb.create_sheet(SHEET_QX)
        ws_q["A1"] = "age"
        ws_q["B1"] = "qx"
        ages = np.asarray(spec.mortality.ages, dtype=int)
        qx = np.asarray(spec.mortality.qx, dtype=float)
        for i in range(int(ages.shape[0])):
            r = 2 + i
            ws_q.cell(row=r, column=1, value=int(ages[i]))
            ws_q.cell(row=r, column=2, value=float(qx[i]))

    nm_ref = _n_months_cell()
    last_cap_row = 3 + MYGA_PROJ_MAX_ROWS
    first = 4
    sp_ref = _in_addr("B", _IN_ROW_SINGLE_PREMIUM)
    rate_ref = _in_addr("B", _IN_ROW_DECLARED_RATE)
    freq_ref = _in_addr("B", _IN_ROW_FREQ)
    issue_age_ref = _in_addr("B", _IN_ROW_ISSUE_AGE)
    exp_m_ref = _in_addr("B", 11)  # Monthly expense row inside Inputs
    exp_inf_ref = _in_addr("B", 12)
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"
    guar_months = int(spec.contract.guarantee_years) * 12

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "MYGA liability cashflows & pricing (formula-driven)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={issue_age_ref}"

    hdr = (
        "Month",  # A
        "t_years",  # B
        "AttainedAge",  # C
        "SurvivalEnd",  # D
        "SurvivalStart",  # E
        "MonthDeathProb",  # F
        "AV_end",  # G
        "DeathCF",  # H
        "MaturityCF",  # I
        "ExpBenefitCF",  # J
        "ExpExpenseCF",  # K
        "ExpTotalCF",  # L  <-- RILA-style accumulation: ExpTotalCF in column L
        "ExpTotalCFAlt",  # M  (mirror -- liability_layouts has total_cf_col="M")
        "",  # N
        "DiscountFactor",  # O
        "PVBenefitCF",  # P
        "PVExpenseCF",  # Q
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{freq_ref})')
        ws_pr.cell(
            row=r,
            column=3,
            value=f'=IF({a}="","",{issue_age_ref}+({a}-1)/{freq_ref})',
        )
        # SurvivalEnd: piecewise constant force from the qx table
        ws_pr.cell(
            row=r,
            column=4,
            value=(
                f'=IF({a}="","",IFERROR(POWER(1-MIN(MAX(IFERROR(INDEX({SHEET_QX}!$B$2:$B$200,'
                f"MATCH(INT({issue_age_ref}+({a}-1)/12),{SHEET_QX}!$A$2:$A$200,0)),0),0),0.999),"
                f"{a}/12)*1,0))"
            ),
        )
        if r == first:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",1)')
        else:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",D{r - 1})')
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        # AV_end: SP * (1+rate)^(t/12)
        ws_pr.cell(
            row=r,
            column=7,
            value=f'=IF({a}="","",{sp_ref}*POWER(1+{rate_ref},{a}/12))',
        )
        # DeathCF = AV * P(death this month)
        ws_pr.cell(row=r, column=8, value=f'=IF({a}="",0,G{r}*F{r})')
        # MaturityCF = AV[T]*survival_end[T-1] only on month T; else 0
        ws_pr.cell(
            row=r,
            column=9,
            value=f'=IF({a}="",0,IF({a}={guar_months},G{r}*D{r},0))',
        )
        # ExpBenefitCF = DeathCF + MaturityCF
        ws_pr.cell(row=r, column=10, value=f'=IF({a}="",0,H{r}+I{r})')
        # Expense (monthly $ * inflation factor * survival_start)
        ws_pr.cell(
            row=r,
            column=11,
            value=f'=IF({a}="",0,{exp_m_ref}*POWER(1+(POWER(1+{exp_inf_ref},1/12)-1),{a}-1)*E{r})',
        )
        # ExpTotalCF (column L) = J + K  (benefit + expense)
        ws_pr.cell(row=r, column=12, value=f'=IF({a}="",0,J{r}+K{r})')
        # ExpTotalCFAlt (column M, the layout-canonical column) = L mirror
        ws_pr.cell(row=r, column=13, value=f'=IF({a}="",0,L{r})')
        # Discount factor from MonthlyCurve_recalc
        ws_pr.cell(
            row=r,
            column=15,
            value=(
                f'=IF({a}="","",IFERROR(INDEX({mc_ref},'
                f'MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))'
            ),
        )
        # PV(benefit), PV(expense)
        ws_pr.cell(row=r, column=16, value=f'=IF({a}="",0,J{r}*O{r})')
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,K{r}*O{r})')

    money_cols = (7, 8, 9, 10, 11, 12, 13, 16, 17)
    ratio_cols = (2, 3, 4, 5, 6, 15)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in ratio_cols:
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV benefits", f"=SUM(P{first}:P{last_cap_row})"),
                (5, "PV expenses", f"=SUM(Q{first}:Q{last_cap_row})"),
                (
                    6,
                    "Σ l_end · v (annuity-style factor)",
                    f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})",
                ),
                (7, "PV total (ben+exp)", "=X4+X5"),
                (8, "Single premium (input)", f"={sp_ref}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    sp_value = float(res.single_premium)

    from annuity_model.excel_builder_helpers import ExcelPythonSnapshot

    snap_py = ExcelPythonSnapshot(
        pv_benefit=pv_b,
        pv_monthly_expenses=pv_e,
        pv_monthly_total=pv_t,
        single_premium=sp_value,
        annuity_factor=float(res.annuity_factor),
    )
    rows_mc: list[tuple[str, float, str, str]] = [
        ("PV benefits", pv_b, f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV expenses", pv_e, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV total (ben+exp)", pv_t, f"={LIABILITY_SHEET_NAME}!X7", "money"),
        ("Single premium", sp_value, f"={LIABILITY_SHEET_NAME}!X8", "money"),
        (
            "Annuity factor (Σ l_end · v)",
            float(res.annuity_factor),
            f"={LIABILITY_SHEET_NAME}!X6",
            "factor",
        ),
    ]
    write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=None,
        alm_snapshot=None,
        pricing_rows=rows_mc,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=(
            "MYGA: column B is Python at export; column C references Liabilities summary. "
            "Single premium is the input; Excel reproduces AV * survival per month and "
            "the maturity payout at month T = guarantee_years * 12."
        ),
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
