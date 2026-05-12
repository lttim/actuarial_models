"""Workbook sheets that carry run-ledger and assumption evidence."""

from __future__ import annotations

import json
from collections.abc import Mapping, Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

RUN_LEDGER_SHEET = "RunLedger"
ASSUMPTION_EVIDENCE_SHEET = "AssumptionEvidence"


def _json_cell(value: Any) -> str:
    return json.dumps(value, sort_keys=True, default=str, separators=(",", ":"))


def _replace_sheet(wb: Workbook, title: str) -> Any:
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def write_run_evidence_sheets(
    wb: Workbook,
    run_summary: Mapping[str, Any] | None,
) -> None:
    """Write ledger and assumption evidence sheets when a run summary exists."""
    if not run_summary:
        return

    ws = _replace_sheet(wb, RUN_LEDGER_SHEET)
    ws["A1"] = "Run ledger summary"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    fields = [
        "run_id",
        "product",
        "scenario_id",
        "created_at",
        "input_hash",
        "parity_status",
        "validation_status",
        "waiver_status",
        "model_version",
        "git_commit",
    ]
    row = 3
    for field in fields:
        if field in run_summary:
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=str(run_summary[field]))
            row += 1
    for field in ("output_metrics", "output_paths", "metadata", "assumption_evidence"):
        if field in run_summary:
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=_json_cell(run_summary[field]))
            row += 1

    rows = run_summary.get("assumption_artifacts")
    if isinstance(rows, Sequence) and not isinstance(rows, (str, bytes, bytearray)):
        _write_assumption_evidence_sheet(wb, rows)


def _write_assumption_evidence_sheet(wb: Workbook, rows: Sequence[Any]) -> None:
    ws = _replace_sheet(wb, ASSUMPTION_EVIDENCE_SHEET)
    headers = (
        "role",
        "mode",
        "artifact_name",
        "version",
        "status",
        "intended_use",
        "approval_id",
        "requires_waiver_for_release",
        "warning",
        "path",
        "sha256",
    )
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    for row_idx, raw in enumerate(rows, start=2):
        item = raw if isinstance(raw, Mapping) else {"artifact_name": str(raw)}
        for col, header in enumerate(headers, start=1):
            value = item.get(header)
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            ws.cell(row=row_idx, column=col, value=value)


__all__ = [
    "ASSUMPTION_EVIDENCE_SHEET",
    "RUN_LEDGER_SHEET",
    "write_run_evidence_sheets",
]
