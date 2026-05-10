"""VUL Excel workbook builder.

Same structure as UL but with the sub-account return as the credit
source. AV path is exported as Python literals; per-month cashflows
use Excel formulas referencing the literals.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from annuity_model import pricing_projection as sp
from annuity_model import vul_projection as vul
from annuity_model.excel_builder_helpers import (
    ExcelPythonSnapshot,
    InputsSheetSpec,
    LiabilitySummaryBlockSpec,
    write_inputs_sheet,
    write_liability_summary_block,
    write_model_check_sheet,
)
from annuity_model.excel_workbook_validator import validate_workbook_or_raise
from annuity_model.mortality_2017_cso import MortalityTable2017CSO
from annuity_model.recalc_excel_shared import (
    RECALC_MONTHLY_CURVE_SHEET,
    write_monthly_curve_logdf,
    write_yield_curve_sheet,
)

LIABILITY_SHEET_NAME = "Liabilities"
SHEET_INPUTS = "Inputs"
SHEET_QX = "QxTable"
SHEET_INDEX = "IndexScenario"
VUL_PROJ_MAX_ROWS = 1000

_IN_ROW_ISSUE_AGE = 3
_IN_ROW_FACE = 4
_IN_ROW_SP = 5
_IN_ROW_FREQ = 6
_IN_ROW_VAL = 7
_IN_ROW_HORIZON = 8
_IN_ROW_SPREAD = 9
_IN_ROW_LOAD = 10
_IN_ROW_EXP_CHG = 11
_IN_ROW_NMONTHS = 18


def _in_addr(col: str, row: int) -> str:
    return f"{SHEET_INPUTS}!${col}${row}"


@dataclass(frozen=True)
class VULExcelBuildSpec:
    contract: vul.VULContract
    yield_curve: sp.YieldCurve
    mortality: sp.MortalityTableQx | sp.MortalityTableRP2014MP2016 | MortalityTable2017CSO | None
    horizon_age: int
    spread: float
    valuation_year: int
    expenses: sp.ExpenseAssumptions
    yield_mode_label: str
    mortality_mode_label: str
    expense_mode_label: str
    expense_annual_inflation: float
    index_s0: float
    index_levels_payment: np.ndarray


def vul_excel_spec_from_launcher(
    *,
    contract: vul.VULContract,
    yield_curve: sp.YieldCurve,
    mortality,
    horizon_age: int,
    spread: float,
    valuation_year: int,
    expenses: sp.ExpenseAssumptions,
    yield_mode_label: str,
    mortality_mode_label: str,
    expense_mode_label: str,
    index_s0: float,
    index_levels_at_payment: np.ndarray,
    expense_annual_inflation: float,
    **_kw: object,
) -> VULExcelBuildSpec:
    return VULExcelBuildSpec(
        contract=contract,
        yield_curve=yield_curve,
        mortality=mortality,
        horizon_age=int(horizon_age),
        spread=float(spread),
        valuation_year=int(valuation_year),
        expenses=expenses,
        yield_mode_label=str(yield_mode_label),
        mortality_mode_label=str(mortality_mode_label),
        expense_mode_label=str(expense_mode_label),
        expense_annual_inflation=float(expense_annual_inflation),
        index_s0=float(index_s0),
        index_levels_payment=np.asarray(index_levels_at_payment, dtype=float),
    )


def build_vul_workbook_from_spec(
    spec: VULExcelBuildSpec,
    *,
    out_path: str | Path | None = None,
) -> bytes:
    res = vul.price_vul_single_premium(
        contract=spec.contract,
        yield_curve=spec.yield_curve,
        mortality=spec.mortality,
        horizon_age=spec.horizon_age,
        spread=spec.spread,
        valuation_year=(
            spec.valuation_year
            if isinstance(spec.mortality, sp.MortalityTableRP2014MP2016)
            else None
        ),
        expenses=spec.expenses,
        index_s0=float(spec.index_s0),
        index_levels_payment=np.asarray(spec.index_levels_payment, dtype=float),
        expense_annual_inflation=float(spec.expense_annual_inflation),
    )

    mort_for_qx = spec.mortality
    if mort_for_qx is None:
        mort_for_qx = MortalityTable2017CSO.load(
            sex=spec.contract.sex, smoker_class=spec.contract.smoker_class
        )
    if isinstance(mort_for_qx, MortalityTable2017CSO):
        qx_table = mort_for_qx.table
    elif isinstance(mort_for_qx, sp.MortalityTableRP2014MP2016):
        ages = np.asarray(mort_for_qx.base_qx_2014.ages, dtype=int)
        qx_arr = np.array(
            [
                float(
                    mort_for_qx.qx_at_int_age_and_calendar_year(
                        age_int=int(a),
                        calendar_year=int(spec.valuation_year),
                    )
                )
                for a in ages
            ],
            dtype=float,
        )
        qx_table = sp.MortalityTableQx(ages, qx_arr)
    else:
        qx_table = mort_for_qx

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = SHEET_INPUTS
    rows = [
        ("Issue age", spec.contract.issue_age),
        ("Face amount", float(spec.contract.face_amount)),
        ("Single premium", float(spec.contract.single_premium)),
        ("Payment frequency (per year)", 12),
        ("Valuation year", spec.valuation_year),
        ("Horizon age", spec.horizon_age),
        ("Spread (added to zero rate)", spec.spread),
        ("Premium load (decimal)", float(spec.contract.premium_load_pct)),
        ("Monthly expense charge ($)", float(spec.contract.monthly_expense_charge)),
        ("Sub-account drift (annual)", float(spec.contract.subaccount_drift_annual)),
        ("Sub-account vol (annual)", float(spec.contract.subaccount_vol_annual)),
        ("Smoker class", spec.contract.smoker_class),
        ("Sex", spec.contract.sex),
        ("Monthly maintenance expense ($)", float(spec.expenses.monthly_expense_dollars)),
        ("Expense annual inflation", float(spec.expense_annual_inflation)),
        ("Yield mode (documentation)", spec.yield_mode_label),
        ("Mortality mode (documentation)", spec.mortality_mode_label),
        ("Expense mode (documentation)", spec.expense_mode_label),
    ]
    write_inputs_sheet(
        ws_in,
        InputsSheetSpec(title="Variable UL Inputs (matches model launcher / Python)", rows=rows),
    )
    nm = (
        f"=MIN(MAX(1,ROUND(({_in_addr('B', _IN_ROW_HORIZON)}"
        f"-{_in_addr('B', _IN_ROW_ISSUE_AGE)})*{_in_addr('B', _IN_ROW_FREQ)},0)),"
        f"{VUL_PROJ_MAX_ROWS})"
    )
    ws_in[f"A{_IN_ROW_NMONTHS}"] = "Model months (formula)"
    ws_in[f"B{_IN_ROW_NMONTHS}"] = nm
    ws_in[f"B{_IN_ROW_NMONTHS}"].number_format = "0"

    ycdf = pd.DataFrame(
        {
            "maturity_years": np.asarray(spec.yield_curve.maturities_years, dtype=float),
            "zero_rate": np.asarray(spec.yield_curve.zero_rates, dtype=float),
        }
    )
    _, y_last_row = write_yield_curve_sheet(wb, ycdf)

    ws_mc_curve = wb.create_sheet(RECALC_MONTHLY_CURVE_SHEET)
    write_monthly_curve_logdf(ws_mc_curve, n_months=VUL_PROJ_MAX_ROWS, y_last_row=y_last_row)

    ws_q = wb.create_sheet(SHEET_QX)
    ws_q["A1"] = "age"
    ws_q["B1"] = "qx"
    ages = np.asarray(qx_table.ages, dtype=int)
    qx_arr = np.asarray(qx_table.qx, dtype=float)
    for i in range(int(ages.shape[0])):
        r = 2 + i
        ws_q.cell(row=r, column=1, value=int(ages[i]))
        ws_q.cell(row=r, column=2, value=float(qx_arr[i]))

    ws_ix = wb.create_sheet(SHEET_INDEX)
    ws_ix["A1"] = "month"
    ws_ix["B1"] = "index_level"
    s0_py = float(spec.index_s0)
    levels_py = np.asarray(spec.index_levels_payment, dtype=float)
    L = np.zeros(levels_py.shape[0] + 1, dtype=float)
    L[0] = s0_py
    L[1:] = levels_py
    for j in range(L.shape[0]):
        ws_ix.cell(row=2 + j, column=1, value=int(j))
        ws_ix.cell(row=2 + j, column=2, value=float(L[j]))

    nm_ref = _in_addr("B", _IN_ROW_NMONTHS)
    last_cap_row = 3 + VUL_PROJ_MAX_ROWS
    first = 4
    issue_age_ref = _in_addr("B", _IN_ROW_ISSUE_AGE)
    freq_ref = _in_addr("B", _IN_ROW_FREQ)
    sp_ref = _in_addr("B", _IN_ROW_SP)
    exp_m_ref = _in_addr("B", 14)
    exp_inf_ref = _in_addr("B", 15)
    mc_ref = f"{RECALC_MONTHLY_CURVE_SHEET}!$L:$L"

    ws_pr = wb.create_sheet(LIABILITY_SHEET_NAME)
    ws_pr["A1"] = "Variable UL liability cashflows & pricing (formula-driven where tractable)"
    ws_pr["A1"].font = Font(bold=True, size=12)
    ws_pr["A2"] = "ReserveAtT0"
    ws_pr["B2"] = 0
    ws_pr["C2"] = f"={issue_age_ref}"

    hdr = (
        "Month",  # A
        "t_years",  # B
        "AttainedAge",  # C
        "SurvivalEnd",  # D
        "SurvivalStart",  # E
        "MonthDeathProb",  # F
        "qx_monthly",  # G
        "IndexLevel_m",  # H
        "MonthlySubReturn",  # I  (formula: H[r]/H[r-1] - 1)
        "AV_end_PY",  # J  (Python literal)
        "DB_PY",  # K  (Python literal)
        "DeathCF",  # L  (formula: K * F)
        "",  # M
        "",  # N
        "DiscountFactor",  # O
        "",  # P
        "ExpBenefitCF",  # Q
        "ExpExpenseCF",  # R
        "ExpTotalCF",  # S
        "PVBenefitCF",  # T
        "PVExpenseCF",  # U
        "PVNetOutflow",  # V
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_pr.cell(row=3, column=c, value=h if h else None)
        cell.font = Font(bold=True)

    av_end_arr = np.asarray(res.account_value_end_month, dtype=float)
    db_arr = np.asarray(res.db_end_month, dtype=float)
    n_months_used = int(av_end_arr.shape[0])

    for r in range(first, last_cap_row + 1):
        a = f"A{r}"
        m_idx = r - first
        if m_idx < n_months_used:
            av_py = float(av_end_arr[m_idx])
            db_py = float(db_arr[m_idx])
        else:
            av_py = 0.0
            db_py = 0.0
        ws_pr.cell(row=r, column=1, value=f'=IF(ROW()-3>{nm_ref},"",ROW()-3)')
        ws_pr.cell(row=r, column=2, value=f'=IF({a}="","",{a}/{freq_ref})')
        ws_pr.cell(row=r, column=3, value=f'=IF({a}="","",{issue_age_ref}+({a}-1)/{freq_ref})')
        ws_pr.cell(
            row=r,
            column=4,
            value=(
                f'=IF({a}="","",IFERROR(POWER(1-MIN(MAX(IFERROR(INDEX({SHEET_QX}!$B$2:$B$200,'
                f"MATCH(INT({issue_age_ref}+({a}-1)/12),{SHEET_QX}!$A$2:$A$200,0)),0),0),0.999),"
                f"{a}/12)*1,0))"
            ),
        )
        if r == first:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",1)')
        else:
            ws_pr.cell(row=r, column=5, value=f'=IF({a}="","",D{r - 1})')
        ws_pr.cell(row=r, column=6, value=f'=IF({a}="","",MAX(0,MIN(1,E{r}-D{r})))')
        ws_pr.cell(row=r, column=7, value=f'=IF({a}="","",IF(E{r}>0,(E{r}-D{r})/E{r},0))')
        ws_pr.cell(
            row=r,
            column=8,
            value=f'=IF({a}="","",IFERROR(INDEX({SHEET_INDEX}!$B$2:$B$10000,MATCH({a},{SHEET_INDEX}!$A$2:$A$10000,0)),""))',
        )
        ws_pr.cell(
            row=r,
            column=9,
            value=(
                f'=IF({a}="","",IF({a}=1,'
                f"IFERROR(INDEX({SHEET_INDEX}!$B$2:$B$10000,MATCH(1,{SHEET_INDEX}!$A$2:$A$10000,0))/"
                f"INDEX({SHEET_INDEX}!$B$2:$B$10000,MATCH(0,{SHEET_INDEX}!$A$2:$A$10000,0))-1,0),"
                f"IFERROR(INDEX({SHEET_INDEX}!$B$2:$B$10000,MATCH({a},{SHEET_INDEX}!$A$2:$A$10000,0))/"
                f"INDEX({SHEET_INDEX}!$B$2:$B$10000,MATCH({a}-1,{SHEET_INDEX}!$A$2:$A$10000,0))-1,0)))"
            ),
        )
        ws_pr.cell(row=r, column=10, value=float(av_py))
        ws_pr.cell(row=r, column=11, value=float(db_py))
        ws_pr.cell(row=r, column=12, value=f'=IF({a}="",0,K{r}*F{r})')
        ws_pr.cell(
            row=r,
            column=15,
            value=f'=IF({a}="","",IFERROR(INDEX({mc_ref},MATCH({a},{RECALC_MONTHLY_CURVE_SHEET}!$A:$A,0)),""))',
        )
        ws_pr.cell(row=r, column=17, value=f'=IF({a}="",0,L{r})')
        ws_pr.cell(
            row=r,
            column=18,
            value=f'=IF({a}="",0,{exp_m_ref}*POWER(1+(POWER(1+{exp_inf_ref},1/12)-1),{a}-1)*E{r})',
        )
        ws_pr.cell(row=r, column=19, value=f'=IF({a}="",0,Q{r}+R{r})')
        ws_pr.cell(row=r, column=20, value=f'=IF({a}="",0,Q{r}*O{r})')
        ws_pr.cell(row=r, column=21, value=f'=IF({a}="",0,R{r}*O{r})')
        ws_pr.cell(row=r, column=22, value=f'=IF({a}="",0,S{r}*O{r})')

    money_cols = (8, 10, 11, 12, 17, 18, 19, 20, 21, 22)
    for r in range(first, last_cap_row + 1):
        for c in money_cols:
            ws_pr.cell(row=r, column=c).number_format = "#,##0.00"
        for c in (2, 3, 4, 5, 6, 7, 9, 15):
            ws_pr.cell(row=r, column=c).number_format = "0.000000"

    write_liability_summary_block(
        ws_pr,
        LiabilitySummaryBlockSpec(
            rows=(
                (4, "PV claims (DB × death-prob)", f"=SUM(T{first}:T{last_cap_row})"),
                (5, "PV expenses", f"=SUM(U{first}:U{last_cap_row})"),
                (
                    6,
                    "Σ l_end · v (annuity-style factor)",
                    f"=SUMPRODUCT(D{first}:D{last_cap_row},O{first}:O{last_cap_row})",
                ),
                (7, "PV total (claims + expenses)", "=X4+X5"),
                (8, "Single premium (input)", f"={sp_ref}"),
                (9, "Reserve at t=0", "=X7"),
            ),
        ),
    )

    pv_b = float(np.sum(res.expected_benefit_cashflows * res.discount_factors))
    pv_e = float(np.sum(res.expected_expense_cashflows * res.discount_factors))
    pv_t = float(np.sum(res.expected_total_cashflows * res.discount_factors))
    snap_py = ExcelPythonSnapshot(
        pv_benefit=pv_b,
        pv_monthly_expenses=pv_e,
        pv_monthly_total=pv_t,
        single_premium=float(res.single_premium),
        annuity_factor=float(res.annuity_factor),
    )
    vul_rows = [
        ("PV claims", pv_b, f"={LIABILITY_SHEET_NAME}!X4", "money"),
        ("PV expenses", pv_e, f"={LIABILITY_SHEET_NAME}!X5", "money"),
        ("PV total (claims + expenses)", pv_t, f"={LIABILITY_SHEET_NAME}!X7", "money"),
        (
            "Single premium (input)",
            float(res.single_premium),
            f"={LIABILITY_SHEET_NAME}!X8",
            "money",
        ),
        ("Annuity factor", float(res.annuity_factor), f"={LIABILITY_SHEET_NAME}!X6", "factor"),
    ]
    write_model_check_sheet(
        wb,
        snap_py,
        alm_layout=None,
        alm_snapshot=None,
        pricing_rows=vul_rows,
        sheet_title=f"Python snapshot vs Excel ({LIABILITY_SHEET_NAME})",
        subtitle=(
            "VUL: like UL but with monthly sub-account return as the credit. "
            "AV path is exported as Python literals (the source of truth); "
            "DeathCF = DB × death-prob this month."
        ),
    )

    validate_workbook_or_raise(wb)
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out_path is not None:
        Path(out_path).write_bytes(data)
    return data
