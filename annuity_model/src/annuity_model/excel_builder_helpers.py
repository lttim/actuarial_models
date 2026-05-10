"""Public surface for Excel-builder helpers shared across products.

Historically the Term and RILA builders reached into the SPIA builder for
``_write_alm_projection_sheet``, ``_write_model_check_sheet``,
``alm_excel_downsample_snapshot`` and friends. The leading underscore was a
lie -- these helpers are part of the cross-product Excel-build contract --
and the cross-imports made the dependency direction unobvious.

This shim re-exports the same callables under explicit public names. New
builders SHOULD import from here. The original ``_`` names in
:mod:`build_pricing_excel_workbook` are kept temporarily for
backwards-compatibility but should be considered private; remove them once
all callers move over.

P1 (2026-04) added the declarative writers below
(:class:`InputsSheetSpec`, :class:`LiabilitySummaryBlockSpec`,
:func:`write_inputs_sheet`, :func:`write_liability_summary_block`). They
collapse the three near-identical chunks of code that previously lived
inline in the SPIA / Term / RILA builders. New products MUST use these
helpers; the old inline copies were bug magnets (a column letter typo in
one builder went unnoticed for two months in 2026-Q1).

Phase 3 will fold these into a real ``BaseProductBuilder`` class with an
overridable ``liability_layout``.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from typing import Any

from openpyxl.styles import Font

# ALM ladder + downsampler (the SPIA builder owns the canonical impl)
from annuity_model.alm_excel_ladder import ALM_ENGINE_SHEET, write_alm_engine_sheet
from annuity_model.build_pricing_excel_workbook import (
    ALM_ENGINE_STEP_MONTHS,
    ALM_EXCEL_PATH_MONTH_CAP,
    ALM_PROJECTION_FIRST_DATA_ROW,
    LIABILITY_SHEET_NAME,
    ALMExcelSnapshot,
    ExcelPythonSnapshot,
    alm_excel_downsample_snapshot,
    alm_excel_truncate_snapshot,
    inject_alm_projection_formula_cached_values,
)

# Public aliases for the historically-`_private` helpers. Builders MUST use
# these names so a future internal rename does not silently break them.
from annuity_model.build_pricing_excel_workbook import (
    _write_alm_projection_sheet as write_alm_projection_sheet,
)
from annuity_model.build_pricing_excel_workbook import (
    _write_model_check_sheet as write_model_check_sheet,
)

# Layouts + validator: convenience re-exports so a builder needs only one
# import line to pull in the entire shared toolkit.
from annuity_model.excel_workbook_validator import validate_workbook, validate_workbook_or_raise
from annuity_model.liability_layouts import LIABILITY_LAYOUTS, LiabilityLayout, liability_layout_for

# ---------------------------------------------------------------------------
# Declarative Inputs-sheet writer (P1, 2026-04).
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class InputsSheetSpec:
    """Declarative description of the canonical "Inputs" sheet.

    Every product workbook has the same skeleton:

    * A1 holds a bold 12pt title.
    * Rows starting at row 3 are (label, value) in columns A and B.
    * Optional "Notes" pane in column D, header at D3.

    Extra rows that aren't simple key-value pairs (e.g. SPIA's "Derived"
    block, Term's formula-driven "Model months" row) stay inline in the
    builder; the helper only owns the boilerplate the three builders
    have always shared verbatim.

    Attributes
    ----------
    title:
        Cell A1 contents (bold 12pt).
    rows:
        Sequence of ``(label, value)`` tuples written starting at row 3.
        ``value`` may be a Python scalar or an Excel formula string
        (``=...``); openpyxl handles both.
    notes:
        Optional sequence of note strings written into column D starting
        at D3. Pass an empty sequence (or omit) to skip the Notes pane.
    notes_header_row:
        Row at which to write "Notes" header in column D. Defaults to 3.
    notes_first_row:
        Row at which to start writing the note bodies. Defaults to 4.

    Examples
    --------
    >>> spec = InputsSheetSpec(
    ...     title="SPIA Inputs (matches model launcher / Python)",
    ...     rows=(("Issue Age", 65), ("Sex", "male")),
    ...     notes=("Rates are decimals.", "Recalculate after edits."),
    ... )
    >>> write_inputs_sheet(ws, spec)
    """

    title: str
    rows: Sequence[tuple[str, object]]
    notes: Sequence[str] = ()
    notes_header_row: int = 3
    notes_first_row: int = 4
    notes_column: str = "D"


def write_inputs_sheet(ws: Any, spec: InputsSheetSpec) -> None:
    """Render *spec* onto worksheet *ws*.

    Sets ``ws.title`` to ``"Inputs"`` (the canonical name; if the caller
    needs a different title they should set it before invoking this
    helper -- ``write_inputs_sheet`` only owns the contents, not the
    sheet name in that case). Bold 12pt is applied to A1.

    The helper is intentionally minimal: it does NOT format value cells
    (number formats vary per builder and are applied by the caller after
    the fact), and it does NOT clear any existing cells. This matches
    the previous inline behavior so the refactor is byte-for-byte
    equivalent.
    """
    ws["A1"] = spec.title
    ws["A1"].font = Font(bold=True, size=12)
    for i, (label, value) in enumerate(spec.rows, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
    if spec.notes:
        col = spec.notes_column
        ws[f"{col}{spec.notes_header_row}"] = "Notes"
        for j, note in enumerate(spec.notes, start=spec.notes_first_row):
            ws[f"{col}{j}"] = note


# ---------------------------------------------------------------------------
# Declarative Liabilities-sheet "Summary" block writer (P1, 2026-04).
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class LiabilitySummaryBlockSpec:
    """Declarative description of the W/X "Summary" block on the
    Liabilities sheet shared by all three product builders.

    The block looks like::

        W3   "Summary"   (bold)
        W4   <label>     X4 <formula>
        W5   <label>     X5 <formula>
        ...

    Each builder previously hand-wrote nine assignment statements per
    product. Centralizing the block-shape here means a future change
    (e.g. adding a "Reserve sensitivity" row) lands in one place.

    Attributes
    ----------
    label_column:
        Column letter for labels (canonical: ``"W"``).
    value_column:
        Column letter for values/formulas (canonical: ``"X"``).
    header_row:
        Row at which "Summary" header is written. Defaults to 3.
    rows:
        Sequence of ``(row, label, value)`` tuples. ``value`` is usually
        an Excel formula string (``"=SUM(T4:T100)"``).
    """

    rows: Sequence[tuple[int, str, str]]
    label_column: str = "W"
    value_column: str = "X"
    header_row: int = 3


def write_liability_summary_block(ws: Any, spec: LiabilitySummaryBlockSpec) -> None:
    """Render the summary block described by *spec* onto *ws*.

    The "Summary" header at ``{label_column}{header_row}`` is bold; the
    body cells take whatever font the workbook already has set (matches
    the previous inline behavior).
    """
    header_cell = f"{spec.label_column}{spec.header_row}"
    ws[header_cell] = "Summary"
    ws[header_cell].font = Font(bold=True)
    for row, label, value in spec.rows:
        ws[f"{spec.label_column}{row}"] = label
        ws[f"{spec.value_column}{row}"] = value


def write_kv_rows(ws: Any, rows: Iterable[tuple[str, object]], *, start_row: int = 3) -> int:
    """Lower-level helper: write ``(label, value)`` pairs into A/B columns.

    Useful when a builder needs the row-loop without the title + Notes
    machinery of :func:`write_inputs_sheet` (e.g. extending an existing
    sheet with extra rows). Returns the row index *after* the last
    written row.
    """
    last = start_row - 1
    for i, (label, value) in enumerate(rows, start=start_row):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        last = i
    return last + 1


__all__ = [
    "ALM_ENGINE_SHEET",
    "ALM_ENGINE_STEP_MONTHS",
    "ALM_EXCEL_PATH_MONTH_CAP",
    "ALM_PROJECTION_FIRST_DATA_ROW",
    "ALMExcelSnapshot",
    "ExcelPythonSnapshot",
    "InputsSheetSpec",
    "LIABILITY_LAYOUTS",
    "LIABILITY_SHEET_NAME",
    "LiabilityLayout",
    "LiabilitySummaryBlockSpec",
    "alm_excel_downsample_snapshot",
    "alm_excel_truncate_snapshot",
    "inject_alm_projection_formula_cached_values",
    "liability_layout_for",
    "validate_workbook",
    "validate_workbook_or_raise",
    "write_alm_engine_sheet",
    "write_alm_projection_sheet",
    "write_inputs_sheet",
    "write_kv_rows",
    "write_liability_summary_block",
    "write_model_check_sheet",
]
