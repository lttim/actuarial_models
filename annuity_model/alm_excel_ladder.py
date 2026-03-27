"""
Excel-native Treasury ladder (ALM_Engine). Mirrors ``run_alm_projection`` for
``rebalance_policy == liquidity_only``. Uses only common worksheet functions
(INDEX/MATCH/IF/EXP) so desktop Excel recognizes formulas — no LET/LAMBDA.

Pro-rata reinvestment matches Python: it runs only in months with bond maturities and
no outstanding borrowing after the first repayment leg — not whenever cash exceeds target.

See ``write_alm_engine_sheet`` for supported reinvest / disinvest / borrowing modes.
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass

import numpy as np
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import pricing_projection as sp

ALM_ENGINE_SHEET = "ALM_Engine"


def _L(c: int) -> str:
    return get_column_letter(c)


def _excel_df_flat(*, t_cell: str, y_last_row: int) -> str:
    """df(t)=exp(-(z+s)*t) with log-linear z between YieldCurve nodes; Excel 2013+."""
    ya = f"YieldCurve!$A$4:$A${y_last_row}"
    zb = f"YieldCurve!$B$4:$B${y_last_row}"
    sp = "Inputs!$B$9"
    br = (
        f"IF({t_cell}<=INDEX({ya},1),1,"
        f"IF({t_cell}>=INDEX({ya},ROWS({ya})),ROWS({ya})-1,MATCH({t_cell},{ya},1)))"
    )
    lo = f"INDEX({ya},{br})"
    hi = f"INDEX({ya},({br})+1)"
    zlo = f"INDEX({zb},{br})"
    zhi = f"INDEX({zb},({br})+1)"
    w = (
        f"IF(ABS(({hi})-({lo}))<1E-15,0,"
        f"(({t_cell})-({lo}))/MAX(ABS(({hi})-({lo})),1E-15))"
    )
    ldf_lo = f"(-(({zlo})+{sp})*({lo}))"
    ldf_hi = f"(-(({zhi})+{sp})*({hi}))"
    ldf_mid = f"({ldf_lo})+({w})*((({ldf_hi})-({ldf_lo})))"
    ldf = (
        f"IF({t_cell}<=INDEX({ya},1),-(INDEX({zb},1)+{sp})*{t_cell},"
        f"IF({t_cell}>=INDEX({ya},ROWS({ya})),-(INDEX({zb},ROWS({zb}))+{sp})*{t_cell},"
        f"{ldf_mid}))"
    )
    return f"EXP({ldf})"


@dataclass(frozen=True)
class ALMEngineLayout:
    first_data_row: int
    last_data_row: int
    col_mv_cash: int
    col_mv_bond_start: int
    col_debt_eom: int
    n_bonds: int
    # (Excel column letter, grid header text, full definition) sorted by column index
    field_guide_rows: tuple[tuple[str, str, str], ...]


def write_alm_engine_sheet(
    ws,
    *,
    period_end_months_1based: list[int],
    n_projection_months: int,
    y_last_row: int,
    asm: sp.ALMAssumptions,
    initial_aum: float,
    snap_bucket_names: tuple[str, ...],
    engine_step_months: int = 1,
    liability_sheet_name: str = "Liabilities",
) -> ALMEngineLayout:
    if asm.rebalance_policy != "liquidity_only":
        raise ValueError(
            "Excel ALM ladder requires rebalance_policy='liquidity_only'. "
            "Select Hold-to-maturity bias in ALM settings for this export."
        )
    if not period_end_months_1based:
        raise ValueError("period_end_months_1based must be non-empty.")
    if period_end_months_1based[-1] > n_projection_months or period_end_months_1based[0] < 1:
        raise ValueError(
            f"period_end_months_1based out of range 1..{n_projection_months}: "
            f"{period_end_months_1based[0]}..{period_end_months_1based[-1]}"
        )

    variable_period = int(engine_step_months) > 1
    buckets = asm.allocation.buckets
    w = np.asarray(asm.allocation.weights, dtype=float)
    nb = len(buckets) - 1
    if nb < 1:
        raise ValueError("Need at least one bond bucket besides cash.")
    if tuple(str(b.name) for b in buckets) != snap_bucket_names:
        raise ValueError("Allocation buckets must match ALMExcelSnapshot.bucket_names.")

    sh = str(liability_sheet_name)

    nom = np.array([float(buckets[k + 1].tenor_years) for k in range(nb)], dtype=float)
    wsum_b = float(np.sum(w[1:]))
    wnorm = [float(w[k + 1]) / wsum_b for k in range(nb)] if wsum_b > 1e-15 else [1.0 / nb] * nb

    ws["A1"] = f"ALM_Engine — Treasury ladder (YieldCurve + {sh} cash outflows)"
    ws["A1"].font = Font(bold=True, size=12)
    if variable_period:
        ws["A2"] = (
            f"Multi-month rows: **dt_y** = period length (yr); **cf** = SUM({sh} col ExpTotalCF) in period. "
            "Data grid starts just below bond parameters. liquidity_only."
        )
        ws["A6"], ws["B6"] = ("Δt note", "(use dt_y col; B6 unused)")
    else:
        ws["A2"] = (
            f"Monthly rows; **B6** = Δt (yr); **cf** = liability outflow for that month ({sh} col ExpTotalCF). "
            "Collapse column groups for ladder detail. liquidity_only."
        )
        ws["A6"], ws["B6"] = "Δt (years)", 1.0 / 12.0
    ws.merge_cells("A2:J2")

    ws["A5"], ws["B5"] = "Initial AUM ($)", float(initial_aum)
    ws["A7"], ws["B7"] = "Borrow: 1=scenario-linked rate, 0=fixed annual", int(
        1 if asm.borrowing_rate_mode == "scenario_linked" else 0
    )
    ws["A8"], ws["B8"] = (
        "Tenor (y) if scenario / fixed annual borrow rate (dec.)",
        float(asm.borrowing_rate_tenor_years if asm.borrowing_rate_mode == "scenario_linked" else asm.borrowing_rate_annual),
    )
    ws["A9"], ws["B9"] = "Borrow spread (dec.) if scenario-linked", float(asm.borrowing_spread_annual)
    ws["A10"], ws["B10"] = "1 = borrow before selling", int(1 if asm.borrowing_policy == "borrow_before_selling" else 0)
    ws["A11"], ws["B11"] = "1 = reinvest maturities pro_rata", int(1 if asm.reinvest_rule == "pro_rata" else 0)
    ws["A12"], ws["B12"] = "1 = disinvest shortest tenor first; 0 = pro_rata MV", int(
        1 if asm.disinvest_rule == "shortest_first" else 0
    )

    df_borrow = _excel_df_flat(t_cell="$B$8", y_last_row=y_last_row)
    ws["A13"] = "Borrow rate (for exp(r*dt))"
    if asm.borrowing_rate_mode == "scenario_linked":
        ws["B13"] = f"=IF($B$7=1,MAX(0,-LN(({df_borrow}))/$B$8+$B$9),$B$8)"
    else:
        ws["B13"] = "=$B$8"

    w_row0 = 16
    ws["A15"] = "Cash weight w0"
    ws["B15"] = float(w[0])
    ws["F14"] = "DF @ issue tenor (zero-coupon factor)"
    ws["F14"].font = Font(bold=True)
    for k in range(nb):
        ws.cell(row=w_row0 + k, column=1, value=f"w bond {k + 1}").font = Font(bold=True)
        ws.cell(row=w_row0 + k, column=2, value=float(w[k + 1]))
        ws.cell(row=w_row0 + k, column=4, value=f"Nominal tenor (y) {k + 1}").font = Font(bold=True)
        ws.cell(row=w_row0 + k, column=5, value=float(nom[k]))
        # One DF per bond — avoids inlining huge formulas into month-0 cash + maturity columns.
        ws.cell(
            row=w_row0 + k,
            column=6,
            value=f"={_excel_df_flat(t_cell=f'$E${w_row0 + k}', y_last_row=y_last_row)}",
        )

    last_param_row = w_row0 + nb - 1
    engine_section_row = last_param_row + 2
    engine_hdr_row = engine_section_row + 1
    engine_data_first_row = engine_hdr_row + 1

    # ---- Column layout (1-based) + header row directly under parameters (no fixed blank gap) ----
    c = 1
    col: dict[str, int | list[int]] = {}
    hdr_by_col: dict[int, str] = {}
    col_gloss: dict[int, str] = {}

    def take1(name: str, header: str, gloss: str) -> int:
        nonlocal c
        cc = c
        hdr_by_col[cc] = header
        col_gloss[cc] = gloss
        col[name] = cc
        c += 1
        return cc

    def taken(
        name: str,
        n: int,
        header_for_index: Callable[[int], str],
        gloss_for_index: Callable[[int], str],
    ) -> list[int]:
        nonlocal c
        cols = list(range(c, c + n))
        for j, cc in enumerate(cols):
            hdr_by_col[cc] = header_for_index(j)
            col_gloss[cc] = gloss_for_index(j)
        col[name] = cols
        c += n
        return cols

    take1(
        "mon",
        f"Month index (matches {sh}! column A)",
        "Calendar month index (1-based) for this engine row; must align with the liability cashflow sheet month column.",
    )
    if variable_period:
        take1(
            "m0",
            "Period start month",
            "First calendar month included in this row's cashflow aggregation (multi-month steps only).",
        )
        take1(
            "dt_y",
            "Period length (years)",
            "Length of the period in years used for interest accrual and tenor roll (sum of months / 12).",
        )
        take1(
            "cf",
            f"Liability outflows in period ($, {sh})",
            f"Sum of expected total cash outflows from **{sh}** column ExpTotalCF for all months in the period.",
        )
    else:
        take1(
            "cf",
            f"Liability outflow, $ ({sh} ExpTotalCF)",
            f"Expected monthly liability cash outflow from **{sh}** column ExpTotalCF (same as column S).",
        )
    take1(
        "d_acc",
        "Borrowing after accrual ($)",
        "Outstanding borrowing balance after accruing interest at the scenario or fixed borrow rate for Δt.",
    )
    t_pm = taken(
        "t_pm",
        nb,
        lambda i: f"Bond {i + 1}: tenor after roll (y)",
        lambda i: (
            f"Remaining maturity in years for bond bucket {i + 1} after rolling time forward by Δt "
            "(zero when the bond pays at month-end)."
        ),
    )
    mat = taken(
        "mat",
        nb,
        lambda i: f"Bond {i + 1}: maturity principal ($)",
        lambda i: f"Face value paid in cash when bucket {i + 1}'s tenor reaches zero this month.",
    )
    f_pm = taken(
        "f_pm",
        nb,
        lambda i: f"Bond {i + 1}: face after maturities ($)",
        lambda i: f"Face amount remaining after maturity payouts for bucket {i + 1}.",
    )
    take1(
        "cash_m",
        "Cash + maturity proceeds ($)",
        "Beginning-of-step cash (with growth rules) plus all principal returned from maturing bonds.",
    )
    take1(
        "rep1",
        "Debt repayment, pass 1 ($)",
        "Cash used to repay borrowing before reinvestment (minimum of cash and debt after accrual).",
    )
    take1(
        "cash_r1",
        "Cash after repayment 1 ($)",
        "Cash remaining after the first debt repayment leg.",
    )
    take1(
        "debt_r1",
        "Debt after repayment 1 ($)",
        "Borrowing balance after the first repayment leg.",
    )
    df_pm = taken(
        "df_pm",
        nb,
        lambda i: f"Bond {i + 1}: discount factor",
        lambda i: (
            f"Zero-coupon discount factor for bucket {i + 1} at its post-roll tenor using **YieldCurve** "
            "and **Inputs** spread (continuously compounded log-linear zeros)."
        ),
    )
    mv_pm = taken(
        "mv_pm",
        nb,
        lambda i: f"Bond {i + 1}: market value ($)",
        lambda i: f"Mark-to-market value of bucket {i + 1} (face after maturities × discount factor).",
    )
    take1(
        "aum_re",
        "Total assets before reinvest ($)",
        "Cash after repayment 1 plus sum of bond market values — AUM before target-weight reinvestment.",
    )
    take1(
        "xsr",
        "Cash surplus vs target ($)",
        "Actual cash minus policy target cash (cash weight × total AUM); drives reinvestment into bonds when positive.",
    )
    defc = taken(
        "defc",
        nb,
        lambda i: f"Bond {i + 1}: gap to target MV ($)",
        lambda i: (
            f"Dollar shortfall of bucket {i + 1} versus its policy weight times total AUM "
            "(zero if bond is overweight)."
        ),
    )
    take1(
        "dsum",
        "Sum of bond gaps ($)",
        "Total of all bond bucket gaps; used to split reinvestment when rebalancing toward targets.",
    )
    split = taken(
        "split",
        nb,
        lambda i: f"Bond {i + 1}: reinvest share",
        lambda i: f"Fraction of aggregate reinvestment allocated to bucket {i + 1} (pro-rata to gaps or weights).",
    )
    dmv = taken(
        "dmv",
        nb,
        lambda i: f"Bond {i + 1}: buy from cash MV ($)",
        lambda i: (
            f"Market-value dollars spent from cash to purchase bucket {i + 1} when underweight. "
            "Cash reduces by this amount; face purchased = dmv / DF (computed in f_re column)."
        ),
    )
    take1(
        "cash_re",
        "Cash after reinvestment ($)",
        "Cash after spending on reinvestment purchases (hold-cash rule skips purchases).",
    )
    f_re = taken(
        "f_re",
        nb,
        lambda i: f"Bond {i + 1}: face after reinvest ($)",
        lambda i: f"Face amount after reinvestment for bucket {i + 1}.",
    )
    t_re = taken(
        "t_re",
        nb,
        lambda i: f"Bond {i + 1}: tenor after reinvest (y)",
        lambda i: f"Tenor in years after reinvestment for bucket {i + 1} (new purchases at bucket nominal tenor if flat).",
    )
    take1(
        "cash_cf",
        "Cash net of liability outflow ($)",
        "Cash after reinvestment minus this row's liability cash outflow.",
    )
    take1(
        "need_raw",
        "Liquidity need before policy ($)",
        "Shortfall if liability outflow exceeds cash (before borrow-first or sell rules).",
    )
    take1(
        "cash_bb",
        "Cash after borrow-first ($)",
        "Cash if borrowing before selling is enabled: need may be funded by increasing debt first.",
    )
    take1(
        "debt_bb",
        "Debt after borrow-first ($)",
        "Borrowing balance after optional borrow-before-sell leg.",
    )
    take1(
        "need_dis",
        "Need met by selling bonds ($)",
        "Remaining liquidity need to be covered by disinvesting bonds (zero if borrow-first cleared it).",
    )
    # One DF @ post-reinvest tenor per bond; disinvest block references these cells
    # (inlining DF formulas hits Excel's 8192-char limit).
    dfd = taken(
        "dfd",
        nb,
        lambda i: f"Bond {i + 1}: DF for disinvest",
        lambda i: (
            f"Discount factor at post-reinvest tenor for bucket {i + 1}, used to price bond sales when raising cash."
        ),
    )

    n_dis = nb + 2  # pro_rata disinvest may need > nb peels
    dis_need: list[int] = []
    dis_cash: list[int] = []
    dis_face: list[list[int]] = []
    for di in range(n_dis):
        dis_need.append(
            take1(
                f"nd{di}",
                f"Disinvest pass {di + 1}: need left ($)",
                f"Remaining liquidity need after pass {di + 1} of bond sales (shortest-first or pro-rata by MV).",
            )
        )
        dis_cash.append(
            take1(
                f"cd{di}",
                f"Disinvest pass {di + 1}: cash ($)",
                f"Cash after pass {di + 1} of bond sales.",
            )
        )
        dis_face.append(
            taken(
                f"fd{di}",
                nb,
                lambda i, d=di: f"Pass {d + 1}: bond {i + 1} face ($)",
                lambda i, d=di: (
                    f"Face remaining in bucket {i + 1} after disinvestment pass {d + 1}; "
                    f"extra passes support pro-rata peels."
                ),
            )
        )

    take1(
        "cash_pd",
        "Cash after disinvest ($)",
        "Cash after all disinvestment passes and borrow-first effects.",
    )
    take1(
        "debt_pb",
        "Debt after borrow / sell ($)",
        "Borrowing after disinvestment and intermediate policy steps.",
    )
    take1(
        "need_b2",
        "Borrow need, pass 2 ($)",
        "Residual shortfall if cash is still negative after disinvest (second borrow leg when not borrow-first).",
    )
    take1(
        "cash_br2",
        "Cash before repay 2 ($)",
        "Cash after optional second borrowing.",
    )
    take1(
        "debt_br2",
        "Debt before repay 2 ($)",
        "Debt before final scheduled repayment.",
    )
    take1(
        "rep2",
        "Debt repayment, pass 2 ($)",
        "Second repayment of debt from available cash.",
    )
    take1(
        "cash_af2",
        "Cash end before carry ($)",
        "Cash after second repayment, before end-of-month carry.",
    )
    take1(
        "debt_af2",
        "Debt end before carry ($)",
        "Debt after second repayment.",
    )
    de = take1(
        "de_e",
        "Borrowing balance EOM ($)",
        "Closing borrowing balance for the month.",
    )
    ce = take1(
        "ca_e",
        "Cash balance EOM ($)",
        "Closing cash balance for the month.",
    )
    fe = taken(
        "fe",
        nb,
        lambda i: f"Bond {i + 1}: face EOM ($)",
        lambda i: f"Closing face amount for bucket {i + 1}.",
    )
    te = taken(
        "te",
        nb,
        lambda i: f"Bond {i + 1}: tenor EOM (y)",
        lambda i: f"Closing tenor in years for bucket {i + 1}.",
    )
    mv0 = take1(
        "mv0",
        "Cash for ALM_Projection ($)",
        "Closing cash — links to **ALM_Projection** bucket column for cash.",
    )
    mvb = taken(
        "mvb",
        nb,
        lambda i: f"Bond {i + 1}: MV EOM ($)",
        lambda i: f"Closing market value for bucket {i + 1} (face EOM × discount factor at tenor EOM).",
    )
    last_col = c - 1

    def C(name: str) -> int | list[int]:
        return col[name]

    field_guide_rows = tuple(
        (get_column_letter(cc), hdr_by_col[cc], col_gloss[cc]) for cc in range(1, last_col + 1)
    )

    ws.merge_cells(
        start_row=engine_section_row,
        start_column=1,
        end_row=engine_section_row,
        end_column=last_col,
    )
    sec = ws.cell(row=engine_section_row, column=1)
    sec.value = (
        "Data grid (read left to right): accrue borrowing → roll tenors / maturities → repay debt → "
        "mark bonds → rebalance toward targets → fund liability outflow → disinvest if needed → final borrow/repay → EOM. "
        "Bond 1…n = Treasury buckets. Use Excel outline (collapse ▶) to hide inner columns."
    )
    sec.font = Font(bold=True, size=11)
    sec.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[engine_section_row].height = 36

    for cc in range(1, last_col + 1):
        hcell = ws.cell(row=engine_hdr_row, column=cc, value=hdr_by_col.get(cc, ""))
        hcell.font = Font(bold=True, size=9)
        hcell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[engine_hdr_row].height = 44
    for cc in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 10.5
    tpm_cols = C("t_pm")
    if not isinstance(tpm_cols, list):
        raise RuntimeError("t_pm")
    last_fd = col[f"fd{n_dis - 1}"]
    if not isinstance(last_fd, list):
        raise RuntimeError("fd")
    o_lo, o_hi = int(tpm_cols[0]), int(last_fd[-1])
    for cc in range(o_lo, o_hi + 1):
        ws.column_dimensions[get_column_letter(cc)].outline_level = 1
    ws.freeze_panes = None

    R0 = engine_data_first_row
    WBASE = w_row0
    n_periods = len(period_end_months_1based)
    mon_c = int(C("mon"))
    cf_i = int(C("cf"))
    if variable_period:
        m0_c = int(C("m0"))
        dt_c = int(C("dt_y"))

    for i in range(n_periods):
        r = R0 + i
        rp = r - 1
        ws.cell(row=r, column=mon_c, value=int(period_end_months_1based[i]))
        if variable_period:
            ws.cell(row=r, column=m0_c, value=f"=IF(ROW()={R0},1,{_L(mon_c)}{rp}+1)")
            ws.cell(row=r, column=dt_c, value=f"=({_L(mon_c)}{r}-{_L(m0_c)}{r}+1)/12")
            ws.cell(
                row=r,
                column=cf_i,
                value=f'=SUM(INDIRECT("{sh}!S"&(3+{_L(m0_c)}{r})&":S"&(3+{_L(mon_c)}{r})))',
            )
        else:
            ws.cell(
                row=r,
                column=cf_i,
                value=f"=INDEX({sh}!$S:$S,3+{_L(mon_c)}{r})",
            )

        d_acc = int(C("d_acc"))
        de_c = int(C("de_e"))
        ca_c = int(C("ca_e"))
        fe = C("fe")
        te = C("te")
        if not isinstance(fe, list) or not isinstance(te, list):
            raise RuntimeError("bad fe/te")

        debt_p = f"IF(ROW()={R0},0,{_L(de_c)}{rp})"
        cash_p = f"IF(ROW()={R0},$B$15*$B$5,{_L(ca_c)}{rp})"
        face_p = [
            (
                f"$B${WBASE + k}*$B$5/$F${WBASE + k}"
                if i == 0
                else f"{_L(fe[k])}{rp}"
            )
            for k in range(nb)
        ]
        trem_p = [f"$E${WBASE + k}" if i == 0 else f"{_L(te[k])}{rp}" for k in range(nb)]

        dt_cell = f"{_L(dt_c)}{r}" if variable_period else "$B$6"
        ws.cell(row=r, column=d_acc, value=f"=({debt_p})*EXP($B$13*{dt_cell})")

        for k in range(nb):
            ws.cell(row=r, column=t_pm[k], value=f"=MAX(0,{trem_p[k]}-{dt_cell})")

        m_parts = []
        for k in range(nb):
            fk = face_p[k]
            Tk = f"{_L(t_pm[k])}{r}"
            m_parts.append(f"IF(AND({fk}>1E-9,{Tk}<=1E-9),{fk},0)")
        msum = "+".join(m_parts)

        for k in range(nb):
            fk = face_p[k]
            Tk = f"{_L(t_pm[k])}{r}"
            ws.cell(row=r, column=mat[k], value=f"=IF(AND({fk}>1E-9,{Tk}<=1E-9),{fk},0)")

        for k in range(nb):
            fk = face_p[k]
            Tk = f"{_L(t_pm[k])}{r}"
            ws.cell(row=r, column=f_pm[k], value=f"=IF(AND({fk}>1E-9,{Tk}<=1E-9),0,{fk})")

        cash_m = int(C("cash_m"))
        ws.cell(row=r, column=cash_m, value=f"=({cash_p})+({msum})")

        rep1 = int(C("rep1"))
        cash_r1 = int(C("cash_r1"))
        debt_r1 = int(C("debt_r1"))
        ws.cell(row=r, column=rep1, value=f"=MIN({_L(cash_m)}{r},{_L(d_acc)}{r})")
        ws.cell(row=r, column=cash_r1, value=f"={_L(cash_m)}{r}-{_L(rep1)}{r}")
        ws.cell(row=r, column=debt_r1, value=f"={_L(d_acc)}{r}-{_L(rep1)}{r}")

        # Align with Python ``run_alm_projection``: pro-rata reinvestment runs only when at least
        # one bond matures this month and borrowing is repaid — not every month excess cash > target.
        mtot = "+".join(f"{_L(mat[k])}{r}" for k in range(nb))

        for k in range(nb):
            ws.cell(
                row=r,
                column=df_pm[k],
                value=f"={_excel_df_flat(t_cell=f'{_L(t_pm[k])}{r}', y_last_row=y_last_row)}",
            )
        for k in range(nb):
            ws.cell(row=r, column=mv_pm[k], value=f"={_L(f_pm[k])}{r}*{_L(df_pm[k])}{r}")

        aum_re = int(C("aum_re"))
        mv_sum = "+".join(f"{_L(mv_pm[k])}{r}" for k in range(nb))
        ws.cell(row=r, column=aum_re, value=f"={_L(cash_r1)}{r}+{mv_sum}")

        xsr = int(C("xsr"))
        ws.cell(row=r, column=xsr, value=f"={_L(cash_r1)}{r}-$B$15*{_L(aum_re)}{r}")

        for k in range(nb):
            ws.cell(
                row=r,
                column=defc[k],
                value=f"=MAX(0,$B${WBASE + k}*{_L(aum_re)}{r}-{_L(mv_pm[k])}{r})",
            )
        dsum_i = int(C("dsum"))
        ds = "+".join(f"{_L(defc[j])}{r}" for j in range(nb))
        ws.cell(row=r, column=dsum_i, value=f"={ds}")

        for k in range(nb):
            wnk = wnorm[k]
            ws.cell(
                row=r,
                column=split[k],
                value=f"=IF({_L(dsum_i)}{r}>1E-9,{_L(defc[k])}{r}/{_L(dsum_i)}{r},{wnk})",
            )

        for k in range(nb):
            wnk = wnorm[k]
            # dmv[k] is the MV dollar amount to purchase for bucket k.
            # cash_re subtracts sum(dmv) and f_re adds dmv/DF to convert MV→face.
            # Do NOT divide by DF here — that conversion happens only in f_re.
            ws.cell(
                row=r,
                column=dmv[k],
                value=(
                    f"=IF(OR($B$11=0,({mtot})<=1E-9,{_L(debt_r1)}{r}>1E-9,{_L(xsr)}{r}<=1E-6),0,"
                    f"{_L(xsr)}{r}*IF({_L(dsum_i)}{r}>1E-9,{_L(defc[k])}{r}/{_L(dsum_i)}{r},{wnk}))"
                ),
            )

        cash_re = int(C("cash_re"))
        dmv_sum = "+".join(f"{_L(dmv[j])}{r}" for j in range(nb))
        ws.cell(
            row=r,
            column=cash_re,
            value=(
                f"=IF(AND($B$11=1,({mtot})>1E-9,{_L(debt_r1)}{r}<=1E-9),"
                f"{_L(cash_r1)}{r}-({dmv_sum}),{_L(cash_r1)}{r})"
            ),
        )

        for k in range(nb):
            t_use = f"IF({_L(t_pm[k])}{r}>1E-9,{_L(t_pm[k])}{r},$E${WBASE + k})"
            denom = _excel_df_flat(t_cell=f"({t_use})", y_last_row=y_last_row)
            ws.cell(
                row=r,
                column=f_re[k],
                value=(
                    f"=IF(AND($B$11=1,({mtot})>1E-9,{_L(debt_r1)}{r}<=1E-9),"
                    f"{_L(f_pm[k])}{r}+IF({_L(dmv[k])}{r}<=0,0,{_L(dmv[k])}{r}/"
                    f"MAX(({denom}),1E-15)),{_L(f_pm[k])}{r})"
                ),
            )
        for k in range(nb):
            ws.cell(
                row=r,
                column=t_re[k],
                value=(
                    f"=IF(AND($B$11=1,({mtot})>1E-9,{_L(debt_r1)}{r}<=1E-9),"
                    f"IF(AND({_L(f_pm[k])}{r}<=1E-9,{_L(t_pm[k])}{r}<=1E-9),$E${WBASE + k},{_L(t_pm[k])}{r}),"
                    f"{_L(t_pm[k])}{r})"
                ),
            )

        dfd_cols = C("dfd")
        if not isinstance(dfd_cols, list):
            raise RuntimeError("dfd")
        for k in range(nb):
            ws.cell(
                row=r,
                column=dfd_cols[k],
                value=f"={_excel_df_flat(t_cell=f'{_L(t_re[k])}{r}', y_last_row=y_last_row)}",
            )

        ccf = int(C("cash_cf"))
        ws.cell(row=r, column=ccf, value=f"={_L(cash_re)}{r}-{_L(cf_i)}{r}")

        need_raw = int(C("need_raw"))
        ws.cell(row=r, column=need_raw, value=f"=MAX(0,-{_L(ccf)}{r})")

        c_bb = int(C("cash_bb"))
        d_bb = int(C("debt_bb"))
        need_dis = int(C("need_dis"))
        ws.cell(
            row=r,
            column=c_bb,
            value=f"=IF(AND($B$10=1,{_L(need_raw)}{r}>0),{_L(ccf)}{r}+{_L(need_raw)}{r},{_L(ccf)}{r})",
        )
        ws.cell(
            row=r,
            column=d_bb,
            value=f"=IF(AND($B$10=1,{_L(need_raw)}{r}>0),{_L(debt_r1)}{r}+{_L(need_raw)}{r},{_L(debt_r1)}{r})",
        )
        ws.cell(row=r, column=need_dis, value=f"=IF($B$10=1,0,{_L(need_raw)}{r})")

        tref = [f"{_L(t_re[k])}{r}" for k in range(nb)]
        df_dis = [f"{_L(dfd_cols[k])}{r}" for k in range(nb)]

        for ir in range(n_dis):
            nc, cc, fk_list = dis_need[ir], dis_cash[ir], dis_face[ir]
            if ir == 0:
                n0 = f"{_L(need_dis)}{r}"
                c0 = f"{_L(c_bb)}{r}"
                f0 = [f"{_L(f_re[k])}{r}" for k in range(nb)]
            else:
                pnc, pcc = dis_need[ir - 1], dis_cash[ir - 1]
                pfk = dis_face[ir - 1]
                n0 = f"{_L(pnc)}{r}"
                c0 = f"{_L(pcc)}{r}"
                f0 = [f"{_L(pfk[k])}{r}" for k in range(nb)]

            trnk = [f"({tref[k]}+{(k + 1)}*1E-9)" for k in range(nb)]
            tmin = f"MIN({','.join(f'IF({f0[k]}>1E-9,{trnk[k]},999)' for k in range(nb))})"
            mv_tot = "+".join(f"MAX(0,{f0[k]})*({df_dis[k]})" for k in range(nb))

            sells: list[str] = []
            for k in range(nb):
                sf = (
                    f"IF(AND($B$12=1,{f0[k]}>1E-9,ABS({trnk[k]}-({tmin}))<5E-10),"
                    f"MIN({f0[k]},{n0}/MAX(({df_dis[k]}),1E-15)),0)"
                )
                pr = (
                    f"IF(AND($B$12=0,{mv_tot}>1E-9,{f0[k]}>1E-9),"
                    f"MIN({f0[k]},{n0}*(MAX(0,{f0[k]})*({df_dis[k]}))/MAX(({mv_tot}),1E-15)),0)"
                )
                sells.append(f"IF($B$12=1,{sf},{pr})")

            pay = "+".join(f"({sells[k]})*({df_dis[k]})" for k in range(nb))
            ws.cell(row=r, column=nc, value=f"=MAX(0,{n0}-({pay}))")
            ws.cell(row=r, column=cc, value=f"={c0}+({pay})")
            for k in range(nb):
                ws.cell(row=r, column=fk_list[k], value=f"=MAX(0,{f0[k]}-({sells[k]}))")

        last_fk = dis_face[-1]
        last_cc = dis_cash[-1]
        c_pd = int(C("cash_pd"))
        d_pb = int(C("debt_pb"))
        ws.cell(row=r, column=c_pd, value=f"={_L(last_cc)}{r}")
        ws.cell(row=r, column=d_pb, value=f"={_L(d_bb)}{r}")

        need_b2 = int(C("need_b2"))
        ws.cell(row=r, column=need_b2, value=f"=MAX(0,-{_L(c_pd)}{r})")

        c_br2 = int(C("cash_br2"))
        d_br2 = int(C("debt_br2"))
        ws.cell(
            row=r,
            column=c_br2,
            value=f"=IF(AND($B$10=0,{_L(need_b2)}{r}>0),{_L(c_pd)}{r}+{_L(need_b2)}{r},{_L(c_pd)}{r})",
        )
        ws.cell(
            row=r,
            column=d_br2,
            value=f"=IF(AND($B$10=0,{_L(need_b2)}{r}>0),{_L(d_pb)}{r}+{_L(need_b2)}{r},{_L(d_pb)}{r})",
        )

        rep2 = int(C("rep2"))
        cash_af2 = int(C("cash_af2"))
        debt_af2 = int(C("debt_af2"))
        ws.cell(row=r, column=rep2, value=f"=MIN({_L(c_br2)}{r},{_L(d_br2)}{r})")
        ws.cell(row=r, column=cash_af2, value=f"={_L(c_br2)}{r}-{_L(rep2)}{r}")
        ws.cell(row=r, column=debt_af2, value=f"={_L(d_br2)}{r}-{_L(rep2)}{r}")

        for k in range(nb):
            ws.cell(row=r, column=fe[k], value=f"={_L(last_fk[k])}{r}")
        ws.cell(row=r, column=de_c, value=f"={_L(debt_af2)}{r}")
        ws.cell(row=r, column=ca_c, value=f"={_L(cash_af2)}{r}")
        for k in range(nb):
            ws.cell(row=r, column=te[k], value=f"={tref[k]}")
        mv0 = int(C("mv0"))
        mvb = C("mvb")
        if not isinstance(mvb, list):
            raise RuntimeError("mvb")
        ws.cell(row=r, column=mv0, value=f"={_L(ca_c)}{r}")
        for k in range(nb):
            ws.cell(
                row=r,
                column=mvb[k],
                value=f"={_L(fe[k])}{r}*({_excel_df_flat(t_cell=f'{_L(te[k])}{r}', y_last_row=y_last_row)})",
            )

    last_r = R0 + n_periods - 1
    mvb_l = C("mvb")
    if not isinstance(mvb_l, list):
        raise RuntimeError("mvb")
    return ALMEngineLayout(
        first_data_row=R0,
        last_data_row=last_r,
        col_mv_cash=int(C("mv0")),
        col_mv_bond_start=int(mvb_l[0]),
        col_debt_eom=int(C("de_e")),
        n_bonds=nb,
        field_guide_rows=field_guide_rows,
    )
