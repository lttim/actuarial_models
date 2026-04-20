"""Build a portfolio workbook (per-policy CF grid + liability rollups + ModelCheck)."""

from __future__ import annotations

import io
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from excel_workbook_validator import validate_workbook_or_raise
from liability_dispatch import liability_path_for
from portfolio import PortfolioResult, ProductTypeRollupScalars
from product_registry import ProductType


def _sorted_product_types(rollups: dict[ProductType, object]) -> tuple[ProductType, ...]:
    return tuple(sorted(rollups, key=lambda p: p.value))


def _policy_cf_row(pricing: object, n_months: int) -> list[float]:
    lp = liability_path_for(pricing)
    cf = np.asarray(lp.expected_total_cashflows, dtype=float).ravel()
    out = [float(cf[i]) if i < cf.size else 0.0 for i in range(n_months)]
    return out


def _policy_cf_rows(policy_results: list[object], n_months: int) -> list[list[float]]:
    """Precompute one liability cashflow vector per policy."""
    rows: list[list[float]] = []
    for pr in policy_results:
        pricing = getattr(pr, "pricing")
        rows.append(_policy_cf_row(pricing, n_months))
    return rows


def build_portfolio_workbook_bytes(res: PortfolioResult) -> bytes:
    """Create an .xlsx with ``PolicyCashflows`` grid, formula-linked ``LiabilityAggregate`` total, and ``ModelCheck``."""
    wb = Workbook()
    # --- Inputs ---
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in["A1"] = "Portfolio workbook (v1)"
    ws_in["A1"].font = Font(bold=True, size=12)
    ws_in["A2"] = (
        "Per-policy liability CF are Python literals on PolicyCashflows; "
        "LiabilityAggregate total_cf sums those columns; ModelCheck reconciles Excel to Python."
    )

    # --- PolicyRegister ---
    ws_pr = wb.create_sheet("PolicyRegister")
    headers = ("policy_id", "product_type", "single_premium")
    for c, h in enumerate(headers, start=1):
        ws_pr.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, pr in enumerate(res.policy_results, start=2):
        ws_pr.cell(row=r, column=1, value=pr.policy_id)
        ws_pr.cell(row=r, column=2, value=pr.product_type.value)
        prem = getattr(pr.pricing, "single_premium", None)
        ws_pr.cell(row=r, column=3, value=float(prem) if prem is not None else None)

    total = res.liability_path_total
    n = len(total.expected_total_cashflows)
    n_pol = len(res.policy_results)
    first_pc = 3
    last_pc = 2 + max(1, n_pol)
    first_letter = get_column_letter(first_pc)
    last_letter = get_column_letter(last_pc)

    # --- PolicyCashflows (month, t_years, one column per policy) ---
    ws_pc = wb.create_sheet("PolicyCashflows")
    ws_pc["A1"] = "month"
    ws_pc["B1"] = "t_years"
    ws_pc["A1"].font = Font(bold=True)
    ws_pc["B1"].font = Font(bold=True)
    for j, pr in enumerate(res.policy_results, start=first_pc):
        hdr = str(pr.policy_id).replace("[", "").replace("]", "").replace("*", "").replace("?", "")[:200]
        ws_pc.cell(row=1, column=j, value=hdr or f"policy_{j}").font = Font(bold=True)
    policy_cf_rows = _policy_cf_rows(list(res.policy_results), n)
    for i in range(n):
        rr = 2 + i
        ws_pc.cell(row=rr, column=1, value=i + 1)
        ws_pc.cell(row=rr, column=2, value=float(total.times_years[i]))
        for j, series in enumerate(policy_cf_rows, start=first_pc):
            ws_pc.cell(row=rr, column=j, value=float(series[i]))

    # --- ProductTypeRollups ---
    ws_rt = wb.create_sheet("ProductTypeRollups")
    ws_rt["A1"] = "product_type"
    ws_rt["B1"] = "policy_count"
    ws_rt["C1"] = "sum_single_premium"
    ws_rt["D1"] = "sum_undiscounted_cf"
    for c in range(1, 5):
        ws_rt.cell(row=1, column=c).font = Font(bold=True)
    r = 2
    for pt in _sorted_product_types(dict(res.rollups_by_product_type)):
        scal: ProductTypeRollupScalars = res.product_type_scalar_rollups[pt]
        path = res.rollups_by_product_type[pt]
        sum_cf = float(path.expected_total_cashflows.sum()) if len(path.expected_total_cashflows) else 0.0
        ws_rt.cell(row=r, column=1, value=pt.value)
        ws_rt.cell(row=r, column=2, value=scal.policy_count)
        ws_rt.cell(
            row=r,
            column=3,
            value=scal.sum_single_premium if scal.sum_single_premium is not None else None,
        )
        ws_rt.cell(
            row=r,
            column=4,
            value=scal.sum_undiscounted_cashflows if scal.sum_undiscounted_cashflows is not None else sum_cf,
        )
        r += 1

    # --- LiabilityAggregate ---
    ws_la = wb.create_sheet("LiabilityAggregate")
    types = _sorted_product_types(dict(res.rollups_by_product_type))
    ws_la["A1"] = "month"
    ws_la["B1"] = "t_years"
    ws_la["C1"] = "total_cf"
    for j, pt in enumerate(types, start=4):
        ws_la.cell(row=1, column=j, value=f"cf_{pt.value}").font = Font(bold=True)
    for c in range(1, 4):
        ws_la.cell(row=1, column=c).font = Font(bold=True)

    last_type_col = 3 + len(types)
    last_type_letter = get_column_letter(last_type_col)

    for i in range(n):
        rr = 2 + i
        ws_la.cell(row=rr, column=1, value=i + 1)
        ws_la.cell(row=rr, column=2, value=float(total.times_years[i]))
        if n_pol >= 1:
            ws_la.cell(
                row=rr,
                column=3,
                value=f"=SUM(PolicyCashflows!{first_letter}{rr}:{last_letter}{rr})",
            )
        else:
            ws_la.cell(row=rr, column=3, value=0.0)
        for j, pt in enumerate(types, start=4):
            cf_j = res.rollups_by_product_type[pt].expected_total_cashflows
            v = float(cf_j[i]) if i < len(cf_j) else 0.0
            ws_la.cell(row=rr, column=j, value=v)

    # --- ModelCheck ---
    ws_mc = wb.create_sheet("ModelCheck")
    ws_mc["A1"] = "month"
    ws_mc["B1"] = "rollup_minus_total"
    ws_mc["C1"] = "python_total_cf"
    ws_mc["D1"] = "excel_total_cf"
    ws_mc["E1"] = "diff_excel_minus_python"
    for c in range(1, 6):
        ws_mc.cell(row=1, column=c).font = Font(bold=True)
    for i in range(n):
        rr = 2 + i
        ws_mc.cell(row=rr, column=1, value=i + 1)
        first_t = get_column_letter(4)
        ws_mc.cell(
            row=rr,
            column=2,
            value=f"=SUM(LiabilityAggregate!{first_t}{rr}:{last_type_letter}{rr})-LiabilityAggregate!C{rr}",
        )
        py_tot = float(total.expected_total_cashflows[i])
        ws_mc.cell(row=rr, column=3, value=py_tot)
        ws_mc.cell(row=rr, column=4, value=f"=LiabilityAggregate!C{rr}")
        ws_mc.cell(row=rr, column=5, value=f"=D{rr}-C{rr}")

    # --- README ---
    ws_rm = wb.create_sheet("README")
    ws_rm["A1"] = "Portfolio v1 workbook"
    ws_rm["A2"] = "Per-policy pricing is seriatim in Python; total_cf on LiabilityAggregate sums PolicyCashflows."
    ws_rm["A3"] = "ALM is not embedded here (v1); run ALM from the app or CLI on the aggregate path."

    validate_workbook_or_raise(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_portfolio_workbook_to_path(res: PortfolioResult, out_path: str | Path) -> Path:
    p = Path(out_path)
    p.write_bytes(build_portfolio_workbook_bytes(res))
    return p


__all__ = ["build_portfolio_workbook_bytes", "build_portfolio_workbook_to_path"]
